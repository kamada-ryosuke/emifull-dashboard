# -*- coding: utf-8 -*-
"""出力Excel(.xlsx)に行を追記する。月次シート(YYMM【デビット】)を自動作成し日付順にソート。

行スキーマ(既存.xlsmと同じA-K列):
  A: 日付            B: 借方勘定科目   C: 借方税区分    D: 借方品目     E: 借方部門
  F: 借方金額        G: 貸方勘定科目   H: 貸方品目      I: 貸方部門     J: 貸方金額    K: 摘要
"""
from __future__ import annotations
import json
from datetime import date, datetime
from pathlib import Path
from typing import Iterable

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

HEADERS = [
    "日付", "借方勘定科目", "借方税区分", "借方品目", "借方部門", "借方金額",
    "貸方勘定科目", "貸方品目", "貸方部門", "貸方金額", "摘要", "検証",
]
COL_WIDTHS = {1: 12, 2: 14, 3: 18, 4: 14, 5: 24, 6: 12, 7: 16, 8: 14, 9: 14, 10: 12, 11: 60, 12: 30}


def _sheet_name_for_date(d: date) -> str:
    """2026-05-01 → '2605【デビット】'"""
    yy = f"{d.year % 100:02d}"
    mm = f"{d.month:02d}"
    return f"{yy}{mm}【デビット】"


def _ensure_workbook(path: Path) -> Workbook:
    if path.exists():
        return openpyxl.load_workbook(path)
    wb = Workbook()
    # デフォルトの空シートは削除（最初の月次シート作成時にちゃんと作る）
    default = wb.active
    wb.remove(default)
    return wb


def _ensure_sheet(wb: Workbook, name: str):
    if name in wb.sheetnames:
        return wb[name]
    ws = wb.create_sheet(name)
    # ヘッダー行
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="305496")
    for i, h in enumerate(HEADERS, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
    for col, w in COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A2"
    return ws


def _row_key(row: tuple) -> tuple:
    """ソートキー: (日付, 借方部門, 摘要)"""
    d = row[0]
    if isinstance(d, datetime):
        d = d.date()
    if d is None:
        d = date.min
    return (d, row[4] or "", row[10] or "")


def append_rows(excel_path: Path, rows: list[dict], credit_account: str = "関西みらい銀行") -> dict:
    """rows を該当月のシートに追記し、日付順にソート。

    rows の各要素キー:
      date(date|datetime|str), account, tax, item(任意), facility_code, amount, summary
    """
    excel_path.parent.mkdir(parents=True, exist_ok=True)
    wb = _ensure_workbook(excel_path)

    # 月別にグルーピング
    by_month: dict[str, list[tuple]] = {}
    for r in rows:
        d = r["date"]
        if isinstance(d, str):
            d = datetime.fromisoformat(d).date()
        elif isinstance(d, datetime):
            d = d.date()
        sheet_name = _sheet_name_for_date(d)
        amount = int(r["amount"])
        # 検証列の文字列(needs_reviewでなければ ✓ チェック済)
        nr = r.get("needs_review")
        if nr:
            note = r.get("notes") or "要確認"
            verify_text = f"⚠ {note}"[:200]
        else:
            verify_text = "✓ 3重チェック OK"
        tup = (
            d,
            r["account"],
            r.get("tax", ""),
            r.get("item", "") or "",
            r["facility_code"],
            amount,
            credit_account,
            "",
            "",
            amount,
            r["summary"],
            verify_text,
        )
        by_month.setdefault(sheet_name, []).append(tup)

    appended_per_sheet: dict[str, int] = {}
    for sheet_name, new_rows in by_month.items():
        ws = _ensure_sheet(wb, sheet_name)
        # 既存行を読み出してマージ（重複してもキーで判別はしない: 呼び出し側で重複チェック済みの想定）
        existing: list[tuple] = []
        for r in range(2, ws.max_row + 1):
            vals = tuple(ws.cell(r, c).value for c in range(1, len(HEADERS) + 1))
            if any(v is not None for v in vals):
                existing.append(vals)
        merged = existing + new_rows
        merged.sort(key=_row_key)

        # 既存データを全消去して書き直し
        if ws.max_row >= 2:
            ws.delete_rows(2, ws.max_row - 1)
        warn_fill = PatternFill("solid", fgColor="FFF2CC")  # 薄い黄色
        for ri, row in enumerate(merged, start=2):
            verify_cell_val = row[11] if len(row) >= 12 else ""
            is_warn = isinstance(verify_cell_val, str) and verify_cell_val.startswith("⚠")
            for ci, val in enumerate(row, start=1):
                cell = ws.cell(row=ri, column=ci, value=val)
                if ci == 1:
                    cell.number_format = "yyyy/mm/dd"
                elif ci in (6, 10):
                    cell.number_format = "#,##0"
                if is_warn:
                    cell.fill = warn_fill
        appended_per_sheet[sheet_name] = len(new_rows)

    wb.save(excel_path)
    return appended_per_sheet


if __name__ == "__main__":
    import sys
    sys.stdout.reconfigure(encoding="utf-8")
    test_path = Path(__file__).resolve().parent.parent / "output" / "_test.xlsx"
    rows = [
        {"date": "2026-05-01", "account": "食材費", "tax": "課対仕入8%（軽）",
         "facility_code": "5.2.UMIEてんり", "amount": 8644, "summary": "イオンビッグ㈱　おやつ代"},
        {"date": "2026-04-15", "account": "消耗品費", "tax": "課対仕入10%",
         "facility_code": "5.1.SORATOてんり", "amount": 1980, "summary": "ダイソー　支援用　教材費"},
    ]
    result = append_rows(test_path, rows)
    print(f"Test wrote: {result} → {test_path}")
    test_path.unlink(missing_ok=True)
