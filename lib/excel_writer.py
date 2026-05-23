"""デビットExcel追記モジュール

役割:
  - レシートOCR結果をデビットExcel (`{法人}_デビット.xlsx`) の該当シートに1行追加
  - シート名は `YYMM【デビット】` (例: 2605【デビット】)
  - 該当YYMMシートが無ければ既存シートのヘッダをコピーして新規作成

主要API:
  resolve_debit_xlsx(corporation, gdrive_dir) -> Path | None
  append_debit_row(xlsx_path, row, sheet_yyyymm=None, dry_run=False) -> dict
"""
from __future__ import annotations

from copy import copy
from datetime import datetime, date
from pathlib import Path
from typing import Any

import openpyxl


# 行構造 (debit_parser.py の列定義と一致)
# A〜K の11列のみ書き込む。L列(検証/3重チェック)は既存数式や手動入力を保護するため触らない。
COLUMN_ORDER = [
    "date",            # A 日付
    "debit_account",   # B 借方勘定科目
    "tax_class",       # C 借方税区分
    "debit_item",      # D 借方品目
    "department",      # E 借方部門
    "amount",          # F 借方金額
    "credit_account",  # G 貸方勘定科目
    "credit_item",     # H 貸方品目
    "credit_dept",     # I 貸方部門
    "credit_amount",   # J 貸方金額
    "description",     # K 摘要
]

# A列(日付)の表示書式 (yyyy/mm/dd ゼロ埋め)
DATE_NUMBER_FORMAT = "yyyy/mm/dd"

DEFAULT_CREDIT_ACCOUNT = "関西みらい銀行"


def _detect_corp_from_filename(name: str) -> str:
    if not name:
        return ""
    n = name.lower()
    if "npo" in n:
        return "NPO法人"
    if "医療法人" in name or "医療" in name:
        return "医療法人"
    return ""


def resolve_debit_xlsx(corporation: str, gdrive_dir: Path) -> Path | None:
    """`G:\\マイドライブ\\…\\01.デビッドカード清算\\` 直下から
    法人ラベルに合致する `*デビット*.xlsx` を返す。"""
    if not gdrive_dir.exists():
        return None
    for p in gdrive_dir.iterdir():
        if not p.is_file() or p.suffix.lower() != ".xlsx":
            continue
        if "デビット" not in p.name and "デビッド" not in p.name:
            continue
        if _detect_corp_from_filename(p.name) == corporation:
            return p
    return None


def _ymd_to_sheet_name(ym: str) -> str:
    """'2026-05' → '2605【デビット】'"""
    y, m = ym.split("-")
    return f"{y[2:]}{m}【デビット】"


def _find_yyyymm_sheet(wb: openpyxl.Workbook, ym: str) -> str | None:
    """ '2026-05' に対応する既存シート名を返す (例: '2605【デビット】')。"""
    yy = ym.split("-")[0][2:]
    mm = ym.split("-")[1]
    target = f"{yy}{mm}"
    for sn in wb.sheetnames:
        if target in sn:
            return sn
    return None


def _copy_sheet_format(
    wb: openpyxl.Workbook, src_name: str, new_name: str
) -> openpyxl.worksheet.worksheet.Worksheet:
    """既存シートのヘッダ行(R1) と列幅をコピーした新規シートを作る。"""
    src = wb[src_name]
    ws = wb.create_sheet(title=new_name)

    # 列幅
    for col_letter, dim in src.column_dimensions.items():
        if dim.width:
            ws.column_dimensions[col_letter].width = dim.width

    # ヘッダ (R1)
    for cell in src[1]:
        new_cell = ws.cell(row=1, column=cell.column, value=cell.value)
        if cell.has_style:
            new_cell.font = copy(cell.font)
            new_cell.fill = copy(cell.fill)
            new_cell.alignment = copy(cell.alignment)
            new_cell.border = copy(cell.border)
            new_cell.number_format = cell.number_format
    return ws


def _next_empty_row(ws) -> int:
    """A列を見て最後にデータがある行 + 1 を返す (ヘッダR1は必ず存在前提)。"""
    last = 1
    for row_idx in range(ws.max_row, 0, -1):
        if ws.cell(row=row_idx, column=1).value not in (None, ""):
            last = row_idx
            break
    return last + 1


def _to_excel_date(v: Any):
    """'2026-05-07' / datetime / date を datetime にして返す。"""
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v
    if isinstance(v, date):
        return datetime(v.year, v.month, v.day)
    s = str(v).strip()
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def append_debit_row(
    xlsx_path: Path,
    row: dict,
    sheet_yyyymm: str | None = None,
    dry_run: bool = False,
) -> dict:
    """xlsx の指定YYMMシートに1行追加。

    Args:
        xlsx_path: 対象 xlsx
        row: 入力辞書。最低限 {date, debit_account, tax_class, department,
             amount} があれば良い。credit_account/credit_amount は省略時に補完。
        sheet_yyyymm: 'YYYY-MM'。省略時は row['date'] から推定。
        dry_run: True なら保存しない (検証用)。

    Returns:
        {'ok': True, 'sheet': '2605【デビット】', 'row': 7, 'created_sheet': False}
        失敗時 {'ok': False, 'error': '...'}
    """
    if not xlsx_path.exists():
        return {"ok": False, "error": f"ファイル不在: {xlsx_path}"}

    # 月特定
    if not sheet_yyyymm:
        d = _to_excel_date(row.get("date"))
        if d is None:
            return {"ok": False, "error": "日付が空または不正です"}
        sheet_yyyymm = d.strftime("%Y-%m")

    # 入力検証
    amount = row.get("amount")
    if amount is None or (isinstance(amount, str) and not amount.strip()):
        return {"ok": False, "error": "金額が空です"}
    try:
        amount_int = int(amount)
    except (ValueError, TypeError):
        return {"ok": False, "error": f"金額が数値ではありません: {amount}"}
    if amount_int <= 0:
        return {"ok": False, "error": "金額は1以上である必要があります"}

    department = (row.get("department") or "").strip()
    if not department:
        return {"ok": False, "error": "部門(借方部門)が空です"}

    debit_account = (row.get("debit_account") or "").strip()
    if not debit_account:
        return {"ok": False, "error": "借方勘定科目が空です"}

    # ファイルロック検査 (Excel で開いていると保存できない)
    try:
        with open(xlsx_path, "rb"):
            pass
    except PermissionError:
        return {
            "ok": False,
            "error": f"{xlsx_path.name} が他で開かれています。閉じてから再実行してください。",
        }

    try:
        wb = openpyxl.load_workbook(xlsx_path)
    except Exception as e:
        return {"ok": False, "error": f"xlsx読込失敗: {e}"}

    # シート特定 / 必要なら作成
    sheet_name = _find_yyyymm_sheet(wb, sheet_yyyymm)
    created_sheet = False
    if sheet_name is None:
        target_name = _ymd_to_sheet_name(sheet_yyyymm)
        # ヘッダコピー元: 一番左の【デビット】シート
        src = next(
            (sn for sn in wb.sheetnames if "デビット" in sn),
            wb.sheetnames[0] if wb.sheetnames else None,
        )
        if src is None:
            return {"ok": False, "error": "ヘッダのコピー元シートが見つかりません"}
        _copy_sheet_format(wb, src, target_name)
        sheet_name = target_name
        created_sheet = True

    ws = wb[sheet_name]

    # 重複検出: 同 (日付, 部門, 金額) の行が既にあればスキップ
    excel_date_check = _to_excel_date(row.get("date"))
    if excel_date_check is not None:
        for row_idx in range(2, ws.max_row + 1):
            v_date = ws.cell(row=row_idx, column=1).value
            v_dept = ws.cell(row=row_idx, column=5).value
            v_amt = ws.cell(row=row_idx, column=6).value
            try:
                v_amt_int = int(v_amt) if v_amt is not None else None
            except (ValueError, TypeError):
                v_amt_int = None
            same_date = (
                isinstance(v_date, datetime)
                and v_date.date() == excel_date_check.date()
            )
            if (
                same_date
                and (v_dept or "").strip() == department
                and v_amt_int == amount_int
            ):
                return {
                    "ok": False,
                    "duplicate": True,
                    "error": (
                        f"同じ日付/部門/金額の行が既に存在 "
                        f"(R{row_idx}: {department} ¥{amount_int:,})"
                    ),
                    "sheet": sheet_name,
                    "row": row_idx,
                }

    next_row = _next_empty_row(ws)

    # 行データ生成
    excel_date = _to_excel_date(row.get("date"))
    description = (row.get("description") or "").strip()
    items_text = (row.get("items_text") or "").strip()
    if items_text and items_text not in description:
        # 摘要を組み立て: '{vendor}　{purpose} / {items}'
        vendor = (row.get("vendor") or "").strip()
        purpose = (row.get("purpose") or "").strip()
        head = "　".join([s for s in (vendor, purpose) if s])
        if head:
            description = f"{head} / {items_text}"
        else:
            description = items_text

    values = {
        "date": excel_date,
        "debit_account": debit_account,
        "tax_class": (row.get("tax_class") or "").strip(),
        "debit_item": (row.get("debit_item") or "").strip(),
        "department": department,
        "amount": amount_int,
        "credit_account": (row.get("credit_account") or DEFAULT_CREDIT_ACCOUNT).strip(),
        "credit_item": (row.get("credit_item") or "").strip(),
        "credit_dept": (row.get("credit_dept") or "").strip(),
        "credit_amount": int(row.get("credit_amount") or amount_int),
        "description": description,
    }

    # スタイル継承: 直前の行 (R{next_row-1}) の書式を新行にコピー
    # L列(検証)は触らないので A〜K のみループ
    style_src_row = next_row - 1 if next_row > 2 else None
    for col_idx, key in enumerate(COLUMN_ORDER, start=1):
        cell = ws.cell(row=next_row, column=col_idx, value=values[key])
        if style_src_row:
            src_cell = ws.cell(row=style_src_row, column=col_idx)
            if src_cell.has_style:
                cell.font = copy(src_cell.font)
                cell.fill = copy(src_cell.fill)
                cell.alignment = copy(src_cell.alignment)
                cell.border = copy(src_cell.border)
                cell.number_format = src_cell.number_format

    # 日付セルは yyyy/mm/dd 固定 (既存行と同じゼロ埋め書式)
    if excel_date is not None:
        ws.cell(row=next_row, column=1).number_format = DATE_NUMBER_FORMAT

    # 既存行の日付書式も統一 (ユーザ手入力で yyyy/m/d 等が混在するケース対策)
    _normalize_date_format_in_sheet(ws)

    # 日付降順ソート (A〜K のみ、L列は不変)。
    # 新規行は次の行に追記済みなので、ソート後に正しい位置へ移動する。
    sort_sheet_by_date_desc(ws)
    final_row = _locate_row_by_content(
        ws, excel_date, department, amount_int,
    ) or next_row

    if dry_run:
        return {
            "ok": True,
            "sheet": sheet_name,
            "row": final_row,
            "created_sheet": created_sheet,
            "dry_run": True,
        }

    try:
        wb.save(xlsx_path)
    except PermissionError:
        return {
            "ok": False,
            "error": f"{xlsx_path.name} が他で開かれています。閉じてから再実行してください。",
        }
    except Exception as e:
        return {"ok": False, "error": f"xlsx保存失敗: {e}"}

    # ---- 書込検証: 保存後に再読込して一致確認 ------------------
    verify = _verify_written_row(
        xlsx_path, sheet_name, final_row, values, excel_date,
    )
    if not verify["ok"]:
        return {
            "ok": False,
            "error": f"書込検証失敗: {verify['error']}",
            "sheet": sheet_name,
            "row": final_row,
            "path": str(xlsx_path),
        }

    return {
        "ok": True,
        "sheet": sheet_name,
        "row": final_row,
        "created_sheet": created_sheet,
        "path": str(xlsx_path),
        "verified": verify["values"],
    }


def update_debit_row(
    xlsx_path: Path,
    sheet_name: str,
    row_idx: int,
    new_values: dict,
) -> dict:
    """既存行の指定セルを上書き更新。

    new_values: {date, debit_account, tax_class, debit_item, department,
                 amount, credit_account, credit_amount, description} のうち
                 提供されたキーのみ更新。
    """
    if not xlsx_path.exists():
        return {"ok": False, "error": f"ファイル不在: {xlsx_path}"}
    try:
        with open(xlsx_path, "rb"):
            pass
    except PermissionError:
        return {
            "ok": False,
            "error": f"{xlsx_path.name} が他で開かれています。閉じてから再実行してください。",
        }

    try:
        wb = openpyxl.load_workbook(xlsx_path)
    except Exception as e:
        return {"ok": False, "error": f"xlsx読込失敗: {e}"}

    if sheet_name not in wb.sheetnames:
        return {"ok": False, "error": f"シートが見つかりません: {sheet_name}"}
    ws = wb[sheet_name]

    if row_idx < 2 or row_idx > ws.max_row:
        return {"ok": False, "error": f"行番号が範囲外: R{row_idx}"}

    # 各列を更新
    col_map = {
        "date": 1,
        "debit_account": 2,
        "tax_class": 3,
        "debit_item": 4,
        "department": 5,
        "amount": 6,
        "credit_account": 7,
        "credit_item": 8,
        "credit_dept": 9,
        "credit_amount": 10,
        "description": 11,
    }
    changed: list[str] = []
    for key, col in col_map.items():
        if key not in new_values:
            continue
        v = new_values[key]
        if key == "date":
            v = _to_excel_date(v)
            if v is None:
                continue
            ws.cell(row=row_idx, column=col).value = v
            ws.cell(row=row_idx, column=col).number_format = DATE_NUMBER_FORMAT
        elif key in ("amount", "credit_amount"):
            try:
                ws.cell(row=row_idx, column=col).value = int(v)
            except (ValueError, TypeError):
                continue
        else:
            ws.cell(row=row_idx, column=col).value = (
                str(v) if v is not None else None
            )
        changed.append(key)

    # 念のため日付書式を統一
    _normalize_date_format_in_sheet(ws)

    # 日付降順ソート (日付編集時の行位置維持のため)
    new_row_idx = row_idx
    if "date" in changed or "amount" in changed or "department" in changed:
        # 更新後の値を使って並び替え後の新位置を特定
        d_val = ws.cell(row=row_idx, column=1).value
        d_obj = d_val if isinstance(d_val, datetime) else (
            datetime(d_val.year, d_val.month, d_val.day)
            if isinstance(d_val, date) else None
        )
        dept_val = str(ws.cell(row=row_idx, column=5).value or '').strip()
        try:
            amt_val = int(ws.cell(row=row_idx, column=6).value)
        except (ValueError, TypeError):
            amt_val = None
        sort_sheet_by_date_desc(ws)
        if d_obj is not None and amt_val is not None:
            located = _locate_row_by_content(ws, d_obj, dept_val, amt_val)
            if located:
                new_row_idx = located
    else:
        sort_sheet_by_date_desc(ws)

    try:
        wb.save(xlsx_path)
    except PermissionError:
        return {
            "ok": False,
            "error": f"{xlsx_path.name} が他で開かれています。閉じてから再実行してください。",
        }
    except Exception as e:
        return {"ok": False, "error": f"xlsx保存失敗: {e}"}

    return {"ok": True, "sheet": sheet_name, "row": new_row_idx, "changed": changed}


def _sort_key_from_cell(v) -> datetime:
    """A列セル値からソート用キー (datetime) を作る。日付でない値は最小値扱い。"""
    if isinstance(v, datetime):
        return v
    if isinstance(v, date):
        return datetime(v.year, v.month, v.day)
    # 文字列の日付も救う
    if isinstance(v, str):
        d = _to_excel_date(v)
        if d is not None:
            return d
    return datetime.min


def sort_sheet_by_date_desc(ws) -> dict:
    """シートのA列(日付)を降順 (新→旧) に並び替える。

    挙動:
      - 並び替え対象: R2 〜 (A列に値がある最後の行)
      - A〜K の値とスタイルを行ごと移動。L列(検証/3重チェック) は不変
        (行位置に固定された検証式や手動入力を保護)。
      - 日付として解釈できない行は最後尾へ
      - 同日付の行は元の順序を保持 (stable sort)

    Returns: {sorted: 並び替えた行数}
    """
    if ws.max_row < 3:
        return {'sorted': 0}

    # データブロックの最終行 (A列に値がある最後の行)
    last_data_row = 1
    for r in range(ws.max_row, 1, -1):
        if ws.cell(row=r, column=1).value not in (None, ''):
            last_data_row = r
            break
    if last_data_row < 3:
        return {'sorted': 0}

    rows: list[tuple[datetime, int, list[dict]]] = []
    for r in range(2, last_data_row + 1):
        date_v = ws.cell(row=r, column=1).value
        key = _sort_key_from_cell(date_v)
        cells: list[dict] = []
        for c in range(1, 12):  # A=1 .. K=11
            src = ws.cell(row=r, column=c)
            cells.append({
                'value': src.value,
                'number_format': src.number_format,
                'has_style': src.has_style,
                'font': copy(src.font) if src.has_style else None,
                'fill': copy(src.fill) if src.has_style else None,
                'alignment': copy(src.alignment) if src.has_style else None,
                'border': copy(src.border) if src.has_style else None,
            })
        rows.append((key, r, cells))

    # 降順: 日付の新しい順。同日は元順序(r)を保つ
    rows.sort(key=lambda x: (x[0], -x[1]), reverse=True)

    # 既に降順ならスキップ (無用な保存を避ける)
    original_order = list(range(2, last_data_row + 1))
    new_order = [r for _k, r, _c in rows]
    if new_order == original_order:
        return {'sorted': 0, 'already_sorted': True}

    # 書き戻し
    for new_idx, (_k, _orig, cells) in enumerate(rows, start=2):
        for c, info in enumerate(cells, start=1):
            tgt = ws.cell(row=new_idx, column=c)
            tgt.value = info['value']
            if info['has_style']:
                tgt.font = info['font']
                tgt.fill = info['fill']
                tgt.alignment = info['alignment']
                tgt.border = info['border']
            tgt.number_format = info['number_format']

    return {'sorted': len(rows)}


def _locate_row_by_content(
    ws,
    date_obj: datetime | None,
    department: str,
    amount_int: int,
) -> int | None:
    """日付/部門/金額の組み合わせで一致する行を探す。
    並び替え後の新規行の位置特定に使う。
    """
    if date_obj is None:
        return None
    for r in range(2, ws.max_row + 1):
        v_date = ws.cell(row=r, column=1).value
        if not isinstance(v_date, datetime):
            continue
        if v_date.date() != date_obj.date():
            continue
        v_dept = str(ws.cell(row=r, column=5).value or '').strip()
        if v_dept != department:
            continue
        try:
            v_amt = int(ws.cell(row=r, column=6).value)
        except (ValueError, TypeError):
            continue
        if v_amt == amount_int:
            return r
    return None


def _normalize_date_format_in_sheet(ws) -> int:
    """シートA列(日付)の表示書式を yyyy/mm/dd に統一する。
    値そのものは変更せず、書式のみ変更。R1(ヘッダ)はスキップ。
    Returns: 変更したセル数。"""
    changed = 0
    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=1)
        if cell.value is None:
            continue
        if cell.number_format != DATE_NUMBER_FORMAT:
            cell.number_format = DATE_NUMBER_FORMAT
            changed += 1
    return changed


def normalize_all_date_formats(xlsx_path: Path) -> dict:
    """xlsx 内の全シートのA列日付書式を yyyy/mm/dd に統一。

    書込はせず書式のみ変更するので、データは保護される。
    Returns: {ok: True, sheets: [{name, changed}], total_changed: N}
    """
    if not xlsx_path.exists():
        return {"ok": False, "error": f"ファイル不在: {xlsx_path}"}
    try:
        with open(xlsx_path, "rb"):
            pass
    except PermissionError:
        return {
            "ok": False,
            "error": f"{xlsx_path.name} が他で開かれています。閉じてから再実行してください。",
        }
    try:
        wb = openpyxl.load_workbook(xlsx_path)
    except Exception as e:
        return {"ok": False, "error": f"xlsx読込失敗: {e}"}

    out = {"ok": True, "sheets": [], "total_changed": 0}
    for sn in wb.sheetnames:
        ws = wb[sn]
        n = _normalize_date_format_in_sheet(ws)
        out["sheets"].append({"name": sn, "changed": n})
        out["total_changed"] += n

    if out["total_changed"] == 0:
        return out  # 変更なしなら保存不要

    try:
        wb.save(xlsx_path)
    except Exception as e:
        return {"ok": False, "error": f"xlsx保存失敗: {e}"}
    return out


def sort_debit_xlsx_by_date_desc(xlsx_path: Path) -> dict:
    """xlsx 内の全 `【デビット】` シートを A列日付の降順 (新→旧) に並び替えて保存。

    A〜K のみ移動、L列(検証)は不変。書き込みは1回だけ実施。
    Returns: {ok, sheets: [{name, sorted, already_sorted}], total_sorted: N}
    """
    if not xlsx_path.exists():
        return {"ok": False, "error": f"ファイル不在: {xlsx_path}"}
    try:
        with open(xlsx_path, "rb"):
            pass
    except PermissionError:
        return {
            "ok": False,
            "error": f"{xlsx_path.name} が他で開かれています。閉じてから再実行してください。",
        }
    try:
        wb = openpyxl.load_workbook(xlsx_path)
    except Exception as e:
        return {"ok": False, "error": f"xlsx読込失敗: {e}"}

    out = {"ok": True, "sheets": [], "total_sorted": 0}
    any_changed = False
    for sn in wb.sheetnames:
        if "デビット" not in sn and "デビッド" not in sn:
            continue
        ws = wb[sn]
        res = sort_sheet_by_date_desc(ws)
        out["sheets"].append({"name": sn, **res})
        if res.get("sorted", 0) > 0:
            any_changed = True
            out["total_sorted"] += res["sorted"]

    if not any_changed:
        return out  # 何も変わってなければ保存不要

    try:
        wb.save(xlsx_path)
    except PermissionError:
        return {
            "ok": False,
            "error": f"{xlsx_path.name} が他で開かれています。閉じてから再実行してください。",
        }
    except Exception as e:
        return {"ok": False, "error": f"xlsx保存失敗: {e}"}
    return out


def _verify_written_row(
    xlsx_path: Path,
    sheet_name: str,
    row_idx: int,
    expected_values: dict,
    expected_date: datetime | None,
) -> dict:
    """保存後の xlsx を再オープンして、追記行が正しく書けているかチェック。"""
    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    except Exception as e:
        return {"ok": False, "error": f"再読込失敗: {e}"}
    if sheet_name not in wb.sheetnames:
        return {"ok": False, "error": f"シート消失: {sheet_name}"}
    ws = wb[sheet_name]

    actual: dict = {}
    for col_idx, key in enumerate(COLUMN_ORDER, start=1):
        actual[key] = ws.cell(row=row_idx, column=col_idx).value

    # 必須項目の照合
    checks = [
        ("date", expected_date, actual.get("date")),
        ("debit_account",
         expected_values["debit_account"], actual.get("debit_account")),
        ("department",
         expected_values["department"], actual.get("department")),
        ("amount",
         expected_values["amount"], actual.get("amount")),
        ("credit_account",
         expected_values["credit_account"], actual.get("credit_account")),
        ("credit_amount",
         expected_values["credit_amount"], actual.get("credit_amount")),
    ]
    for label, exp, got in checks:
        if label == "date":
            if exp is None and got is None:
                continue
            if exp is None or got is None:
                return {"ok": False,
                        "error": f"{label} 不一致: 期待={exp} 実={got}"}
            if isinstance(got, datetime) and got.date() != exp.date():
                return {"ok": False,
                        "error": f"{label} 不一致: 期待={exp.date()} 実={got.date()}"}
            continue
        if label in ("amount", "credit_amount"):
            try:
                if int(got) != int(exp):
                    return {"ok": False,
                            "error": f"{label} 不一致: 期待={exp} 実={got}"}
            except (ValueError, TypeError):
                return {"ok": False,
                        "error": f"{label} 数値変換失敗: {got!r}"}
            continue
        if str(got or "").strip() != str(exp or "").strip():
            return {"ok": False,
                    "error": f"{label} 不一致: 期待={exp!r} 実={got!r}"}

    # 日付の表示書式
    cell_a = ws.cell(row=row_idx, column=1)
    if expected_date is not None and cell_a.number_format != DATE_NUMBER_FORMAT:
        # 書式が違うだけなら警告レベル (致命ではない) → notes だけ残す
        return {
            "ok": True, "values": actual,
            "warning": f"日付書式: 期待={DATE_NUMBER_FORMAT} 実={cell_a.number_format}",
        }
    return {"ok": True, "values": actual}
