"""一般経費精算書 (xlsx / PDF) 出力モジュール

役割:
  - 既存の `01.現金出納帳 振替伝票 経費精算書（EMI/のじ）.xlsm` の date シート (例: 260408) を
    テンプレートとして読み出し、新規シートを生成する。
  - 「交通費」セクションは空のまま、「一般経費」セクション (r13-r37) のみ埋める。
  - 単独の xlsx として保存。PDF 化は openpyxl の機能では出来ないため、
    呼び出し側 (Streamlit) で Excel 出力 → ユーザが必要に応じ PDF 化、もしくは
    ReportLab を使って同等レイアウトの PDF を直接生成する。

経費精算書フォーマット (1-indexed):
  r1: タイトル行 ("医療法人社団EMIFULL ... 交通費・一般経費精算書")
  r2: 申請日
  r3: 所属
  r4: 氏名
  r5-r12: 交通費セクション (空欄のまま)
  r13: ラベル "一般経費"
  r14: ヘッダー (月日 | 支払先 | 支払事由 | 支払金額)
  r15-r36: 明細 (最大22行)
  r37: 合計B
  r39: 総合計 (A+B)
"""
from __future__ import annotations

from copy import copy
from datetime import datetime, date
from pathlib import Path

import openpyxl


# 一般経費明細の最大行数
# テンプレシート上は r15-r36 の 22 行だが、行を挿入して最大 25 行まで対応
MAX_DETAIL_ROWS = 25
DETAIL_START_ROW = 15
TEMPLATE_END_ROW = 36          # テンプレ上の元の最終明細行
TEMPLATE_DETAIL_ROWS = TEMPLATE_END_ROW - DETAIL_START_ROW + 1  # 22
TOTAL_B_ROW_TEMPLATE = 37      # 拡張前の合計B行
GRAND_TOTAL_ROW_TEMPLATE = 39  # 拡張前の総合計行

# 列 (A=1, H=8)
COL_DESCRIPTION = 1   # A
COL_AMOUNT = 8        # H


def _to_wareki(d) -> str:
    """日付 → '令和N年M月D日'。令和は 2019/05/01〜。"""
    if d is None:
        return ''
    if isinstance(d, str):
        try:
            d = datetime.strptime(d, '%Y-%m-%d').date()
        except ValueError:
            return d
    if isinstance(d, datetime):
        d = d.date()
    if not isinstance(d, date):
        return str(d)
    # 令和: 2019/05/01〜  (令和元年=2019)
    if d.year >= 2019:
        if d.year == 2019 and d.month < 5:
            year_label = '平成31年'
        else:
            r = d.year - 2018
            year_label = '令和元年' if r == 1 else f'令和{r}年'
    else:
        # ざっくり平成
        y = d.year - 1988
        year_label = f'平成{y}年'
    return f'{year_label}{d.month}月{d.day}日'


def _format_md(d) -> str:
    """日付 → 'M月D日'。"""
    if d is None:
        return ''
    if isinstance(d, str):
        try:
            d = datetime.strptime(d, '%Y-%m-%d').date()
        except ValueError:
            return d
    if isinstance(d, datetime):
        d = d.date()
    if not isinstance(d, date):
        return str(d)
    return f'{d.month}月{d.day}日'


def find_template_sheet(wb: openpyxl.Workbook) -> str | None:
    """テンプレートに使える経費精算書シートを探す。
    数字のみのシート名 (例: '260408') の中で、最もデータ件数の少ない (フォーマットが綺麗な) ものを採用。"""
    candidates = []
    for sn in wb.sheetnames:
        clean = sn.replace('-', '').replace('ｰ', '').strip()
        if clean.isdigit() and 5 <= len(clean) <= 8:
            ws = wb[sn]
            title = ws.cell(row=1, column=1).value or ''
            if '経費精算' in str(title):
                count = sum(1 for r in range(
                    DETAIL_START_ROW, TEMPLATE_END_ROW + 1)
                            if ws.cell(row=r, column=COL_AMOUNT).value)
                candidates.append((count, sn))
    if not candidates:
        return None
    candidates.sort()
    return candidates[0][1]


def _shift_merged_cells(ws, insert_at: int, n_rows: int) -> None:
    """insert_rows 時に merged_cells が連動しないので、MergedCellRange を
    直接書き換えてシフトする (unmerge/merge API は失敗するため使えない)。

    insert_at 以降に始まる範囲: 全体を下に n_rows シフト
    insert_at をまたぐ範囲      : 下端を n_rows 拡張
    """
    if n_rows <= 0:
        return
    for mr in ws.merged_cells.ranges:
        if mr.min_row >= insert_at:
            mr.min_row += n_rows
            mr.max_row += n_rows
        elif mr.max_row >= insert_at:
            mr.max_row += n_rows


def _expand_detail_rows(ws, n_extra_rows: int) -> int:
    """明細領域を `n_extra_rows` 行だけ広げる。
    Returns: 拡張後の最終明細行番号 (1-indexed)。"""
    if n_extra_rows <= 0:
        return TEMPLATE_END_ROW
    insert_at = TEMPLATE_END_ROW
    ws.insert_rows(insert_at, amount=n_extra_rows)
    # openpyxl は insert_rows で merged_cells を動かさないので手動で再構築
    _shift_merged_cells(ws, insert_at, n_extra_rows)

    # 新しい空行に r35 (テンプレの直前) の書式をコピー
    src_row_idx = insert_at - 1
    for new_r in range(insert_at, insert_at + n_extra_rows):
        for c_idx in range(1, 16):
            src = ws.cell(row=src_row_idx, column=c_idx)
            tgt = ws.cell(row=new_r, column=c_idx)
            if src.has_style:
                tgt.font = copy(src.font)
                tgt.border = copy(src.border)
                tgt.fill = copy(src.fill)
                tgt.alignment = copy(src.alignment)
                tgt.number_format = src.number_format
            tgt.value = None
    return TEMPLATE_END_ROW + n_extra_rows


def make_expense_form_workbook(template_xlsm_path,
                                applicant_name: str,
                                request_date,
                                department: str = '障がい事業部',
                                items: list[dict] = None,
                                ) -> openpyxl.Workbook:
    """テンプレートのExcelファイルから単一シートの経費精算書ブックを生成。

    Args:
        template_xlsm_path: 元のxlsmファイルのパス (テンプレートシート抽出用)
        applicant_name: 氏名
        request_date: 申請日 (date / datetime / 'YYYY-MM-DD')
        department: 所属
        items: [{'date': 'YYYY-MM-DD', 'vendor': str, 'purpose': str, 'amount': int}, ...]
            上限 22 件 (超過分は警告対象)

    Returns:
        単一シートの新規 Workbook (呼び出し側で .save() 等)
    """
    items = items or []
    if len(items) > MAX_DETAIL_ROWS:
        items = items[:MAX_DETAIL_ROWS]

    src_wb = openpyxl.load_workbook(
        str(template_xlsm_path), keep_vba=False, data_only=False)
    template_name = find_template_sheet(src_wb)
    if template_name is None:
        raise ValueError('テンプレートとなる経費精算書シートが見つかりません')

    src_ws = src_wb[template_name]

    new_sheet_name = (
        f'経費精算_{applicant_name}_'
        f'{_format_md(request_date).replace("月", "-").replace("日", "")}'
    )
    if len(new_sheet_name) > 31:
        new_sheet_name = new_sheet_name[:31]
    copied = src_wb.copy_worksheet(src_ws)
    copied.title = new_sheet_name

    # ---- 明細領域の拡張 (22行 → 最大25行) ----
    n_items = len(items)
    extra = max(0, n_items - TEMPLATE_DETAIL_ROWS)
    detail_end = _expand_detail_rows(copied, extra)
    # 行挿入で TOTAL_B / GRAND_TOTAL の行番号がシフトしている
    total_b_row = TOTAL_B_ROW_TEMPLATE + extra
    grand_total_row = GRAND_TOTAL_ROW_TEMPLATE + extra

    # 申請者情報を上書き
    copied.cell(row=2, column=8).value = _to_wareki(request_date)
    copied.cell(row=3, column=8).value = department
    copied.cell(row=4, column=8).value = applicant_name

    # 一般経費セクションをまずクリア
    for r in range(DETAIL_START_ROW, detail_end + 1):
        copied.cell(row=r, column=COL_DESCRIPTION).value = None
        copied.cell(row=r, column=COL_AMOUNT).value = None

    # 明細書込
    total_b = 0
    for i, item in enumerate(items):
        r_idx = DETAIL_START_ROW + i
        date_label = _format_md(item.get('date'))
        vendor = (item.get('vendor') or '').strip()
        purpose = (item.get('purpose') or '').strip()
        parts = []
        if date_label:
            parts.append(date_label)
        if vendor:
            parts.append(vendor)
        if purpose:
            parts.append(purpose)
        copied.cell(row=r_idx, column=COL_DESCRIPTION).value = (
            '　'.join(parts))
        amount = int(item.get('amount') or 0)
        copied.cell(row=r_idx, column=COL_AMOUNT).value = amount
        total_b += amount

    copied.cell(row=total_b_row, column=COL_AMOUNT).value = total_b
    copied.cell(row=grand_total_row, column=COL_AMOUNT).value = total_b

    try:
        copied.cell(row=42 + extra, column=6).value = _to_wareki(request_date)
    except Exception:
        pass
    try:
        copied.cell(row=45 + extra, column=8).value = applicant_name
    except Exception:
        pass

    # 不要なシートを削除して、新シート1枚だけ残す
    keep = copied.title
    for sn in list(src_wb.sheetnames):
        if sn != keep:
            del src_wb[sn]

    return src_wb
