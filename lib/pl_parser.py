"""部門別損益計算書 Excel/CSV パーサ

設計:
- 列順は毎月変わりうる前提。部門名（excel_name または display_name）でマッチする。
- 部門名の prefix (例 '4.1.') は除去してから照合する。
- 行（科目）は科目名（column A／CSVは column B）で照合する。
- 不要な部門列・不明な科目行は warn＋skip する。

主要API:
    parse_pl_workbook(file_or_path, fiscal_start_ym, subunit_lookup, known_accounts)
        -> ParseResult   (Excel: 複数月シート)
    parse_pl_csv(file_or_path, filename, subunit_lookup, known_accounts, year_month=None)
        -> SheetParseResult  (CSV: 1月分)

subunit_lookup: dict[str, str]
    認識可能な部門名（excel_name または display_name 等）→ 正規(excel_name)へのマップ。
    DB から `{s['excel_name']: s['excel_name'], s['display_name']: s['excel_name']}` のように作る。
"""
from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass, field
from typing import Iterable

import openpyxl


SHEET_MONTH_RE = re.compile(r'(\d{1,2})\s*月')

# 部門名の先頭にある番号 prefix を除去する
# 例: "4.1.SORATOいなみ" / "01　のじぎく高砂" / "001  名前"
SUBUNIT_PREFIX_RES = [
    re.compile(r'^\s*\d+\.\d+\.\s*'),         # '4.1.'
    re.compile(r'^\s*\d+[\s　]+'),         # '01　' / '01 '
]

# CSVファイル名から年月を抽出するパターン (例: 「期間：2023年06月」)
FILENAME_YM_RE = re.compile(r'(\d{4})\s*年\s*(\d{1,2})\s*月')


def _strip_subunit_prefix(name: str) -> str:
    """部門名先頭の番号 prefix を除去。'4.1.SORATOいなみ' / '01　のじぎく高砂' に対応。"""
    if not name:
        return ''
    s = name
    for pat in SUBUNIT_PREFIX_RES:
        s = pat.sub('', s)
    return s.strip()


def _extract_ym_from_filename(filename: str) -> str | None:
    """ファイル名から '2023-06' 形式の年月を抽出。"""
    if not filename:
        return None
    m = FILENAME_YM_RE.search(filename)
    if not m:
        return None
    return f"{int(m.group(1))}-{int(m.group(2)):02d}"


def _resolve_subunit(name: str, subunit_lookup: dict[str, str]) -> str | None:
    """部門名 → 正規(excel_name)。prefix除去・空白正規化を試みる。"""
    if not name:
        return None
    candidates = [name.strip(), _strip_subunit_prefix(name)]
    for c in candidates:
        if c in subunit_lookup:
            return subunit_lookup[c]
    return None


@dataclass
class SheetParseResult:
    sheet_name: str
    year_month: str | None  # 'YYYY-MM' or None (skipped)
    matched_subunits: list[str] = field(default_factory=list)   # excel_name
    unknown_subunit_columns: list[str] = field(default_factory=list)
    unknown_account_rows: list[str] = field(default_factory=list)
    # entries: subunit_excel_name -> [(account_name, amount), ...]
    entries: dict[str, list[tuple[str, int]]] = field(default_factory=dict)
    skipped_reason: str | None = None


@dataclass
class ParseResult:
    fiscal_start_ym: str
    sheets: list[SheetParseResult] = field(default_factory=list)
    error: str | None = None

    @property
    def matched_year_months(self) -> list[str]:
        return sorted({s.year_month for s in self.sheets if s.year_month})

    @property
    def total_entry_count(self) -> int:
        return sum(
            sum(len(v) for v in s.entries.values())
            for s in self.sheets if s.year_month
        )


def _add_months(yyyymm: str, n: int) -> str:
    y, m = int(yyyymm[:4]), int(yyyymm[5:7])
    total = y * 12 + (m - 1) + n
    return f"{total // 12}-{(total % 12) + 1:02d}"


def _normalize_str(v) -> str:
    if v is None:
        return ''
    return str(v).strip()


def _resolve_year_month(sheet_name: str, fiscal_start_ym: str) -> str | None:
    """シート名から月を抽出し、期首年月を起点に YYYY-MM を返す。"""
    m = SHEET_MONTH_RE.search(sheet_name or '')
    if not m:
        return None
    target_month = int(m.group(1))
    if not (1 <= target_month <= 12):
        return None
    fy_y, fy_m = int(fiscal_start_ym[:4]), int(fiscal_start_ym[5:7])
    # fy_m から始まり12ヶ月。target_month を fy_m を 0-index として何ヶ月目かを決める
    offset = (target_month - fy_m) % 12
    return _add_months(fiscal_start_ym, offset)


def _find_header_row(ws, subunit_lookup: dict[str, str], scan_rows: int = 10) -> int | None:
    """既知のサブ部門名（prefix 付きでも可）または '合計' を含む行を探す。"""
    max_col = ws.max_column
    best_row, best_hits = None, 0
    for r in range(1, min(scan_rows, ws.max_row) + 1):
        hits = 0
        for c in range(1, max_col + 1):
            v = _normalize_str(ws.cell(row=r, column=c).value)
            if not v:
                continue
            if v == '合計' or v == '部門合計':
                hits += 1
            elif _resolve_subunit(v, subunit_lookup):
                hits += 1
        if hits > best_hits:
            best_hits = hits
            best_row = r
    return best_row if best_hits >= 1 else None


def _build_column_map(ws, header_row: int, subunit_lookup: dict[str, str],
                      first_data_col: int = 1) -> tuple[dict[str, int], list[str]]:
    """header_row の各列を走査し、{excel_name: col_index} と未知の列名一覧を返す。
    first_data_col は部門列が始まる列 (1=column A, 2=column B など)。"""
    col_map: dict[str, int] = {}
    unknown: list[str] = []
    for c in range(first_data_col, ws.max_column + 1):
        v = _normalize_str(ws.cell(row=header_row, column=c).value)
        if not v or v == '合計' or v == '部門合計':
            continue
        canonical = _resolve_subunit(v, subunit_lookup)
        if canonical:
            col_map[canonical] = c
        else:
            unknown.append(v)
    return col_map, unknown


def _build_account_row_map(ws, header_row: int, known_account_names: set[str]) -> tuple[dict[str, int], list[str]]:
    """column A を header_row より下に走査し、{account_name: row_index} を返す。"""
    row_map: dict[str, int] = {}
    unknown: list[str] = []
    for r in range(header_row + 1, ws.max_row + 1):
        v = _normalize_str(ws.cell(row=r, column=1).value)
        if not v:
            continue
        if v in known_account_names:
            row_map[v] = r
        else:
            unknown.append(v)
    return row_map, unknown


def _read_amount(cell_value) -> int:
    if cell_value is None or cell_value == '':
        return 0
    try:
        return int(round(float(cell_value)))
    except (ValueError, TypeError):
        return 0


def parse_pl_workbook(
    file_or_path,
    fiscal_start_ym: str,
    subunit_lookup: dict[str, str],
    known_accounts: Iterable[str],
) -> ParseResult:
    """Excel ファイルをパースし、各シート毎の取込候補データを返す。
    subunit_lookup: 認識可能な部門名(excel_name または display_name) → 正規 excel_name のマップ。
    """
    known_accounts_set = set(known_accounts)

    result = ParseResult(fiscal_start_ym=fiscal_start_ym)

    try:
        if hasattr(file_or_path, 'read'):
            data = file_or_path.read()
            wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True, read_only=False)
        else:
            wb = openpyxl.load_workbook(file_or_path, data_only=True, read_only=False)
    except Exception as e:
        result.error = f"Excelファイルを開けません: {e}"
        return result

    for sheet_name in wb.sheetnames:
        sheet_norm = sheet_name.strip()
        sr = SheetParseResult(sheet_name=sheet_name, year_month=None)

        if not SHEET_MONTH_RE.search(sheet_norm):
            sr.skipped_reason = "月度シートではない"
            result.sheets.append(sr)
            continue

        ym = _resolve_year_month(sheet_norm, fiscal_start_ym)
        if ym is None:
            sr.skipped_reason = "シート名から年月を判定できない"
            result.sheets.append(sr)
            continue

        ws = wb[sheet_name]

        header_row = _find_header_row(ws, subunit_lookup)
        if header_row is None:
            sr.skipped_reason = "ヘッダ行（部門名行）が見つからない"
            sr.year_month = ym
            result.sheets.append(sr)
            continue

        col_map, unknown_cols = _build_column_map(ws, header_row, subunit_lookup, first_data_col=2)
        row_map, unknown_rows = _build_account_row_map(ws, header_row, known_accounts_set)

        sr.year_month = ym
        sr.matched_subunits = sorted(col_map.keys())
        sr.unknown_subunit_columns = unknown_cols
        sr.unknown_account_rows = unknown_rows

        for excel_name, col in col_map.items():
            entries: list[tuple[str, int]] = []
            for account_name, row in row_map.items():
                cell = ws.cell(row=row, column=col).value
                amount = _read_amount(cell)
                entries.append((account_name, amount))
            sr.entries[excel_name] = entries

        result.sheets.append(sr)

    return result


def parse_pl_csv(
    file_or_path,
    filename: str,
    subunit_lookup: dict[str, str],
    known_accounts: Iterable[str],
    year_month: str | None = None,
    encoding: str = 'cp932',
) -> 'SheetParseResult':
    """1ファイル＝1月分のCSV(Shift-JIS等)をパース。
    - 行構成: 1行目タイトル、2行目ヘッダ(列A=勘定科目コード、列B=科目名、列C以降=部門)
    - 列ヘッダの数字prefix（'4.1.' 等）は除去してマッチング。
    """
    sr = SheetParseResult(sheet_name=filename or 'csv', year_month=year_month)

    if year_month is None:
        year_month = _extract_ym_from_filename(filename or '')
    if year_month is None:
        sr.skipped_reason = "ファイル名から年月を判定できない"
        return sr
    sr.year_month = year_month

    # 読み込み（bytes/file-like/path どれでも）
    try:
        if hasattr(file_or_path, 'read'):
            raw = file_or_path.read()
            if isinstance(raw, bytes):
                text = raw.decode(encoding, errors='replace')
            else:
                text = raw
        else:
            with open(file_or_path, encoding=encoding) as f:
                text = f.read()
    except Exception as e:
        sr.skipped_reason = f"CSV読み込みエラー: {e}"
        return sr

    # CSV パース（universal newlines）
    try:
        rows = list(csv.reader(text.splitlines()))
    except Exception as e:
        sr.skipped_reason = f"CSVパースエラー: {e}"
        return sr

    if len(rows) < 2:
        sr.skipped_reason = "CSV内容が空"
        return sr

    known_accounts_set = set(known_accounts)

    # ヘッダ行を探索（合計列または既知部門名を含む行）
    header_row_idx: int | None = None
    for r_idx, row in enumerate(rows[:10]):
        for cell in row:
            cv = (cell or '').strip()
            if cv in ('合計', '部門合計'):
                header_row_idx = r_idx
                break
            if _resolve_subunit(cv, subunit_lookup):
                header_row_idx = r_idx
                break
        if header_row_idx is not None:
            break

    if header_row_idx is None:
        sr.skipped_reason = "ヘッダ行が見つからない"
        return sr

    header = rows[header_row_idx]
    col_map: dict[str, int] = {}     # excel_name -> col_index
    unknown_cols: list[str] = []
    for c_idx, cell in enumerate(header):
        cv = (cell or '').strip()
        if not cv or cv in ('合計', '部門合計', '勘定科目コード'):
            continue
        canonical = _resolve_subunit(cv, subunit_lookup)
        if canonical:
            col_map[canonical] = c_idx
        else:
            unknown_cols.append(cv)
    sr.matched_subunits = sorted(col_map.keys())
    sr.unknown_subunit_columns = unknown_cols

    if not col_map:
        sr.skipped_reason = "認識できる部門列が無い"
        return sr

    # データ行を走査（科目名は column B = index 1）
    unknown_rows: list[str] = []
    for r in rows[header_row_idx + 1:]:
        if len(r) < 2:
            continue
        account_name = (r[1] or '').strip()
        if not account_name:
            continue
        if account_name not in known_accounts_set:
            unknown_rows.append(account_name)
            continue
        for excel_name, c_idx in col_map.items():
            if c_idx >= len(r):
                continue
            cell = (r[c_idx] or '').strip()
            if cell == '':
                amount = 0
            else:
                try:
                    amount = int(round(float(cell.replace(',', ''))))
                except (ValueError, TypeError):
                    amount = 0
            sr.entries.setdefault(excel_name, []).append((account_name, amount))

    sr.unknown_account_rows = unknown_rows
    return sr


def build_subunit_lookup(subunits) -> dict[str, str]:
    """DBから取得したサブ部門レコード一覧から、認識用ルックアップを構築。
    excel_name と display_name の両方を recognized name として登録し、値は excel_name。"""
    lookup: dict[str, str] = {}
    for s in subunits:
        en = s['excel_name']
        dn = s.get('display_name') or en
        lookup[en] = en
        if dn != en:
            lookup[dn] = en
    return lookup


def import_parse_result_to_db(parse_result: ParseResult, db_module) -> dict:
    """ParseResult を DB に書き込む。
    戻り値: {'sheets': N, 'subunits': N, 'entries': N, 'year_months': [...], 'errors': [...]}
    """
    summary = {'sheets': 0, 'subunits': 0, 'entries': 0,
               'year_months': [], 'errors': []}
    if parse_result.error:
        summary['errors'].append(parse_result.error)
        return summary

    # 名前 → ID 解決を一度キャッシュ
    accounts = {a['name']: a['id'] for a in db_module.list_pl_accounts()}
    subunits = {s['excel_name']: s['id'] for s in db_module.list_pl_subunits()}

    yms_seen = set()
    for sr in parse_result.sheets:
        if sr.year_month is None or not sr.entries:
            continue
        summary['sheets'] += 1
        yms_seen.add(sr.year_month)

        for excel_name, name_amount_pairs in sr.entries.items():
            sub_id = subunits.get(excel_name)
            if sub_id is None:
                summary['errors'].append(
                    f"{sr.year_month}: サブ部門マスタに無い: {excel_name}"
                )
                continue
            # 同じ科目が複数行ある場合は最後の出現で上書き（CSV側のセクション小計重複対策）
            seen_by_acc: dict[int, int] = {}
            for account_name, amount in name_amount_pairs:
                acc_id = accounts.get(account_name)
                if acc_id is None:
                    summary['errors'].append(
                        f"{sr.year_month}/{excel_name}: 不明科目: {account_name}"
                    )
                    continue
                seen_by_acc[acc_id] = amount
            entries_resolved = list(seen_by_acc.items())
            n = db_module.replace_pl_entries(sr.year_month, sub_id, entries_resolved)
            summary['subunits'] += 1
            summary['entries'] += n

    summary['year_months'] = sorted(yms_seen)
    return summary
