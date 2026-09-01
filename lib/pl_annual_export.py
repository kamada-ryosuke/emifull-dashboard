"""損益ダッシュボードの年度別PL Excel出力。"""

from __future__ import annotations

import io
import re
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


_INVALID_SHEET_CHARS = re.compile(r"[\\/*?:\[\]]")


def fiscal_year_months(fiscal_year: int) -> list[str]:
    """4月開始の年度に属する12か月を返す。"""
    return [
        f"{fiscal_year if month >= 4 else fiscal_year + 1}-{month:02d}"
        for month in list(range(4, 13)) + list(range(1, 4))
    ]


def _unique_sheet_title(label: str, used: set[str]) -> str:
    base = _INVALID_SHEET_CHARS.sub("_", str(label)).strip(" '") or "PL"
    base = base[:31]
    title = base
    suffix_no = 2
    while title.lower() in used:
        suffix = f"_{suffix_no}"
        title = f"{base[:31 - len(suffix)]}{suffix}"
        suffix_no += 1
    used.add(title.lower())
    return title


def build_annual_pl_xlsx(
    fiscal_year: int,
    entries: list[dict],
    accounts: list[dict],
    sheet_specs: list[dict],
    category_labels: dict[str, str] | None = None,
) -> bytes:
    """施設・法人ごとに1シートを持つ年度PLブックを生成する。

    ``sheet_specs`` は ``label`` と ``subunit_ids`` を持つ辞書の配列。
    データが未取込の月は空欄、取込済み月の未計上科目は0として出力する。
    """
    months = fiscal_year_months(int(fiscal_year))
    category_labels = category_labels or {}
    ordered_accounts = sorted(accounts, key=lambda a: (a.get("display_order", 0), a.get("id", 0)))

    amount_by_key: dict[tuple[int, int, str], int] = defaultdict(int)
    months_by_subunit: dict[int, set[str]] = defaultdict(set)
    for entry in entries:
        ym = entry.get("year_month")
        if ym not in months:
            continue
        subunit_id = int(entry["subunit_id"])
        account_id = int(entry["account_id"])
        amount_by_key[(subunit_id, account_id, ym)] += int(entry.get("amount") or 0)
        months_by_subunit[subunit_id].add(ym)

    wb = Workbook()
    wb.remove(wb.active)
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"

    navy = "1F4E78"
    blue = "D9EAF7"
    pale_blue = "EAF3F8"
    total_fill = "FFF2CC"
    white = "FFFFFF"
    gray = "64748B"
    thin_gray = Side(style="thin", color="CBD5E1")
    border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)
    used_titles: set[str] = set()

    for spec in sheet_specs:
        label = str(spec.get("label") or "PL")
        subunit_ids = {int(sid) for sid in spec.get("subunit_ids", [])}
        ws = wb.create_sheet(_unique_sheet_title(label, used_titles))
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "C5"
        ws.auto_filter.ref = f"A4:{get_column_letter(3 + len(months))}{4 + len(ordered_accounts)}"

        last_col = 3 + len(months)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
        title_cell = ws.cell(1, 1, f"{fiscal_year}年度 月次損益計算書　{label}")
        title_cell.font = Font(size=16, bold=True, color=white)
        title_cell.fill = PatternFill("solid", fgColor=navy)
        title_cell.alignment = Alignment(vertical="center")
        ws.row_dimensions[1].height = 29

        first_ym, last_ym = months[0], months[-1]
        imported_months = [
            ym for ym in months
            if any(ym in months_by_subunit.get(sid, set()) for sid in subunit_ids)
        ]
        status = (
            f"対象期間: {first_ym} ～ {last_ym}　単位: 円　"
            f"取込済: {len(imported_months)}/12か月"
        )
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_col)
        ws.cell(2, 1, status).font = Font(size=10, color=gray)
        ws.cell(2, 1).alignment = Alignment(vertical="center")
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=last_col)
        ws.cell(3, 1, "※ 空欄の月は、対象施設の損益データが未取込です。年度合計は取込済み月のみを合算しています。")
        ws.cell(3, 1).font = Font(size=9, italic=True, color=gray)

        headers = ["区分", "勘定科目", *[f"{int(ym[:4])}年{int(ym[5:])}月" for ym in months], "年度合計"]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(4, col_idx, header)
            cell.font = Font(bold=True, color=white)
            cell.fill = PatternFill("solid", fgColor=navy)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        ws.row_dimensions[4].height = 24

        for row_idx, account in enumerate(ordered_accounts, start=5):
            account_id = int(account["id"])
            is_total = bool(account.get("is_total"))
            category = str(account.get("category") or "")
            row_fill = total_fill if is_total else (pale_blue if row_idx % 2 else white)

            values: list[int | None] = []
            for ym in months:
                month_is_imported = any(
                    ym in months_by_subunit.get(sid, set()) for sid in subunit_ids
                )
                if not month_is_imported:
                    values.append(None)
                    continue
                values.append(sum(amount_by_key[(sid, account_id, ym)] for sid in subunit_ids))

            row_values = [
                category_labels.get(category, category),
                account.get("name") or "",
                *values,
                sum(value for value in values if value is not None),
            ]
            for col_idx, value in enumerate(row_values, start=1):
                cell = ws.cell(row_idx, col_idx, value)
                cell.border = border
                cell.fill = PatternFill("solid", fgColor=row_fill)
                cell.alignment = Alignment(
                    horizontal="right" if col_idx >= 3 else "left",
                    vertical="center",
                )
                if col_idx >= 3:
                    cell.number_format = '#,##0;[Red]-#,##0;0'
                if is_total:
                    cell.font = Font(bold=True)

        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["B"].width = 26
        for col_idx in range(3, last_col + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 15
        ws.auto_filter.ref = f"A4:{get_column_letter(last_col)}{4 + len(ordered_accounts)}"
        ws.print_title_rows = "1:4"
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.oddFooter.center.text = "Page &P / &N"
        ws.oddFooter.right.text = f"{fiscal_year}年度"

    if not wb.worksheets:
        ws = wb.create_sheet("PL")
        ws["A1"] = "出力対象がありません。"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
