"""給与台帳からダウンロードするExcel帳票。"""

from __future__ import annotations

import io

import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def build_overtime_targets_excel(
    employees: pd.DataFrame,
    *,
    period_label: str,
    corporation_label: str,
) -> bytes:
    """残業対象者一覧を見やすい色付きExcelとして返す。"""
    columns = ['順位', '所属', '氏名', '申請残業(h)']
    missing = [column for column in columns if column not in employees.columns]
    if missing:
        raise ValueError(f"Excel出力に必要な列がありません: {', '.join(missing)}")

    wb = Workbook()
    ws = wb.active
    ws.title = '残業対象者'
    ws.sheet_view.showGridLines = False

    navy = '1E3A5F'
    blue = 'DCE6F1'
    pale_blue = 'EDF4FB'
    white = 'FFFFFF'
    gray = '64748B'
    line = 'CBD5E1'
    gold = 'FFD966'
    silver = 'D9E2F3'
    bronze = 'F4B183'
    font_name = 'BIZ UDPゴシック'

    ws.merge_cells('A1:D1')
    title = ws['A1']
    title.value = '残業対象者一覧'
    title.fill = PatternFill('solid', fgColor=navy)
    title.font = Font(name=font_name, size=16, bold=True, color=white)
    title.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 30

    ws.merge_cells('A2:D2')
    subtitle = ws['A2']
    subtitle.value = (
        f"対象期間: {period_label}　法人: {corporation_label}　"
        f"対象者: {len(employees)}名"
    )
    subtitle.font = Font(name=font_name, size=10, color=gray)
    subtitle.alignment = Alignment(vertical='center')
    ws.row_dimensions[2].height = 23

    header_row = 4
    thin = Side(style='thin', color=line)
    border = Border(right=thin, bottom=thin)
    for column_index, column in enumerate(columns, start=1):
        cell = ws.cell(row=header_row, column=column_index, value=column)
        cell.fill = PatternFill('solid', fgColor=blue)
        cell.font = Font(name=font_name, size=10, bold=True, color=navy)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(top=thin, right=thin, bottom=thin)
    ws.row_dimensions[header_row].height = 26

    first_data_row = header_row + 1
    for row_offset, values in enumerate(
        employees[columns].itertuples(index=False, name=None),
    ):
        row_index = first_data_row + row_offset
        for column_index, value in enumerate(values, start=1):
            cell = ws.cell(row=row_index, column=column_index, value=value)
            cell.font = Font(name=font_name, size=10)
            cell.border = border
            cell.alignment = Alignment(
                horizontal=(
                    'center' if column_index == 1
                    else 'right' if column_index == 4
                    else 'left'
                ),
                vertical='center',
            )
            if row_offset % 2:
                cell.fill = PatternFill('solid', fgColor=pale_blue)
        ws.cell(row=row_index, column=4).number_format = '0.00'
        ws.row_dimensions[row_index].height = 23

        rank_fill = {1: gold, 2: silver, 3: bronze}.get(int(values[0]))
        if rank_fill:
            rank_cell = ws.cell(row=row_index, column=1)
            rank_cell.fill = PatternFill('solid', fgColor=rank_fill)
            rank_cell.font = Font(name=font_name, size=10, bold=True, color=navy)

    last_data_row = header_row + len(employees)
    if last_data_row >= first_data_row:
        hours_range = f'D{first_data_row}:D{last_data_row}'
        ws.conditional_formatting.add(
            hours_range,
            ColorScaleRule(
                start_type='min', start_color='E2F0D9',
                mid_type='percentile', mid_value=50, mid_color='FFF2CC',
                end_type='max', end_color='F4CCCC',
            ),
        )
        for row_index in range(first_data_row, last_data_row + 1):
            ws.cell(row=row_index, column=4).font = Font(
                name=font_name, size=10, bold=True,
            )
        ws.auto_filter.ref = f'A{header_row}:D{last_data_row}'

    widths = {'A': 8, 'B': 25, 'C': 18, 'D': 16}
    for column, width in widths.items():
        ws.column_dimensions[column].width = width
    ws.freeze_panes = f'A{first_data_row}'
    ws.print_title_rows = f'1:{header_row}'
    ws.print_area = f'A1:{get_column_letter(len(columns))}{max(last_data_row, header_row)}'
    ws.page_setup.orientation = 'portrait'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.sheet_properties.outlinePr.summaryBelow = True
    ws.oddFooter.center.text = 'Page &P / &N'
    ws.oddFooter.center.size = 9
    ws.oddFooter.center.color = gray

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
