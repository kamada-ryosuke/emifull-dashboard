"""売上収支予測表の全施設詳細を、閲覧しやすいExcelに変換する。"""

from __future__ import annotations

from io import BytesIO

import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter


NAVY = "17365D"
BLUE = "DDEBF7"
TEAL = "DDEFEA"
GRAY = "E7E6E6"
GREEN = "E2F0D9"
GOLD = "FFF2CC"
PALE_RED = "FCE8E6"
PALE_ORANGE = "FCE4D6"
PALE_GREEN = "E2F0D9"
PALE_BLUE = "D9EAF7"
WHITE = "FFFFFF"
TEXT = "243447"
MUTED = "667085"
BORDER = "CCD5E0"
RED = "C00000"
DARK_GREEN = "006100"

AMOUNT_COLUMNS = {
    "予定売上(千円)", "着地予測売上(千円)", "売上差(千円)",
    "予定販管費(千円)", "着地予測販管費(千円)",
    "予定利益(千円)", "着地予測利益(千円)", "利益差(千円)",
}
PERCENT_COLUMNS = {"予定利益率", "着地予測利益率"}
INTEGER_COLUMNS = {"月間予定延べ利用回数", "月末着地予測利用回数", "利用回数差"}

FACILITY_LAYOUT = [
    (
        (
            "SORATOいなみ", "UMIEいなみ", "UMIEいなみ第二教室",
            "SORATOいなみ第二教室", "BLOOMいなみ",
        ),
        "いなみエリア合計",
    ),
    (
        (
            "SORATOてんり", "UMIEてんり", "BLOOMてんり",
            "カラダキッズてんり",
        ),
        "てんりエリア合計",
    ),
    (("ジョブカレッジかこがわ", "カラダキッズかこがわ"), "かこがわエリア"),
    (("Hinodeシェアホーム加古川",), None),
    (("Hinodeシェアホーム天理",), None),
    (("のじぎく高砂", "のじぎく稲美", "のじぎく加古川"), "のじぎく合計"),
]
AREA_TOTAL_LABELS = {total for _, total in FACILITY_LAYOUT if total}


def _normalize_facility_name(value) -> str:
    return str(value or "").replace("　", "").replace(" ", "").casefold()


def _area_total_row(group_df: pd.DataFrame, label: str) -> dict:
    total = {column: "" for column in group_df.columns}
    total["施設名"] = label
    for column in AMOUNT_COLUMNS | INTEGER_COLUMNS:
        if column in group_df.columns:
            values = pd.to_numeric(group_df[column], errors="coerce")
            value = values.sum(min_count=1)
            total[column] = None if pd.isna(value) else int(value)

    planned_revenue = total.get("予定売上(千円)")
    planned_profit = total.get("予定利益(千円)")
    landing_revenue = total.get("着地予測売上(千円)")
    landing_profit = total.get("着地予測利益(千円)")
    total["予定利益率"] = (
        planned_profit / planned_revenue * 100
        if planned_revenue not in (None, 0) and planned_profit is not None else None
    )
    total["着地予測利益率"] = (
        landing_profit / landing_revenue * 100
        if landing_revenue not in (None, 0) and landing_profit is not None else None
    )
    if "状態" in total:
        total["状態"] = "エリア合計"
    return total


def prepare_revenue_forecast_detail_df(df: pd.DataFrame) -> pd.DataFrame:
    """指定された施設順に並べ、対象エリアの合計行を挿入する。"""
    if df.empty or "施設名" not in df.columns:
        return df.copy()

    rows_by_name = {
        _normalize_facility_name(row["施設名"]): row.to_dict()
        for _, row in df.iterrows()
    }
    used = set()
    output_rows = []
    for facility_names, total_label in FACILITY_LAYOUT:
        group_rows = []
        for facility_name in facility_names:
            key = _normalize_facility_name(facility_name)
            row = rows_by_name.get(key)
            if row is None:
                continue
            group_rows.append(row)
            output_rows.append(row)
            used.add(key)
        if total_label and group_rows:
            group_df = pd.DataFrame(group_rows, columns=df.columns)
            output_rows.append(_area_total_row(group_df, total_label))

    for _, row in df.iterrows():
        key = _normalize_facility_name(row["施設名"])
        if key not in used:
            output_rows.append(row.to_dict())

    return pd.DataFrame(output_rows, columns=df.columns)


def _excel_value(column: str, value):
    if pd.isna(value):
        return None
    if column in PERCENT_COLUMNS:
        return float(value) / 100
    if column in AMOUNT_COLUMNS or column in INTEGER_COLUMNS:
        return float(value) if isinstance(value, float) else int(value)
    return value


def _formula_sum(column_letter: str, first_row: int, last_row: int) -> str:
    if last_row < first_row:
        return "=0"
    return f"=SUBTOTAL(109,{column_letter}{first_row}:{column_letter}{last_row})"


def _formula_sum_rows(column_letter: str, row_numbers: list[int]) -> str:
    if not row_numbers:
        return "=0"
    segments = []
    start = previous = row_numbers[0]
    for row_number in row_numbers[1:]:
        if row_number == previous + 1:
            previous = row_number
            continue
        segments.append((start, previous))
        start = previous = row_number
    segments.append((start, previous))
    formulas = [
        f"SUBTOTAL(109,{column_letter}{start}:{column_letter}{end})"
        for start, end in segments
    ]
    if len(formulas) == 1:
        return f"={formulas[0]}"
    return f"=SUM({','.join(formulas)})"


def build_revenue_forecast_xlsx(df: pd.DataFrame, target_ym: str) -> bytes:
    """詳細テーブルを色付きのExcelワークブックとして返す。"""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "全施設詳細"
    sheet.sheet_view.showGridLines = False

    columns = list(df.columns)
    last_col = len(columns)
    last_col_letter = get_column_letter(last_col)
    header_row = 8
    first_data_row = header_row + 1
    last_data_row = first_data_row + len(df) - 1
    total_row = max(first_data_row, last_data_row + 1)
    facility_excel_rows = [
        first_data_row + offset
        for offset, (_, row) in enumerate(df.iterrows())
        if str(row.get("施設名") or "") not in AREA_TOTAL_LABELS
    ]

    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    title = sheet.cell(1, 1, "売上収支予測表｜全施設 詳細テーブル")
    title.fill = PatternFill("solid", fgColor=NAVY)
    title.font = Font(name="Yu Gothic", size=17, bold=True, color=WHITE)
    title.alignment = Alignment(horizontal="left", vertical="center")
    sheet.row_dimensions[1].height = 32

    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_col)
    subtitle = sheet.cell(
        2, 1,
        f"対象年月：{target_ym}　｜　差異＝着地予測－予定　｜　金額単位：千円",
    )
    subtitle.font = Font(name="Yu Gothic", size=10, color=MUTED)
    subtitle.alignment = Alignment(vertical="center")
    sheet.row_dimensions[2].height = 22

    index_by_name = {name: idx + 1 for idx, name in enumerate(columns)}
    cards = [
        (1, 4, "予定売上合計", "予定売上(千円)", BLUE),
        (5, 8, "着地予測売上合計", "着地予測売上(千円)", TEAL),
        (9, 12, "着地予測利益合計", "着地予測利益(千円)", GREEN),
        (13, last_col, "施設数", None, GOLD),
    ]
    for start_col, end_col, label, source_column, color in cards:
        if start_col > last_col:
            continue
        end_col = min(end_col, last_col)
        sheet.merge_cells(start_row=4, start_column=start_col, end_row=4, end_column=end_col)
        sheet.merge_cells(start_row=5, start_column=start_col, end_row=5, end_column=end_col)
        label_cell = sheet.cell(4, start_col, label)
        value_cell = sheet.cell(5, start_col)
        if source_column:
            source_letter = get_column_letter(index_by_name[source_column])
            value_cell.value = _formula_sum_rows(source_letter, facility_excel_rows)
            value_cell.number_format = '#,##0" 千円";[Red]-#,##0" 千円";－'
        else:
            value_cell.value = len(facility_excel_rows)
            value_cell.number_format = '0"施設"'
        for row in (4, 5):
            for col in range(start_col, end_col + 1):
                cell = sheet.cell(row, col)
                cell.fill = PatternFill("solid", fgColor=color)
                cell.border = Border(
                    top=Side(style="thin", color=BORDER),
                    bottom=Side(style="thin", color=BORDER),
                )
        label_cell.font = Font(name="Yu Gothic", size=9, bold=True, color=MUTED)
        value_cell.font = Font(name="Yu Gothic", size=15, bold=True, color=TEXT)
        label_cell.alignment = value_cell.alignment = Alignment(
            horizontal="center", vertical="center"
        )
    sheet.row_dimensions[4].height = 21
    sheet.row_dimensions[5].height = 29

    groups = [
        (1, 1, "施設", NAVY),
        (2, 4, "利用回数", "4472C4"),
        (5, 7, "売上（千円）", "1F7A8C"),
        (8, 9, "販管費（千円）", "7F8C8D"),
        (10, 14, "利益・利益率", "548235"),
        (15, last_col, "運営状況", "BF9000"),
    ]
    for start_col, end_col, label, color in groups:
        if start_col > last_col:
            continue
        end_col = min(end_col, last_col)
        if start_col != end_col:
            sheet.merge_cells(
                start_row=7, start_column=start_col, end_row=7, end_column=end_col
            )
        cell = sheet.cell(7, start_col, label)
        cell.fill = PatternFill("solid", fgColor=color)
        cell.font = Font(name="Yu Gothic", size=10, bold=True, color=WHITE)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        for col in range(start_col, end_col + 1):
            sheet.cell(7, col).fill = PatternFill("solid", fgColor=color)
    sheet.row_dimensions[7].height = 22

    thin_bottom = Side(style="thin", color=BORDER)
    for col_idx, column in enumerate(columns, start=1):
        cell = sheet.cell(header_row, col_idx, column)
        cell.fill = PatternFill("solid", fgColor="F2F4F7")
        cell.font = Font(name="Yu Gothic", size=9, bold=True, color=TEXT)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=Side(style="medium", color=NAVY))
    sheet.row_dimensions[header_row].height = 42

    for row_offset, (_, record) in enumerate(df.iterrows(), start=first_data_row):
        is_area_total = str(record.get("施設名") or "") in AREA_TOTAL_LABELS
        status = str(record.get("状態") or "")
        row_fill = (
            PALE_BLUE if is_area_total else
            PALE_RED if "赤字予測" in status else
            PALE_ORANGE if "実績未入力" in status else None
        )
        for col_idx, column in enumerate(columns, start=1):
            cell = sheet.cell(row_offset, col_idx, _excel_value(column, record[column]))
            cell.font = Font(name="Yu Gothic", size=9, color=TEXT, bold=is_area_total)
            cell.border = Border(bottom=thin_bottom)
            cell.alignment = Alignment(
                horizontal="left" if column in {"施設名", "状態"} else "right",
                vertical="center",
            )
            if row_fill:
                cell.fill = PatternFill("solid", fgColor=row_fill)
            if column in AMOUNT_COLUMNS:
                cell.number_format = '#,##0;[Red]-#,##0;－'
            elif column in INTEGER_COLUMNS:
                cell.number_format = '#,##0;[Red]-#,##0;－'
            elif column in PERCENT_COLUMNS:
                cell.number_format = '0.0%;[Red]-0.0%;－'
        sheet.row_dimensions[row_offset].height = 21

    for col_idx, column in enumerate(columns, start=1):
        cell = sheet.cell(total_row, col_idx)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.font = Font(name="Yu Gothic", size=9, bold=True, color=WHITE)
        cell.border = Border(top=Side(style="medium", color=NAVY))
        cell.alignment = Alignment(horizontal="right", vertical="center")
        letter = get_column_letter(col_idx)
        if col_idx == 1:
            cell.value = "表示中の合計"
            cell.alignment = Alignment(horizontal="left", vertical="center")
        elif column in AMOUNT_COLUMNS or column in INTEGER_COLUMNS:
            cell.value = _formula_sum_rows(letter, facility_excel_rows)
            cell.number_format = '#,##0;[Red]-#,##0;－'
        elif column == "予定利益率":
            profit = get_column_letter(index_by_name["予定利益(千円)"])
            revenue = get_column_letter(index_by_name["予定売上(千円)"])
            profit_formula = _formula_sum_rows(profit, facility_excel_rows)[1:]
            revenue_formula = _formula_sum_rows(revenue, facility_excel_rows)[1:]
            cell.value = f'=IFERROR({profit_formula}/{revenue_formula},0)'
            cell.number_format = '0.0%'
        elif column == "着地予測利益率":
            profit = get_column_letter(index_by_name["着地予測利益(千円)"])
            revenue = get_column_letter(index_by_name["着地予測売上(千円)"])
            profit_formula = _formula_sum_rows(profit, facility_excel_rows)[1:]
            revenue_formula = _formula_sum_rows(revenue, facility_excel_rows)[1:]
            cell.value = f'=IFERROR({profit_formula}/{revenue_formula},0)'
            cell.number_format = '0.0%'
    sheet.row_dimensions[total_row].height = 24

    if len(df):
        table_ref = f"A{header_row}:{last_col_letter}{last_data_row}"
        table = Table(displayName="RevenueForecastTable", ref=table_ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        sheet.add_table(table)

        green_fill = PatternFill("solid", fgColor=PALE_GREEN)
        red_fill = PatternFill("solid", fgColor=PALE_RED)
        green_font = Font(color=DARK_GREEN, bold=True)
        red_font = Font(color=RED, bold=True)
        for column in ("利用回数差", "売上差(千円)", "利益差(千円)"):
            col_letter = get_column_letter(index_by_name[column])
            target = f"{col_letter}{first_data_row}:{col_letter}{last_data_row}"
            sheet.conditional_formatting.add(
                target, CellIsRule(operator="greaterThan", formula=["0"], fill=green_fill, font=green_font)
            )
            sheet.conditional_formatting.add(
                target, CellIsRule(operator="lessThan", formula=["0"], fill=red_fill, font=red_font)
            )
        for column in ("予定利益(千円)", "着地予測利益(千円)", "予定利益率", "着地予測利益率"):
            col_letter = get_column_letter(index_by_name[column])
            target = f"{col_letter}{first_data_row}:{col_letter}{last_data_row}"
            sheet.conditional_formatting.add(
                target, CellIsRule(operator="lessThan", formula=["0"], fill=red_fill, font=red_font)
            )
        status_letter = get_column_letter(index_by_name["状態"])
        sheet.conditional_formatting.add(
            f"{status_letter}{first_data_row}:{status_letter}{last_data_row}",
            FormulaRule(
                formula=[f'ISNUMBER(SEARCH("順調",{status_letter}{first_data_row}))'],
                fill=green_fill,
                font=green_font,
            ),
        )

    widths = {
        "施設名": 25, "月間予定延べ利用回数": 15, "月末着地予測利用回数": 17,
        "利用回数差": 12, "予定売上(千円)": 14, "着地予測売上(千円)": 16,
        "売上差(千円)": 13, "予定販管費(千円)": 15, "着地予測販管費(千円)": 17,
        "予定利益(千円)": 14, "着地予測利益(千円)": 16, "利益差(千円)": 13,
        "予定利益率": 12, "着地予測利益率": 14, "営業日数": 10, "－の日数": 10,
        "実績入力状況": 13, "最終更新日": 20, "状態": 24,
    }
    for col_idx, column in enumerate(columns, start=1):
        sheet.column_dimensions[get_column_letter(col_idx)].width = widths.get(column, 14)

    sheet.freeze_panes = "B9"
    sheet.auto_filter.ref = f"A{header_row}:{last_col_letter}{last_data_row}" if len(df) else None
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.print_title_rows = f"1:{header_row}"
    sheet.sheet_view.zoomScale = 85
    sheet.oddHeader.center.text = f"売上収支予測表 {target_ym}"
    sheet.oddFooter.right.text = "Page &P / &N"

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
