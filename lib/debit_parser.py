"""デビットカードExcel(`{法人}_デビット.xlsx`)のパーサ。

Dropbox: `00障がい事業部/{法人}_デビット.xlsx`
シート名: `2605【デビット】` 等 (YYMM)
列構成:
  A: 日付
  B: 借方勘定科目
  C: 借方税区分
  D: 借方品目
  E: 借方部門          (例: '5.1.SORATOてんり', '01　のじぎく高砂')
  F: 借方金額
  G: 貸方勘定科目      (例: '関西みらい銀行')
  H: 貸方品目
  I: 貸方部門
  J: 貸方金額
  K: 摘要              (例: '㈱吉田石油　ガソリン代（送迎車） / 軽油 49.22L @143¥7,038')
  L: 検証

摘要の構造:
  `{購入先(vendor)}　{用途(purpose)} / {品目1}¥{価格1}, {品目2}¥{価格2}, ...`
"""
from __future__ import annotations

import hashlib
import io
import re
from datetime import datetime, date
from typing import Iterable

import openpyxl

from lib.journal_parser import _strip_dept_prefix, _normalize_dept


# シート名の年月パターン  例: '2605【デビット】' → 2026-05
_SHEET_YM_RE = re.compile(r'(\d{2})(\d{2})')

# 摘要 → 品目分解  例: 'ポテコ¥1,746' → name='ポテコ', price=1746
_ITEM_PRICE_RE = re.compile(r'(.+?)[¥￥]([0-9,]+)\s*$')


def _safe_int(v) -> int:
    if v is None:
        return 0
    if isinstance(v, (int, float)):
        return int(v)
    s = (str(v).strip()
         .replace(',', '')
         .replace('¥', '').replace('￥', '')
         .replace('＄', '').replace('$', '')   # 全角・半角ドル記号
         .replace('€', '').replace('￡', '').replace('£', ''))  # その他通貨記号
    if not s:
        return 0
    try:
        return int(float(s))
    except ValueError:
        return 0


def parse_sheet_year_month(sheet_name: str) -> str | None:
    """'2605【デビット】' → '2026-05'。判定不能なら None。"""
    if not sheet_name:
        return None
    m = _SHEET_YM_RE.search(sheet_name)
    if not m:
        return None
    yy = int(m.group(1))
    mm = int(m.group(2))
    if not (1 <= mm <= 12):
        return None
    # 50以下 → 2000年代、それ以外 → 1900年代
    year = 2000 + yy if yy <= 50 else 1900 + yy
    return f"{year:04d}-{mm:02d}"


def _to_date_str(v) -> str | None:
    """セル値→ 'YYYY-MM-DD' 文字列。"""
    if v is None or v == '':
        return None
    if isinstance(v, datetime):
        return v.strftime('%Y-%m-%d')
    if isinstance(v, date):
        return v.strftime('%Y-%m-%d')
    s = str(v).strip()
    # 'YYYY/MM/DD' or 'YYYY-MM-DD' 等を正規化
    s = s.replace('/', '-')
    m = re.match(r'^(\d{4})-(\d{1,2})-(\d{1,2})', s)
    if m:
        y, mo, d = map(int, m.groups())
        try:
            return f"{y:04d}-{mo:02d}-{d:02d}"
        except Exception:
            return None
    return None


def parse_description(desc: str) -> dict:
    """摘要から { 'vendor', 'purpose', 'items': [{name, price}], 'items_text' } を抽出。

    例:
      '㈱吉田石油　ガソリン代（送迎車） / 軽油 49.22L @143¥7,038'
      → vendor='㈱吉田石油', purpose='ガソリン代（送迎車）',
         items=[{'name': '軽油 49.22L @143', 'price': 7038}]
    """
    if not desc:
        return {'vendor': '', 'purpose': '', 'items': [], 'items_text': ''}

    text = str(desc).strip()
    # 左右を ' / ' で分離（最初の出現のみ）
    if ' / ' in text:
        left, right = text.split(' / ', 1)
    else:
        left, right = text, ''

    # 左側: 全角空白（　）で vendor / purpose を分離。
    # 半角スペースは vendor 名内部にも使われる ("Claude Code", "フレッシュ石守 稲美店" 等)
    # ため、全角スペースのみをセパレータとする。
    # 全角スペースがない場合は先頭が vendor 全体。
    left = left.strip()
    if '　' in left:  # 全角スペース
        vendor, purpose = left.split('　', 1)
        vendor = vendor.strip()
        purpose = purpose.strip()
    else:
        vendor = left.strip()
        purpose = ''

    # 右側: ', ' (カンマ+空白) でアイテム分解。
    # 価格内のカンマ (例 '¥7,038') を区切りと誤認しないため、空白付きカンマで分ける。
    items: list[dict] = []
    if right:
        # フォールバック: '、' 全角カンマ も区切りに含める
        chunks = re.split(r',\s+|、', right)
        for chunk in chunks:
            chunk = chunk.strip().rstrip(',').strip()
            if not chunk:
                continue
            m = _ITEM_PRICE_RE.search(chunk)
            if m:
                name = m.group(1).strip()
                price = _safe_int(m.group(2))
                items.append({'name': name, 'price': price})
            else:
                items.append({'name': chunk, 'price': 0})

    return {
        'vendor': vendor,
        'purpose': purpose,
        'items': items,
        'items_text': right.strip(),
    }


def parse_debit_workbook(file_or_path,
                         corporation: str,
                         subunit_lookup: dict[str, int] | None = None) -> dict:
    """デビットExcelをパースし、行リストとメタ情報を返す。

    Args:
        file_or_path: バイト/ファイルオブジェクト/パス
        corporation: '医療法人' or 'NPO法人' 等のラベル
        subunit_lookup: 借方部門名 → subunit_id ルックアップ

    Returns:
        {
            'rows': [...],            # DB INSERT 用
            'file_hash': str,
            'sheet_summaries': [{'sheet': str, 'year_month': str|None, 'rows': int}],
            'unknown_departments': set[str],
            'errors': [str],
        }
    """
    sub_lookup = subunit_lookup or {}

    # バイト取得とハッシュ
    if hasattr(file_or_path, 'read'):
        data = file_or_path.read()
        if isinstance(data, str):
            data = data.encode('utf-8')
    else:
        with open(file_or_path, 'rb') as f:
            data = f.read()

    file_hash = hashlib.sha256(data).hexdigest()
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)

    rows: list[dict] = []
    summaries: list[dict] = []
    unknown_depts: set[str] = set()
    errors: list[str] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        ym = parse_sheet_year_month(sheet_name)
        sheet_rows = 0

        for r_idx in range(2, ws.max_row + 1):
            try:
                date_v = ws.cell(row=r_idx, column=1).value
                tx_date = _to_date_str(date_v)
                if not tx_date:
                    continue

                debit_acc = (ws.cell(row=r_idx, column=2).value or '')
                tax_class = (ws.cell(row=r_idx, column=3).value or '')
                debit_item = (ws.cell(row=r_idx, column=4).value or '')
                debit_dept_raw = (ws.cell(row=r_idx, column=5).value or '')
                amount = _safe_int(ws.cell(row=r_idx, column=6).value)
                credit_acc = (ws.cell(row=r_idx, column=7).value or '')
                description = (ws.cell(row=r_idx, column=11).value or '')
                check = (ws.cell(row=r_idx, column=12).value or '')

                debit_acc = str(debit_acc).strip()
                tax_class = str(tax_class).strip()
                debit_item = str(debit_item).strip()
                debit_dept_raw = str(debit_dept_raw).strip()
                credit_acc = str(credit_acc).strip()
                description = str(description).strip()
                check = str(check).strip()

                # 空行（金額0かつ科目なし）はスキップ
                if amount == 0 and not debit_acc and not description:
                    continue

                # 部門の解決
                dept_clean = _strip_dept_prefix(debit_dept_raw)
                subunit_id = None
                for cand in [debit_dept_raw, dept_clean, _normalize_dept(debit_dept_raw)]:
                    if cand and cand in sub_lookup:
                        subunit_id = sub_lookup[cand]
                        break
                if subunit_id is None and debit_dept_raw:
                    unknown_depts.add(debit_dept_raw)

                # 摘要解析
                parsed = parse_description(description)

                rows.append({
                    'corporation': corporation,
                    'transaction_date': tx_date,
                    'year_month': tx_date[:7],
                    'debit_account': debit_acc,
                    'tax_class': tax_class,
                    'debit_item': debit_item,
                    'department_raw': debit_dept_raw,
                    'department_clean': dept_clean,
                    'subunit_id': subunit_id,
                    'amount': amount,
                    'credit_account': credit_acc,
                    'description': description,
                    'vendor': parsed['vendor'],
                    'purpose': parsed['purpose'],
                    'items_text': parsed['items_text'],
                    'items_count': len(parsed['items']),
                    'check_status': check,
                    'sheet_name': sheet_name,
                })
                sheet_rows += 1
            except Exception as e:
                errors.append(f"{sheet_name} 行{r_idx}: {e}")

        summaries.append({
            'sheet': sheet_name, 'year_month': ym, 'rows': sheet_rows,
        })

    return {
        'rows': rows,
        'file_hash': file_hash,
        'sheet_summaries': summaries,
        'unknown_departments': unknown_depts,
        'errors': errors,
    }


def collect_items(rows: Iterable[dict]) -> list[dict]:
    """全行から品目を展開して返す。各行の `description` をパースし
    [{transaction_date, vendor, item_name, price, debit_account, department_raw, subunit_id}, ...] を返す。
    """
    out: list[dict] = []
    for r in rows:
        parsed = parse_description(r.get('description') or '')
        for it in parsed['items']:
            out.append({
                'transaction_date': r.get('transaction_date'),
                'year_month': r.get('year_month'),
                'corporation': r.get('corporation'),
                'vendor': parsed['vendor'],
                'purpose': parsed['purpose'],
                'item_name': it['name'],
                'price': it['price'],
                'debit_account': r.get('debit_account'),
                'department_raw': r.get('department_raw'),
                'department_clean': r.get('department_clean'),
                'subunit_id': r.get('subunit_id'),
            })
    return out
