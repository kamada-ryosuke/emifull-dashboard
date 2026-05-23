"""現金立替清算 (現金出納帳 / 一般経費精算書) パーサ & 書込モジュール
"""
from __future__ import annotations

import hashlib
import io
import json
import re
from datetime import datetime, date
from pathlib import Path

import openpyxl

from lib.journal_parser import _strip_dept_prefix, _normalize_dept


# シート名の年月パターン (例: '2604【出納帳】' → 2026-04)
_SHEET_YM_RE = re.compile(r'(\d{2})(\d{2})')

# 出納帳シート判定: '【出納帳】' を含むシート名
CASHBOOK_SHEET_MARKER = '【出納帳】'

# 出納帳の列インデックス (1-indexed)
CASHBOOK_COL = {
    'date': 1,           # A
    'item': 4,           # D
    'account': 6,        # F
    'tax_class': 7,      # G
    'department': 8,     # H
    'description': 9,    # I
    'amount_in': 10,     # J
    'amount_out': 11,    # K
    'balance': 12,       # L
}

# データ開始行 (前月繰越が r8 なので r9 から)
CASHBOOK_DATA_START_ROW = 9


# ============================================================
# 施設マスタ (facilities.json) ヘルパー
# ============================================================

_FACILITIES_PATH = Path(__file__).resolve().parent.parent / 'config' / 'facilities.json'


def _load_facilities_master() -> dict:
    if not _FACILITIES_PATH.exists():
        return {}
    try:
        return json.loads(_FACILITIES_PATH.read_text(encoding='utf-8'))
    except Exception:
        return {}


def list_facility_codes(corp_label: str | None = None) -> list[str]:
    """facilities.json から code リストを取得。
    corp_label = '医療法人' / 'NPO法人' / None(=全部)。
    """
    master = _load_facilities_master()
    out: list[str] = []
    for key, body in master.items():
        if not isinstance(body, dict) or 'facilities' not in body:
            continue
        if corp_label is not None:
            corp_name = body.get('_corp_name', '')
            if corp_label == '医療法人' and not corp_name.startswith('医療法人'):
                continue
            if corp_label == 'NPO法人' and not (
                corp_name.startswith('NPO') or
                corp_name.startswith('特定非営利') or
                'NPO' in corp_name
            ):
                continue
        for f in body['facilities']:
            code = f.get('code')
            if code:
                out.append(code)
    return out


def list_facility_codes_by_corp() -> dict[str, list[str]]:
    """{'医療法人': [...], 'NPO法人': [...]} を返す。"""
    return {
        '医療法人': list_facility_codes('医療法人'),
        'NPO法人': list_facility_codes('NPO法人'),
    }


def ensure_facility_folders(root_dir: Path,
                             include_status_dirs: bool = True) -> dict:
    """root_dir 配下に facilities.json で定義された施設フォルダを作成。

    Args:
        root_dir: 作成先 (例: G:/.../02.現金立替レシート)
        include_status_dirs: '_処理済み' '_要確認' を作るか

    Returns:
        {'created': [...], 'existing': [...], 'errors': [...]}
    """
    created: list[str] = []
    existing: list[str] = []
    errors: list[str] = []
    if not root_dir.exists():
        try:
            root_dir.mkdir(parents=True, exist_ok=True)
        except Exception as e:
            return {'created': [], 'existing': [], 'errors': [f'{root_dir}: {e}']}

    targets: list[str] = list_facility_codes(None)
    if include_status_dirs:
        targets += ['_処理済み', '_要確認']

    for name in targets:
        d = root_dir / name
        try:
            if d.exists():
                existing.append(name)
            else:
                d.mkdir(parents=True, exist_ok=True)
                created.append(name)
        except Exception as e:
            errors.append(f'{name}: {e}')
    return {'created': created, 'existing': existing, 'errors': errors}


def detect_corp_from_facility_code(facility_code: str) -> str | None:
    """施設コードから法人ラベルを返す。"""
    if not facility_code:
        return None
    master = _load_facilities_master()
    for key, body in master.items():
        if not isinstance(body, dict) or 'facilities' not in body:
            continue
        for f in body['facilities']:
            if f.get('code') == facility_code:
                corp_name = body.get('_corp_name', '')
                if corp_name.startswith('医療法人'):
                    return '医療法人'
                if (corp_name.startswith('NPO') or
                    corp_name.startswith('特定非営利') or
                    'NPO' in corp_name):
                    return 'NPO法人'
                return corp_name
    return None


# ============================================================
# 出納帳パース (xlsm の【出納帳】シート → DB行)
# ============================================================

def _safe_int(v) -> int:
    if v is None:
        return 0
    if isinstance(v, (int, float)):
        return int(v)
    s = str(v).strip().replace(',', '').replace('¥', '').replace('￥', '')
    if not s:
        return 0
    try:
        return int(float(s))
    except ValueError:
        return 0


def _to_date_str(v) -> str | None:
    if v is None or v == '':
        return None
    if isinstance(v, datetime):
        return v.strftime('%Y-%m-%d')
    if isinstance(v, date):
        return v.strftime('%Y-%m-%d')
    s = str(v).strip().replace('/', '-')
    m = re.match(r'^(\d{4})-(\d{1,2})-(\d{1,2})', s)
    if m:
        y, mo, d = map(int, m.groups())
        return f"{y:04d}-{mo:02d}-{d:02d}"
    return None


def parse_sheet_year_month(sheet_name: str) -> str | None:
    """'2604【出納帳】' → '2026-04'"""
    if not sheet_name:
        return None
    m = _SHEET_YM_RE.search(sheet_name)
    if not m:
        return None
    yy = int(m.group(1))
    mm = int(m.group(2))
    if not (1 <= mm <= 12):
        return None
    year = 2000 + yy if yy <= 50 else 1900 + yy
    return f"{year:04d}-{mm:02d}"


def detect_corporation_from_filename(filename: str) -> str:
    """ファイル名から法人を推定。'EMI' or '医療法人' を含めば医療法人、'のじ' or 'NPO' なら NPO法人。"""
    if not filename:
        return ''
    name = filename
    low = name.lower()
    if 'のじ' in name or 'npo' in low:
        return 'NPO法人'
    if 'EMI' in name or '医療法人' in name or '医療' in name:
        return '医療法人'
    return ''


def _parse_description(description: str) -> dict:
    """備考(I列)から vendor / purpose / payee を推定 (best-effort)。

    例: '4月30日　ガソリン代立替分　鎌田亮介'
        → vendor='', purpose='ガソリン代立替分', payee='鎌田亮介'
    例: '㈱エイプルパーキング　駐車料金'
        → vendor='㈱エイプルパーキング', purpose='駐車料金', payee=''
    """
    if not description:
        return {'vendor': '', 'purpose': '', 'payee': ''}
    # 全角空白で分割
    text = str(description).strip()
    parts = re.split(r'[　]+', text)  # 全角空白
    # 先頭のトークンが「N月N日」なら除去 (日付情報は別カラム参照)
    if parts and re.match(r'^\d{1,2}月\d{1,2}日$', parts[0]):
        parts = parts[1:]
    # 末尾のトークンが氏名っぽい (漢字2-4文字 + 漢字1-2文字 のような) なら payee として抜く
    payee = ''
    if parts:
        last = parts[-1].strip()
        if re.match(r'^[一-龥ぁ-んァ-ヴー]{2,8}$', last) and len(last) <= 8 and last != parts[0]:
            payee = last
            parts = parts[:-1]
    # 残りを 1 or 2 トークンに分割
    rest = parts
    if len(rest) >= 2:
        vendor = rest[0]
        purpose = '　'.join(rest[1:])
    elif len(rest) == 1:
        vendor = ''
        purpose = rest[0]
    else:
        vendor = ''
        purpose = ''
    return {'vendor': vendor, 'purpose': purpose, 'payee': payee}


def _classify_entry_kind(description: str, account: str,
                         amount_in: int, amount_out: int) -> str:
    """エントリ種別を粗く分類。
    - opening: 前月繰越 (金額0で備考が前月繰越)
    - income: 入金 > 0
    - expense: 出金 > 0 + 勘定科目あり
    - transfer: それ以外 (振替・残高調整)
    """
    desc = (description or '').strip()
    if '前月より繰越' in desc or '前月繰越' in desc:
        return 'opening'
    if amount_in > 0 and amount_out == 0:
        return 'income'
    if amount_out > 0 and amount_in == 0:
        return 'expense'
    return 'transfer'


def parse_cashbook_workbook(file_or_path,
                             corporation: str,
                             subunit_lookup: dict | None = None) -> dict:
    """xlsm/xlsx から `*【出納帳】` シートを全部読み出す。

    Returns:
        {
            'rows': [...],           # cash_advance_entries 用
            'file_hash': str,
            'sheet_summaries': [{'sheet': str, 'year_month': str|None, 'rows': int}],
            'unknown_departments': set,
            'errors': [str],
        }
    """
    sub_lookup = subunit_lookup or {}

    if hasattr(file_or_path, 'read'):
        data = file_or_path.read()
        if isinstance(data, str):
            data = data.encode('utf-8')
    else:
        with open(file_or_path, 'rb') as f:
            data = f.read()

    file_hash = hashlib.sha256(data).hexdigest()
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True, keep_vba=False)

    rows: list[dict] = []
    summaries: list[dict] = []
    unknown_depts: set[str] = set()
    errors: list[str] = []

    for sheet_name in wb.sheetnames:
        if CASHBOOK_SHEET_MARKER not in sheet_name:
            continue
        ws = wb[sheet_name]
        sheet_ym = parse_sheet_year_month(sheet_name)
        sheet_rows = 0

        for r_idx in range(CASHBOOK_DATA_START_ROW, ws.max_row + 1):
            try:
                date_v = ws.cell(row=r_idx, column=CASHBOOK_COL['date']).value
                tx_date = _to_date_str(date_v)
                amount_in = _safe_int(ws.cell(row=r_idx, column=CASHBOOK_COL['amount_in']).value)
                amount_out = _safe_int(ws.cell(row=r_idx, column=CASHBOOK_COL['amount_out']).value)
                debit_acc = (ws.cell(row=r_idx, column=CASHBOOK_COL['account']).value or '')
                tax_class = (ws.cell(row=r_idx, column=CASHBOOK_COL['tax_class']).value or '')
                debit_item = (ws.cell(row=r_idx, column=CASHBOOK_COL['item']).value or '')
                debit_dept = (ws.cell(row=r_idx, column=CASHBOOK_COL['department']).value or '')
                description = (ws.cell(row=r_idx, column=CASHBOOK_COL['description']).value or '')

                # 全部空ならスキップ
                if (not tx_date and amount_in == 0 and amount_out == 0
                        and not str(description).strip() and not str(debit_acc).strip()):
                    continue

                # 日付なし行はスキップ (入出金合計欄など)
                if not tx_date:
                    continue

                debit_acc = str(debit_acc).strip()
                tax_class = str(tax_class).strip()
                debit_item = str(debit_item).strip()
                debit_dept_raw = str(debit_dept).strip()
                description = str(description).strip()

                dept_clean = _strip_dept_prefix(debit_dept_raw) if debit_dept_raw else ''
                subunit_id = None
                for cand in [debit_dept_raw, dept_clean, _normalize_dept(debit_dept_raw)]:
                    if cand and cand in sub_lookup:
                        subunit_id = sub_lookup[cand]
                        break
                if subunit_id is None and debit_dept_raw:
                    unknown_depts.add(debit_dept_raw)

                parsed = _parse_description(description)
                kind = _classify_entry_kind(description, debit_acc, amount_in, amount_out)

                rows.append({
                    'corporation': corporation,
                    'transaction_date': tx_date,
                    'year_month': tx_date[:7],
                    'debit_account': debit_acc,
                    'tax_class': tax_class,
                    'debit_item': debit_item,
                    'department_raw': debit_dept_raw,
                    'department_clean': dept_clean,
                    'subunit_id': subunit_id,
                    'amount_in': amount_in,
                    'amount_out': amount_out,
                    'description': description,
                    'vendor': parsed['vendor'],
                    'purpose': parsed['purpose'],
                    'payee': parsed['payee'],
                    'entry_kind': kind,
                    'split_total_amount': None,
                    'split_facility_count': 1,
                    'sheet_name': sheet_name,
                    'receipt_id': None,
                })
                sheet_rows += 1
            except Exception as e:
                errors.append(f"{sheet_name} 行{r_idx}: {e}")

        summaries.append({
            'sheet': sheet_name, 'year_month': sheet_ym, 'rows': sheet_rows,
        })

    return {
        'rows': rows,
        'file_hash': file_hash,
        'sheet_summaries': summaries,
        'unknown_departments': unknown_depts,
        'errors': errors,
    }


# ============================================================
# 出納帳 Excel 書込
# ============================================================

def _ymd_to_sheet_name(ym: str) -> str:
    """'2026-05' → '2605【出納帳】'"""
    y, m = ym.split('-')
    return f"{y[2:]}{m}{CASHBOOK_SHEET_MARKER}"


def _find_yyyymm_sheet(wb, ym: str) -> str | None:
    """'2026-05' に対応する出納帳シート名を返す。"""
    yy = ym.split('-')[0][2:]
    mm = ym.split('-')[1]
    target = f"{yy}{mm}"
    for sn in wb.sheetnames:
        if target in sn and CASHBOOK_SHEET_MARKER in sn:
            return sn
    return None


def append_cashbook_rows(xlsx_path, rows: list[dict]) -> dict:
    """出納帳 xlsx に複数行を追記。各 row は cash_advance_entries の dict。

    シートが存在しなければ作らずエラー (フォーマット崩れ防止)。
    日付降順ソートは行わない (Excel のフォーマット保持優先で末尾追加)。

    Returns:
        {'ok': bool, 'sheet': str, 'appended': int, 'error': str|None}
    """
    if not rows:
        return {'ok': True, 'sheet': '', 'appended': 0, 'error': None}

    # 全行が同月か確認
    yms = sorted({r['year_month'] for r in rows if r.get('year_month')})
    if len(yms) != 1:
        return {'ok': False, 'sheet': '', 'appended': 0,
                'error': f'複数月混在は非対応: {yms}'}
    ym = yms[0]

    try:
        wb = openpyxl.load_workbook(str(xlsx_path), keep_vba=True)
    except Exception as e:
        return {'ok': False, 'sheet': '', 'appended': 0,
                'error': f'Excel読込失敗: {e}'}

    sheet_name = _find_yyyymm_sheet(wb, ym)
    if sheet_name is None:
        return {'ok': False, 'sheet': '', 'appended': 0,
                'error': f'出納帳シートが見つかりません: {ym}'}

    ws = wb[sheet_name]
    # 末尾の最終データ行を探す
    last_row = CASHBOOK_DATA_START_ROW - 1
    for r_idx in range(CASHBOOK_DATA_START_ROW, ws.max_row + 1):
        if ws.cell(row=r_idx, column=CASHBOOK_COL['date']).value is not None:
            last_row = r_idx
        else:
            # 日付セルがあるが他列に値がある可能性も - 念のため
            if any(ws.cell(row=r_idx, column=c).value is not None
                   for c in CASHBOOK_COL.values()):
                last_row = r_idx
    write_row = last_row + 1

    # 直近行から残高を取得 (差引残高 = L列)
    prev_balance = _safe_int(
        ws.cell(row=last_row, column=CASHBOOK_COL['balance']).value
    )
    appended = 0
    for r in rows:
        try:
            d = r.get('transaction_date')
            if isinstance(d, str):
                d = datetime.strptime(d, '%Y-%m-%d')
            ws.cell(row=write_row, column=CASHBOOK_COL['date']).value = d
            ws.cell(row=write_row, column=CASHBOOK_COL['item']).value = r.get('debit_item') or ''
            ws.cell(row=write_row, column=CASHBOOK_COL['account']).value = r.get('debit_account') or ''
            ws.cell(row=write_row, column=CASHBOOK_COL['tax_class']).value = r.get('tax_class') or ''
            ws.cell(row=write_row, column=CASHBOOK_COL['department']).value = r.get('department_raw') or ''
            ws.cell(row=write_row, column=CASHBOOK_COL['description']).value = r.get('description') or ''
            ai = _safe_int(r.get('amount_in'))
            ao = _safe_int(r.get('amount_out'))
            ws.cell(row=write_row, column=CASHBOOK_COL['amount_in']).value = ai if ai > 0 else None
            ws.cell(row=write_row, column=CASHBOOK_COL['amount_out']).value = ao if ao > 0 else None
            prev_balance = prev_balance + ai - ao
            ws.cell(row=write_row, column=CASHBOOK_COL['balance']).value = prev_balance
            write_row += 1
            appended += 1
        except Exception as e:
            wb.close()
            return {'ok': False, 'sheet': sheet_name, 'appended': appended,
                    'error': f'書込失敗 行{write_row}: {e}'}

    try:
        wb.save(str(xlsx_path))
    except Exception as e:
        return {'ok': False, 'sheet': sheet_name, 'appended': appended,
                'error': f'保存失敗 (Excel側で開いていませんか?): {e}'}
    finally:
        wb.close()

    return {'ok': True, 'sheet': sheet_name, 'appended': appended, 'error': None}
