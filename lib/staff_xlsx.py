"""職員台帳v1.xlsx を Streamlit 用 DataFrame に変換するローダー

データソース: C:\\売上入金管理ツール\\output\\職員台帳整備\\職員台帳v1.xlsx
Notion 互換のスキーマ (load_seishain_df / load_paato_df 互換) を提供する。
"""
from __future__ import annotations

import re
from pathlib import Path
from typing import Optional

import openpyxl
import pandas as pd

DAICHO_PATH = Path(r'C:\売上入金管理ツール\output\職員台帳整備\職員台帳v1.xlsx')


def _to_int_or_none(v):
    if v is None: return None
    if isinstance(v, bool): return None
    if isinstance(v, (int, float)):
        try: return int(v)
        except: return None
    s = str(v).strip().replace(',', '')
    s = re.sub(r'[^\d\-]', '', s)
    return int(s) if s and s != '-' else None


def _split_list(v) -> list:
    """資格名称等のスペース/カンマ区切り文字列をリスト化"""
    if not v: return []
    if isinstance(v, list): return v
    s = str(v).strip()
    if not s: return []
    parts = re.split(r'[ 　、,／/]+', s)
    return [p for p in parts if p]


def _parse_tenure(v) -> Optional[float]:
    """「3年4ヶ月」→ 3.33 / 数値→そのまま"""
    if v is None: return None
    if isinstance(v, (int, float)): return float(v)
    s = str(v)
    m_y = re.search(r'(\d+)\s*年', s)
    m_m = re.search(r'(\d+)\s*ヶ?月', s)
    if m_y or m_m:
        years = int(m_y.group(1)) if m_y else 0
        months = int(m_m.group(1)) if m_m else 0
        return round(years + months / 12, 1)
    try: return float(s)
    except: return None


def _normalize_date(v) -> str:
    """生年月日・入社日を YYYY-MM-DD 形式に正規化(できなければ元の文字列)"""
    if v is None: return ''
    if hasattr(v, 'strftime'): return v.strftime('%Y-%m-%d')
    s = str(v).strip()
    # 和暦変換: S60.6.6 / R5.4.1 / H1.4.4 など
    m = re.match(r'^([SHRMshrm昭平令])\s*(\d{1,2})[\.\/年]\s*(\d{1,2})[\.\/月]\s*(\d{1,2})', s)
    if m:
        era_char = m.group(1).upper()
        y = int(m.group(2))
        mo = int(m.group(3))
        d = int(m.group(4))
        if era_char == 'S' or era_char == '昭': ad = 1925 + y
        elif era_char == 'H' or era_char == '平': ad = 1988 + y
        elif era_char == 'R' or era_char == '令': ad = 2018 + y
        else: ad = y
        return f'{ad:04d}-{mo:02d}-{d:02d}'
    # YYYY/MM/DD or YYYY-MM-DD
    m = re.match(r'^(\d{4})[\.\/年\-]\s*(\d{1,2})[\.\/月\-]\s*(\d{1,2})', s)
    if m:
        return f'{int(m.group(1)):04d}-{int(m.group(2)):02d}-{int(m.group(3)):02d}'
    return s


def load_daicho_df() -> pd.DataFrame:
    """①現役職員台帳 → DataFrame (Streamlit互換スキーマ)"""
    if not DAICHO_PATH.exists():
        return pd.DataFrame()
    wb = openpyxl.load_workbook(DAICHO_PATH, data_only=True)
    ws = wb['①現役職員台帳']
    HEADER = 5
    headers = [ws.cell(row=HEADER, column=c).value for c in range(1, ws.max_column + 1)]
    idx = {h: i for i, h in enumerate(headers) if h}

    rows = []
    for r in range(HEADER + 1, ws.max_row + 1):
        cell_values = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        if not cell_values[idx.get('氏名', 0)]:
            continue

        def gv(col):
            i = idx.get(col)
            return cell_values[i] if i is not None else None

        # 雇用区分判定: 'パートナー職員', '時給制' は時給扱い、それ以外は月給
        koyou = gv('雇用区分') or '正社員'
        kyuyo = gv('給与体系') or ''
        is_part = ('パート' in str(koyou)) or ('時給' in str(kyuyo))

        # 所属(list化)
        sho = gv('所属')
        kenmu = gv('兼務')
        facilities = []
        if sho: facilities.append(str(sho).strip())
        if kenmu: facilities.append(str(kenmu).strip())

        rec = {
            # 基本情報
            '職員番号': _to_int_or_none(gv('職員番号')),
            '氏名': str(gv('氏名') or '').strip(),
            'フリガナ': gv('フリガナ') or '',
            '所属施設': facilities,
            '雇用区分': '正社員' if not is_part else 'パート',
            '職種': gv('職種') or '',
            '役職': gv('役職') or '',
            '等級': gv('等級') or '',
            '号棒': gv('号俸'),  # ※xlsxは「号俸」、Notion互換は「号棒」
            '住所': gv('住所') or '',
            '電話番号': gv('電話') or '',
            'メールアドレス': gv('メール') or '',
            '生年月日': _normalize_date(gv('生年月日')),
            '年齢': _to_int_or_none(gv('年齢')),
            '入社日': _normalize_date(gv('入職年月日')),
            '退職日': '',  # 現役シートなので空
            '勤続年数': _parse_tenure(gv('勤続年数')),
            '保有資格': _split_list(gv('資格名称')),
            'ステータス': '在職中',
            'メモ': gv('備考') or '',
            'notion_url': '',  # Notion未連携
            # 給与系(月給制)
            '基本給': _to_int_or_none(gv('基本給')),
            '職位手当': _to_int_or_none(gv('職位手当')),
            '役職手当': _to_int_or_none(gv('役職手当')),
            '地域手当': _to_int_or_none(gv('地域手当')),
            '業務手当': _to_int_or_none(gv('業務手当')),
            '保育手当': _to_int_or_none(gv('保育手当')),
            '住宅手当': _to_int_or_none(gv('住宅手当')),
            '処遇改善手当': _to_int_or_none(gv('臨時手当（処遇改善金）')),
            '資格手当': _to_int_or_none(gv('資格手当')),
            '年収保証手当': _to_int_or_none(gv('年収保証手当/年')),
            '月給合計': _to_int_or_none(gv('月給')),
            '年収概算': _to_int_or_none(gv('年収')),
            # 給与系(時給制) — 現状xlsxには時給細分列なし
            '基本時給①': None,
            '資格時給②': None,
            '処遇改善時給③': None,
            '時給合計': None,
            # 給与体系・契約書・履歴書
            '給与体系': gv('給与体系') or '',
            '給与体系履歴': gv('給与体系履歴') or '',
            '雇用契約書ファイル': gv('雇用契約書ファイル') or '',
            '履歴書ファイル': gv('履歴書ファイル') or '',
            '保育経験(年)': gv('保育経験(年)'),
            '児童経験(年)': gv('児童経験(年)'),
            '保育5年以上': gv('保育5年以上') or '',
            '児童5年以上': gv('児童5年以上') or '',
            '保育10年以上': gv('保育10年以上') or '',
            '児童10年以上': gv('児童10年以上') or '',
            '経歴メモ': gv('経歴メモ') or '',
            '_source_xlsx': str(DAICHO_PATH),
        }
        rows.append(rec)

    return pd.DataFrame(rows)


def load_seishain_df() -> pd.DataFrame:
    """正社員(月給制) — Notion互換"""
    df = load_daicho_df()
    if df.empty: return df
    return df[df['雇用区分'] == '正社員'].reset_index(drop=True)


def load_paato_df() -> pd.DataFrame:
    """パート(時給制) — Notion互換"""
    df = load_daicho_df()
    if df.empty: return df
    return df[df['雇用区分'] != '正社員'].reset_index(drop=True)
