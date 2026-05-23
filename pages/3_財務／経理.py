"""財務／経理 - デビットカード & 現金立替レシート

階層構成:
  💳 デビットカード      - デビット明細・レシート (既存)
    └ 取込 / サマリ / 施設別 / 勘定科目別 / 購入先別 / 購入明細 / レシート
  💴 現金立替レシート    - 現金立替清算 (出納帳 / 経費精算書)
    └ 取込 / サマリ / 施設別 / 勘定科目別 / 購入先別 / 出納帳
        / レシート / 経費精算書出力
"""
from __future__ import annotations

import io
from datetime import datetime
from pathlib import Path

import openpyxl
import pandas as pd
import streamlit as st

from lib import (
    db, styling, auth, debit_parser, receipt_ocr, excel_writer,
    receipt_processor, cash_advance_parser, expense_form,
)

# HEIC/HEIF を Pillow で開けるようにする（iPhone 写真対応）
# 毎回の呼び出しで pillow_heif を ensure する方式。
# Streamlit 起動後に `pip install pillow-heif` した場合も次の呼び出しで効く。
def _ensure_heif_opener() -> bool:
    """pillow_heif の register_heif_opener を実行。成否を返す。

    モジュール起動時には pillow_heif が無くても、後からインストールされていれば
    動的にここで読み込めるよう、毎回 import を試みる。"""
    try:
        from pillow_heif import register_heif_opener
        register_heif_opener()
        return True
    except Exception:
        return False


# 起動時に一度だけ試す (グローバル状態の参照用)
_HEIF_OK = _ensure_heif_opener()

styling.inject_global_css()
auth.require_admin()
auth.render_sidebar_navigation()

st.title("財務／経理")
st.markdown(
    "<p style='color:#64748b; font-size:15px;'>"
    "デビット明細と現金立替レシートを取り込み、施設別・勘定科目別・購入先別に費用を可視化します。"
    "</p>",
    unsafe_allow_html=True,
)

# ============================================================
# 親タブ: デビットカード / 現金立替レシート
# ============================================================
parent_debit, parent_cash = st.tabs([
    "💳 デビットカード", "💴 現金立替レシート",
])


# ============================================================
# 共通ヘルパー
# ============================================================

# 📂 デビット明細・レシート格納先（Google Drive for Desktop）
# https://drive.google.com/drive/folders/1FToAeulgwYV58yNMlTeMJ2eekoBlUfnL
#
# 構造のバリエーション:
#   旧: …\管理者提出物\01.デビッドカード清算\           (Excel と レシート同階層)
#   新: …\管理者提出物\01.経費処理フォルダ\              (Excel)
#                                  └ 01.デビッドカード清算\  (レシート)
GDRIVE_PARENT = Path(r"G:\マイドライブ\管理者提出物")
RECEIPT_FOLDER_PATTERN = "デビッドカード清算"


def _load_debit_folder_config() -> dict:
    """drive_config.json からデビットフォルダ参照情報を読み込む。"""
    cfg_path = Path(__file__).resolve().parent.parent / 'config' / 'drive_config.json'
    out = {'url': None, 'id': None, 'name': None}
    if cfg_path.exists():
        try:
            import json
            cfg = json.loads(cfg_path.read_text(encoding='utf-8'))
            out['url'] = cfg.get('debit_receipt_folder_url')
            out['id'] = cfg.get('debit_receipt_folder_id')
            out['name'] = cfg.get('debit_receipt_folder_name')
        except Exception:
            pass
    return out


DEBIT_FOLDER_REF = _load_debit_folder_config()

# 後方互換のため旧 Dropbox パスも保持（フォールバック用）
DROPBOX_LEGACY_DIR = Path(
    r"C:\Users\user01\株式会社ＥＭＩＦＵＬＬ Dropbox"
    r"\医療法人社団EMIFULL\障がい事業部\00障がい事業部"
)


def _find_gdrive_dir() -> Path | None:
    """`管理者提出物` 配下から、デビットExcelの格納フォルダを動的に検索。

    判定ロジック:
      1. 直下に `*デビット*.xlsx` があるフォルダを優先 (新: `01.経費処理フォルダ`)
      2. その他の場合は `*デビッドカード清算*` フォルダを使用 (旧構造)
    検索は最大2階層まで。"""
    if not GDRIVE_PARENT.exists():
        return None

    def _has_debit_xlsx(d: Path) -> bool:
        try:
            return any(
                p.is_file()
                and p.suffix.lower() == '.xlsx'
                and ('デビット' in p.name or 'デビッド' in p.name)
                for p in d.iterdir()
            )
        except (PermissionError, OSError):
            return False

    candidates: list[Path] = []

    # 1) 直下に Excel を持つフォルダを優先
    for p in GDRIVE_PARENT.iterdir():
        if p.is_dir() and _has_debit_xlsx(p):
            candidates.append(p)
    if candidates:
        candidates.sort(key=lambda p: p.name)
        return candidates[0]

    # 2) `デビッドカード清算` を含むフォルダ (旧構造 or サブ)
    for p in GDRIVE_PARENT.iterdir():
        if p.is_dir() and RECEIPT_FOLDER_PATTERN in p.name:
            candidates.append(p)
    if candidates:
        candidates.sort(key=lambda p: p.name)
        return candidates[0]

    # 3) 親直下になければ 1 階層深くまで再帰検索
    for p in GDRIVE_PARENT.iterdir():
        if not p.is_dir():
            continue
        try:
            for sub in p.iterdir():
                if sub.is_dir() and (
                    _has_debit_xlsx(sub)
                    or RECEIPT_FOLDER_PATTERN in sub.name
                ):
                    candidates.append(sub)
        except (PermissionError, OSError):
            continue
    if candidates:
        candidates.sort(key=lambda p: p.name)
        return candidates[0]
    return None


def _resolve_storage_dir() -> Path:
    """デビット格納先のパスを解決。Drive優先、無ければ旧Dropbox。"""
    drive = _find_gdrive_dir()
    if drive is not None:
        return drive
    return DROPBOX_LEGACY_DIR


# 既存コードからの参照名（後方互換）
GDRIVE_DIR = _find_gdrive_dir() or (GDRIVE_PARENT / RECEIPT_FOLDER_PATTERN)
DROPBOX_DIR = _resolve_storage_dir()

# 法人候補（ファイル名から推定）
DEFAULT_CORPORATIONS = ["医療法人", "NPO法人"]

CORP_COLORS = {
    "医療法人": ("#dbeafe", "#1e3a8a"),
    "NPO法人": ("#dcfce7", "#14532d"),
}


def _format_yen(v):
    if v is None or pd.isna(v):
        return ''
    try:
        return f"¥{int(v):,}"
    except (ValueError, TypeError):
        return ''


def _detect_corporation_from_filename(filename: str) -> str:
    """`医療法人EMIFULL_デビット.xlsx` → `医療法人` を推定。"""
    if not filename:
        return ''
    if 'NPO' in filename or 'npo' in filename.lower():
        return 'NPO法人'
    if '医療法人' in filename or '医療' in filename:
        return '医療法人'
    return ''


def _list_dropbox_files() -> list[dict]:
    """格納フォルダ（Drive / 旧Dropbox）上の `*デビット*.xlsx` を返す。"""
    storage = _resolve_storage_dir()
    if not storage.exists():
        return []
    out = []
    for p in storage.iterdir():
        if not p.is_file():
            continue
        if p.suffix.lower() != '.xlsx':
            continue
        if 'デビット' not in p.name and 'デビッド' not in p.name:
            continue
        stat = p.stat()
        out.append({
            'path': p,
            'name': p.name,
            'corporation': _detect_corporation_from_filename(p.name),
            'size': stat.st_size,
            'mtime': stat.st_mtime,
        })
    out.sort(key=lambda x: x['name'])
    return out


def _corp_badge(corp: str) -> str:
    bg, fg = CORP_COLORS.get(corp, ("#f1f5f9", "#475569"))
    return (f"<span style='background:{bg}; color:{fg}; padding:3px 10px; "
            f"border-radius:6px; font-weight:700; font-size:12px;'>{corp}</span>")


def _safe_topn(label: str, total: int, prefer: int, key: str,
               floor: int = 5) -> int:
    """データ件数に応じて TopN を決める。
    件数が少ないときはスライダーを出さず全件表示する。"""
    if total <= floor:
        st.caption(f"（{label}: 全 {total} 件を表示）")
        return total
    return st.slider(label, floor, total, min(prefer, total), key=key)


# レシート対応拡張子（写メ / PDF / その他画像形式）
RECEIPT_IMAGE_EXTS = {'.jpg', '.jpeg', '.png', '.heic', '.heif',
                      '.webp', '.bmp', '.gif', '.tiff', '.tif'}
RECEIPT_PDF_EXTS = {'.pdf'}
RECEIPT_ALL_EXTS = RECEIPT_IMAGE_EXTS | RECEIPT_PDF_EXTS


def _is_receipt_file(path: Path) -> bool:
    return path.suffix.lower() in RECEIPT_ALL_EXTS


def _receipt_kind(path: Path) -> str:
    s = path.suffix.lower()
    if s in RECEIPT_PDF_EXTS:
        return 'pdf'
    if s in RECEIPT_IMAGE_EXTS:
        return 'image'
    return 'other'


# 内側のレシート格納サブフォルダ名（新旧対応）
RECEIPT_SUBFOLDER_NAMES = ('01.デビッドカード清算', 'デビッドカード')


def _list_receipt_files(storage_dir: Path) -> list[dict]:
    """デビットカード清算レシート（写メ / PDF）のみを再帰列挙。

    対象: `{storage_dir}/01.デビッドカード清算/...` (または旧 `デビッドカード/...`)
    現金立替レシート (`02.現金立替レシート/`) など他用途のフォルダは除外する。

    storage_dir 自身がレシート格納フォルダ (旧構造) の場合は storage_dir をルートにする。
    """
    if not storage_dir.exists():
        return []

    # 1) サブフォルダにレシートフォルダがある場合 (新構造): そこだけスキャン
    receipt_root: Path | None = None
    sub_prefix: str = ''  # 相対パス算出用に保存
    for sub_name in RECEIPT_SUBFOLDER_NAMES:
        sub = storage_dir / sub_name
        if sub.exists() and sub.is_dir():
            receipt_root = sub
            sub_prefix = sub_name
            break

    # 2) 無ければ storage_dir 自体がレシートフォルダ (旧構造) かチェック
    if receipt_root is None:
        if storage_dir.name in RECEIPT_SUBFOLDER_NAMES:
            receipt_root = storage_dir
            sub_prefix = storage_dir.name
        else:
            return []  # デビットレシートフォルダが見つからない

    out: list[dict] = []
    seen_paths: set = set()
    for p in receipt_root.rglob('*'):
        try:
            if not p.is_file():
                continue
            if not _is_receipt_file(p):
                continue
            if p in seen_paths:
                continue
            seen_paths.add(p)

            # 施設名は receipt_root からの相対で判定
            try:
                rel_to_root = p.relative_to(receipt_root)
            except ValueError:
                continue
            parts = rel_to_root.parts
            # parts パターン:
            #   {施設}/file
            #   _処理済み/{月}/{施設}/file
            #   _要確認/{施設}/file
            facility = ''
            status_label = ''
            year_month_label = ''
            if len(parts) >= 2 and parts[0] in ('_処理済み', '_要確認'):
                status_label = parts[0].lstrip('_')
                if parts[0] == '_処理済み' and len(parts) >= 4:
                    year_month_label = parts[1]
                    facility = parts[2]
                elif parts[0] == '_要確認' and len(parts) >= 3:
                    facility = parts[1]
                else:
                    facility = parts[1] if len(parts) >= 2 else ''
            elif len(parts) >= 2:
                facility = parts[0]
            else:
                facility = '(未分類)'

            # 表示用 rel パス (storage_dir からのパス)
            try:
                rel_display = p.relative_to(storage_dir)
            except ValueError:
                rel_display = rel_to_root

            stat = p.stat()
            out.append({
                'path': p,
                'name': p.name,
                'rel': str(rel_display),
                'facility': facility,
                'status': status_label,
                'year_month_folder': year_month_label,
                'kind': _receipt_kind(p),
                'ext': p.suffix.lower(),
                'size': stat.st_size,
                'mtime': stat.st_mtime,
            })
        except (PermissionError, OSError):
            continue
    out.sort(key=lambda x: (x['facility'], -x['mtime']))
    return out


def _guess_corporation_from_facility(facility: str) -> str:
    """部門名から法人ラベルを推定。
    のじぎく系 → NPO法人 / それ以外 → 医療法人 (現行データに基づく既定値)。
    確定情報ではないのでフォーム側で変更可能とする。"""
    if not facility:
        return '医療法人'
    if 'のじぎく' in facility:
        return 'NPO法人'
    return '医療法人'


@st.cache_data(show_spinner=False, max_entries=128)
def _heic_to_jpeg_bytes(
    file_bytes: bytes, max_side: int = 800,
) -> tuple[bytes | None, str | None]:
    """HEIC/HEIF バイト列を JPEG バイト列に変換（プレビュー用）。

    Returns: (jpeg_bytes, error_message)。成功時 error_message=None。
    キャッシュは bytes キーで効くが、ライブラリ未導入時は None を返してエラー文字列も付ける。
    """
    # ここで都度 ensure する (pip install 後も新しい呼び出しで反映される)
    if not _ensure_heif_opener():
        return None, (
            "pillow-heif がインストールされていません。"
            "管理者にコマンド `pip install pillow-heif` の実行を依頼してください。"
        )
    try:
        from PIL import Image, ImageOps
        with Image.open(io.BytesIO(file_bytes)) as im:
            im = ImageOps.exif_transpose(im)  # iPhone の縦横回転を反映
            if im.mode not in ('RGB', 'L'):
                im = im.convert('RGB')
            # プレビュー軽量化のため長辺を縮小 (デフォルト 800px)
            w, h = im.size
            if max(w, h) > max_side:
                im.thumbnail((max_side, max_side))
            buf = io.BytesIO()
            im.save(buf, format='JPEG', quality=82, optimize=True)
            return buf.getvalue(), None
    except Exception as e:
        return None, f"{type(e).__name__}: {e}"


def _format_size(n: int) -> str:
    if n < 1024:
        return f"{n} B"
    if n < 1024 * 1024:
        return f"{n / 1024:.1f} KB"
    return f"{n / 1024 / 1024:.1f} MB"


# ============================================================
# 🤖 レシートOCR + デビットExcel自動追記
# ============================================================
_OCR_ACCOUNT_OPTIONS = [
    '食材費', '消耗品費', '燃料費', '旅費交通費',
    '会議費', '福利厚生費', '事務用品費', '通信費',
    '雑費', '車両費', '修繕費', '（その他・直接入力）',
]
_OCR_TAX_OPTIONS = [
    '課対仕入10%', '課対仕入8%（軽）',
]


def _render_ocr_popover(r: dict, file_bytes: bytes, idx: int) -> None:
    """レシートカード末尾に出すOCR + Excel追記用ポップオーバー。"""
    # 画像のみ対応 (PDFは未対応)
    if r['kind'] != 'image':
        return

    key_prefix = f"ocrf_{idx}"

    with st.popover("🤖 OCRで解析 & Excel追記", width='stretch'):
        if not receipt_ocr.is_available():
            st.warning(
                "**ANTHROPIC_API_KEY が未設定です**\n\n"
                "下記いずれかの方法で設定してください:\n"
                "1. 環境変数 `ANTHROPIC_API_KEY` を設定して再起動\n"
                "2. `config/anthropic.json` を作成: "
                "`{\"api_key\": \"sk-ant-...\"}`\n\n"
                "API キーは https://console.anthropic.com/ から取得できます。"
            )
            return

        cache_key = f"ocr_cache::{r['path']}"
        existing = st.session_state.get(cache_key)

        run_clicked = st.button(
            "🔍 OCR解析を実行" if existing is None else "🔁 OCR再解析",
            key=f"{key_prefix}_run", width='stretch',
        )
        if run_clicked:
            with st.spinner("Claude Vision で解析中..."):
                result = receipt_ocr.analyze_receipt(
                    file_bytes,
                    ext=r['ext'],
                    facility_hint=r.get('facility') or None,
                )
            st.session_state[cache_key] = result
            existing = result

        if not existing:
            st.caption("「OCR解析を実行」を押すと内容を自動抽出します。")
            return

        if not existing.get('ok'):
            st.error(f"OCR失敗: {existing.get('error', '不明なエラー')}")
            if existing.get('raw_text'):
                with st.expander("生レスポンスを表示"):
                    st.code(existing['raw_text'])
            return

        data = existing['data']
        usage = existing.get('usage', {})

        st.caption(
            f"信頼度: **{data.get('confidence', '?')}** "
            f"／ tokens: in={usage.get('input_tokens', 0)} "
            f"out={usage.get('output_tokens', 0)}"
        )
        if data.get('notes'):
            st.info(f"💡 {data['notes']}")

        # ---- 編集フォーム ----
        with st.form(f"{key_prefix}_form", clear_on_submit=False):
            corp_default = _guess_corporation_from_facility(r.get('facility', ''))
            corp = st.selectbox(
                "法人 (= 追記先Excel)",
                options=['医療法人', 'NPO法人'],
                index=0 if corp_default == '医療法人' else 1,
                key=f"{key_prefix}_corp",
            )

            col_a, col_b = st.columns(2)
            with col_a:
                # 日付
                from datetime import datetime as _dt2, date as _date2
                date_str = data.get('date') or ''
                try:
                    default_date = _dt2.strptime(date_str, '%Y-%m-%d').date()
                except (ValueError, TypeError):
                    default_date = _date2.fromtimestamp(r['mtime'])
                date_val = st.date_input(
                    "日付", value=default_date, key=f"{key_prefix}_date",
                )
                amount_val = st.number_input(
                    "金額 (税込合計)",
                    min_value=0, step=1,
                    value=int(data.get('amount') or 0),
                    key=f"{key_prefix}_amount",
                )
                vendor_val = st.text_input(
                    "店舗 (摘要先頭)",
                    value=data.get('vendor') or '',
                    key=f"{key_prefix}_vendor",
                )

            with col_b:
                # 借方部門 = レシートの施設フォルダ
                department_val = st.text_input(
                    "借方部門",
                    value=r.get('facility', ''),
                    key=f"{key_prefix}_dept",
                    help="デビットExcelのE列。施設名そのまま入れます。",
                )
                acct_default = data.get('suggested_account') or ''
                if acct_default in _OCR_ACCOUNT_OPTIONS:
                    acct_idx = _OCR_ACCOUNT_OPTIONS.index(acct_default)
                else:
                    acct_idx = len(_OCR_ACCOUNT_OPTIONS) - 1
                acct_choice = st.selectbox(
                    "借方勘定科目",
                    options=_OCR_ACCOUNT_OPTIONS,
                    index=acct_idx,
                    key=f"{key_prefix}_acct",
                )
                if acct_choice == '（その他・直接入力）':
                    acct_val = st.text_input(
                        "勘定科目 (直接入力)",
                        value=acct_default,
                        key=f"{key_prefix}_acct_free",
                    )
                else:
                    acct_val = acct_choice

                tax_default = data.get('suggested_tax_class') or '課対仕入10%'
                tax_idx = (
                    _OCR_TAX_OPTIONS.index(tax_default)
                    if tax_default in _OCR_TAX_OPTIONS
                    else 0
                )
                tax_val = st.selectbox(
                    "借方税区分",
                    options=_OCR_TAX_OPTIONS, index=tax_idx,
                    key=f"{key_prefix}_tax",
                )

            purpose_val = st.text_input(
                "用途 (摘要中の補足)",
                value=data.get('purpose') or '',
                key=f"{key_prefix}_purpose",
            )

            items_text = receipt_ocr.items_to_text(data.get('items') or [])
            items_val = st.text_area(
                "品目 (摘要末尾。OCR抽出を編集可)",
                value=items_text, height=70,
                key=f"{key_prefix}_items",
            )

            submit = st.form_submit_button(
                "📝 デビットExcelに追記", type='primary', width='stretch',
            )

        if submit:
            if amount_val <= 0:
                st.error("金額は1以上を入力してください。")
                return

            xlsx_path = excel_writer.resolve_debit_xlsx(corp, GDRIVE_DIR)
            if xlsx_path is None:
                st.error(
                    f"{corp} のデビットExcelが見つかりません: {GDRIVE_DIR}"
                )
                return

            row_dict = {
                'date': date_val.strftime('%Y-%m-%d'),
                'debit_account': acct_val,
                'tax_class': tax_val,
                'department': department_val,
                'amount': int(amount_val),
                'vendor': vendor_val,
                'purpose': purpose_val,
                'items_text': items_val,
            }
            with st.spinner(f"{xlsx_path.name} に追記中..."):
                res = excel_writer.append_debit_row(xlsx_path, row_dict)

            if not res.get('ok'):
                st.error(f"追記失敗: {res.get('error')}")
                return

            extra = ''
            if res.get('created_sheet'):
                extra = f"（新規シート `{res['sheet']}` を作成）"
            st.success(
                f"✅ 追記完了: `{xlsx_path.name}` シート `{res['sheet']}` "
                f"R{res['row']} {extra}"
            )

            # receipt_processed に記録 (重複処理防止)
            try:
                ocr_data = existing.get('data', {}) if existing else {}
                ocr_data_aug = dict(ocr_data)
                # ユーザ編集後の最終値で上書き
                ocr_data_aug.update({
                    'date': date_val.strftime('%Y-%m-%d'),
                    'amount': int(amount_val),
                    'vendor': vendor_val,
                    'suggested_account': acct_val,
                    'suggested_tax_class': tax_val,
                    'purpose': purpose_val,
                })
                rec = receipt_processor._build_base_record(
                    r, status='success',
                    corporation=corp,
                    ocr_data=ocr_data_aug,
                    ocr_meta=existing,
                    write_result=res,
                )
                db.upsert_receipt_processed(rec)
            except Exception as e:
                st.warning(f"処理ログ保存に失敗: {e}")

            # DB に自動取込してサマリ等を更新
            with st.spinner("デビットExcelをDBに再取込中..."):
                ref = receipt_processor.refresh_debit_db(GDRIVE_DIR)
            st.info(
                f"📊 DB取込: 新規 {ref['inserted']} 件 / "
                f"スキップ {ref['skipped']} 件 → "
                "サマリ／施設別／勘定科目別／購入先別／購入明細 に反映済み"
            )
            st.session_state.pop(cache_key, None)


# ============================================================
# タブ構成 (デビットカード親タブの子)
# ============================================================
(tab_import, tab_summary, tab_fac, tab_acc, tab_vendor,
 tab_detail, tab_receipts) = parent_debit.tabs([
    "📥 取込",
    "📊 サマリ",
    "🏢 施設別",
    "📚 勘定科目別",
    "🏪 購入先別",
    "🛒 購入明細",
    "📷 レシート",
])


# ============================================================
# 📥 取込タブ
# ============================================================
with tab_import:
    existing_yms = db.list_debit_year_months()
    existing_corps = db.list_debit_corporations()

    storage_dir = _resolve_storage_dir()
    is_gdrive = str(storage_dir).startswith(str(GDRIVE_PARENT))

    # ---- ① レシート格納先 (Drive ソース) -----------------------------
    src_url = DEBIT_FOLDER_REF.get('url') or ''
    src_name = DEBIT_FOLDER_REF.get('name') or '01.デビッドカード清算'
    name_html = (
        f"<a href='{src_url}' target='_blank' "
        f"style='color:#1e40af; font-weight:600;'>{src_name}</a>"
        if src_url else f"<b>{src_name}</b>"
    )
    if storage_dir.exists():
        st.markdown(
            f"<div style='background:#f0f9ff; border-left:4px solid #3b82f6; "
            f"padding:10px 16px; border-radius:6px; margin:0 0 16px;'>"
            f"📁 <b>レシート格納先 (Google Drive)</b>: {name_html}"
            f"<br><span style='color:#475569; font-size:12px;'>"
            f"📂 ローカル: <code>{storage_dir}</code>"
            f"</span></div>",
            unsafe_allow_html=True,
        )
    else:
        # 検出失敗: 診断情報 + 再試行ボタン
        gp_exists = GDRIVE_PARENT.exists()
        st.markdown(
            f"<div style='background:#fef3c7; border-left:4px solid #f59e0b; "
            f"padding:12px 16px; border-radius:6px; margin:0 0 12px;'>"
            f"⚠️ <b>Drive上のレシート格納先が見つかりませんでした</b><br>"
            f"📁 Drive URL: {name_html}"
            f"</div>",
            unsafe_allow_html=True,
        )

        diag1, diag2 = st.columns([1, 1])
        with diag1:
            st.caption(
                f"**親フォルダ**: `{GDRIVE_PARENT}`<br>"
                f"&nbsp;&nbsp;検出: {'✅ 存在' if gp_exists else '❌ 不在'}",
                unsafe_allow_html=True,
            )
        with diag2:
            st.caption(
                f"**検索条件**: 「{RECEIPT_FOLDER_PATTERN}」 or `*デビット*.xlsx`<br>"
                f"&nbsp;&nbsp;最終フォールバック: <code>{storage_dir}</code>",
                unsafe_allow_html=True,
            )

        if not gp_exists:
            st.info(
                "📌 **Google Drive for Desktop の状態を確認してください**\n\n"
                "1. タスクトレイの Drive アイコンが緑/同期中か\n"
                "2. サインイン中アカウントに `01.経費処理フォルダ` または "
                "`01.デビッドカード清算` が同期されているか\n"
                "3. 起動直後で同期が完了していない場合は数秒待つ\n\n"
                f"ブラウザで Drive を開いて確認: {src_url}"
            )
        else:
            # 親はあるが子フォルダで検出できないケース
            try:
                children = sorted(
                    p.name for p in GDRIVE_PARENT.iterdir() if p.is_dir()
                )
                with st.expander(
                    f"🔎 親フォルダの直下: {len(children)}件 (名前を確認)",
                    expanded=False,
                ):
                    st.code('\n'.join(children) if children else '(空)')
                st.info(
                    "親フォルダは存在しますが、デビットExcelやレシート用フォルダが見つかりません。"
                    "上記リストに `01.経費処理フォルダ` や `01.デビッドカード清算` が"
                    "ある場合は、ボタンで再検索してください。"
                )
            except Exception as e:
                st.error(f"親フォルダ走査エラー: {e}")

        retry_col, _ = st.columns([1, 3])
        with retry_col:
            if st.button(
                "🔄 Drive 再検索",
                key='debit_retry_drive',
                width='stretch',
                help='ページを再読み込みして Drive を再検出します',
            ):
                st.cache_data.clear()
                st.rerun()

    # ---- ② 取込実行ボタン (主要アクション) ---------------------------
    files = _list_dropbox_files() if storage_dir.exists() else []
    target_idx = list(range(len(files)))  # デフォルト: 全ファイル

    run_col, opt_col = st.columns([3, 1])
    with run_col:
        run_clicked = st.button(
            "🚀 取込実行（Drive上の全ファイル）",
            type='primary', key='debit_dropbox_run',
            width='stretch',
            disabled=not files,
        )
    with opt_col:
        overwrite = st.checkbox(
            "同月分を上書き（再取込）", value=False,
            help="チェック時は対象月のデータを削除してから取込",
            key='debit_dropbox_overwrite',
        )

    if storage_dir.exists() and not files:
        st.info(f"デビット xlsx が見つかりません: `{storage_dir}`")

    # ---- ③ 詳細・オプション (折りたたみ) ----------------------------
    if files:
        with st.expander(
            "⚙️ 詳細・オプション "
            "(取込状態 / Excelダウンロード / 個別ファイル選択)",
            expanded=False,
        ):
            # 取込状態 (KPI)
            cstat1, cstat2, cstat3 = st.columns(3)
            cstat1.metric("取込済 月数", f"{len(existing_yms)}")
            cstat2.metric("法人数", f"{len(existing_corps)}")
            cstat3.metric("最新月", existing_yms[0] if existing_yms else "—")

            # Excel 最終更新サマリ
            _excel_latest = max(f['mtime'] for f in files)
            _excel_latest_str = datetime.fromtimestamp(
                _excel_latest
            ).strftime('%Y-%m-%d %H:%M')
            _last_import = db.get_debit_last_imported_at()
            if _last_import is not None:
                _last_import_str = _last_import.strftime('%Y-%m-%d %H:%M')
                if _excel_latest <= _last_import.timestamp() + 1:
                    st.caption(
                        f"📅 Excel最終更新: **{_excel_latest_str}** ／ "
                        f"前回取込: {_last_import_str} "
                        "(Excelは前回取込から変更なし)"
                    )
                else:
                    st.caption(
                        f"📅 Excel最終更新: **{_excel_latest_str}** ／ "
                        f"前回取込: {_last_import_str} "
                        "(Excelの方が新しい → 取込推奨)"
                    )
            else:
                st.caption(
                    f"📅 Excel最終更新: **{_excel_latest_str}** ／ "
                    "まだ取込実績はありません"
                )

            # Excelダウンロード
            st.markdown("##### 📥 Excelをそのままダウンロード")
            dl_cols = st.columns(max(2, len(files)))
            for i, f in enumerate(files):
                with dl_cols[i % len(dl_cols)]:
                    try:
                        with open(f['path'], 'rb') as fp:
                            file_bytes = fp.read()
                        corp_label = f['corporation'] or '(未判定)'
                        st.download_button(
                            label=f"⬇️ {corp_label} ／ {f['name']}",
                            data=file_bytes,
                            file_name=f['name'],
                            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                            key=f'debit_dl_{i}',
                            width='stretch',
                        )
                        mtime_str = datetime.fromtimestamp(
                            f['mtime']
                        ).strftime('%Y-%m-%d %H:%M')
                        st.caption(
                            f"サイズ: {f['size'] / 1024:.1f} KB ／ "
                            f"📅 最終更新: {mtime_str}"
                        )
                    except Exception as e:
                        st.error(f"読み込み失敗: {f['name']}: {e}")

            # ファイル一覧 + 個別選択
            st.markdown("##### 📋 取込対象ファイル一覧")
            file_rows = pd.DataFrame([
                {
                    '法人': f['corporation'] or '(未判定)',
                    'ファイル名': f['name'],
                    'サイズ': f"{f['size']/1024:.1f} KB",
                    '最終更新': datetime.fromtimestamp(
                        f['mtime']
                    ).strftime('%Y-%m-%d %H:%M'),
                }
                for f in files
            ])
            st.dataframe(file_rows, hide_index=True, width='stretch')
            target_idx = st.multiselect(
                "個別に取込するファイルを絞る場合（既定: 全ファイル）",
                options=list(range(len(files))),
                default=list(range(len(files))),
                format_func=lambda i: files[i]['name'],
                key='debit_dropbox_select',
            )

            # ---- 既存データの日付降順並び替え ----
            st.markdown("##### 🔃 既存シートを日付降順に並び替え")
            st.caption(
                "各 `YYMM【デビット】` シート内の行を日付の新しい順 (上から) に "
                "並び替えます。A〜K列のみ移動し、L列(検証)は触りません。"
            )
            if st.button(
                "🔃 全デビットExcelを日付降順にソート",
                key='debit_sort_all_btn',
                width='stretch',
            ):
                sort_summary = []
                with st.spinner("デビットExcelを並び替え中..."):
                    for f in files:
                        try:
                            res = excel_writer.sort_debit_xlsx_by_date_desc(
                                Path(f['path'])
                            )
                            sort_summary.append({
                                'corporation': f['corporation'] or '(未判定)',
                                'name': f['name'],
                                'ok': res.get('ok', False),
                                'total': res.get('total_sorted', 0),
                                'sheets': res.get('sheets', []),
                                'error': res.get('error'),
                            })
                        except Exception as e:
                            sort_summary.append({
                                'corporation': f['corporation'] or '(未判定)',
                                'name': f['name'],
                                'ok': False,
                                'error': str(e),
                            })
                for s in sort_summary:
                    if not s['ok']:
                        st.error(
                            f"❌ {s['corporation']} / `{s['name']}`: "
                            f"{s.get('error')}"
                        )
                        continue
                    if s['total'] == 0:
                        st.info(
                            f"✓ {s['corporation']} / `{s['name']}`: "
                            "既に降順 (変更なし)"
                        )
                    else:
                        st.success(
                            f"✅ {s['corporation']} / `{s['name']}`: "
                            f"計 {s['total']} 行を並び替え"
                        )
                        with st.expander("シート別の詳細", expanded=False):
                            for sh in s['sheets']:
                                if sh.get('sorted', 0) > 0:
                                    st.write(
                                        f"- `{sh['name']}`: {sh['sorted']} 行"
                                    )
                                elif sh.get('already_sorted'):
                                    st.write(f"- `{sh['name']}`: 既に降順")

    # ---- 取込実行ロジック ---------------------------------------------
    if files:
        if True:  # 旧コードのインデント維持用ダミー (削除予定)

                if run_clicked:
                    sub_lookup = {}
                    for s in db.list_pl_subunits():
                        sub_lookup[s['excel_name']] = s['id']
                        if s.get('display_name'):
                            sub_lookup[s['display_name']] = s['id']

                    # ----- Phase 0: 過去に処理済みなのに移動されていないレシートを整理 -----
                    receipts_storage = _resolve_storage_dir()
                    all_receipts = (
                        _list_receipt_files(receipts_storage)
                        if receipts_storage.exists() else []
                    )
                    if all_receipts:
                        with st.spinner("処理済みレシートを整理中..."):
                            recon = receipt_processor.reconcile_processed_locations(
                                all_receipts, GDRIVE_DIR,
                            )
                        if recon['moved'] > 0:
                            st.info(
                                f"📦 過去に処理済みのレシート **{recon['moved']}件** を "
                                f"`_処理済み/{{月}}/{{施設}}/` に整理しました"
                            )
                        if recon['errors']:
                            for em in recon['errors']:
                                st.warning(f"⚠️ {em}")
                        # 移動したのでファイル一覧を再取得
                        if recon['moved'] > 0:
                            all_receipts = _list_receipt_files(receipts_storage)

                    # ----- Phase 1: 未処理レシートをOCR→デビットExcelに自動追記 -----
                    unprocessed = receipt_processor.detect_unprocessed(all_receipts)
                    receipt_summary = None
                    if unprocessed and receipt_ocr.is_available():
                        progress_bar = st.progress(
                            0.0, text=f"レシートOCR処理中 (0/{len(unprocessed)})..."
                        )

                        def _on_progress(i, total, item):
                            progress_bar.progress(
                                i / max(total, 1),
                                text=f"レシートOCR {i}/{total}: {item['file_name']}",
                            )

                        receipt_summary = receipt_processor.auto_process_batch(
                            unprocessed, GDRIVE_DIR, on_progress=_on_progress,
                        )
                        progress_bar.empty()
                    elif unprocessed and not receipt_ocr.is_available():
                        st.warning(
                            f"未処理レシート {len(unprocessed)} 件ありますが、"
                            "ANTHROPIC_API_KEY が未設定のためOCRをスキップします。"
                        )

                    # ----- Phase 2: デビットExcel → DB 取込 -----
                    total_inserted = 0
                    total_skipped = 0
                    all_yms = set()
                    with st.spinner("デビットExcel取込中..."):
                        # レシート追記後の最新Excelを再読込するため files を取り直す
                        latest_files = _list_dropbox_files()
                        files_to_import = (
                            latest_files if latest_files else files
                        )
                        for f in files_to_import:
                            corp = f['corporation'] or 'その他'
                            with open(f['path'], 'rb') as fp:
                                result = debit_parser.parse_debit_workbook(
                                    fp, corporation=corp, subunit_lookup=sub_lookup,
                                )
                            if overwrite:
                                yms_in_file = sorted({
                                    r['year_month'] for r in result['rows']
                                    if r.get('year_month')
                                })
                                for ym in yms_in_file:
                                    db.delete_debit_by_year_month(ym, corporation=corp)
                            ins = db.insert_debit_entries(
                                result['rows'], filename=f['name'],
                                corporation=corp, file_hash=result['file_hash'],
                            )
                            total_inserted += ins['inserted']
                            total_skipped += ins['skipped']
                            all_yms.update(ins['year_months'])

                            st.write(
                                f"- **{corp}** / `{f['name']}`: "
                                f"新規 {ins['inserted']} / スキップ {ins['skipped']}"
                            )
                            if result['unknown_departments']:
                                with st.expander(
                                    f"⚠️ {f['name']} : 未マッチ部門 "
                                    f"{len(result['unknown_departments'])}件"
                                ):
                                    for d in sorted(result['unknown_departments']):
                                        st.text(d)

                    # ----- 結果サマリ -----
                    if receipt_summary and receipt_summary['success'] > 0:
                        st.success(
                            f"✅ 取込完了 — Excel新規 {total_inserted} / "
                            f"スキップ {total_skipped} ／ "
                            f"レシート自動追記 {receipt_summary['success']} 件"
                        )
                    else:
                        st.success(
                            f"✅ 取込完了 — 新規 {total_inserted} / "
                            f"スキップ {total_skipped} / 対象月 {sorted(all_yms)}"
                        )

                    # 確認画面用に session_state に保存 (rerun せず下で表示)
                    if receipt_summary and (
                        receipt_summary['success'] > 0
                        or receipt_summary['manual_pending'] > 0
                        or receipt_summary['failed'] > 0
                    ):
                        st.session_state['receipt_review_pending'] = (
                            receipt_summary['results']
                        )
                    else:
                        st.session_state.pop('receipt_review_pending', None)

    # ---- 📤 手動アップロード (詳細・オプション扱い) ----
    with st.expander("📤 手動アップロード (Excelを直接ファイル指定)", expanded=False):
        c1, c2 = st.columns([2, 1])
        with c1:
            uploaded = st.file_uploader(
                "デビットExcel（.xlsx）",
                type=['xlsx'], accept_multiple_files=True,
                key='debit_uploader',
            )
        with c2:
            corp_input = st.selectbox(
                "法人ラベル",
                options=DEFAULT_CORPORATIONS + ["その他（手動入力）"],
                key='debit_corp_select',
            )
            if corp_input == "その他（手動入力）":
                corp_input = st.text_input(
                    "法人名", value="", key='debit_corp_manual',
                )

        # 取込実行ボタン（最上部に表示）
        run_col, opt_col = st.columns([2, 1])
        with run_col:
            run_manual = st.button(
                "🚀 取込実行（アップロードしたファイル）",
                type='primary', key='debit_manual_run',
                width='stretch',
                disabled=not (uploaded and corp_input),
            )
        with opt_col:
            overwrite_manual = st.checkbox(
                "同月分を上書き（再取込）", value=False,
                key='debit_manual_overwrite',
            )

        if uploaded and corp_input:
            sub_lookup = {}
            for s in db.list_pl_subunits():
                sub_lookup[s['excel_name']] = s['id']
                if s.get('display_name'):
                    sub_lookup[s['display_name']] = s['id']

            with st.spinner("解析中..."):
                results = []
                for f in uploaded:
                    auto_corp = _detect_corporation_from_filename(f.name) or corp_input
                    result = debit_parser.parse_debit_workbook(
                        f, corporation=auto_corp, subunit_lookup=sub_lookup,
                    )
                    results.append((f, auto_corp, result))

            with st.expander("📋 解析プレビュー", expanded=False):
                preview = []
                unknown_dep_all = set()
                for f, corp, result in results:
                    preview.append({
                        '法人': corp,
                        'ファイル': f.name,
                        'シート': ', '.join(s['sheet'] for s in result['sheet_summaries']),
                        '行数': len(result['rows']),
                        '対象月': ', '.join(sorted({r['year_month'] for r in result['rows']})),
                    })
                    unknown_dep_all.update(result['unknown_departments'])
                st.dataframe(pd.DataFrame(preview), hide_index=True, width='stretch')

                if unknown_dep_all:
                    st.markdown(f"⚠️ **未マッチ部門 {len(unknown_dep_all)}件**")
                    st.write(sorted(unknown_dep_all))
                    st.caption("（未マッチでも行は保存されます。施設マスタに追加すると以降紐付きます。）")

            if run_manual:
                total_inserted = 0
                total_skipped = 0
                all_yms = set()
                for f, corp, result in results:
                    if overwrite_manual:
                        yms_in_file = sorted({
                            r['year_month'] for r in result['rows']
                            if r.get('year_month')
                        })
                        for ym in yms_in_file:
                            db.delete_debit_by_year_month(ym, corporation=corp)
                    ins = db.insert_debit_entries(
                        result['rows'], filename=f.name,
                        corporation=corp, file_hash=result['file_hash'],
                    )
                    total_inserted += ins['inserted']
                    total_skipped += ins['skipped']
                    all_yms.update(ins['year_months'])

                st.success(
                    f"✅ 取込完了 — 新規 {total_inserted} / "
                    f"スキップ {total_skipped} / 対象月 {sorted(all_yms)}"
                )
                st.rerun()

    # ---- 📝 取り込んだレシート内容の確認・修正 ----
    _review_results = st.session_state.get('receipt_review_pending')
    if _review_results:
        st.markdown("---")
        with st.container(border=True):
            st.markdown(
                "### 📝 このレシートを取り込みました — 内容のご確認をお願いします"
            )
            st.caption(
                "OCRが自動抽出した結果(✅取込済)、自動追記できなかったもの(⚠️要編集)、"
                "新規手入力(➕)のすべてを下のテーブルで編集できます。"
                "金額・勘定科目・税区分などを確認・修正 → 「✅ 確定して反映」を押してください。"
            )

            review_rows: list[dict] = []
            for rr in _review_results:
                ex = rr.get('excel') or {}
                excel_path = ex.get('path', '')
                excel_sheet = ex.get('sheet', '')
                excel_row = int(ex.get('row', 0)) if ex.get('row') else 0
                # 法人推定: Excel名 → ダメなら施設名から
                corp_label = ''
                if excel_path:
                    corp_label = excel_writer._detect_corp_from_filename(
                        Path(excel_path).name
                    )
                if not corp_label:
                    fac = rr.get('facility') or ''
                    corp_label = (
                        'NPO法人' if 'のじぎく' in fac else '医療法人'
                    )

                # Excel既存行の現値で補完 (success の場合のみ)
                tax_class_val = '課対仕入10%'
                description_val = ''
                if excel_path and excel_row > 0:
                    try:
                        wb_v = openpyxl.load_workbook(excel_path, data_only=True)
                        ws_v = wb_v[excel_sheet]
                        tax_class_val = (
                            ws_v.cell(row=excel_row, column=3).value
                            or '課対仕入10%'
                        )
                        description_val = (
                            ws_v.cell(row=excel_row, column=11).value or ''
                        )
                    except Exception:
                        pass

                # 状態ラベル
                if rr['status'] == 'success':
                    state_label = '✅取込済'
                elif rr['status'] == 'manual_pending':
                    state_label = '⚠️要編集'
                else:
                    state_label = '❌OCR失敗'

                review_rows.append({
                    '状態': state_label,
                    'レシート': rr['file_name'],
                    '法人': corp_label or '医療法人',
                    '日付': rr.get('date') or '',
                    '部門': rr.get('facility') or '',
                    '勘定科目': rr.get('account') or '',
                    '税区分': tax_class_val,
                    '金額': int(rr.get('amount') or 0),
                    '摘要': description_val,
                    'OCRエラー': rr.get('reason') or '',
                    '_excel_path': excel_path,
                    '_excel_sheet': excel_sheet,
                    '_excel_row': excel_row,
                })

            review_df = pd.DataFrame(review_rows)
            failed_reasons = [
                str(r.get('reason') or '').strip()
                for r in _review_results
                if r.get('status') == 'failed' and str(r.get('reason') or '').strip()
            ]
            if failed_reasons:
                with st.expander("OCR失敗の理由", expanded=True):
                    for msg in sorted(set(failed_reasons)):
                        st.error(msg)

            # 編集テーブル (動的=ユーザが行追加可能)
            edited = st.data_editor(
                review_df,
                column_config={
                    '状態': st.column_config.TextColumn(
                        '状態', disabled=True, width='small',
                        help='✅取込済=Excelに自動追記済 / ⚠️要編集=要入力 / ❌OCR失敗 / ➕新規',
                    ),
                    'レシート': st.column_config.TextColumn(
                        'レシート', disabled=True, width='small',
                    ),
                    '法人': st.column_config.SelectboxColumn(
                        '法人',
                        options=['医療法人', 'NPO法人'],
                        required=True,
                        help='デビットExcelの行き先 (NPO法人/医療法人 のどちらの.xlsxに書くか)',
                    ),
                    '日付': st.column_config.TextColumn(
                        '日付', help='YYYY-MM-DD 形式',
                    ),
                    '部門': st.column_config.TextColumn(
                        '部門', help='例: 5.2.UMIEてんり / 16.1.カラダキッズてんり',
                    ),
                    '勘定科目': st.column_config.SelectboxColumn(
                        '勘定科目',
                        options=_OCR_ACCOUNT_OPTIONS[:-1],
                        required=False,
                    ),
                    '税区分': st.column_config.SelectboxColumn(
                        '税区分',
                        options=_OCR_TAX_OPTIONS,
                        required=False,
                    ),
                    '金額': st.column_config.NumberColumn(
                        '金額', format='%d', min_value=0,
                    ),
                    '摘要': st.column_config.TextColumn('摘要', width='large'),
                    'OCRエラー': st.column_config.TextColumn(
                        'OCRエラー', disabled=True, width='large',
                    ),
                    '_excel_path': None,
                    '_excel_sheet': None,
                    '_excel_row': None,
                },
                hide_index=True,
                width='stretch',
                num_rows='dynamic',  # ユーザが行を追加できる
                key='receipt_review_editor',
            )

            st.caption(
                "💡 一番下の **「+」ボタンで新規行を追加** すると、"
                "OCRを介さず手入力でデビットExcelに行を追加できます (法人を選んで日付/部門/金額を入力)。"
            )

            conf_col1, conf_col2 = st.columns([1, 1])
            with conf_col1:
                confirm = st.button(
                    "✅ 確定して反映 → サマリ等へ",
                    type='primary', width='stretch',
                    key='receipt_review_confirm',
                )
            with conf_col2:
                cancel = st.button(
                    "↩ 確認画面を閉じる",
                    width='stretch',
                    help="編集を破棄せず、画面のみ閉じます。次回取込時に再表示されます。",
                    key='receipt_review_cancel',
                )

            if confirm:
                update_count = 0
                append_count = 0
                skip_count = 0
                error_msgs: list[str] = []
                affected_yms: set[str] = set()
                affected_corps: set[str] = set()

                for idx, (_, row) in enumerate(edited.iterrows()):
                    # 必須項目チェック
                    date_str = str(row.get('日付') or '').strip()
                    dept = str(row.get('部門') or '').strip()
                    amount_v = row.get('金額')
                    debit_account = str(row.get('勘定科目') or '').strip()
                    corp = str(row.get('法人') or '').strip() or '医療法人'

                    try:
                        amount_int = int(amount_v) if amount_v else 0
                    except (ValueError, TypeError):
                        amount_int = 0

                    if not date_str or not dept or amount_int <= 0 or not debit_account:
                        skip_count += 1
                        if any([date_str, dept, amount_int, debit_account]):
                            error_msgs.append(
                                f"行{idx+1}: 日付/部門/勘定科目/金額のいずれかが未入力のためスキップ"
                            )
                        continue

                    # 月集計用
                    if len(date_str) >= 7:
                        affected_yms.add(date_str[:7])
                    affected_corps.add(corp)

                    excel_row_val = row.get('_excel_row')
                    is_existing = (
                        pd.notna(excel_row_val)
                        and int(excel_row_val) > 0
                        and row.get('_excel_path')
                    )

                    if is_existing:
                        # ----- 既存行を更新 -----
                        orig = review_df[review_df['レシート'] == row['レシート']]
                        if len(orig) == 0:
                            continue
                        orig_row = orig.iloc[0]
                        changed = {}
                        if date_str != str(orig_row['日付']):
                            changed['date'] = date_str
                        if dept != str(orig_row['部門']):
                            changed['department'] = dept
                        if debit_account != str(orig_row['勘定科目']):
                            changed['debit_account'] = debit_account
                        if str(row['税区分']) != str(orig_row['税区分']):
                            changed['tax_class'] = str(row['税区分'])
                        if amount_int != int(orig_row['金額']):
                            changed['amount'] = amount_int
                            changed['credit_amount'] = amount_int
                        if str(row['摘要']) != str(orig_row['摘要']):
                            changed['description'] = str(row['摘要'])

                        if not changed:
                            continue
                        upd = excel_writer.update_debit_row(
                            Path(str(row['_excel_path'])),
                            str(row['_excel_sheet']),
                            int(excel_row_val),
                            changed,
                        )
                        if upd.get('ok'):
                            update_count += 1
                        else:
                            error_msgs.append(
                                f"{row.get('レシート','?')}: {upd.get('error')}"
                            )
                    else:
                        # ----- 新規 (要編集 / OCR失敗 / 手動入力) → Excelに追記 -----
                        xlsx_path = excel_writer.resolve_debit_xlsx(
                            corp, GDRIVE_DIR,
                        )
                        if xlsx_path is None:
                            error_msgs.append(
                                f"行{idx+1}: {corp} のデビットExcelが見つかりません"
                            )
                            continue
                        write_res = excel_writer.append_debit_row(xlsx_path, {
                            'date': date_str,
                            'debit_account': debit_account,
                            'tax_class': str(row.get('税区分') or '課対仕入10%'),
                            'department': dept,
                            'amount': amount_int,
                            'description': str(row.get('摘要') or ''),
                        })
                        if write_res.get('ok'):
                            append_count += 1
                            # receipt_processed を success に更新 (再処理回避)
                            file_name = row.get('レシート') or ''
                            if file_name and file_name != 'nan':
                                rr_orig = next(
                                    (rr for rr in _review_results
                                     if rr['file_name'] == file_name),
                                    None,
                                )
                                if rr_orig and rr_orig.get('file_path'):
                                    db.upsert_receipt_processed({
                                        'file_path': rr_orig['file_path'],
                                        'file_name': file_name,
                                        'status': 'success',
                                        'corporation': corp,
                                        'transaction_date': date_str,
                                        'amount': amount_int,
                                        'debit_account': debit_account,
                                        'excel_path': str(xlsx_path),
                                        'excel_sheet': write_res['sheet'],
                                        'excel_row': write_res['row'],
                                    })
                        elif write_res.get('duplicate'):
                            skip_count += 1
                        else:
                            error_msgs.append(
                                f"行{idx+1} ({corp}): {write_res.get('error')}"
                            )

                # ----- DB 再取込 -----
                with st.spinner("DBに反映中..."):
                    if update_count > 0 or append_count > 0:
                        # 影響月をクリアしてからフルリインポート
                        for ym in affected_yms:
                            for corp in affected_corps:
                                db.delete_debit_by_year_month(ym, corporation=corp)
                    ref = receipt_processor.refresh_debit_db(GDRIVE_DIR)

                # ----- 結果表示 -----
                msg_parts = []
                if update_count > 0:
                    msg_parts.append(f"既存行更新 **{update_count}**")
                if append_count > 0:
                    msg_parts.append(f"新規追記 **{append_count}**")
                if skip_count > 0:
                    msg_parts.append(f"スキップ {skip_count}")
                summary_msg = " / ".join(msg_parts) if msg_parts else "変更なし"
                st.success(
                    f"✅ 確定 — {summary_msg} ／ DB取込 新規 {ref['inserted']} "
                    f"/ スキップ {ref['skipped']}\n\n"
                    "サマリ／施設別／勘定科目別／購入先別／購入明細 に反映済みです。"
                )
                if error_msgs:
                    for em in error_msgs:
                        st.warning(em)

                st.balloons()
                st.info(
                    "👉 「📊 サマリ」タブをクリックすると最新データが表示されます。"
                )
                st.session_state.pop('receipt_review_pending', None)

            if cancel:
                st.session_state.pop('receipt_review_pending', None)
                st.rerun()

    # ---- 取込履歴 ----
    imports = db.list_debit_imports(limit=10)
    if imports:
        with st.expander(f"📜 取込履歴（直近{len(imports)}件）", expanded=False):
            st.dataframe(pd.DataFrame([
                {
                    '取込日時': r['imported_at'],
                    '法人': r['corporation'],
                    'ファイル': r['source_filename'],
                    '行数': r['row_count'],
                    '新規': r['inserted_count'],
                    'スキップ': r['skipped_count'],
                    '対象月': r['year_months'],
                }
                for r in imports
            ]), hide_index=True, width='stretch')


# ============================================================
# 💴 現金立替レシート 親タブ
# ============================================================
# デビットカード側の `st.stop()` より前に描画する必要があるため、
# このブロックは関数化せずに `with parent_cash:` でインライン描画する。
# 詳細実装は lib/cash_advance_parser.py / lib/expense_form.py を使う。
# ============================================================

# 経費処理フォルダ (Drive)
CASH_GDRIVE_PARENT = Path(r"G:\マイドライブ\管理者提出物\01.経費処理フォルダ")
CASH_GDRIVE_FOLDER_NAME = "02.現金立替レシート"   # 施設別レシート格納先

# 出納帳ダウンロード後の Drive 共有フォルダ (ユーザに案内するURL)
CASH_DRIVE_OUTPUT_URL = (
    "https://drive.google.com/drive/folders/"
    "1FToAeulgwYV58yNMlTeMJ2eekoBlUfnL"
)

# 経費精算書テンプレート (xlsm) の保管場所
CASH_DROPBOX_DIR = Path(
    r"C:\Users\user01\株式会社ＥＭＩＦＵＬＬ Dropbox\障がい事業部"
    r"\99.ゴミ箱\【⑤】2026年度データ\2604"
)
CASH_XLSM_PATTERN = "01.現金出納帳"

# 法人ラベル → ファイル名キーワード
CASH_XLSM_BY_CORP = {
    "医療法人": "EMI",
    "NPO法人": "のじ",
}


def _find_cash_xlsm(corp_label: str) -> Path | None:
    if not CASH_DROPBOX_DIR.exists():
        return None
    needle = CASH_XLSM_BY_CORP.get(corp_label, "")
    for p in CASH_DROPBOX_DIR.iterdir():
        if not p.is_file() or p.suffix.lower() != ".xlsm":
            continue
        if CASH_XLSM_PATTERN in p.name and needle and needle in p.name:
            return p
    return None


def _list_cash_xlsm_files() -> list[dict]:
    if not CASH_DROPBOX_DIR.exists():
        return []
    out = []
    for p in CASH_DROPBOX_DIR.iterdir():
        if not p.is_file() or p.suffix.lower() != ".xlsm":
            continue
        if CASH_XLSM_PATTERN not in p.name:
            continue
        corp = cash_advance_parser.detect_corporation_from_filename(p.name)
        stat = p.stat()
        out.append({
            'path': p, 'name': p.name,
            'corporation': corp or '(未判定)',
            'size': stat.st_size, 'mtime': stat.st_mtime,
        })
    out.sort(key=lambda x: x['name'])
    return out


def _resolve_cash_receipt_root() -> Path | None:
    if not CASH_GDRIVE_PARENT.exists():
        return None
    root = CASH_GDRIVE_PARENT / CASH_GDRIVE_FOLDER_NAME
    if not root.exists():
        return None
    return root


def _cash_render_receipt_popover(r: dict, file_bytes: bytes,
                                  rcpt_db: dict | None,
                                  idx_key: str) -> None:
    """現金立替レシートカードに付ける OCR + 取込ポップオーバー。
    OCR を実行 → cash_advance_receipts テーブルにメタデータを upsert。
    ユーザは値を編集可。「経費精算書」タブから候補として参照される。"""
    label = "🤖 OCR & 取込登録"
    if rcpt_db and rcpt_db.get('ocr_status') == 'success':
        label = "✏️ 取込内容を編集"

    with st.popover(label, width='stretch'):
        if not receipt_ocr.is_available():
            st.warning(
                "ANTHROPIC_API_KEY が未設定のため OCR は使えません。"
                "config/anthropic.json または環境変数を設定してください。"
            )

        cache_key = f"cash_ocr_cache::{r['path']}"
        # DB 既存値があればキャッシュとして利用
        existing = st.session_state.get(cache_key)
        if not existing and rcpt_db and rcpt_db.get('ocr_result_json'):
            import json as _json
            try:
                cached_data = _json.loads(rcpt_db['ocr_result_json'])
                existing = {'ok': True, 'data': cached_data, 'usage': {}}
                st.session_state[cache_key] = existing
            except Exception:
                pass

        # OCR 実行ボタン
        if receipt_ocr.is_available():
            run_label = "🔍 OCR実行" if existing is None else "🔁 OCR再解析"
            if st.button(run_label, key=f"{idx_key}_ocrrun",
                          width='stretch'):
                with st.spinner("Claude Vision で解析中..."):
                    ocr_res = receipt_ocr.analyze_receipt(
                        file_bytes,
                        ext=r['ext'],
                        facility_hint=r.get('facility') or None,
                    )
                st.session_state[cache_key] = ocr_res
                existing = ocr_res

        # OCR 結果を表示してフォームを出す
        ocr_data = (existing.get('data') or {}) if existing else {}

        # 既存DB値があればフォームの初期値として優先
        from datetime import datetime as _dt2, date as _date2
        date_default = None
        if rcpt_db and rcpt_db.get('transaction_date'):
            try:
                date_default = _dt2.strptime(
                    rcpt_db['transaction_date'], '%Y-%m-%d').date()
            except Exception:
                pass
        if date_default is None and ocr_data.get('date'):
            try:
                date_default = _dt2.strptime(
                    ocr_data['date'], '%Y-%m-%d').date()
            except Exception:
                pass
        if date_default is None:
            date_default = _date2.fromtimestamp(r['mtime'])

        amount_default = (
            (rcpt_db or {}).get('total_amount')
            or ocr_data.get('amount') or 0
        )
        vendor_default = (
            (rcpt_db or {}).get('vendor')
            or ocr_data.get('vendor') or ''
        )
        purpose_default = (
            (rcpt_db or {}).get('purpose')
            or ocr_data.get('purpose') or ''
        )
        payee_default = (rcpt_db or {}).get('payee') or ''

        # 法人推定
        corp_guess = (
            (rcpt_db or {}).get('corporation')
            or cash_advance_parser.detect_corp_from_facility_code(
                r.get('facility'))
            or _guess_corporation_from_facility(r.get('facility') or '')
        )
        with st.form(f"{idx_key}_form", clear_on_submit=False):
            corp = st.selectbox(
                "法人", options=['医療法人', 'NPO法人'],
                index=0 if corp_guess == '医療法人' else 1,
                key=f"{idx_key}_corp",
            )
            c1, c2 = st.columns(2)
            with c1:
                rcp_date = st.date_input(
                    "日付", value=date_default,
                    key=f"{idx_key}_date",
                )
                rcp_amt = st.number_input(
                    "金額 (税込合計)",
                    min_value=0, step=1,
                    value=int(amount_default or 0),
                    key=f"{idx_key}_amt",
                )
                rcp_vendor = st.text_input(
                    "支払先", value=vendor_default,
                    key=f"{idx_key}_vendor",
                )
            with c2:
                rcp_purpose = st.text_input(
                    "用途", value=purpose_default,
                    key=f"{idx_key}_purpose",
                )
                rcp_payee = st.text_input(
                    "立替者氏名", value=payee_default,
                    key=f"{idx_key}_payee",
                )
                # 一律科目のヒント (OCR 推定)
                acc_default = ocr_data.get('suggested_account') or ''
                rcp_acc = st.selectbox(
                    "推定勘定科目（参考）",
                    options=[
                        '', '食材費', '消耗品費', '燃料費', '会議費',
                        '福利厚生費', '事務用品費', '通信費',
                        '雑費', '車両費', '修繕費', '研修費',
                        '水道光熱費', '租税公課',
                    ],
                    index=([
                        '', '食材費', '消耗品費', '燃料費', '会議費',
                        '福利厚生費', '事務用品費', '通信費',
                        '雑費', '車両費', '修繕費', '研修費',
                        '水道光熱費', '租税公課',
                    ].index(acc_default) if acc_default in [
                        '', '食材費', '消耗品費', '燃料費', '会議費',
                        '福利厚生費', '事務用品費', '通信費',
                        '雑費', '車両費', '修繕費', '研修費',
                        '水道光熱費', '租税公課',
                    ] else 0),
                    key=f"{idx_key}_acc",
                )

            submit = st.form_submit_button(
                "💾 取込内容を保存（経費精算書の候補に登録）",
                type='primary', width='stretch',
            )

        if submit:
            if rcp_amt <= 0:
                st.error("金額は1以上を入力してください")
            else:
                import json as _json
                ocr_result_json = (
                    _json.dumps({
                        **ocr_data,
                        # ユーザ編集後の値を上書きで保存
                        'date': rcp_date.strftime('%Y-%m-%d'),
                        'amount': int(rcp_amt),
                        'vendor': rcp_vendor,
                        'purpose': rcp_purpose,
                        'suggested_account': rcp_acc,
                    }, ensure_ascii=False)
                )
                rec = {
                    'file_path': str(r['path']),
                    'file_name': r['name'],
                    'file_size': r['size'],
                    'file_mtime': r['mtime'],
                    'corporation': corp,
                    'facility_folder': r['facility'],
                    'transaction_date': rcp_date.strftime('%Y-%m-%d'),
                    'year_month': rcp_date.strftime('%Y-%m'),
                    'vendor': rcp_vendor,
                    'purpose': rcp_purpose,
                    'total_amount': int(rcp_amt),
                    'payee': rcp_payee,
                    'payment_status': 'pending',
                    'ocr_status': 'success' if ocr_data else 'manual',
                    'ocr_confidence': ocr_data.get('confidence'),
                    'ocr_model': existing.get('model') if existing else None,
                    'ocr_input_tokens': (existing or {}).get(
                        'usage', {}).get('input_tokens'),
                    'ocr_output_tokens': (existing or {}).get(
                        'usage', {}).get('output_tokens'),
                    'ocr_result_json': ocr_result_json,
                    'note': None,
                }
                rid = db.upsert_cash_receipt(rec)
                st.success(
                    f"✅ レシート #{rid} を登録しました ／ "
                    f"{rcp_date} {rcp_vendor} ¥{rcp_amt:,}"
                )


def _list_cash_receipt_files(storage_dir: Path) -> list[dict]:
    if not storage_dir.exists():
        return []
    out: list[dict] = []
    seen = set()
    for p in storage_dir.rglob('*'):
        try:
            if not p.is_file():
                continue
            if not _is_receipt_file(p):
                continue
            if p in seen:
                continue
            seen.add(p)
            rel = p.relative_to(storage_dir)
            parts = rel.parts
            facility = ''
            status_label = ''
            year_month_label = ''
            if len(parts) >= 1:
                first = parts[0]
                if first in ('_処理済み', '_要確認'):
                    status_label = first.lstrip('_')
                    if first == '_処理済み' and len(parts) >= 4:
                        year_month_label = parts[1]
                        facility = parts[2]
                    elif len(parts) >= 3:
                        facility = parts[1]
                    else:
                        facility = '(未分類)'
                else:
                    facility = first
            stat = p.stat()
            out.append({
                'path': p, 'name': p.name, 'rel': str(rel),
                'facility': facility,
                'status': status_label,
                'year_month_folder': year_month_label,
                'kind': _receipt_kind(p),
                'ext': p.suffix.lower(),
                'size': stat.st_size, 'mtime': stat.st_mtime,
            })
        except (PermissionError, OSError):
            continue
    out.sort(key=lambda x: (x['facility'], -x['mtime']))
    return out


with parent_cash:
    st.markdown(
        "<div style='background:#fef3c7; border-left:4px solid #f59e0b; "
        "padding:10px 14px; border-radius:6px; margin-bottom:14px;'>"
        "<b>💴 現金立替レシート清算</b>"
        "<br><span style='color:#78350f; font-size:13px;'>"
        "施設フォルダにアップロードした立替レシートをOCRで取り込み、"
        "出納帳行と一般経費精算書を生成します（たたき台）。"
        "</span></div>",
        unsafe_allow_html=True,
    )

    (cash_tab_import, cash_tab_summary, cash_tab_fac, cash_tab_acc,
     cash_tab_vendor, cash_tab_book, cash_tab_receipts,
     cash_tab_form) = st.tabs([
        "📥 取込",
        "📊 サマリ",
        "🏢 施設別",
        "📚 勘定科目別",
        "🏪 購入先別",
        "📒 出納帳",
        "📷 レシート",
        "📄 経費精算書",
    ])

    # ----- 📥 取込 -----
    with cash_tab_import:
        existing_cash_yms = db.list_cash_year_months()
        existing_cash_corps = db.list_cash_corporations()
        cc1, cc2, cc3 = st.columns(3)
        cc1.metric("取込済 月数", f"{len(existing_cash_yms)}")
        cc2.metric("法人数", f"{len(existing_cash_corps)}")
        cc3.metric("最新月", existing_cash_yms[0] if existing_cash_yms else "—")

        # ----- 施設フォルダ自動作成ボタン -----
        st.markdown("### 📁 レシート格納用 施設フォルダ")
        cash_root_path = CASH_GDRIVE_PARENT / CASH_GDRIVE_FOLDER_NAME
        cur_subdirs = []
        if cash_root_path.exists():
            cur_subdirs = sorted(
                [p.name for p in cash_root_path.iterdir() if p.is_dir()]
            )
        master_codes = cash_advance_parser.list_facility_codes(None)

        fc1, fc2 = st.columns([3, 1])
        with fc1:
            st.caption(
                f"**格納先**: `{cash_root_path}`  ／  "
                f"既存フォルダ: {len(cur_subdirs)} 件 ／ "
                f"施設マスタ: {len(master_codes)} 件"
            )
        with fc2:
            if st.button("🏗️ 施設フォルダを作成/更新",
                          key='cash_make_folders', width='stretch'):
                with st.spinner("フォルダを作成中..."):
                    res = cash_advance_parser.ensure_facility_folders(
                        cash_root_path, include_status_dirs=True,
                    )
                if res.get('errors'):
                    for e in res['errors']:
                        st.error(e)
                st.success(
                    f"✅ 新規作成 {len(res['created'])} / "
                    f"既存 {len(res['existing'])}"
                )
                if res['created']:
                    with st.expander("新規作成フォルダ", expanded=False):
                        for n in res['created']:
                            st.text(n)
                st.rerun()

        if cur_subdirs:
            with st.expander(
                f"📂 既存施設フォルダ一覧 ({len(cur_subdirs)}件)",
                expanded=False,
            ):
                # マスタにあるか/ないかで色分け表示
                mset = set(master_codes) | {'_処理済み', '_要確認'}
                lines = []
                for n in cur_subdirs:
                    mark = '✅' if n in mset else '⚠️ (マスタ未登録)'
                    lines.append(f"- {mark}  {n}")
                st.markdown('\n'.join(lines))

        st.markdown("---")

        st.markdown("### 取込方法を選択")
        cash_src = st.radio(
            "取込ソース",
            ["📁 Dropbox 上の xlsm から取込", "📤 手動アップロード"],
            horizontal=True, key='cash_src',
        )

        if cash_src.startswith("📁"):
            xlsm_files = _list_cash_xlsm_files()
            if not xlsm_files:
                st.info(
                    f"出納帳 xlsm が見つかりません: {CASH_DROPBOX_DIR}"
                )
            else:
                st.markdown(
                    f"<div style='margin:8px 0;'>"
                    f"<span style='background:#dbeafe; color:#1e3a8a; "
                    f"padding:4px 12px; border-radius:6px; font-weight:700; "
                    f"font-size:13px;'>📂 Dropbox</span>&nbsp;&nbsp;"
                    f"<code>{CASH_DROPBOX_DIR}</code></div>",
                    unsafe_allow_html=True,
                )
                run_col, opt_col = st.columns([2, 1])
                with run_col:
                    cash_run = st.button(
                        "🚀 取込実行（出納帳シートのみ）",
                        type='primary', key='cash_dropbox_run',
                        width='stretch',
                    )
                with opt_col:
                    cash_overwrite = st.checkbox(
                        "同月分を上書き", value=False, key='cash_overwrite',
                        help="チェック時は対象月のデータを削除してから取込",
                    )

                preview_rows = []
                for f in xlsm_files:
                    mtime_str = datetime.fromtimestamp(
                        f['mtime']).strftime('%Y-%m-%d %H:%M')
                    preview_rows.append({
                        '法人': f['corporation'],
                        'ファイル': f['name'],
                        'サイズ': f"{f['size']/1024:.1f} KB",
                        '最終更新': mtime_str,
                    })
                st.dataframe(
                    pd.DataFrame(preview_rows), hide_index=True, width='stretch',
                )

                if cash_run:
                    sub_lookup = {}
                    for s in db.list_pl_subunits():
                        sub_lookup[s['excel_name']] = s['id']
                        if s.get('display_name'):
                            sub_lookup[s['display_name']] = s['id']

                    total_inserted = 0
                    total_skipped = 0
                    all_yms_taken = set()
                    with st.spinner("出納帳シートを解析中..."):
                        for f in xlsm_files:
                            corp = f['corporation']
                            if corp == '(未判定)':
                                continue
                            with open(f['path'], 'rb') as fp:
                                result = cash_advance_parser.parse_cashbook_workbook(
                                    fp, corporation=corp,
                                    subunit_lookup=sub_lookup,
                                )
                            if cash_overwrite:
                                yms_in_file = sorted({
                                    r['year_month'] for r in result['rows']
                                    if r.get('year_month')
                                })
                                for ym in yms_in_file:
                                    db.delete_cash_by_year_month(
                                        ym, corporation=corp)
                            ins = db.insert_cash_entries(
                                result['rows'], filename=f['name'],
                                corporation=corp, file_hash=result['file_hash'],
                            )
                            total_inserted += ins['inserted']
                            total_skipped += ins['skipped']
                            all_yms_taken.update(ins['year_months'])

                            st.write(
                                f"- **{corp}** / `{f['name']}`: "
                                f"新規 {ins['inserted']} / "
                                f"スキップ {ins['skipped']}"
                            )
                            if result['unknown_departments']:
                                with st.expander(
                                    f"⚠️ {f['name']} : 未マッチ部門 "
                                    f"{len(result['unknown_departments'])}件"
                                ):
                                    for d in sorted(result['unknown_departments']):
                                        st.text(d)

                    st.success(
                        f"✅ 取込完了 — 新規 {total_inserted} / "
                        f"スキップ {total_skipped} / 対象月 "
                        f"{sorted(all_yms_taken)}"
                    )
                    st.rerun()

        else:
            cu1, cu2 = st.columns([2, 1])
            with cu1:
                cash_uploaded = st.file_uploader(
                    "現金出納帳 xlsm/xlsx",
                    type=['xlsm', 'xlsx'], accept_multiple_files=True,
                    key='cash_uploader',
                )
            with cu2:
                cash_corp_input = st.selectbox(
                    "法人ラベル", options=['医療法人', 'NPO法人'],
                    key='cash_corp_select',
                )

            run_manual = st.button(
                "🚀 取込実行（アップロードファイル）",
                type='primary', key='cash_manual_run',
                width='stretch',
                disabled=not (cash_uploaded and cash_corp_input),
            )
            if run_manual:
                sub_lookup = {}
                for s in db.list_pl_subunits():
                    sub_lookup[s['excel_name']] = s['id']
                    if s.get('display_name'):
                        sub_lookup[s['display_name']] = s['id']

                total_inserted = 0
                total_skipped = 0
                all_yms_taken = set()
                with st.spinner("解析中..."):
                    for f in cash_uploaded:
                        auto_corp = cash_advance_parser.detect_corporation_from_filename(
                            f.name) or cash_corp_input
                        result = cash_advance_parser.parse_cashbook_workbook(
                            f, corporation=auto_corp, subunit_lookup=sub_lookup,
                        )
                        ins = db.insert_cash_entries(
                            result['rows'], filename=f.name,
                            corporation=auto_corp,
                            file_hash=result['file_hash'],
                        )
                        total_inserted += ins['inserted']
                        total_skipped += ins['skipped']
                        all_yms_taken.update(ins['year_months'])

                st.success(
                    f"✅ 取込完了 — 新規 {total_inserted} / "
                    f"スキップ {total_skipped} / 対象月 {sorted(all_yms_taken)}"
                )
                st.rerun()

        cash_imports = db.list_cash_imports(limit=10)
        if cash_imports:
            with st.expander(
                f"📜 取込履歴（直近{len(cash_imports)}件）",
                expanded=False,
            ):
                st.dataframe(pd.DataFrame([
                    {
                        '取込日時': r['imported_at'],
                        '法人': r['corporation'],
                        'ファイル': r['source_filename'],
                        '行数': r['row_count'],
                        '新規': r['inserted_count'],
                        'スキップ': r['skipped_count'],
                        '対象月': r['year_months'],
                    }
                    for r in cash_imports
                ]), hide_index=True, width='stretch')

    # ---- 共通フィルタ・データ取得 ----
    cash_yms_all = db.list_cash_year_months()
    cash_corps_all = db.list_cash_corporations()

    def _cash_filter_box(key_prefix: str) -> dict:
        yms_asc = sorted(cash_yms_all)
        cols = st.columns([2, 2, 2])
        with cols[0]:
            start_ym = st.selectbox(
                "期間（開始月）", yms_asc, index=0,
                key=f'{key_prefix}_start',
            )
        with cols[1]:
            end_ym = st.selectbox(
                "期間（終了月）", yms_asc,
                index=len(yms_asc) - 1, key=f'{key_prefix}_end',
            )
        with cols[2]:
            corp = st.selectbox(
                "法人", options=['（全法人）'] + cash_corps_all,
                key=f'{key_prefix}_corp',
            )
        if start_ym > end_ym:
            start_ym, end_ym = end_ym, start_ym
        sel = [ym for ym in yms_asc if start_ym <= ym <= end_ym]
        return {
            'year_months': sel,
            'corporation': None if corp == '（全法人）' else corp,
            'start_ym': start_ym, 'end_ym': end_ym,
        }

    def _cash_query_df(year_months, corporation,
                       departments=None, accounts=None, keyword=None):
        rows = db.query_cash_entries(
            year_months=list(year_months) if year_months else None,
            corporation=corporation,
            departments=list(departments) if departments else None,
            accounts=list(accounts) if accounts else None,
            keyword=keyword,
        )
        return pd.DataFrame(rows)

    if not cash_yms_all:
        for _t in (cash_tab_summary, cash_tab_fac, cash_tab_acc,
                    cash_tab_vendor, cash_tab_book, cash_tab_form):
            with _t:
                st.info(
                    "📭 まだ取込済みの現金立替明細がありません。"
                    "「📥 取込」タブから出納帳xlsmを取り込んでください。"
                )

    # ----- 📊 サマリ (出金/入金切替 + 施設別) -----
    with cash_tab_summary:
        if cash_yms_all:
            st.markdown("### 📊 現金立替 サマリ")
            f = _cash_filter_box('cash_sum')
            view_mode = st.radio(
                "集計ベース",
                ['💸 出金ベース', '💰 入金ベース'],
                horizontal=True, key='cash_sum_view',
            )
            is_expense_view = view_mode.startswith('💸')

            df = _cash_query_df(tuple(f['year_months']), f['corporation'])
            if df.empty:
                st.info("条件に該当するデータがありません。")
            else:
                exp_df = df[df['entry_kind'] == 'expense']
                inc_df = df[df['entry_kind'] == 'income']
                tot_out = int(exp_df['amount_out'].sum())
                tot_in = int(inc_df['amount_in'].sum())

                c1, c2, c3, c4, c5 = st.columns(5)
                c1.metric("出金合計", _format_yen(tot_out))
                c2.metric("入金合計", _format_yen(tot_in))
                c3.metric("出金件数", f"{len(exp_df):,}")
                c4.metric("入金件数", f"{len(inc_df):,}")
                c5.metric("差引", _format_yen(tot_in - tot_out))

                target_df = exp_df if is_expense_view else inc_df
                amt_col = 'amount_out' if is_expense_view else 'amount_in'
                amt_label = '出金合計' if is_expense_view else '入金合計'

                if target_df.empty:
                    st.info(f"{view_mode}に該当するデータがありません。")
                else:
                    st.markdown(f"#### 月別推移（{view_mode}）")
                    monthly = target_df.groupby(
                        'year_month', as_index=False).agg(
                        金額=(amt_col, 'sum'),
                        件数=(amt_col, 'count'),
                    ).sort_values('year_month')
                    mdisp = monthly.copy()
                    mdisp['金額'] = mdisp['金額'].apply(_format_yen)
                    mdisp['件数'] = mdisp['件数'].apply(lambda v: f"{v:,}")
                    mdisp.columns = ['年月', amt_label, '件数']
                    st.dataframe(mdisp, hide_index=True, width='stretch')

                    # 施設別 (財務／経理マスタ準拠)
                    st.markdown(f"#### 🏢 施設別（{view_mode}）")
                    master_codes = (
                        cash_advance_parser.list_facility_codes(
                            f['corporation'])
                        if f['corporation']
                        else cash_advance_parser.list_facility_codes(None)
                    )
                    fac_agg = target_df.groupby(
                        'department_raw', as_index=False).agg(
                        合計金額=(amt_col, 'sum'),
                        件数=(amt_col, 'count'),
                    )
                    fac_agg = fac_agg[fac_agg['department_raw'] != '']
                    amt_lookup = dict(zip(
                        fac_agg['department_raw'], fac_agg['合計金額']))
                    cnt_lookup = dict(zip(
                        fac_agg['department_raw'], fac_agg['件数']))
                    rows_disp = []
                    seen = set()
                    for code in master_codes:
                        rows_disp.append({
                            '施設': code,
                            amt_label: int(amt_lookup.get(code, 0)),
                            '件数': int(cnt_lookup.get(code, 0)),
                        })
                        seen.add(code)
                    for d in fac_agg['department_raw']:
                        if d not in seen:
                            rows_disp.append({
                                '施設': f'⚠️ {d} (マスタ未登録)',
                                amt_label: int(amt_lookup.get(d, 0)),
                                '件数': int(cnt_lookup.get(d, 0)),
                            })
                    fdf = pd.DataFrame(rows_disp).sort_values(
                        amt_label, ascending=False).reset_index(drop=True)
                    fdf_disp = fdf.copy()
                    fdf_disp[amt_label] = fdf_disp[amt_label].apply(
                        lambda v: _format_yen(v) if v else '—')
                    fdf_disp['件数'] = fdf_disp['件数'].apply(
                        lambda v: f"{v:,}" if v else '—')
                    st.dataframe(
                        fdf_disp, hide_index=True, width='stretch',
                        height=min(60 + 36 * len(fdf_disp), 700),
                    )
                    st.caption(
                        f"📌 マスタ施設 {len(master_codes)}件 ／ "
                        f"対象期間 {f['start_ym']} 〜 {f['end_ym']}"
                    )

    # ----- 🏢 施設別 (出金/入金両軸) -----
    with cash_tab_fac:
        if cash_yms_all:
            st.markdown("### 🏢 施設別 費用分析（現金立替）")
            f = _cash_filter_box('cash_fac')
            view = st.radio(
                "集計ベース",
                ['💸 出金ベース', '💰 入金ベース', '🔁 出入金併記'],
                horizontal=True, key='cash_fac_view',
            )
            df = _cash_query_df(tuple(f['year_months']), f['corporation'])
            if df.empty:
                st.info("該当データなし")
            else:
                exp_df = df[df['entry_kind'] == 'expense']
                inc_df = df[df['entry_kind'] == 'income']

                if view.startswith('💸'):
                    agg = exp_df.groupby(
                        'department_raw', as_index=False).agg(
                        合計金額=('amount_out', 'sum'),
                        件数=('amount_out', 'count'),
                        平均=('amount_out', 'mean'),
                    ).sort_values('合計金額', ascending=False)
                    agg = agg[agg['department_raw'] != '']
                    if agg.empty:
                        st.info("出金データなし")
                    else:
                        agg['平均'] = agg['平均'].fillna(0).astype(int)
                        disp = agg.copy().reset_index(drop=True)
                        disp.insert(0, '順位', disp.index + 1)
                        disp['合計金額'] = disp['合計金額'].apply(_format_yen)
                        disp['平均'] = disp['平均'].apply(_format_yen)
                        disp['件数'] = disp['件数'].apply(
                            lambda v: f"{v:,}")
                        disp.columns = [
                            '順位', '施設', '出金合計', '件数', '平均単価']
                        st.dataframe(disp, hide_index=True, width='stretch')
                elif view.startswith('💰'):
                    agg = inc_df.groupby(
                        'department_raw', as_index=False).agg(
                        合計金額=('amount_in', 'sum'),
                        件数=('amount_in', 'count'),
                    ).sort_values('合計金額', ascending=False)
                    agg = agg[agg['department_raw'] != '']
                    if agg.empty:
                        st.info("入金データなし")
                    else:
                        disp = agg.copy().reset_index(drop=True)
                        disp.insert(0, '順位', disp.index + 1)
                        disp['合計金額'] = disp['合計金額'].apply(_format_yen)
                        disp['件数'] = disp['件数'].apply(
                            lambda v: f"{v:,}")
                        disp.columns = ['順位', '施設', '入金合計', '件数']
                        st.dataframe(disp, hide_index=True, width='stretch')
                else:  # 出入金併記
                    out_agg = exp_df.groupby(
                        'department_raw')['amount_out'].sum()
                    in_agg = inc_df.groupby(
                        'department_raw')['amount_in'].sum()
                    facs = sorted(set(out_agg.index) | set(in_agg.index))
                    rows_d = []
                    for d in facs:
                        if not d:
                            continue
                        rows_d.append({
                            '施設': d,
                            '出金': int(out_agg.get(d, 0)),
                            '入金': int(in_agg.get(d, 0)),
                            '差引': int(in_agg.get(d, 0))
                            - int(out_agg.get(d, 0)),
                        })
                    fdf = pd.DataFrame(rows_d).sort_values(
                        '出金', ascending=False).reset_index(drop=True)
                    fdf_disp = fdf.copy()
                    for c in ('出金', '入金', '差引'):
                        fdf_disp[c] = fdf_disp[c].apply(
                            lambda v: _format_yen(v) if v else '—')
                    st.dataframe(fdf_disp, hide_index=True, width='stretch')

    # ----- 📚 勘定科目別 (施設別ドリルダウン + クロス集計) -----
    with cash_tab_acc:
        if cash_yms_all:
            st.markdown("### 📚 勘定科目別（現金立替）")
            f = _cash_filter_box('cash_acc')
            view = st.radio(
                "集計ベース",
                ['💸 出金', '💰 入金'], horizontal=True,
                key='cash_acc_view',
            )
            is_exp = view.startswith('💸')
            amt_col = 'amount_out' if is_exp else 'amount_in'
            kind_target = 'expense' if is_exp else 'income'
            amt_label = '出金合計' if is_exp else '入金合計'

            df = _cash_query_df(tuple(f['year_months']), f['corporation'])
            df = df[df['entry_kind'] == kind_target]
            if df.empty:
                st.info("該当データなし")
            else:
                agg = df.groupby('debit_account', as_index=False).agg(
                    合計金額=(amt_col, 'sum'),
                    件数=(amt_col, 'count'),
                ).sort_values('合計金額', ascending=False)
                agg = agg[agg['debit_account'] != '']
                disp = agg.copy().reset_index(drop=True)
                disp.insert(0, '順位', disp.index + 1)
                disp['合計金額'] = disp['合計金額'].apply(_format_yen)
                disp['件数'] = disp['件数'].apply(lambda v: f"{v:,}")
                disp.columns = ['順位', '勘定科目', amt_label, '件数']
                st.dataframe(disp, hide_index=True, width='stretch')

                # 施設別ドリルダウン
                st.markdown("#### 🔍 施設別ドリルダウン")
                fac_options = ['（全施設）'] + sorted([
                    o for o in df['department_raw'].dropna().unique() if o
                ])
                sel_fac = st.selectbox(
                    "施設を選択", options=fac_options,
                    key='cash_acc_fac_drill',
                )
                if sel_fac != '（全施設）':
                    sub = df[df['department_raw'] == sel_fac]
                    sub_agg = sub.groupby(
                        'debit_account', as_index=False).agg(
                        金額=(amt_col, 'sum'),
                        件数=(amt_col, 'count'),
                    ).sort_values('金額', ascending=False)
                    sub_agg = sub_agg[sub_agg['debit_account'] != '']
                    if sub_agg.empty:
                        st.info("該当行なし")
                    else:
                        sub_disp = sub_agg.copy()
                        sub_disp['金額'] = sub_disp['金額'].apply(_format_yen)
                        sub_disp['件数'] = sub_disp['件数'].apply(
                            lambda v: f"{v:,}")
                        sub_disp.columns = ['勘定科目', amt_label, '件数']
                        st.dataframe(
                            sub_disp, hide_index=True, width='stretch')

                # 科目 × 施設のクロス集計表
                st.markdown(f"#### 🔢 科目 × 施設 クロス集計（{view}）")
                top_n = st.slider(
                    "対象科目数（上位）", 3, max(3, len(agg)),
                    min(8, len(agg)), key='cash_acc_topn',
                )
                top_accs = list(agg.head(top_n)['debit_account'])
                pivot = df[df['debit_account'].isin(top_accs)].pivot_table(
                    index='department_raw', columns='debit_account',
                    values=amt_col, aggfunc='sum', fill_value=0,
                )
                pivot = pivot[pivot.index != '']
                if not pivot.empty:
                    pivot['合計'] = pivot.sum(axis=1)
                    pivot = pivot.sort_values('合計', ascending=False)
                    pivot = pivot.reset_index()
                    pivot.rename(columns={'department_raw': '施設'},
                                  inplace=True)
                    pdisp = pivot.copy()
                    for c in pdisp.columns:
                        if c == '施設':
                            continue
                        pdisp[c] = pdisp[c].apply(
                            lambda v: f"{int(v):,}" if v else '—')
                    st.dataframe(
                        pdisp, hide_index=True, width='stretch',
                        height=min(60 + 36 * len(pdisp), 700),
                    )

    # ----- 🏪 購入先別 (施設別ドリルダウン) -----
    with cash_tab_vendor:
        if cash_yms_all:
            st.markdown("### 🏪 購入先別（現金立替・備考から推定）")
            f = _cash_filter_box('cash_vendor')
            view = st.radio(
                "集計ベース",
                ['💸 出金', '💰 入金'], horizontal=True,
                key='cash_vendor_view',
            )
            is_exp = view.startswith('💸')
            amt_col = 'amount_out' if is_exp else 'amount_in'
            kind_target = 'expense' if is_exp else 'income'
            amt_label = '出金合計' if is_exp else '入金合計'

            df = _cash_query_df(tuple(f['year_months']), f['corporation'])
            df = df[(df['entry_kind'] == kind_target) & (df['vendor'] != '')]
            if df.empty:
                st.info("購入先が抽出できる行がありません。")
            else:
                agg = df.groupby('vendor', as_index=False).agg(
                    合計金額=(amt_col, 'sum'),
                    件数=(amt_col, 'count'),
                ).sort_values('合計金額', ascending=False).head(50)
                disp = agg.copy().reset_index(drop=True)
                disp.insert(0, '順位', disp.index + 1)
                disp['合計金額'] = disp['合計金額'].apply(_format_yen)
                disp['件数'] = disp['件数'].apply(lambda v: f"{v:,}")
                disp.columns = ['順位', '購入先', amt_label, '件数']
                st.dataframe(disp, hide_index=True, width='stretch')

                # 施設別ドリルダウン
                st.markdown("#### 🔍 施設別ドリルダウン")
                fac_options = ['（全施設）'] + sorted([
                    o for o in df['department_raw'].dropna().unique() if o
                ])
                sel_fac = st.selectbox(
                    "施設を選択", options=fac_options,
                    key='cash_vendor_fac_drill',
                )
                if sel_fac != '（全施設）':
                    sub = df[df['department_raw'] == sel_fac]
                    sub_agg = sub.groupby('vendor', as_index=False).agg(
                        金額=(amt_col, 'sum'),
                        件数=(amt_col, 'count'),
                    ).sort_values('金額', ascending=False).head(30)
                    if sub_agg.empty:
                        st.info("該当行なし")
                    else:
                        sub_disp = sub_agg.copy()
                        sub_disp['金額'] = sub_disp['金額'].apply(_format_yen)
                        sub_disp['件数'] = sub_disp['件数'].apply(
                            lambda v: f"{v:,}")
                        sub_disp.columns = ['購入先', amt_label, '件数']
                        st.dataframe(
                            sub_disp, hide_index=True, width='stretch')

    # ----- 📒 出納帳 (法人別タブ + 手入力UI + Excel ダウンロード) -----
    with cash_tab_book:
        if cash_yms_all:
            st.markdown("### 📒 出納帳")
            st.caption(
                "📤 ダウンロードした Excel は次の Drive 共有フォルダに"
                f"アップロードしてください: {CASH_DRIVE_OUTPUT_URL}"
            )

            # 法人別タブ（医療法人 / NPO法人）
            book_subtabs = st.tabs([f"🏥 {c}" for c in cash_corps_all])

            for tab_idx, sub_tab in enumerate(book_subtabs):
                book_corp = cash_corps_all[tab_idx]
                with sub_tab:
                    yms_asc = sorted(cash_yms_all)
                    book_ym = st.selectbox(
                        "対象月",
                        options=list(reversed(yms_asc)),
                        index=0, key=f'cash_book_ym_{tab_idx}',
                    )

                    # ===== 手入力フォーム =====
                    with st.expander(
                        "✍️ 出納帳に1行を手入力で追加", expanded=False,
                    ):
                        # 過去データから候補リストを取得
                        all_accs = db.list_cash_accounts()
                        # 施設候補: 法人マスタから
                        fac_codes = cash_advance_parser.list_facility_codes(
                            book_corp)
                        if not fac_codes:
                            fac_codes = sorted([
                                d for d in db.list_cash_departments()
                            ])

                        with st.form(
                            f"cash_manual_form_{tab_idx}",
                            clear_on_submit=True,
                        ):
                            mc1, mc2, mc3 = st.columns([1, 2, 2])
                            with mc1:
                                m_date = st.date_input(
                                    "日付",
                                    value=datetime.now().date(),
                                    key=f'm_date_{tab_idx}',
                                )
                            with mc2:
                                m_fac = st.selectbox(
                                    "施設（部門）",
                                    options=fac_codes,
                                    key=f'm_fac_{tab_idx}',
                                )
                            with mc3:
                                acct_options = (
                                    all_accs
                                    + ['（その他・直接入力）']
                                ) if all_accs else ['（その他・直接入力）']
                                m_acc_sel = st.selectbox(
                                    "勘定科目",
                                    options=acct_options,
                                    key=f'm_acc_sel_{tab_idx}',
                                )
                                if m_acc_sel == '（その他・直接入力）':
                                    m_acc = st.text_input(
                                        "勘定科目（直接入力）",
                                        key=f'm_acc_{tab_idx}',
                                    )
                                else:
                                    m_acc = m_acc_sel

                            mc4, mc5 = st.columns([3, 1])
                            with mc4:
                                m_desc = st.text_input(
                                    "備考（商品/支払先/用途等）",
                                    key=f'm_desc_{tab_idx}',
                                )
                            with mc5:
                                m_tax = st.selectbox(
                                    "税区分",
                                    options=[
                                        '課対仕入10%', '課対仕入8%（軽）',
                                        '対象外',
                                    ],
                                    key=f'm_tax_{tab_idx}',
                                )

                            mc6, mc7, mc8 = st.columns(3)
                            with mc6:
                                m_in = st.number_input(
                                    "入金 (¥)", min_value=0, step=1,
                                    value=0, key=f'm_in_{tab_idx}',
                                )
                            with mc7:
                                m_out = st.number_input(
                                    "出金 (¥)", min_value=0, step=1,
                                    value=0, key=f'm_out_{tab_idx}',
                                )
                            with mc8:
                                m_payee = st.text_input(
                                    "立替者（任意）",
                                    key=f'm_payee_{tab_idx}',
                                )

                            submit = st.form_submit_button(
                                "➕ 出納帳に追加", type='primary',
                                width='stretch',
                            )

                        if submit:
                            if m_in == 0 and m_out == 0:
                                st.error(
                                    "入金 or 出金 のどちらかを入力してください"
                                )
                            elif not m_desc:
                                st.error("備考を入力してください")
                            else:
                                kind = 'income' if m_in > 0 else 'expense'
                                tx = m_date.strftime('%Y-%m-%d')
                                eid = db.insert_manual_cash_entry({
                                    'corporation': book_corp,
                                    'transaction_date': tx,
                                    'year_month': tx[:7],
                                    'debit_account': m_acc or '',
                                    'tax_class': m_tax,
                                    'department_raw': m_fac,
                                    'amount_in': int(m_in),
                                    'amount_out': int(m_out),
                                    'description': m_desc,
                                    'payee': m_payee,
                                    'entry_kind': kind,
                                })
                                st.success(
                                    f"✅ 行 #{eid} を追加しました"
                                    f"({tx} / {m_fac} / {m_acc} / "
                                    f"{'入金' if m_in else '出金'} "
                                    f"{_format_yen(m_in or m_out)})"
                                )
                                st.rerun()

                    # ===== 出納帳本体 =====
                    rows = db.query_cash_entries(
                        year_months=[book_ym], corporation=book_corp,
                        order_desc=True,
                    )
                    if not rows:
                        st.info("対象月にデータがありません。")
                    else:
                        ordered = sorted(
                            rows,
                            key=lambda r: (r['transaction_date'], r['id']),
                        )
                        bal = 0
                        for r in ordered:
                            bal += (int(r.get('amount_in') or 0)
                                    - int(r.get('amount_out') or 0))
                            r['balance_calc'] = bal
                        ordered_desc = list(reversed(ordered))

                        disp_rows = []
                        for r in ordered_desc:
                            ai = int(r.get('amount_in') or 0)
                            ao = int(r.get('amount_out') or 0)
                            disp_rows.append({
                                '日付': r['transaction_date'],
                                '勘定科目': r.get('debit_account') or '',
                                '税区分': r.get('tax_class') or '',
                                '部門 (施設)': r.get('department_raw') or '',
                                '備考': r.get('description') or '',
                                '入金': _format_yen(ai) if ai else '',
                                '出金': _format_yen(ao) if ao else '',
                                '差引残高': _format_yen(r['balance_calc']),
                            })
                        book_df = pd.DataFrame(disp_rows)
                        st.dataframe(
                            book_df, hide_index=True, width='stretch',
                            height=min(60 + 36 * len(book_df), 700),
                        )
                        st.caption(
                            f"📌 {len(book_df)}件 ／ "
                            f"入金: {_format_yen(sum(int(r.get('amount_in') or 0) for r in ordered))} ／ "
                            f"出金: {_format_yen(sum(int(r.get('amount_out') or 0) for r in ordered))} ／ "
                            f"月末残高: {_format_yen(ordered[-1]['balance_calc'])}"
                        )

                        # ---- Excel ダウンロード (Spreadsheet 用) ----
                        dl_cols = st.columns([2, 2])
                        with dl_cols[0]:
                            # Excel 出力 (openpyxl)
                            import openpyxl
                            wb_out = openpyxl.Workbook()
                            ws_out = wb_out.active
                            ws_out.title = f"{book_corp}_{book_ym}"
                            headers = ['日付', '勘定科目', '税区分',
                                       '部門 (施設)', '備考',
                                       '入金', '出金', '差引残高']
                            for c_idx, h in enumerate(headers, start=1):
                                ws_out.cell(row=1, column=c_idx).value = h
                            for r_idx, r in enumerate(ordered, start=2):
                                ws_out.cell(row=r_idx, column=1).value = (
                                    datetime.strptime(
                                        r['transaction_date'], '%Y-%m-%d')
                                )
                                ws_out.cell(row=r_idx, column=2).value = (
                                    r.get('debit_account') or '')
                                ws_out.cell(row=r_idx, column=3).value = (
                                    r.get('tax_class') or '')
                                ws_out.cell(row=r_idx, column=4).value = (
                                    r.get('department_raw') or '')
                                ws_out.cell(row=r_idx, column=5).value = (
                                    r.get('description') or '')
                                ai = int(r.get('amount_in') or 0)
                                ao = int(r.get('amount_out') or 0)
                                if ai:
                                    ws_out.cell(row=r_idx, column=6).value = ai
                                if ao:
                                    ws_out.cell(row=r_idx, column=7).value = ao
                                ws_out.cell(row=r_idx, column=8).value = (
                                    r['balance_calc'])

                            # 列幅調整
                            for col_letter, w in zip(
                                ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H'],
                                [12, 14, 14, 24, 50, 12, 12, 14],
                            ):
                                ws_out.column_dimensions[col_letter].width = w

                            buf_xlsx = io.BytesIO()
                            wb_out.save(buf_xlsx)
                            buf_xlsx.seek(0)
                            xlsx_name = (
                                f"出納帳_{book_corp}_{book_ym}.xlsx"
                            )
                            st.download_button(
                                "📥 出納帳 Excel をダウンロード",
                                data=buf_xlsx.getvalue(),
                                file_name=xlsx_name,
                                mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                                key=f'cash_book_xlsx_{tab_idx}',
                                width='stretch',
                            )

                        with dl_cols[1]:
                            csv_bytes = book_df.to_csv(
                                index=False).encode('utf-8-sig')
                            st.download_button(
                                "📥 出納帳 CSV ダウンロード",
                                data=csv_bytes,
                                file_name=(
                                    f"出納帳_{book_corp}_{book_ym}.csv"
                                ),
                                mime='text/csv',
                                key=f'cash_book_csv_{tab_idx}',
                                width='stretch',
                            )

                        st.info(
                            "💡 ダウンロードした **Excel** を Drive の指定フォルダに"
                            f"アップロードしてください → "
                            f"[📁 出納帳保存先を開く]({CASH_DRIVE_OUTPUT_URL})"
                        )

    # ----- 📷 レシート -----
    with cash_tab_receipts:
        st.markdown("### 📷 レシート（現金立替）")
        st.caption(
            f"格納先: `{CASH_GDRIVE_PARENT / CASH_GDRIVE_FOLDER_NAME}`"
            "／施設フォルダごとに写メ・PDFを置いてください。"
        )

        cash_storage = _resolve_cash_receipt_root()
        if cash_storage is None:
            st.warning(
                f"格納フォルダが見つかりません: "
                f"{CASH_GDRIVE_PARENT / CASH_GDRIVE_FOLDER_NAME}\n\n"
                "Google Drive for Desktop を起動するか、"
                "フォルダを作成してください。"
            )
        else:
            with st.spinner("レシート読込中..."):
                cash_receipts = _list_cash_receipt_files(cash_storage)

            if not cash_receipts:
                st.info(
                    "レシートファイルがありません。"
                    "施設フォルダ配下に JPG/PNG/HEIC/PDF を置いてください。"
                )
            else:
                n_total = len(cash_receipts)
                n_img = sum(1 for r in cash_receipts if r['kind'] == 'image')
                n_pdf = sum(1 for r in cash_receipts if r['kind'] == 'pdf')
                facilities = sorted({
                    r['facility'] for r in cash_receipts if r['facility']
                })

                rc1, rc2, rc3, rc4 = st.columns(4)
                rc1.metric("レシート総数", f"{n_total:,}")
                rc2.metric("📸 写メ", f"{n_img:,}")
                rc3.metric("📄 PDF", f"{n_pdf:,}")
                rc4.metric("施設数", f"{len(facilities)}")

                fc1, fc2, fc3 = st.columns([2, 2, 2])
                with fc1:
                    sel_fac = st.multiselect(
                        "施設で絞り込み", options=facilities,
                        default=[], placeholder="未選択 = 全施設",
                        key='cash_rcp_fac',
                    )
                with fc2:
                    sel_kind = st.multiselect(
                        "種別", options=['📸 写メ', '📄 PDF'],
                        default=[], key='cash_rcp_kind',
                    )
                with fc3:
                    sort_mode = st.selectbox(
                        "並び順",
                        options=['新しい順', '古い順', '施設順'],
                        index=0, key='cash_rcp_sort',
                    )

                filt = list(cash_receipts)
                if sel_fac:
                    filt = [r for r in filt if r['facility'] in sel_fac]
                if sel_kind:
                    ks = set()
                    if '📸 写メ' in sel_kind:
                        ks.add('image')
                    if '📄 PDF' in sel_kind:
                        ks.add('pdf')
                    filt = [r for r in filt if r['kind'] in ks]
                if sort_mode == '新しい順':
                    filt.sort(key=lambda r: -r['mtime'])
                elif sort_mode == '古い順':
                    filt.sort(key=lambda r: r['mtime'])
                else:
                    filt.sort(key=lambda r: (r['facility'], -r['mtime']))

                page_size = st.select_slider(
                    "1ページの表示件数",
                    options=[6, 12, 24, 48],
                    value=12, key='cash_rcp_page_size',
                )
                total_pages = max(1,
                    (len(filt) + page_size - 1) // page_size)
                page = 1
                if total_pages > 1:
                    page = st.number_input(
                        f"ページ (1〜{total_pages})", min_value=1,
                        max_value=total_pages, value=1, step=1,
                        key='cash_rcp_page',
                    )
                start = (page - 1) * page_size
                end = start + page_size
                items_view = filt[start:end]

                from datetime import datetime as _dt2
                cols_per_row = 3
                for i in range(0, len(items_view), cols_per_row):
                    row_items = items_view[i:i + cols_per_row]
                    cols = st.columns(cols_per_row)
                    for col, r in zip(cols, row_items):
                        with col:
                            with st.container(border=True):
                                mt = _dt2.fromtimestamp(r['mtime']).strftime(
                                    '%Y-%m-%d %H:%M')
                                kind_icon = '📸' if r['kind'] == 'image' else '📄'
                                st.markdown(
                                    f"**{kind_icon} {r['facility']}**<br>"
                                    f"<span style='color:#64748b; "
                                    f"font-size:11px;'>"
                                    f"{mt} ／ {_format_size(r['size'])}</span>",
                                    unsafe_allow_html=True,
                                )
                                try:
                                    with open(r['path'], 'rb') as fp:
                                        fb = fp.read()
                                except Exception as e:
                                    st.error(f"読込失敗: {e}")
                                    continue

                                if r['kind'] == 'image':
                                    if r['ext'] in ('.heic', '.heif'):
                                        jpg, err = _heic_to_jpeg_bytes(fb)
                                        if jpg is not None:
                                            st.image(jpg, width='stretch')
                                        else:
                                            st.info(f"HEIC変換失敗: {err}")
                                    else:
                                        try:
                                            st.image(fb, width='stretch')
                                        except Exception:
                                            pass
                                else:
                                    import base64 as _b64m
                                    try:
                                        b64 = _b64m.b64encode(fb).decode('ascii')
                                        st.markdown(
                                            f'<iframe src="data:application/pdf;'
                                            f'base64,{b64}" width="100%" '
                                            f'height="240" style="border:1px '
                                            f'solid #cbd5e1; border-radius:6px;"'
                                            f'></iframe>',
                                            unsafe_allow_html=True,
                                        )
                                    except Exception:
                                        pass

                                # 既存のOCR/取込状態をDBから取得
                                rcpt_db = db.get_cash_receipt_by_path(
                                    str(r['path']))
                                state_badge = ''
                                if rcpt_db:
                                    if rcpt_db.get('ocr_status') == 'success':
                                        state_badge = (
                                            "<span style='background:#d1fae5; "
                                            "color:#065f46; padding:1px 6px; "
                                            "border-radius:4px; font-size:10px; "
                                            "font-weight:600;'>OCR済</span>"
                                        )
                                    if rcpt_db.get('payment_status') == 'settled':
                                        state_badge += (
                                            "<span style='background:#dbeafe; "
                                            "color:#1e3a8a; padding:1px 6px; "
                                            "border-radius:4px; font-size:10px; "
                                            "font-weight:600; margin-left:4px;'>"
                                            "清算済</span>"
                                        )
                                if state_badge:
                                    st.markdown(state_badge,
                                                unsafe_allow_html=True)
                                st.caption(f"`{r['name']}`")

                                if r['kind'] == 'image':
                                    _cash_render_receipt_popover(
                                        r, fb, rcpt_db,
                                        idx_key=f"cashrcp_{start + items_view.index(r)}",
                                    )

    # ----- 📄 経費精算書 -----
    with cash_tab_form:
        if cash_yms_all:
            st.markdown("### 📄 一般経費精算書 出力")
            st.caption(
                "「一般経費」セクションに最大 25 行まで自動入力します"
                "（交通費は除外）。 既存の xlsm をテンプレートとして使用します。"
            )

            ef1, ef2, ef3 = st.columns([2, 2, 2])
            with ef1:
                form_corp = st.selectbox(
                    "法人", options=cash_corps_all, key='form_corp',
                )
            with ef2:
                form_ym = st.selectbox(
                    "対象月",
                    options=list(reversed(sorted(cash_yms_all))),
                    key='form_ym',
                )
            with ef3:
                form_payee = st.text_input(
                    "申請者氏名", value='鎌田 亮介', key='form_payee',
                )

            # ---- ソース選択: 出納帳 / 取込済レシート / 両方 ----
            source_mode = st.radio(
                "明細のソース",
                ['📒 出納帳から', '📷 取込済レシートから', '🔁 両方'],
                horizontal=True, key='form_source',
            )

            # 出納帳行
            entries_candidates: list[dict] = []
            if source_mode in ('📒 出納帳から', '🔁 両方'):
                entries_candidates = [
                    r for r in db.query_cash_entries(
                        year_months=[form_ym], corporation=form_corp,
                        order_desc=False,
                    )
                    if r.get('entry_kind') == 'expense'
                    and (r.get('debit_account') or '') != '旅費交通費'
                ]
                payee_filter = form_payee.strip().replace(
                    ' ', '').replace('　', '')
                if payee_filter:
                    def _matches(r):
                        p = (r.get('payee') or '').replace(
                            ' ', '').replace('　', '')
                        desc = (r.get('description') or '').replace(
                            ' ', '').replace('　', '')
                        return payee_filter in p or payee_filter in desc
                    pre = [r for r in entries_candidates if _matches(r)]
                    if pre:
                        entries_candidates = pre

            # 取込済レシート (cash_advance_receipts)
            receipts_candidates: list[dict] = []
            if source_mode in ('📷 取込済レシートから', '🔁 両方'):
                with db.get_conn() as _conn:
                    rows = _conn.execute(
                        """SELECT * FROM cash_advance_receipts
                            WHERE corporation = ?
                              AND year_month = ?
                            ORDER BY transaction_date""",
                        (form_corp, form_ym),
                    ).fetchall()
                receipts_candidates = [dict(r) for r in rows]

            if not entries_candidates and not receipts_candidates:
                st.info(
                    "対象月・法人に該当する明細がありません。"
                    "出納帳に取込むか、レシートタブで OCR 取込してください。"
                )
            else:
                # 統合候補表示
                # 出納帳行 → ソース 'cashbook'
                # レシート → ソース 'receipt'
                merged = []
                for r in entries_candidates:
                    merged.append({
                        'src': 'cashbook',
                        '選択': True,
                        '日付': r['transaction_date'],
                        '勘定科目': r.get('debit_account') or '',
                        '部門': r.get('department_raw') or '',
                        '備考': r.get('description') or '',
                        '金額': int(r.get('amount_out') or 0),
                        '_ref': ('cashbook', r['id'], r),
                    })
                for r in receipts_candidates:
                    merged.append({
                        'src': 'receipt',
                        '選択': True,
                        '日付': r['transaction_date'],
                        '勘定科目': '(レシート)',
                        '部門': r.get('facility_folder') or '',
                        '備考': (r.get('vendor') or '') + ' / '
                                + (r.get('purpose') or ''),
                        '金額': int(r.get('total_amount') or 0),
                        '_ref': ('receipt', r['id'], r),
                    })

                st.markdown(f"#### 候補 ({len(merged)}件)")
                df_pick = pd.DataFrame(merged)
                df_disp = df_pick.drop(columns=['_ref'])
                df_disp.insert(
                    0, 'ソース',
                    df_disp['src'].map({
                        'cashbook': '📒 出納帳',
                        'receipt': '📷 レシート',
                    }),
                )
                df_disp = df_disp.drop(columns=['src'])
                edited = st.data_editor(
                    df_disp,
                    hide_index=True, width='stretch',
                    column_config={
                        '選択': st.column_config.CheckboxColumn(
                            '選択', default=True),
                        '金額': st.column_config.NumberColumn(
                            '金額', format='¥%d'),
                    },
                    disabled=['ソース', '日付', '勘定科目', '部門', '備考'],
                    key='form_picker',
                )
                df_pick['選択'] = edited['選択'].values

                selected_rows = []
                for i, sel in enumerate(df_pick['選択']):
                    if sel:
                        selected_rows.append(df_pick.iloc[i]['_ref'])

                if selected_rows:
                    sel_total = sum(
                        int(ref[2].get('amount_out') or
                            ref[2].get('total_amount') or 0)
                        for ref in selected_rows
                    )
                    st.success(
                        f"✅ {len(selected_rows)} 件選択中 ／ "
                        f"合計: {_format_yen(sel_total)}"
                    )
                else:
                    st.warning("選択行がありません。")

                form_date = st.date_input(
                    "申請日", value=datetime.now().date(),
                    key='form_request_date',
                )

                if st.button(
                    "📤 経費精算書 Excel を生成",
                    type='primary', key='form_generate',
                    disabled=not selected_rows,
                ):
                    template_path = _find_cash_xlsm(form_corp)
                    if template_path is None:
                        st.error(
                            f"テンプレートxlsmが見つかりません: "
                            f"{CASH_DROPBOX_DIR}"
                        )
                    elif len(selected_rows) > expense_form.MAX_DETAIL_ROWS:
                        st.error(
                            f"明細行が {expense_form.MAX_DETAIL_ROWS} "
                            f"行を超えています (現在 {len(selected_rows)} 件)。"
                            "選択を減らすか、複数枚に分けてください。"
                        )
                    else:
                        items = []
                        for src, _id, ref in selected_rows:
                            tx_str = ref['transaction_date']
                            tx = datetime.strptime(tx_str, '%Y-%m-%d')
                            if src == 'cashbook':
                                md_prefix = f"{tx.month}月{tx.day}日　"
                                desc_only = (ref.get('description') or '').replace(
                                    md_prefix, '')
                                items.append({
                                    'date': tx_str,
                                    'vendor': ref.get('vendor') or '',
                                    'purpose': desc_only,
                                    'amount': int(ref.get('amount_out') or 0),
                                })
                            else:  # receipt
                                items.append({
                                    'date': tx_str,
                                    'vendor': ref.get('vendor') or '',
                                    'purpose': ref.get('purpose') or '',
                                    'amount': int(ref.get('total_amount') or 0),
                                })
                        try:
                            wb = expense_form.make_expense_form_workbook(
                                template_xlsm_path=template_path,
                                applicant_name=form_payee,
                                request_date=form_date,
                                items=items,
                            )
                            buf = io.BytesIO()
                            wb.save(buf)
                            buf.seek(0)
                            fname = (
                                f"経費精算書_{form_corp}_{form_ym}_"
                                f"{form_payee.replace(' ', '')}_"
                                f"{form_date.strftime('%Y%m%d')}.xlsx"
                            )
                            st.download_button(
                                "⬇️ 経費精算書 Excel をダウンロード",
                                data=buf.getvalue(),
                                file_name=fname,
                                mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                                key='form_download',
                            )
                            st.success(
                                f"✅ 生成完了 ({len(items)}行 / "
                                f"{_format_yen(sum(i['amount'] for i in items))})"
                            )
                        except Exception as e:
                            st.error(f"生成失敗: {e}")


# ============================================================
# 共通フィルタ（取込タブ以外で使う） - デビットカード側
# ============================================================
all_yms = db.list_debit_year_months()
all_corps = db.list_debit_corporations()

if not all_yms:
    with tab_summary:
        st.info("📭 まだ取込済みのデビット明細がありません。「📥 取込」タブから取り込んでください。")
    with tab_fac:
        st.info("データがありません。")
    with tab_acc:
        st.info("データがありません。")
    with tab_vendor:
        st.info("データがありません。")
    with tab_detail:
        st.info("データがありません。")
    st.stop()


def _filter_box(key_prefix: str) -> dict:
    """期間・法人の共通フィルタUI。"""
    yms_sorted_asc = sorted(all_yms)
    cols = st.columns([2, 2, 2])
    with cols[0]:
        start_ym = st.selectbox(
            "期間（開始月）", yms_sorted_asc, index=0, key=f'{key_prefix}_start',
        )
    with cols[1]:
        end_ym = st.selectbox(
            "期間（終了月）", yms_sorted_asc, index=len(yms_sorted_asc) - 1,
            key=f'{key_prefix}_end',
        )
    with cols[2]:
        corp = st.selectbox(
            "法人",
            options=['（全法人）'] + all_corps,
            key=f'{key_prefix}_corp',
        )
    if start_ym > end_ym:
        st.warning("開始月が終了月より後です。")
        start_ym, end_ym = end_ym, start_ym
    selected_yms = [ym for ym in yms_sorted_asc if start_ym <= ym <= end_ym]
    return {
        'year_months': selected_yms,
        'corporation': None if corp == '（全法人）' else corp,
        'start_ym': start_ym,
        'end_ym': end_ym,
    }


@st.cache_data(ttl=60)
def _query_df(year_months: tuple[str, ...],
              corporation: str | None,
              departments: tuple[str, ...] | None = None,
              accounts: tuple[str, ...] | None = None) -> pd.DataFrame:
    rows = db.query_debit_entries(
        year_months=list(year_months) if year_months else None,
        corporation=corporation,
        departments=list(departments) if departments else None,
        accounts=list(accounts) if accounts else None,
    )
    return pd.DataFrame(rows)


# ============================================================
# 📊 サマリタブ
# ============================================================
with tab_summary:
    st.markdown("### 📊 全体サマリ")
    f = _filter_box('summary')
    df = _query_df(tuple(f['year_months']), f['corporation'])

    if df.empty:
        st.info("条件に該当するデータがありません。")
    else:
        total = int(df['amount'].sum())
        cnt = len(df)
        avg = int(df['amount'].mean()) if cnt else 0
        n_facilities = df['department_raw'].replace('', pd.NA).nunique()
        n_vendors = df['vendor'].replace('', pd.NA).nunique()
        n_accounts = df['debit_account'].replace('', pd.NA).nunique()

        c1, c2, c3, c4, c5, c6 = st.columns(6)
        c1.metric("合計金額", _format_yen(total))
        c2.metric("取引件数", f"{cnt:,}")
        c3.metric("平均単価", _format_yen(avg))
        c4.metric("施設数", f"{n_facilities}")
        c5.metric("購入先数", f"{n_vendors}")
        c6.metric("勘定科目数", f"{n_accounts}")

        st.markdown("#### 月別推移")
        monthly = df.groupby('year_month', as_index=False).agg(
            金額=('amount', 'sum'),
            件数=('amount', 'count'),
        ).sort_values('year_month')

        # 法人別の月次内訳もまとめて表示
        has_multi_corp = df['corporation'].nunique() >= 2
        if has_multi_corp:
            corp_piv = df.pivot_table(
                index='year_month', columns='corporation',
                values='amount', aggfunc='sum', fill_value=0,
            ).reindex(monthly['year_month']).reset_index()
            for c in corp_piv.columns:
                if c == 'year_month':
                    continue
                monthly[c] = corp_piv[c].values

        # 前月比 (差額・比率)
        monthly['前月比'] = monthly['金額'].diff().fillna(0).astype(int)
        prev = monthly['金額'].shift(1)
        monthly['前月比率'] = (
            (monthly['金額'] - prev) / prev.replace(0, pd.NA)
        ).fillna(0)

        # 表示用に整形
        disp_cols = ['year_month', '金額', '件数']
        if has_multi_corp:
            disp_cols += [c for c in monthly.columns
                          if c not in ('year_month', '金額', '件数',
                                       '前月比', '前月比率')]
        disp_cols += ['前月比', '前月比率']

        monthly_disp = monthly[disp_cols].copy()
        # 合計行
        total_row = {'year_month': '合計',
                     '金額': monthly['金額'].sum(),
                     '件数': monthly['件数'].sum(),
                     '前月比': monthly['前月比'].sum(),
                     '前月比率': 0}
        if has_multi_corp:
            for c in disp_cols:
                if c not in ('year_month', '金額', '件数', '前月比', '前月比率'):
                    total_row[c] = monthly[c].sum()
        monthly_disp = pd.concat(
            [monthly_disp, pd.DataFrame([total_row])], ignore_index=True,
        )

        # フォーマット
        rename_map = {'year_month': '年月', '金額': '合計金額', '件数': '取引件数'}
        for c in monthly_disp.columns:
            if c in ('金額', '件数', '前月比') or (
                has_multi_corp and c not in rename_map and c not in ('前月比率',)
            ):
                # 数値カラムを整形
                monthly_disp[c] = monthly_disp[c].apply(
                    lambda v: f"{int(v):,}" if pd.notna(v) and v != '' else '—'
                )
        # 前月比率
        def _fmt_pct(v):
            if pd.isna(v) or v == 0:
                return '—'
            arrow = '▲' if v > 0 else '▼'
            return f"{arrow} {abs(v) * 100:.1f}%"
        monthly_disp['前月比率'] = monthly['前月比率'].tolist() + [None]
        monthly_disp['前月比率'] = monthly_disp['前月比率'].apply(_fmt_pct)

        monthly_disp = monthly_disp.rename(columns=rename_map)

        # 行ハイライト（合計行と最大月）
        max_idx = int(monthly['金額'].idxmax()) if len(monthly) else -1

        def _row_style(row):
            styles = [''] * len(row)
            if row.iloc[0] == '合計':
                return ['background-color:#f1f5f9; font-weight:700; '
                        'color:#0f172a; border-top:2px solid #cbd5e1'] * len(row)
            try:
                ridx = int(row.name)
                if ridx == max_idx:
                    return ['background-color:#fef3c7; font-weight:600'] * len(row)
            except (ValueError, TypeError):
                pass
            return styles

        styler = (
            monthly_disp.style
            .apply(_row_style, axis=1)
            .set_properties(**{'font-size': '14px'})
            .set_table_styles([
                {'selector': 'th',
                 'props': 'background-color:#1e3a8a; color:white; '
                          'font-weight:700; text-align:center;'},
                {'selector': 'td',
                 'props': 'text-align:right; padding:6px 12px;'},
                {'selector': 'td:first-child',
                 'props': 'text-align:center; font-weight:600; '
                          'background-color:#f8fafc;'},
            ])
        )
        st.dataframe(styler, hide_index=True, width='stretch',
                     height=min(60 + 36 * len(monthly_disp), 560))
        st.caption("💡 黄色: 期間中で最大の月 ／ 灰色: 合計行")

        # ----------------------------------------------------------------
        # 施設別サマリ（指定勘定科目のクロス集計）
        # ----------------------------------------------------------------
        st.markdown("---")
        st.markdown("### 🏢 施設別サマリ")
        st.caption("対象勘定科目: 食材費／消耗品費／燃料費／旅費交通費／通信費／支払手数料／採用教育費／給食費")

        FACILITY_SUMMARY_ACCOUNTS = [
            "食材費", "消耗品費", "燃料費", "旅費交通費",
            "通信費", "支払手数料", "採用教育費", "給食費",
        ]

        fac_df = df[df['department_raw'] != ''].copy()
        if fac_df.empty:
            st.info("施設情報のあるデータがありません。")
        else:
            target_df = fac_df[fac_df['debit_account'].isin(FACILITY_SUMMARY_ACCOUNTS)]
            if target_df.empty:
                st.info("対象勘定科目（食材費／消耗品費等）のデータがありません。")
            else:
                fac_piv = target_df.pivot_table(
                    index='department_raw',
                    columns='debit_account',
                    values='amount',
                    aggfunc='sum',
                    fill_value=0,
                )
                # 8科目を必ず全列表示（データ無しは0）
                for col in FACILITY_SUMMARY_ACCOUNTS:
                    if col not in fac_piv.columns:
                        fac_piv[col] = 0
                fac_piv = fac_piv[FACILITY_SUMMARY_ACCOUNTS]
                fac_piv['合計'] = fac_piv.sum(axis=1)
                fac_piv = fac_piv.sort_values('合計', ascending=False)
                # 合計行
                total_row = pd.DataFrame(
                    [fac_piv.sum(axis=0)],
                    index=['合計'],
                )
                fac_piv = pd.concat([fac_piv, total_row])
                fac_piv.index.name = '施設'
                fac_piv = fac_piv.reset_index()

                # 表示整形
                disp = fac_piv.copy()
                num_cols = FACILITY_SUMMARY_ACCOUNTS + ['合計']
                for c in num_cols:
                    disp[c] = disp[c].apply(
                        lambda v: f"{int(v):,}" if v else '—'
                    )

                # スタイリング: 施設別最大値のセルを強調 / 合計列は太字 / 合計行は背景
                def _highlight_cells(row):
                    styles = [''] * len(row)
                    if row['施設'] == '合計':
                        return [
                            'background-color:#f1f5f9; font-weight:700; '
                            'color:#0f172a; border-top:2px solid #cbd5e1'
                        ] * len(row)
                    return styles

                def _emphasize_total(col):
                    if col.name == '合計':
                        return [
                            'font-weight:700; color:#1e3a8a; '
                            'background-color:#eff6ff'
                        ] * len(col)
                    return [''] * len(col)

                styler_fac = (
                    disp.style
                    .apply(_highlight_cells, axis=1)
                    .apply(_emphasize_total, axis=0)
                    .set_table_styles([
                        {'selector': 'th',
                         'props': 'background-color:#1e3a8a; color:white; '
                                  'font-weight:700; text-align:center; '
                                  'padding:8px 10px;'},
                        {'selector': 'td',
                         'props': 'text-align:right; padding:6px 10px;'},
                        {'selector': 'td:first-child',
                         'props': 'text-align:left; font-weight:600; '
                                  'background-color:#f8fafc; '
                                  'border-right:2px solid #cbd5e1;'},
                    ])
                )

                st.dataframe(
                    styler_fac, hide_index=True, width='stretch',
                    height=min(60 + 36 * len(disp), 700),
                )
                st.caption(
                    f"📌 施設数: {len(disp) - 1}　／　"
                    f"対象期間 {f['start_ym']} 〜 {f['end_ym']}　／　"
                    f"対象金額合計: {_format_yen(int(target_df['amount'].sum()))}"
                )

                # 科目別の総額カード
                st.markdown("#### 勘定科目別 合計")
                acc_totals = target_df.groupby(
                    'debit_account', as_index=False)['amount'].sum()
                acc_totals = acc_totals.set_index('debit_account')
                cols = st.columns(4)
                for i, acc in enumerate(FACILITY_SUMMARY_ACCOUNTS):
                    val = int(acc_totals['amount'].get(acc, 0))
                    with cols[i % 4]:
                        st.metric(acc, _format_yen(val))


# ============================================================
# 🏢 施設別タブ
# ============================================================
with tab_fac:
    st.markdown("### 🏢 施設別 費用分析")
    f = _filter_box('fac')
    df = _query_df(tuple(f['year_months']), f['corporation'])

    if df.empty:
        st.info("条件に該当するデータがありません。")
    else:
        # 施設(部門)別の集計
        agg = df.groupby('department_raw', as_index=False).agg(
            合計金額=('amount', 'sum'),
            件数=('amount', 'count'),
            平均=('amount', 'mean'),
        ).sort_values('合計金額', ascending=False)
        agg['平均'] = agg['平均'].astype(int)
        agg = agg[agg['department_raw'] != '']

        total_sum = int(agg['合計金額'].sum())
        st.markdown(f"**施設数: {len(agg)} / 総額: {_format_yen(total_sum)}**")

        st.markdown("#### 施設別 合計金額（多い順）")
        top_n = _safe_topn("表示件数", len(agg), 15, key='fac_topn')
        top = agg.head(top_n)

        # 構成比とランクを付けた表
        top_disp = top.copy().reset_index(drop=True)
        top_disp.insert(0, '順位', top_disp.index + 1)
        top_disp['構成比'] = (top_disp['合計金額'] / total_sum * 100).round(1) \
            if total_sum > 0 else 0
        top_disp_fmt = top_disp.copy()
        top_disp_fmt['合計金額'] = top_disp_fmt['合計金額'].apply(_format_yen)
        top_disp_fmt['平均'] = top_disp_fmt['平均'].apply(_format_yen)
        top_disp_fmt['件数'] = top_disp_fmt['件数'].apply(lambda v: f"{v:,}")
        top_disp_fmt['構成比'] = top_disp_fmt['構成比'].astype(str) + '%'
        top_disp_fmt.columns = ['順位', '施設', '合計金額', '件数', '平均単価', '構成比']

        def _rank_style(row):
            if row['順位'] == 1:
                return ['background-color:#fef3c7; font-weight:700'] * len(row)
            if row['順位'] == 2:
                return ['background-color:#fef9c3'] * len(row)
            if row['順位'] == 3:
                return ['background-color:#fefce8'] * len(row)
            return [''] * len(row)

        styler_top = (
            top_disp_fmt.style
            .apply(_rank_style, axis=1)
            .set_table_styles([
                {'selector': 'th',
                 'props': 'background-color:#1e3a8a; color:white; '
                          'font-weight:700; text-align:center; padding:8px 10px;'},
                {'selector': 'td',
                 'props': 'text-align:right; padding:6px 12px;'},
                {'selector': 'td:nth-child(1)',
                 'props': 'text-align:center; font-weight:700;'},
                {'selector': 'td:nth-child(2)',
                 'props': 'text-align:left; font-weight:600; '
                          'background-color:#f8fafc; border-right:2px solid #cbd5e1;'},
            ])
        )
        st.dataframe(
            styler_top, hide_index=True, width='stretch',
            height=min(60 + 36 * len(top_disp_fmt), 700),
        )
        st.caption("💡 上位3件をハイライト")

        st.markdown("#### 施設別 ✕ 月別 集計")
        if len(agg) > 0:
            piv = df[df['department_raw'] != ''].pivot_table(
                index='department_raw', columns='year_month',
                values='amount', aggfunc='sum', fill_value=0,
            )
            piv = piv.loc[agg.head(top_n)['department_raw']]
            piv['合計'] = piv.sum(axis=1)
            piv = piv.sort_values('合計', ascending=False)
            # 月合計行
            month_total = pd.DataFrame([piv.sum(axis=0)], index=['月合計'])
            piv = pd.concat([piv, month_total])
            piv.index.name = '施設＼月'
            piv = piv.reset_index()

            piv_disp = piv.copy()
            num_cols = [c for c in piv.columns if c != '施設＼月']
            for c in num_cols:
                piv_disp[c] = piv_disp[c].apply(
                    lambda v: f"{int(v):,}" if v else '—'
                )

            def _piv_row_style(row):
                if row['施設＼月'] == '月合計':
                    return ['background-color:#f1f5f9; font-weight:700; '
                            'border-top:2px solid #cbd5e1'] * len(row)
                return [''] * len(row)

            def _piv_total_col_style(col):
                if col.name == '合計':
                    return ['font-weight:700; color:#1e3a8a; '
                            'background-color:#eff6ff'] * len(col)
                return [''] * len(col)

            styler_piv = (
                piv_disp.style
                .apply(_piv_row_style, axis=1)
                .apply(_piv_total_col_style, axis=0)
                .set_table_styles([
                    {'selector': 'th',
                     'props': 'background-color:#1e3a8a; color:white; '
                              'font-weight:700; text-align:center; padding:8px 10px;'},
                    {'selector': 'td',
                     'props': 'text-align:right; padding:6px 10px;'},
                    {'selector': 'td:first-child',
                     'props': 'text-align:left; font-weight:600; '
                              'background-color:#f8fafc; border-right:2px solid #cbd5e1;'},
                ])
            )
            st.dataframe(
                styler_piv, hide_index=True, width='stretch',
                height=min(60 + 36 * len(piv_disp), 700),
            )

        st.markdown("#### 施設ドリルダウン")
        sel_fac = st.selectbox(
            "施設を選択", options=list(agg['department_raw']),
            key='fac_drill_select',
        )
        if sel_fac:
            sub = df[df['department_raw'] == sel_fac]
            c1, c2, c3 = st.columns(3)
            c1.metric("合計", _format_yen(int(sub['amount'].sum())))
            c2.metric("件数", f"{len(sub):,}")
            c3.metric("平均", _format_yen(int(sub['amount'].mean())))

            st.markdown("**勘定科目別**")
            ag_acc = sub.groupby('debit_account', as_index=False).agg(
                金額=('amount', 'sum'), 件数=('amount', 'count'),
            ).sort_values('金額', ascending=False)
            ag_acc_disp = ag_acc.copy()
            ag_acc_disp['金額'] = ag_acc_disp['金額'].apply(_format_yen)
            st.dataframe(ag_acc_disp, hide_index=True, width='stretch')

            st.markdown("**購入先トップ10**")
            ag_v = sub[sub['vendor'] != ''].groupby(
                'vendor', as_index=False).agg(
                金額=('amount', 'sum'), 件数=('amount', 'count'),
            ).sort_values('金額', ascending=False).head(10)
            ag_v_disp = ag_v.copy()
            ag_v_disp['金額'] = ag_v_disp['金額'].apply(_format_yen)
            st.dataframe(ag_v_disp, hide_index=True, width='stretch')


# ============================================================
# 📚 勘定科目別タブ
# ============================================================
with tab_acc:
    st.markdown("### 📚 勘定科目別 費用分析")
    f = _filter_box('acc')
    df = _query_df(tuple(f['year_months']), f['corporation'])

    if df.empty:
        st.info("条件に該当するデータがありません。")
    else:
        agg = df.groupby('debit_account', as_index=False).agg(
            合計金額=('amount', 'sum'),
            件数=('amount', 'count'),
            平均=('amount', 'mean'),
        ).sort_values('合計金額', ascending=False)
        agg = agg[agg['debit_account'] != '']
        agg['平均'] = agg['平均'].astype(int)
        agg['構成比'] = (agg['合計金額'] / agg['合計金額'].sum() * 100).round(1)

        total_sum_acc = int(agg['合計金額'].sum())
        st.markdown(f"**科目数: {len(agg)} / 総額: {_format_yen(total_sum_acc)}**")

        st.markdown("#### 科目別 合計（多い順）")
        agg_disp = agg.copy().reset_index(drop=True)
        agg_disp.insert(0, '順位', agg_disp.index + 1)
        agg_fmt = agg_disp.copy()
        agg_fmt['合計金額'] = agg_fmt['合計金額'].apply(_format_yen)
        agg_fmt['平均'] = agg_fmt['平均'].apply(_format_yen)
        agg_fmt['件数'] = agg_fmt['件数'].apply(lambda v: f"{v:,}")
        agg_fmt['構成比'] = agg_fmt['構成比'].astype(str) + '%'
        agg_fmt.columns = ['順位', '勘定科目', '合計金額', '件数', '平均単価', '構成比']

        def _acc_rank_style(row):
            if row['順位'] == 1:
                return ['background-color:#fef3c7; font-weight:700'] * len(row)
            if row['順位'] == 2:
                return ['background-color:#fef9c3'] * len(row)
            if row['順位'] == 3:
                return ['background-color:#fefce8'] * len(row)
            return [''] * len(row)

        styler_acc = (
            agg_fmt.style
            .apply(_acc_rank_style, axis=1)
            .set_table_styles([
                {'selector': 'th',
                 'props': 'background-color:#1e3a8a; color:white; '
                          'font-weight:700; text-align:center; padding:8px 10px;'},
                {'selector': 'td',
                 'props': 'text-align:right; padding:6px 12px;'},
                {'selector': 'td:nth-child(1)',
                 'props': 'text-align:center; font-weight:700;'},
                {'selector': 'td:nth-child(2)',
                 'props': 'text-align:left; font-weight:600; '
                          'background-color:#f8fafc; border-right:2px solid #cbd5e1;'},
            ])
        )
        st.dataframe(
            styler_acc, hide_index=True, width='stretch',
            height=min(60 + 36 * len(agg_fmt), 700),
        )
        st.caption("💡 上位3件をハイライト")

        st.markdown("#### 科目別 ✕ 月別 集計")
        top_n = _safe_topn("対象科目数", len(agg), 8, key='acc_topn', floor=3)
        top_accs = list(agg.head(top_n)['debit_account'])
        piv = df[df['debit_account'].isin(top_accs)].pivot_table(
            index='year_month', columns='debit_account',
            values='amount', aggfunc='sum', fill_value=0,
        ).sort_index()
        if not piv.empty:
            piv = piv[top_accs]  # 表示順を金額多い順に
            piv['月計'] = piv.sum(axis=1)
            month_total = pd.DataFrame([piv.sum(axis=0)], index=['科目計'])
            piv = pd.concat([piv, month_total])
            piv.index.name = '年月＼科目'
            piv = piv.reset_index()
            piv_fmt = piv.copy()
            for c in piv.columns:
                if c == '年月＼科目':
                    continue
                piv_fmt[c] = piv_fmt[c].apply(
                    lambda v: f"{int(v):,}" if v else '—'
                )

            def _acc_piv_row_style(row):
                if row['年月＼科目'] == '科目計':
                    return ['background-color:#f1f5f9; font-weight:700; '
                            'border-top:2px solid #cbd5e1'] * len(row)
                return [''] * len(row)

            def _acc_piv_col_style(col):
                if col.name == '月計':
                    return ['font-weight:700; color:#1e3a8a; '
                            'background-color:#eff6ff'] * len(col)
                return [''] * len(col)

            styler_apiv = (
                piv_fmt.style
                .apply(_acc_piv_row_style, axis=1)
                .apply(_acc_piv_col_style, axis=0)
                .set_table_styles([
                    {'selector': 'th',
                     'props': 'background-color:#1e3a8a; color:white; '
                              'font-weight:700; text-align:center; padding:8px 10px;'},
                    {'selector': 'td',
                     'props': 'text-align:right; padding:6px 10px;'},
                    {'selector': 'td:first-child',
                     'props': 'text-align:center; font-weight:600; '
                              'background-color:#f8fafc; border-right:2px solid #cbd5e1;'},
                ])
            )
            st.dataframe(
                styler_apiv, hide_index=True, width='stretch',
                height=min(60 + 36 * len(piv_fmt), 700),
            )

        st.markdown("#### 勘定科目ドリルダウン")
        sel_acc = st.selectbox(
            "勘定科目を選択", options=list(agg['debit_account']),
            key='acc_drill_select',
        )
        if sel_acc:
            sub = df[df['debit_account'] == sel_acc]
            c1, c2, c3 = st.columns(3)
            c1.metric("合計", _format_yen(int(sub['amount'].sum())))
            c2.metric("件数", f"{len(sub):,}")
            c3.metric("平均", _format_yen(int(sub['amount'].mean())))

            st.markdown(f"**「{sel_acc}」の購入先トップ20**")
            ag_v = sub[sub['vendor'] != ''].groupby(
                'vendor', as_index=False).agg(
                金額=('amount', 'sum'), 件数=('amount', 'count'),
            ).sort_values('金額', ascending=False).head(20)
            ag_v_disp = ag_v.copy()
            ag_v_disp['金額'] = ag_v_disp['金額'].apply(_format_yen)
            st.dataframe(ag_v_disp, hide_index=True, width='stretch')

            st.markdown(f"**「{sel_acc}」の施設別**")
            ag_f = sub[sub['department_raw'] != ''].groupby(
                'department_raw', as_index=False).agg(
                金額=('amount', 'sum'), 件数=('amount', 'count'),
            ).sort_values('金額', ascending=False)
            ag_f_disp = ag_f.copy()
            ag_f_disp['金額'] = ag_f_disp['金額'].apply(_format_yen)
            st.dataframe(ag_f_disp, hide_index=True, width='stretch')


# ============================================================
# 🏪 購入先別タブ
# ============================================================
with tab_vendor:
    st.markdown("### 🏪 購入先別 分析")
    f = _filter_box('vendor')
    df = _query_df(tuple(f['year_months']), f['corporation'])
    df = df[df['vendor'] != '']

    if df.empty:
        st.info("条件に該当するデータがありません。")
    else:
        agg = df.groupby('vendor', as_index=False).agg(
            合計金額=('amount', 'sum'),
            件数=('amount', 'count'),
            平均=('amount', 'mean'),
        ).sort_values('合計金額', ascending=False)
        agg['平均'] = agg['平均'].astype(int)

        total_sum_v = int(agg['合計金額'].sum())
        st.markdown(f"**購入先数: {len(agg)} / 総額: {_format_yen(total_sum_v)}**")

        top_n = _safe_topn("ランキング表示件数", min(50, len(agg)), 20, key='vendor_topn')
        top = agg.head(top_n)

        st.markdown("#### 購入先ランキング（金額）")
        top_disp = top.copy().reset_index(drop=True)
        top_disp.insert(0, '順位', top_disp.index + 1)
        top_disp['構成比'] = (top_disp['合計金額'] / total_sum_v * 100).round(1) \
            if total_sum_v > 0 else 0
        top_fmt = top_disp.copy()
        top_fmt['合計金額'] = top_fmt['合計金額'].apply(_format_yen)
        top_fmt['平均'] = top_fmt['平均'].apply(_format_yen)
        top_fmt['件数'] = top_fmt['件数'].apply(lambda v: f"{v:,}")
        top_fmt['構成比'] = top_fmt['構成比'].astype(str) + '%'
        top_fmt.columns = ['順位', '購入先', '合計金額', '件数', '平均単価', '構成比']

        def _v_rank_style(row):
            if row['順位'] == 1:
                return ['background-color:#fef3c7; font-weight:700'] * len(row)
            if row['順位'] == 2:
                return ['background-color:#fef9c3'] * len(row)
            if row['順位'] == 3:
                return ['background-color:#fefce8'] * len(row)
            return [''] * len(row)

        styler_v = (
            top_fmt.style
            .apply(_v_rank_style, axis=1)
            .set_table_styles([
                {'selector': 'th',
                 'props': 'background-color:#92400e; color:white; '
                          'font-weight:700; text-align:center; padding:8px 10px;'},
                {'selector': 'td',
                 'props': 'text-align:right; padding:6px 12px;'},
                {'selector': 'td:nth-child(1)',
                 'props': 'text-align:center; font-weight:700;'},
                {'selector': 'td:nth-child(2)',
                 'props': 'text-align:left; font-weight:600; '
                          'background-color:#fff7ed; border-right:2px solid #fed7aa;'},
            ])
        )
        st.dataframe(
            styler_v, hide_index=True, width='stretch',
            height=min(60 + 36 * len(top_fmt), 700),
        )
        st.caption("💡 上位3件をハイライト")

        st.markdown("#### 購入先ドリルダウン")
        sel_v = st.selectbox(
            "購入先を選択", options=list(agg['vendor']),
            key='vendor_drill_select',
        )
        if sel_v:
            sub = df[df['vendor'] == sel_v]
            c1, c2, c3, c4 = st.columns(4)
            c1.metric("合計", _format_yen(int(sub['amount'].sum())))
            c2.metric("件数", f"{len(sub):,}")
            c3.metric("平均", _format_yen(int(sub['amount'].mean())))
            c4.metric("利用施設数", f"{sub['department_raw'].replace('', pd.NA).nunique()}")

            col_a, col_b = st.columns(2)
            with col_a:
                st.markdown("**月別 利用額**")
                m = sub.groupby('year_month', as_index=False).agg(
                    金額=('amount', 'sum'), 件数=('amount', 'count'),
                ).sort_values('year_month')
                m_fmt = m.copy()
                m_fmt['金額'] = m_fmt['金額'].apply(_format_yen)
                m_fmt['件数'] = m_fmt['件数'].apply(lambda v: f"{v:,}")
                m_fmt.columns = ['年月', '金額', '件数']
                st.dataframe(m_fmt, hide_index=True, width='stretch')
            with col_b:
                st.markdown("**施設別 利用額**")
                ag_f = sub[sub['department_raw'] != ''].groupby(
                    'department_raw', as_index=False).agg(
                    金額=('amount', 'sum'), 件数=('amount', 'count'),
                ).sort_values('金額', ascending=False).head(15)
                if not ag_f.empty:
                    ag_f_fmt = ag_f.copy()
                    ag_f_fmt['金額'] = ag_f_fmt['金額'].apply(_format_yen)
                    ag_f_fmt['件数'] = ag_f_fmt['件数'].apply(lambda v: f"{v:,}")
                    ag_f_fmt.columns = ['施設', '金額', '件数']
                    st.dataframe(ag_f_fmt, hide_index=True, width='stretch')
                else:
                    st.caption("施設データなし")

            st.markdown(f"**「{sel_v}」の取引明細**")
            sub_disp = sub[[
                'transaction_date', 'department_raw', 'debit_account',
                'amount', 'description',
            ]].copy()
            sub_disp['amount'] = sub_disp['amount'].apply(lambda v: f"{int(v):,}")
            sub_disp.columns = ['日付', '施設', '勘定科目', '金額', '摘要']
            st.dataframe(sub_disp, hide_index=True, width='stretch')


# ============================================================
# 🛒 購入明細タブ
# ============================================================
with tab_detail:
    st.markdown("### 🛒 購入明細 検索／品目展開")
    f = _filter_box('detail')

    cols = st.columns([2, 2, 2])
    with cols[0]:
        dept_filter = st.multiselect(
            "施設で絞り込み",
            options=db.list_debit_departments(),
            default=[],
            key='detail_dept',
        )
    with cols[1]:
        acc_filter = st.multiselect(
            "勘定科目で絞り込み",
            options=db.list_debit_accounts(),
            default=[],
            key='detail_acc',
        )
    with cols[2]:
        keyword = st.text_input(
            "キーワード検索（摘要・購入先）", value="",
            placeholder="例: ガソリン, ダイソー", key='detail_keyword',
        )

    rows = db.query_debit_entries(
        year_months=f['year_months'],
        corporation=f['corporation'],
        departments=dept_filter or None,
        accounts=acc_filter or None,
        keyword=keyword or None,
    )
    df = pd.DataFrame(rows)

    if df.empty:
        st.info("条件に該当するデータがありません。")
    else:
        c1, c2, c3 = st.columns(3)
        c1.metric("件数", f"{len(df):,}")
        c2.metric("合計", _format_yen(int(df['amount'].sum())))
        c3.metric("平均", _format_yen(int(df['amount'].mean())))

        view_mode = st.radio(
            "表示モード",
            ["📋 取引一覧", "🛒 品目展開（摘要から品目を抽出）"],
            horizontal=True, key='detail_mode',
        )

        if view_mode.startswith("📋"):
            disp = df[[
                'transaction_date', 'corporation', 'department_raw',
                'debit_account', 'vendor', 'purpose', 'amount', 'description',
            ]].copy()
            disp['amount'] = disp['amount'].apply(lambda v: f"{int(v):,}")
            disp.columns = ['日付', '法人', '施設', '勘定科目', '購入先',
                            '用途', '金額', '摘要']
            st.dataframe(disp, hide_index=True, width='stretch', height=520)

            csv_bytes = df[[
                'transaction_date', 'corporation', 'department_raw',
                'debit_account', 'vendor', 'purpose', 'amount', 'description',
            ]].to_csv(index=False).encode('utf-8-sig')
            st.download_button(
                "📥 取引明細をCSVダウンロード",
                data=csv_bytes,
                file_name=f"debit_{f['start_ym']}_{f['end_ym']}.csv",
                mime='text/csv',
                key='detail_csv',
            )
        else:
            # 品目展開
            items = debit_parser.collect_items(df.to_dict('records'))
            if not items:
                st.info("品目を抽出できるデータがありません。")
            else:
                idf = pd.DataFrame(items)
                ic1, ic2, ic3 = st.columns(3)
                ic1.metric("品目数", f"{len(idf):,}")
                ic2.metric("合計", _format_yen(int(idf['price'].sum())))
                ic3.metric("ユニーク品目", f"{idf['item_name'].nunique():,}")

                st.markdown("#### 品目ランキング（同名集計）")
                ig = idf.groupby('item_name', as_index=False).agg(
                    回数=('price', 'count'),
                    合計=('price', 'sum'),
                    平均=('price', 'mean'),
                ).sort_values('合計', ascending=False).head(50)
                ig['平均'] = ig['平均'].astype(int)
                ig_disp = ig.copy()
                ig_disp['合計'] = ig_disp['合計'].apply(_format_yen)
                ig_disp['平均'] = ig_disp['平均'].apply(_format_yen)
                ig_disp.columns = ['品目', '購入回数', '合計金額', '平均単価']
                st.dataframe(ig_disp, hide_index=True, width='stretch')

                st.markdown("#### 品目明細（個別）")
                disp = idf[[
                    'transaction_date', 'corporation', 'department_raw',
                    'debit_account', 'vendor', 'item_name', 'price',
                ]].copy()
                disp['price'] = disp['price'].apply(lambda v: f"{int(v):,}")
                disp.columns = ['日付', '法人', '施設', '勘定科目', '購入先',
                                '品目', '価格']
                st.dataframe(disp, hide_index=True, width='stretch', height=480)

                csv_bytes = idf.to_csv(index=False).encode('utf-8-sig')
                st.download_button(
                    "📥 品目明細をCSVダウンロード",
                    data=csv_bytes,
                    file_name=f"debit_items_{f['start_ym']}_{f['end_ym']}.csv",
                    mime='text/csv',
                    key='detail_items_csv',
                )


# ============================================================
# 📷 レシートタブ（写メ / PDF プレビュー）
# ============================================================
import base64

with tab_receipts:
    st.markdown("### 📷 レシート（写メ / PDF）")
    st.caption(
        "Google Drive 上のレシート画像（JPEG/PNG/HEIC等）と PDF を施設別に閲覧できます。"
        " 01.デビッドカード清算／{施設}／ 配下を再帰的にスキャン。"
    )
    if DEBIT_FOLDER_REF.get('url'):
        st.markdown(
            f"<div style='background:#f0f9ff; border-left:4px solid #3b82f6; "
            f"padding:8px 14px; border-radius:4px; margin:6px 0 14px;'>"
            f"📁 <b>格納先 (Google Drive)</b>: "
            f"<a href='{DEBIT_FOLDER_REF['url']}' target='_blank' "
            f"style='color:#1e40af;'>{DEBIT_FOLDER_REF.get('name') or 'Drive フォルダ'}</a>"
            f"</div>",
            unsafe_allow_html=True,
        )

    storage_root = _resolve_storage_dir()
    if not storage_root.exists():
        st.warning(f"格納フォルダが見つかりません: {storage_root}")
    else:
        with st.spinner("レシート読込中..."):
            receipts = _list_receipt_files(storage_root)

        if not receipts:
            st.info(
                "レシートファイルが見つかりません。\n"
                "対応形式: " + ", ".join(sorted(RECEIPT_ALL_EXTS))
            )
        else:
            # ---- KPI ----
            n_total = len(receipts)
            n_img = sum(1 for r in receipts if r['kind'] == 'image')
            n_pdf = sum(1 for r in receipts if r['kind'] == 'pdf')
            facilities = sorted({r['facility'] for r in receipts if r['facility']})
            total_size = sum(r['size'] for r in receipts)

            c1, c2, c3, c4, c5 = st.columns(5)
            c1.metric("レシート総数", f"{n_total:,}")
            c2.metric("📸 写メ", f"{n_img:,}")
            c3.metric("📄 PDF", f"{n_pdf:,}")
            c4.metric("施設数", f"{len(facilities)}")
            c5.metric("総容量", _format_size(total_size))

            # ---- 自動処理 状態表示のみ (実行は「📥 取込」タブから) ----
            n_unproc = len(receipt_processor.detect_unprocessed(receipts))
            n_processed = sum(
                1 for r in receipts
                if r.get('kind') == 'image'
                and db.is_receipt_processed(str(r['path']))
            )
            # _処理済みフォルダにあるがOCR未実行のファイル（孤立ファイル）
            n_orphaned = sum(
                1 for r in receipts
                if r.get('status') == '処理済み'
                and r.get('kind') == 'image'
                and not db.is_receipt_processed(str(r['path']))
            )
            if n_unproc > 0:
                st.info(
                    f"🤖 未処理レシート **{n_unproc}** 件 ／ 処理済 {n_processed} 件 — "
                    "「📥 取込」タブの「🚀 取込実行」ボタンで一括処理されます。"
                )
            elif n_processed > 0:
                st.caption(
                    f"🤖 処理済 {n_processed} 件 ／ 未処理 0 件"
                )
            # 孤立ファイル（_処理済みにあるが未OCR）の警告
            if n_orphaned > 0:
                st.warning(
                    f"⚠️ **`_処理済み`フォルダ内に未OCRのファイルが {n_orphaned} 件あります。**\n\n"
                    "これらは手動で移動されましたが、OCR＆Excel追記がまだ完了していません。\n"
                    "各レシートカード下部の「🤖 OCRで解析 & Excel追記」ボタンから処理してください。"
                )

            # ---- フィルタ ----
            st.markdown("#### 🔎 絞り込み")
            fc1, fc2, fc3, fc4 = st.columns([2, 2, 2, 1])
            with fc1:
                sel_facilities = st.multiselect(
                    "施設で絞り込み",
                    options=facilities,
                    default=[],
                    placeholder="未選択 = 全施設",
                    key='rcp_fac',
                )
            with fc2:
                kind_filter = st.multiselect(
                    "種別で絞り込み",
                    options=['📸 写メ', '📄 PDF'],
                    default=[],
                    placeholder="未選択 = 全種別",
                    key='rcp_kind',
                )
            with fc3:
                status_options = ['（未振分）', '処理済み', '要確認']
                status_filter = st.multiselect(
                    "ステータスで絞り込み",
                    options=status_options,
                    default=[],
                    placeholder="未選択 = 全て",
                    key='rcp_status',
                )
            with fc4:
                sort_mode = st.selectbox(
                    "並び順",
                    options=['新しい順', '古い順', '施設順', 'ファイル名順'],
                    index=0,
                    key='rcp_sort',
                )

            # フィルタ適用
            filtered = list(receipts)
            if sel_facilities:
                filtered = [r for r in filtered if r['facility'] in sel_facilities]
            if kind_filter:
                kind_set = set()
                if '📸 写メ' in kind_filter:
                    kind_set.add('image')
                if '📄 PDF' in kind_filter:
                    kind_set.add('pdf')
                filtered = [r for r in filtered if r['kind'] in kind_set]
            if status_filter:
                target_statuses = set()
                if '（未振分）' in status_filter:
                    target_statuses.add('')
                if '処理済み' in status_filter:
                    target_statuses.add('処理済み')
                if '要確認' in status_filter:
                    target_statuses.add('要確認')
                filtered = [r for r in filtered if r.get('status', '') in target_statuses]

            if sort_mode == '新しい順':
                filtered.sort(key=lambda r: -r['mtime'])
            elif sort_mode == '古い順':
                filtered.sort(key=lambda r: r['mtime'])
            elif sort_mode == '施設順':
                filtered.sort(key=lambda r: (r['facility'], -r['mtime']))
            else:
                filtered.sort(key=lambda r: r['name'])

            st.caption(f"表示中: {len(filtered):,} / 全 {n_total:,} 件")

            # ---- 施設別レシート一覧テーブル ----
            with st.expander("📋 施設別レシート集計", expanded=False):
                fac_rows = []
                for fac in facilities:
                    fac_recs = [r for r in receipts if r['facility'] == fac]
                    fac_rows.append({
                        '施設': fac,
                        '📸 写メ': sum(1 for r in fac_recs if r['kind'] == 'image'),
                        '📄 PDF': sum(1 for r in fac_recs if r['kind'] == 'pdf'),
                        '合計': len(fac_recs),
                        '総容量': _format_size(sum(r['size'] for r in fac_recs)),
                    })
                fac_df = pd.DataFrame(fac_rows).sort_values('合計', ascending=False)
                st.dataframe(fac_df, hide_index=True, width='stretch')

            # ---- ギャラリー表示 ----
            from datetime import datetime as _dt
            st.markdown("#### 🖼️ レシート プレビュー")

            page_size = st.select_slider(
                "1ページの表示件数",
                options=[6, 12, 24, 48, 96],
                value=12,
                key='rcp_page_size',
            )

            total_pages = max(1, (len(filtered) + page_size - 1) // page_size)
            if total_pages > 1:
                page = st.number_input(
                    f"ページ (1〜{total_pages})", min_value=1,
                    max_value=total_pages, value=1, step=1,
                    key='rcp_page',
                )
            else:
                page = 1

            start = (page - 1) * page_size
            end = start + page_size
            page_items = filtered[start:end]

            # ----- 1行に並べる枚数 (大きさ調整) -----
            cols_per_row = st.select_slider(
                "1行に並べる枚数 (写真の大きさ調整)",
                options=[3, 4, 5, 6],
                value=4,
                key='rcp_cols_per_row',
                help='大きい数字ほど 1枚が小さくコンパクトに並びます',
            )
            for i in range(0, len(page_items), cols_per_row):
                row_items = page_items[i:i + cols_per_row]
                cols = st.columns(cols_per_row)
                for col, r in zip(cols, row_items):
                    with col:
                        with st.container(border=True):
                            mtime_str = _dt.fromtimestamp(r['mtime']).strftime('%Y-%m-%d %H:%M')
                            kind_icon = '📸' if r['kind'] == 'image' else '📄'
                            status = r.get('status', '')
                            status_badge = ''
                            if status == '処理済み':
                                status_badge = (
                                    "<span style='background:#d1fae5; color:#065f46; "
                                    "padding:1px 6px; border-radius:4px; font-size:10px; "
                                    "font-weight:600; margin-left:6px;'>処理済</span>"
                                )
                            elif status == '要確認':
                                status_badge = (
                                    "<span style='background:#fee2e2; color:#991b1b; "
                                    "padding:1px 6px; border-radius:4px; font-size:10px; "
                                    "font-weight:600; margin-left:6px;'>要確認</span>"
                                )
                            ym_text = (f" ／ {r['year_month_folder']}"
                                       if r.get('year_month_folder') else '')

                            # 自動処理ステータス (receipt_processed) のバッジ
                            proc = db.get_receipt_processed(str(r['path']))
                            auto_badge = ''
                            if proc:
                                if proc['status'] == 'success':
                                    auto_badge = (
                                        "<span style='background:#dbeafe; color:#1e40af; "
                                        "padding:1px 6px; border-radius:4px; font-size:10px; "
                                        "font-weight:600; margin-left:6px;'>🤖 OCR済</span>"
                                    )
                                elif proc['status'] == 'manual_pending':
                                    auto_badge = (
                                        "<span style='background:#fef3c7; color:#78350f; "
                                        "padding:1px 6px; border-radius:4px; font-size:10px; "
                                        "font-weight:600; margin-left:6px;'>要確認</span>"
                                    )
                                elif proc['status'] == 'failed':
                                    auto_badge = (
                                        "<span style='background:#fee2e2; color:#991b1b; "
                                        "padding:1px 6px; border-radius:4px; font-size:10px; "
                                        "font-weight:600; margin-left:6px;'>OCR失敗</span>"
                                    )

                            st.markdown(
                                f"**{kind_icon} {r['facility']}**{status_badge}{auto_badge}<br>"
                                f"<span style='color:#64748b; font-size:11px;'>"
                                f"{mtime_str}{ym_text} ／ {_format_size(r['size'])}"
                                f"</span>",
                                unsafe_allow_html=True,
                            )

                            try:
                                with open(r['path'], 'rb') as fp:
                                    file_bytes = fp.read()
                            except Exception as e:
                                st.error(f"読込失敗: {e}")
                                continue

                            if r['kind'] == 'image':
                                if r['ext'] in ('.heic', '.heif'):
                                    # HEIC/HEIF は JPEG に変換してプレビュー
                                    jpg, err = _heic_to_jpeg_bytes(file_bytes)
                                    if jpg is not None:
                                        try:
                                            st.image(jpg, width='stretch')
                                        except Exception as e:
                                            st.warning(f"HEIC表示失敗: {e}")
                                    else:
                                        st.warning(
                                            f"🖼️ HEIC変換失敗: {err}\n\n"
                                            "（DLしてご確認ください）"
                                        )
                                else:
                                    try:
                                        st.image(file_bytes, width='stretch')
                                    except Exception as e:
                                        st.warning(f"画像表示失敗: {e}")
                            else:  # pdf
                                # PDFは埋め込み (iframe with base64) で先頭ページ表示
                                try:
                                    b64 = base64.b64encode(file_bytes).decode('ascii')
                                    pdf_html = (
                                        f'<iframe src="data:application/pdf;base64,{b64}" '
                                        f'width="100%" height="280" '
                                        f'style="border:1px solid #cbd5e1; border-radius:6px;">'
                                        f'</iframe>'
                                    )
                                    st.markdown(pdf_html, unsafe_allow_html=True)
                                except Exception as e:
                                    st.warning(f"PDFプレビュー失敗: {e}")

                            st.caption(f"`{r['name']}`")

                            # ダウンロード
                            mime_map = {
                                '.jpg': 'image/jpeg',
                                '.jpeg': 'image/jpeg',
                                '.png': 'image/png',
                                '.heic': 'image/heic',
                                '.heif': 'image/heic',
                                '.webp': 'image/webp',
                                '.bmp': 'image/bmp',
                                '.gif': 'image/gif',
                                '.tiff': 'image/tiff',
                                '.tif': 'image/tiff',
                                '.pdf': 'application/pdf',
                            }
                            st.download_button(
                                f"⬇️ ダウンロード",
                                data=file_bytes,
                                file_name=r['name'],
                                mime=mime_map.get(r['ext'], 'application/octet-stream'),
                                key=f"rcp_dl_{start + page_items.index(r)}",
                                width='stretch',
                            )
                            # OCR & Excel追記ポップオーバー (画像のみ)
                            _render_ocr_popover(
                                r, file_bytes,
                                start + page_items.index(r),
                            )

            if total_pages > 1:
                st.caption(
                    f"📄 ページ {page} / {total_pages} "
                    f"（{start + 1}〜{min(end, len(filtered))} 件目）"
                )
