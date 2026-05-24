"""損益ダッシュボード - 部門別 P&L の管理 / 可視化 / 分析

タブ構成:
  📊 サマリ         単月 or 期間で利益率付きの全体像
  🔍 構成比         売上・販管費の科目別シェア
  📅 比較           当月 vs 前月 vs 前年（単月）
"""
import io
import html
import os
import re
import zipfile
from collections import defaultdict
import streamlit as st
import pandas as pd
from lib import db, styling, auth, pl_parser, journal_parser


def _safe_filename(s: str) -> str:
    """Windows/Mac/Linux いずれでも問題ないファイル名に変換。"""
    s = s.replace('／', '_').replace('（', '(').replace('）', ')').replace('/', '_')
    s = re.sub(r'[\\:*?"<>|]', '_', s)
    s = re.sub(r'\s+', ' ', s).strip()
    return s or 'report'


def _build_pdfs_zip(items, period_tag: str, prefix: str) -> bytes:
    """施設ごとの個別PDFをZIPにまとめる。
    items: [(label, [page_elems]), ...]
    period_tag: ファイル名末尾に付ける期間ラベル（例: '2026-03'）
    prefix: ファイル名先頭の種別（例: '比較', '構成比'）
    """
    from lib import pdf_report as _pdf  # 局所import (登録済フォントの再呼び出しを避ける)
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        seen = {}
        for label, page_elems in items:
            base = _safe_filename(f"{prefix}_{label}_{period_tag}")
            # 同名重複を回避
            n = seen.get(base, 0)
            seen[base] = n + 1
            fname = f"{base}.pdf" if n == 0 else f"{base}_{n+1}.pdf"
            pdf_bytes = _pdf.build_pdf([page_elems])
            zf.writestr(fname, pdf_bytes)
    return buf.getvalue()


def _unique_yms(*ym_groups) -> list[str]:
    """CSV出力対象月を重複なしで並べる。"""
    seen = set()
    out = []
    for group in ym_groups:
        for ym in (group or []):
            if ym and ym not in seen:
                seen.add(ym)
                out.append(ym)
    return out


def _scope_subunit_ids(group_id: int | None = None,
                       subunit_id: int | None = None,
                       subunit_ids: list[int] | None = None) -> list[int] | None:
    """仕訳帳CSV用に、現在の対象範囲に含まれるサブ部門IDをそろえる。"""
    if subunit_ids:
        return list(subunit_ids)
    if subunit_id:
        return [subunit_id]
    if group_id:
        return [s['id'] for s in db.list_pl_subunits(group_id=group_id)]
    return None


def _csv_bytes_from_df(df: pd.DataFrame) -> bytes:
    return df.to_csv(index=False, encoding='utf-8-sig').encode('utf-8-sig')


def _pl_csv_bytes(scope_label: str, year_months: list[str],
                  group_id: int | None = None,
                  subunit_id: int | None = None,
                  subunit_ids: list[int] | None = None) -> bytes:
    fetch_kwargs = {}
    if subunit_ids:
        fetch_kwargs['subunit_ids'] = list(subunit_ids)
    elif subunit_id:
        fetch_kwargs['subunit_ids'] = [subunit_id]
    elif group_id:
        fetch_kwargs['group_ids'] = [group_id]

    rows = db.fetch_pl_entries(year_months=year_months, **fetch_kwargs) if year_months else []
    df = pd.DataFrame(rows)
    if df.empty:
        df = pd.DataFrame(columns=[
            '対象', '年月', '親グループ', 'サブ部門', '科目区分',
            '科目', '金額', '集計行',
        ])
    else:
        df = df.assign(対象=scope_label)
        df = df.rename(columns={
            'year_month': '年月',
            'group_name': '親グループ',
            'subunit_name': 'サブ部門',
            'category': '科目区分',
            'account_name': '科目',
            'amount': '金額',
            'is_total': '集計行',
        })
        keep_cols = ['対象', '年月', '親グループ', 'サブ部門', '科目区分', '科目', '金額', '集計行']
        df = df[[c for c in keep_cols if c in df.columns]]
    return _csv_bytes_from_df(df)


def _journal_csv_bytes(scope_label: str, year_months: list[str],
                       group_id: int | None = None,
                       subunit_id: int | None = None,
                       subunit_ids: list[int] | None = None) -> bytes:
    target_subunit_ids = _scope_subunit_ids(
        group_id=group_id, subunit_id=subunit_id, subunit_ids=subunit_ids
    )
    rows = db.export_journal_entries(
        year_months=year_months,
        subunit_ids=target_subunit_ids,
    ) if year_months else []
    df = pd.DataFrame(rows)
    if df.empty:
        df = pd.DataFrame(columns=[
            '対象', '取引日', '仕訳ID', '仕訳番号', '行番号', '取引内容',
            '借方科目', '借方金額', '借方部門', '借方サブ部門ID', '借方取引先', '借方摘要', '借方品目',
            '貸方科目', '貸方金額', '貸方部門', '貸方サブ部門ID', '貸方取引先', '貸方摘要', '貸方品目',
            '取込日時',
        ])
    else:
        df = df.assign(対象=scope_label)
        df = df.rename(columns={
            'transaction_date': '取引日',
            'journal_id': '仕訳ID',
            'journal_no': '仕訳番号',
            'record_no': '行番号',
            'transaction_content': '取引内容',
            'debit_account': '借方科目',
            'debit_amount': '借方金額',
            'debit_dept_clean': '借方部門',
            'debit_subunit_id': '借方サブ部門ID',
            'debit_vendor': '借方取引先',
            'debit_memo': '借方摘要',
            'debit_item': '借方品目',
            'credit_account': '貸方科目',
            'credit_amount': '貸方金額',
            'credit_dept_clean': '貸方部門',
            'credit_subunit_id': '貸方サブ部門ID',
            'credit_vendor': '貸方取引先',
            'credit_memo': '貸方摘要',
            'credit_item': '貸方品目',
            'imported_at': '取込日時',
        })
        keep_cols = [
            '対象', '取引日', '仕訳ID', '仕訳番号', '行番号', '取引内容',
            '借方科目', '借方金額', '借方部門', '借方サブ部門ID', '借方取引先', '借方摘要', '借方品目',
            '貸方科目', '貸方金額', '貸方部門', '貸方サブ部門ID', '貸方取引先', '貸方摘要', '貸方品目',
            '取込日時',
        ]
        df = df[[c for c in keep_cols if c in df.columns]]
    return _csv_bytes_from_df(df)


def _build_scope_csv_zip(items, period_tag: str, prefix: str,
                         include_journal: bool = True) -> bytes:
    """施設ごとの損益CSVと、あれば仕訳帳CSVをZIPにまとめる。"""
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        seen = {}
        for item in items:
            label = item['label']
            base = _safe_filename(f"{prefix}_{label}_{period_tag}")
            n = seen.get(base, 0)
            seen[base] = n + 1
            suffix = '' if n == 0 else f"_{n+1}"

            pl_bytes = _pl_csv_bytes(
                label, item.get('year_months') or [],
                group_id=item.get('group_id'),
                subunit_id=item.get('subunit_id'),
                subunit_ids=item.get('subunit_ids'),
            )
            zf.writestr(f"{base}{suffix}_損益.csv", pl_bytes)

            if include_journal:
                journal_bytes = _journal_csv_bytes(
                    label, item.get('year_months') or [],
                    group_id=item.get('group_id'),
                    subunit_id=item.get('subunit_id'),
                    subunit_ids=item.get('subunit_ids'),
                )
                zf.writestr(f"{base}{suffix}_仕訳帳.csv", journal_bytes)
    return buf.getvalue()

styling.inject_global_css()
auth.require_login()
auth.render_sidebar_navigation()
_is_admin = auth.is_admin()

st.title("損益ダッシュボード")
st.markdown(
    "<p style='color:#64748b; font-size:15px;'>"
    "部門別損益計算書を取り込み、利益率・構成比・比較で分析できます。"
    "</p>",
    unsafe_allow_html=True,
)
show_import_tools = False
if _is_admin:
    show_import_tools = st.toggle(
        "取込・削除などの管理メニューを表示する",
        value=False,
        help="普段の閲覧を軽くするため、必要な時だけ開きます。",
    )

# =============================================================
# 取込セクション
# =============================================================
existing_yms = db.list_pl_year_months()
data_exists = len(existing_yms) > 0

if show_import_tools:
  with st.expander(
      "📥 Excel取込" + (" (まだデータがありません)" if not data_exists else f" (取込済: {len(existing_yms)}ヶ月)"),
      expanded=not data_exists,
  ):
    st.markdown(
        "**第N期部門別損益計算書.xlsx** を取り込みます。"
        "シート名（例: `6月分`〜`3月分`）から月を判定し、"
        "**期首年月**を起点に YYYY-MM へ変換します。"
    )

    c1, c2 = st.columns([1, 2])
    with c1:
        ym_options = [f"{y}-{m:02d}" for y in range(2020, 2031) for m in range(1, 13)]
        default_idx = ym_options.index("2024-06") if "2024-06" in ym_options else 0
        fiscal_start_ym = st.selectbox(
            "期首年月", ym_options, index=default_idx,
            help="第5期(2024/6開始)なら 2024-06。第6期なら 2025-06。",
        )
    with c2:
        fiscal_label = st.text_input(
            "期ラベル(任意)", value="第5期",
            help="取込履歴の参照用ラベル",
        )

    upload_mode = st.radio(
        "取込モード",
        ["A: ファイル全部（複数シート）", "B: 1シートのみ"],
        horizontal=True,
    )

    uploaded = st.file_uploader("Excelファイル(.xlsx)", type=['xlsx'], key='pl_uploader')

    if uploaded is not None:
        sub_lookup = pl_parser.build_subunit_lookup(db.list_pl_subunits())
        known_accs = [a['name'] for a in db.list_pl_accounts()]

        with st.spinner("解析中..."):
            parse_result = pl_parser.parse_pl_workbook(
                uploaded, fiscal_start_ym=fiscal_start_ym,
                subunit_lookup=sub_lookup, known_accounts=known_accs,
            )

        if parse_result.error:
            st.error(parse_result.error)
        else:
            valid_sheets = [s for s in parse_result.sheets if s.year_month is not None]
            chosen_sheets = valid_sheets

            if upload_mode.startswith("B"):
                if not valid_sheets:
                    st.warning("取込可能な月度シートがありません。")
                    chosen_sheets = []
                else:
                    sheet_options = {
                        f"{s.sheet_name.strip()} → {s.year_month}": s
                        for s in valid_sheets
                    }
                    chosen_label = st.selectbox(
                        "取込対象シート", list(sheet_options.keys()),
                        key='pl_sheet_choice',
                    )
                    chosen_sheets = [sheet_options[chosen_label]]

            st.markdown("##### 解析結果プレビュー")
            preview_rows = []
            for sr in chosen_sheets:
                preview_rows.append({
                    'シート': sr.sheet_name.strip(),
                    '年月': sr.year_month,
                    'マッチ部門数': len(sr.matched_subunits),
                    '不明列(無視)': len(sr.unknown_subunit_columns),
                    '不明科目': len(sr.unknown_account_rows),
                })
            if preview_rows:
                st.dataframe(pd.DataFrame(preview_rows), hide_index=True, width='stretch')

            unknown_cols = set()
            unknown_rows = set()
            for sr in chosen_sheets:
                unknown_cols.update(sr.unknown_subunit_columns)
                unknown_rows.update(sr.unknown_account_rows)
            if unknown_cols:
                with st.expander(f"⚠️ マスタに無い部門列 (取込しない): {len(unknown_cols)}件"):
                    st.write(sorted(unknown_cols))
            if unknown_rows:
                with st.expander(f"⚠️ マスタに無い科目行 (取込しない): {len(unknown_rows)}件"):
                    st.write(sorted(unknown_rows))

            confirmed = st.button("🚀 この内容で取込を実行", type='primary', key='pl_import_run')
            if confirmed and chosen_sheets:
                filtered = pl_parser.ParseResult(fiscal_start_ym=fiscal_start_ym)
                filtered.sheets = chosen_sheets
                summary = pl_parser.import_parse_result_to_db(filtered, db)
                db.record_pl_import(
                    filename=uploaded.name, fiscal_label=fiscal_label,
                    fiscal_start_ym=fiscal_start_ym,
                    sheet_count=summary['sheets'], entry_count=summary['entries'],
                    year_months_list=summary['year_months'],
                )
                st.success(
                    f"✅ 取込完了: {summary['sheets']}シート / "
                    f"{summary['subunits']}部門 / {summary['entries']}エントリ "
                    f"({', '.join(summary['year_months'])})"
                )
                if summary['errors']:
                    with st.expander(f"取込時の警告 {len(summary['errors'])}件"):
                        for e in summary['errors']:
                            st.text(e)
                st.rerun()

# =============================================================
# 旧形式 CSV取込（試算表：損益計算書）
# =============================================================
if show_import_tools:
  with st.expander(
      "📥 旧形式CSV取込 (試算表：損益計算書) — 1ファイル＝1月分", expanded=False,
  ):
    st.markdown(
        "Excelに切替える前の **試算表：損益計算書_…(期間：YYYY年MM月).csv** を取り込みます。"
        "ファイル名から年月を自動判定（例: `期間：2023年06月` → `2023-06`）。"
        "**複数月をまとめてアップロード可**。"
    )

    csv_uploaded = st.file_uploader(
        "CSVファイル(.csv)", type=['csv'], accept_multiple_files=True,
        key='pl_csv_uploader',
    )

    if csv_uploaded:
        sub_lookup = pl_parser.build_subunit_lookup(db.list_pl_subunits())
        known_accs = [a['name'] for a in db.list_pl_accounts()]

        with st.spinner("解析中..."):
            csv_results = []
            for f in csv_uploaded:
                sr = pl_parser.parse_pl_csv(
                    f, filename=f.name,
                    subunit_lookup=sub_lookup,
                    known_accounts=known_accs,
                )
                csv_results.append(sr)

        # プレビュー
        st.markdown("##### 解析結果プレビュー")
        prev_rows = []
        for sr in csv_results:
            prev_rows.append({
                'ファイル': sr.sheet_name,
                '年月': sr.year_month or '(未判定)',
                'マッチ部門数': len(sr.matched_subunits),
                '不明列(無視)': len(sr.unknown_subunit_columns),
                '不明科目': len(sr.unknown_account_rows),
                'スキップ理由': sr.skipped_reason or '',
            })
        st.dataframe(pd.DataFrame(prev_rows), hide_index=True, width='stretch')

        # 不明科目 / 列 をまとめて出す
        unk_cols_all = set()
        unk_rows_all = set()
        for sr in csv_results:
            unk_cols_all.update(sr.unknown_subunit_columns)
            unk_rows_all.update(sr.unknown_account_rows)
        if unk_cols_all:
            with st.expander(f"⚠️ マスタに無い部門列 (取込しない): {len(unk_cols_all)}件"):
                st.write(sorted(unk_cols_all))
        if unk_rows_all:
            with st.expander(f"⚠️ マスタに無い科目行 (取込しない): {len(unk_rows_all)}件"):
                st.write(sorted(unk_rows_all))
                st.caption("※ セクション見出し（売上高・売上原価・販売管理費 等）は値を持たない見出し行のため無視されるのが正常です。")

        # 取込可能なシート（年月判定OKかつエントリあり）
        valid_csv = [sr for sr in csv_results if sr.year_month and sr.entries]
        st.write(f"取込対象: **{len(valid_csv)}** ファイル（{len(csv_uploaded) - len(valid_csv)} はスキップ）")

        if valid_csv and st.button("🚀 CSVを取込実行", type='primary', key='pl_csv_import_run'):
            filtered = pl_parser.ParseResult(fiscal_start_ym='(CSV)')
            filtered.sheets = valid_csv
            summary = pl_parser.import_parse_result_to_db(filtered, db)
            db.record_pl_import(
                filename=f"CSV×{len(valid_csv)}件: " + ', '.join(s.sheet_name[:30] for s in valid_csv[:3]) + ('...' if len(valid_csv) > 3 else ''),
                fiscal_label='(旧CSV)',
                fiscal_start_ym='(CSV)',
                sheet_count=summary['sheets'],
                entry_count=summary['entries'],
                year_months_list=summary['year_months'],
            )
            st.success(
                f"✅ 取込完了: {summary['sheets']}ファイル / "
                f"{summary['subunits']}部門 / {summary['entries']}エントリ "
                f"({', '.join(summary['year_months'])})"
            )
            if summary['errors']:
                with st.expander(f"取込時の警告 {len(summary['errors'])}件"):
                    for e in summary['errors']:
                        st.text(e)
            st.rerun()

# =============================================================
# 仕訳帳CSV取込 (分析レポート用の取引明細データ)
# =============================================================
if show_import_tools:
  with st.expander("📒 仕訳帳CSV取込 — 分析レポートの取引先明細用", expanded=False):
    st.markdown(
        "freee形式の仕訳帳CSV(`仕訳帳（新）CSV ...csv`)を取り込みます。"
        "**比較タブの分析レポート / PDF** に取引先別の詳細明細を表示するために使用。"
        " 借方部門名(例 `01 のじぎく高砂`)を自動認識してサブ部門にマッピング。"
    )

    journal_uploaded = st.file_uploader(
        "仕訳帳CSV (.csv / Shift-JIS)",
        type=['csv'], accept_multiple_files=False, key='journal_csv_uploader',
    )

    if journal_uploaded is not None:
        sub_lookup_j = journal_parser.build_subunit_lookup_for_journal(db.list_pl_subunits())
        with st.spinner("解析中..."):
            j_result = journal_parser.parse_journal_csv(journal_uploaded, sub_lookup_j)

        if j_result['errors']:
            with st.expander(f"⚠️ 解析エラー {len(j_result['errors'])}件"):
                for e in j_result['errors'][:50]:
                    st.text(e)

        st.success(
            f"解析完了: 全{len(j_result['rows'])}行 / "
            f"マッチ部門 {len(j_result['matched_subunits'])} / "
            f"未マッチ部門 {len(j_result['unknown_departments'])}"
        )

        if j_result['unknown_departments']:
            with st.expander(f"⚠️ マスタに無い部門名 (subunit_id 未解決): {len(j_result['unknown_departments'])}件"):
                for d in sorted(j_result['unknown_departments']):
                    st.text(d)
                st.caption(
                    "本部・管理部門・カフェ等は サブ部門マスタに無いため照合できませんが、行は保存されます。"
                )

        if st.button("🚀 この内容で取込を実行", type='primary', key='journal_import_run'):
            res = db.insert_journal_entries(
                j_result['rows'],
                filename=journal_uploaded.name,
                file_hash=j_result['file_hash'],
            )
            st.success(
                f"✅ 取込完了: 新規 {res['inserted']} / スキップ {res['skipped']} / "
                f"対象期間 {res['date_range']}"
            )
            st.rerun()

if show_import_tools:
    journal_imports_log = db.list_journal_imports(limit=5)
    if journal_imports_log:
        with st.expander(f"📜 仕訳帳 取込履歴 (直近{len(journal_imports_log)}件)"):
            st.dataframe(pd.DataFrame([
                {
                    '取込日時': r['imported_at'], 'ファイル': r['source_filename'],
                    '期間': r['date_range'], '行数': r['row_count'],
                    '新規': r['inserted_count'], 'スキップ': r['skipped_count'],
                } for r in journal_imports_log
            ]), hide_index=True, width='stretch')

# =============================================================
# 取込履歴
# =============================================================
if show_import_tools:
    imports_log = db.list_pl_imports(limit=10)
    if imports_log:
        with st.expander(f"📜 取込履歴 (直近{len(imports_log)}件)"):
            st.dataframe(pd.DataFrame([
                {
                    '取込日時': r['imported_at'], 'ファイル': r['source_filename'],
                    '期': r['fiscal_label'], '期首': r['fiscal_start_ym'],
                    'シート数': r['sheet_count'], 'エントリ数': r['entry_count'],
                    '対象月': r['year_months'],
                } for r in imports_log
            ]), hide_index=True, width='stretch')

if not data_exists:
    if _is_admin:
        st.info("👆 上の取込セクションから Excel をアップロードしてください。")
    else:
        st.info("まだ損益データが登録されていません。管理者に取込を依頼してください。")
    st.stop()


# =============================================================
# 共通: ヘルパ
# =============================================================
def color_profit(val):
    """利益はプラスを緑、マイナスを赤。"""
    try:
        n = int(val)
    except (TypeError, ValueError):
        return ''
    if n < 0:
        return 'color:#dc2626; font-weight:700'
    if n > 0:
        return 'color:#15803d; font-weight:700'
    return ''


def color_cost_change(val):
    """コスト増減: 増えたら赤、減ったら緑（利益とは逆）。"""
    try:
        n = int(val)
    except (TypeError, ValueError):
        return ''
    if n > 0:
        return 'color:#dc2626; font-weight:700'
    if n < 0:
        return 'color:#15803d; font-weight:700'
    return ''


def color_pct_change(val):
    """『+5.2%』『-3.1%』形式の文字列を色付け（プラス＝緑）"""
    if not isinstance(val, str):
        return ''
    if val.startswith('-'):
        return 'color:#dc2626; font-weight:700'
    if val.startswith('+') and val != '+0.0%':
        return 'color:#15803d; font-weight:700'
    return ''


def fmt_pct(num, denom, signed=False):
    if not denom:
        return '-'
    v = num / denom * 100
    if signed:
        return f"{v:+.1f}%"
    return f"{v:.1f}%"


def fmt_diff_pct(curr, prev, signed=True):
    """前期比 増減率"""
    if not prev:
        return '-'
    v = (curr - prev) / abs(prev) * 100
    return f"{v:+.1f}%" if signed else f"{v:.1f}%"


def sum_by_cat(entries, cat):
    return sum(e['amount'] for e in entries if e['category'] == cat)


def fetch_pl(year_months, group_id=None, subunit_id=None):
    return db.fetch_pl_entries(
        year_months=year_months,
        group_ids=[group_id] if group_id else None,
        subunit_ids=[subunit_id] if subunit_id else None,
    )


# =============================================================
# 共通フィルタ: グループ/サブ部門 + 会計年度モード
# =============================================================
st.markdown("---")
st.markdown("### 🎯 表示対象")

c_target, c_mode = st.columns([2, 1])

with c_target:
    # 階層型セレクタ: 親(グループ)と子(サブ部門)どちらでも選べる
    groups_all = db.list_pl_groups()
    all_subs = db.list_pl_subunits()
    subs_by_group = defaultdict(list)
    for s in all_subs:
        subs_by_group[s['group_id']].append(s)

    # 選択肢: label -> ('all'|'group'|'subunit', id)
    selection_options: dict[str, tuple[str, int | None]] = {}
    selection_options["📊 全グループ（比較表示）"] = ('all', None)
    for group in groups_all:
        try:
            g_num = int(group['code'])
        except ValueError:
            g_num = group['id']
        parent_label = f"{group['code']}.{group['name']}"
        if group.get('note'):
            parent_label += f"（{group['note']}）"
        selection_options[parent_label] = ('group', group['id'])

        for s_idx, sub in enumerate(subs_by_group.get(group['id'], []), start=1):
            child_label = f"　└ {g_num}.{s_idx}.{sub['display_name']}"
            selection_options[child_label] = ('subunit', sub['id'])

    sel_label = st.selectbox(
        "対象（親グループ または 子サブ部門）",
        list(selection_options.keys()),
        key='pl_target_select',
    )
    sel_type, sel_id = selection_options[sel_label]
    sel_group_id = sel_id if sel_type == 'group' else None
    sel_subunit_id = sel_id if sel_type == 'subunit' else None
    # ヘッダ表示用の整形ラベル（インデント記号を除去）
    sel_display = sel_label.replace('　└ ', '').strip()
    # サブ部門選択時、その親グループIDも保持（販管費比較に使う場合）
    sel_parent_group_id = None
    if sel_type == 'subunit':
        sub_obj = next(s for s in all_subs if s['id'] == sel_id)
        sel_parent_group_id = sub_obj['group_id']

with c_mode:
    fy_mode_label = st.radio(
        "会計年度の基準",
        ["法人決算(6月開始・第N期)", "現場評価(4月開始・年度)"],
        key='pl_fy_mode',
        help="構成比(期累計)／比較タブで使う期の区切り方を切替えます。",
    )
    fy_start_month = 6 if fy_mode_label.startswith("法人") else 4

def fy_label(fy):
    return db.pl_fiscal_year_label(fy, start_month=fy_start_month)

# =============================================================
# 業績会議・PDFレポート 共通定数 (tab実行前に定義)
# =============================================================
PERSONNEL_ACCOUNTS = {
    '管理者給', '指導員給', '法定福利費', '退職給付費用', '賞与', '事務員給',
    # NPO法人EMIFULL(のじぎく系)で人件費に含める科目。
    # B型利用者工賃 は経費扱い(人件費に含めない)。
    'A型利用者給', '賞与引当金繰入',
}

# 業績会議用 行構成: (region_label, [(row_label, [excel_name, ...])])
REPORT_STRUCTURE = [
    ("いなみ", [
        ("SORATOいなみ",         ["SORATOいなみ"]),
        ("UMIEいなみ",           ["UMIEいなみ"]),
        ("UMIEいなみ第二",        ["UMIEいなみ2"]),
        ("SORATOいなみ第二",      ["SORATOいなみ2"]),
        ("BLOOMいなみ",          ["BLOOMいなみ"]),
    ]),
    ("てんり", [
        ("SORATOてんり",         ["SORATOてんり"]),
        ("UMIEてんり",           ["UMIEてんり"]),
        ("BLOOMてんり",          ["BLOOMてんり"]),
        ("カラダキッズてんり",    ["カラダキッズてんり"]),
    ]),
    ("天理", [
        ("シェア天理1.2",         ["シェア天理1.2", "シェアホーム天理1.2"]),
        ("シェア天理3",           ["シェア天理3", "シェアホーム天理3"]),
    ]),
    ("かこがわ", [
        ("ジョブカレッジかこがわ", ["ジョブカレッジかこがわ", "ナインカレッジ"]),
        ("カラダキッズかこがわ",   ["カラダキッズかこがわ", "ナインキッズ"]),
    ]),
    ("加古川", [
        ("シェア加古川",          ["シェアホーム加古川"]),
        ("相談支援NOAH",         ["NOAH（加古川）"]),
    ]),
]

EXCLUDED_ROWS = [
    ("きたはま",          ["SORATOUMIEきたはま"]),
]

# NPO法人EMIFULL (のじぎく系)
NPO_REPORT_STRUCTURE = [
    ("高砂", [
        ("のじぎく高砂", ["のじぎく高砂"]),
    ]),
    ("稲美", [
        ("のじぎく稲美",    ["のじぎく稲美", "キッチン"]),  # 大西キッチンを合算
        ("のじぎく加古川",  ["のじぎく加古川"]),
        ("こすもす稲美",    ["こすもす稲美"]),
        ("だがし屋キューブ", ["だがし屋キューブ"]),
    ]),
]

# excel_name → subunit_id (PDF/業績会議で頻繁に使う)
_pl_subunits_all = db.list_pl_subunits()
subs_by_excel = {s['excel_name']: s['id'] for s in _pl_subunits_all}
subunit_by_id = {s['id']: s for s in _pl_subunits_all}

REPORTER_ROLES = ["部長", "次長", "課長", "係長", "主任", "副主任", "一般職"]


def _report_facilities_for_structure(structure, excluded_rows=None):
    """業績会議の行と同じ単位で、報告対象施設の選択肢を作る。"""
    rows = []
    seen = set()
    for _, subrows in structure:
        for row_label, excel_names in subrows:
            if row_label in seen:
                continue
            sub_ids = [subs_by_excel[n] for n in excel_names if n in subs_by_excel]
            rows.append({
                'id': sub_ids[0] if sub_ids else None,
                'name': row_label,
                'excel_names': excel_names,
            })
            seen.add(row_label)
    for row_label, excel_names in (excluded_rows or []):
        if row_label in seen:
            continue
        sub_ids = [subs_by_excel[n] for n in excel_names if n in subs_by_excel]
        rows.append({
            'id': sub_ids[0] if sub_ids else None,
            'name': row_label,
            'excel_names': excel_names,
        })
        seen.add(row_label)
    return rows


REPORT_FACILITIES = (
    _report_facilities_for_structure(REPORT_STRUCTURE, EXCLUDED_ROWS)
    + _report_facilities_for_structure(NPO_REPORT_STRUCTURE, [])
)
REPORT_FACILITY_BY_NAME = {f['name']: f for f in REPORT_FACILITIES}


def _default_report_month():
    for key in ('pl_summary_ym', 'pl_meeting_ym'):
        if st.session_state.get(key) in existing_yms:
            return st.session_state[key]
    return existing_yms[0]


def _compact_text(text, limit=90):
    text = re.sub(r'\s+', ' ', (text or '')).strip()
    if len(text) <= limit:
        return text
    return text[:limit - 1] + "…"


def _yen_man(v):
    return f"{int(v / 10000):,}万円"


def _pct(v):
    return f"{v:.1f}%" if isinstance(v, (int, float)) else "－"


def _pl_metrics_for_excel_names(target_month, excel_names):
    sub_ids = {subs_by_excel[n] for n in excel_names if n in subs_by_excel}
    entries = [
        e for e in db.fetch_pl_entries(year_months=[target_month])
        if e['subunit_id'] in sub_ids
    ]
    rev = sum(e['amount'] for e in entries if e['category'] == 'revenue_total')
    sga = sum(e['amount'] for e in entries if e['category'] == 'sga_total')
    pers = sum(e['amount'] for e in entries
               if e['category'] == 'sga' and e['account_name'] in PERSONNEL_ACCOUNTS)
    exp = sga - pers
    op = sum(e['amount'] for e in entries if e['category'] == 'op_profit')
    return {
        '売上': rev,
        '人件費率': (pers / rev * 100) if rev else None,
        '経費率': (exp / rev * 100) if rev else None,
        '営業利益': op,
        '利益率': (op / rev * 100) if rev else None,
    }


def _fallback_report_summary(target_month, facility_name, excel_names, reports):
    if not reports:
        return "未提出"
    metrics = _pl_metrics_for_excel_names(target_month, excel_names)
    previous = _compact_text(" / ".join(r.get('previous_review') or '' for r in reports), 70)
    issue = _compact_text(" / ".join(r.get('issue_review') or '' for r in reports), 70)
    actions = _compact_text(" / ".join(r.get('next_actions') or '' for r in reports), 70)
    return (
        f"{target_month}の{facility_name}は、売上{_yen_man(metrics['売上'])}、"
        f"人件費率{_pct(metrics['人件費率'])}、経費率{_pct(metrics['経費率'])}、"
        f"営業利益{_yen_man(metrics['営業利益'])}、利益率{_pct(metrics['利益率'])}で、"
        f"前月の振り返りは{previous or '確認中'}、現在の課題は{issue or '整理中'}であり、"
        f"次月は{actions or '改善策の具体化'}を進める必要がある。"
    )


def _openai_api_key():
    value = os.getenv("OPENAI_API_KEY")
    if value:
        return value
    try:
        return st.secrets.get("OPENAI_API_KEY")
    except Exception:
        return None


def build_facility_report_summary(target_month, facility_name, excel_names, reports,
                                  force_ai=False):
    """報告書と損益数値を合わせて、施設/月単位の一文要約を返す。"""
    if not reports:
        return "未提出"
    if not force_ai:
        saved = next((r.get('ai_summary') for r in reports if r.get('ai_summary')), None)
        if saved:
            return saved

    fallback = _fallback_report_summary(target_month, facility_name, excel_names, reports)
    api_key = _openai_api_key()
    if not force_ai or not api_key:
        return fallback
    try:
        from openai import OpenAI
        metrics = _pl_metrics_for_excel_names(target_month, excel_names)
        body = "\n\n".join(
            f"報告者:{r.get('reporter_role')} {r.get('reporter_name')}\n"
            f"前月の振り返り:{r.get('previous_review') or ''}\n"
            f"現在の課題:{r.get('issue_review')}\n"
            f"次月以降の対策:{r.get('next_actions')}\n"
            f"その他:{r.get('other_notes') or ''}"
            for r in reports
        )
        prompt = (
            "施設ごとの月次収支報告を、会議資料に載せる一文の日本語に要約してください。"
            "一文で簡潔に、売上・人件費率・経費率・営業利益・利益率のうち自然に入る数値を含め、"
            "複数報告者がいる場合は内容を統合してください。\n"
            f"対象月:{target_month}\n施設:{facility_name}\n"
            f"数値:{metrics}\n報告書:\n{body}"
        )
        client = OpenAI(api_key=api_key)
        res = client.chat.completions.create(
            model=os.getenv("OPENAI_MODEL", "gpt-4o-mini"),
            messages=[
                {"role": "system", "content": "あなたは福祉事業の収支会議資料を作る実務者です。"},
                {"role": "user", "content": prompt},
            ],
            temperature=0.2,
            max_tokens=180,
        )
        return (res.choices[0].message.content or fallback).strip().replace("\n", "")
    except Exception:
        return fallback


def _field_source_text(reports, field_name):
    values = []
    for r in reports:
        text = (r.get(field_name) or '').strip()
        if text:
            values.append(text)
    return "\n".join(values)


def _fallback_field_summary(field_label, source_text):
    source_text = re.sub(r'\s+', ' ', (source_text or '')).strip()
    if not source_text:
        return ""
    chunks = re.split(r'(?<=[。！？])\s*', source_text)
    sentences = [c for c in chunks if c.strip()]
    if len(sentences) >= 2:
        return "".join(sentences[:3])
    return _compact_text(source_text, 180)


def build_facility_field_summary(target_month, facility_name, excel_names, reports,
                                 field_name, field_label):
    """会議プレビュー/Excel用に、報告原文を2〜3文へ整える。"""
    if not reports:
        return "未提出"
    source_text = _field_source_text(reports, field_name)
    if not source_text:
        return ""

    fallback = _fallback_field_summary(field_label, source_text)
    api_key = _openai_api_key()
    if not api_key:
        return fallback
    try:
        from openai import OpenAI
        metrics = _pl_metrics_for_excel_names(target_month, excel_names)
        prompt = (
            f"{facility_name}の月次報告の「{field_label}」を、業績会議に載せる文章として"
            "日本語で2〜3文に要約してください。"
            "原文の重要な意味を変えず、読みやすく、箇条書きではなく文章でまとめてください。"
            "必要に応じて売上・人件費率・経費率・営業利益・利益率の数値も自然に含めてください。\n"
            f"対象月:{target_month}\n施設:{facility_name}\n数値:{metrics}\n原文:\n{source_text}"
        )
        client = OpenAI(api_key=api_key)
        res = client.chat.completions.create(
            model=os.getenv("OPENAI_MODEL", "gpt-4o-mini"),
            messages=[
                {"role": "system", "content": "あなたは福祉事業の業績会議資料を読みやすく整える実務者です。"},
                {"role": "user", "content": prompt},
            ],
            temperature=0.2,
            max_tokens=260,
        )
        return (res.choices[0].message.content or fallback).strip()
    except Exception:
        return fallback


def _refresh_facility_month_summary(target_month, facility_name):
    fac = REPORT_FACILITY_BY_NAME.get(facility_name)
    if not fac:
        return None
    reports = db.get_profit_reports(target_month=target_month, facility_name=facility_name)
    if not reports:
        return None
    summary = build_facility_report_summary(
        target_month, facility_name, fac['excel_names'], reports, force_ai=True
    )
    for r in reports:
        db.update_profit_report_ai_summary(r['id'], summary)
    return summary


def _reports_for_month_by_facility(target_month):
    grouped = defaultdict(list)
    for r in db.get_profit_reports(target_month=target_month):
        grouped[r['facility_name']].append(r)
    return grouped


def _report_preview_rows(target_month, facilities=None):
    grouped = _reports_for_month_by_facility(target_month)
    rows = []
    for fac in (facilities or REPORT_FACILITIES):
        reports = grouped.get(fac['name'], [])
        latest = reports[0] if reports else {}
        rows.append({
            '対象月': target_month,
            '施設名': fac['name'],
            '提出状況': '提出済み' if reports else '未提出',
            'AI要約': build_facility_report_summary(
                target_month, fac['name'], fac['excel_names'], reports
            ),
            '前月の振り返り要約': build_facility_field_summary(
                target_month, fac['name'], fac['excel_names'], reports,
                'previous_review', '前月の振り返り'
            ),
            '現在の課題要約': build_facility_field_summary(
                target_month, fac['name'], fac['excel_names'], reports,
                'issue_review', '現在の課題'
            ),
            '対策要約': build_facility_field_summary(
                target_month, fac['name'], fac['excel_names'], reports,
                'next_actions', '次月以降の対策'
            ),
            'その他要約': build_facility_field_summary(
                target_month, fac['name'], fac['excel_names'], reports,
                'other_notes', 'その他'
            ),
            '前月の振り返り': "\n---\n".join(r.get('previous_review') or '' for r in reports),
            '現在の課題': "\n---\n".join(r.get('issue_review') or '' for r in reports),
            '対策': "\n---\n".join(r.get('next_actions') or '' for r in reports),
            'その他': "\n---\n".join(r.get('other_notes') or '' for r in reports),
            '報告者': " / ".join(
                f"{r.get('reporter_role')} {r.get('reporter_name')}" for r in reports
            ),
            '提出日時': latest.get('created_at', ''),
        })
    return rows


def _section_text_from_preview(rows, field_name):
    lines = []
    for r in rows:
        value = (r.get(field_name) or '').strip()
        if not value:
            value = '未提出' if r.get('提出状況') == '未提出' else ''
        if value:
            lines.append(f"{r['施設名']}：{value}")
    return "\n\n".join(lines) or "未提出"


def _render_summary_cards(title, rows, field_name):
    st.markdown(f"###### {title}")
    st.markdown(
        """
        <style>
        .meeting-summary-list {
            display: flex;
            flex-direction: column;
            gap: 10px;
            margin-bottom: 18px;
        }
        .meeting-summary-card {
            background: #ffffff;
            border: 1px solid #dbe3ef;
            border-left: 5px solid #93c5fd;
            border-radius: 8px;
            padding: 12px 14px;
            font-family: "BIZ UDPゴシック", "Yu Gothic", Meiryo, sans-serif;
            color: #172033;
            line-height: 1.75;
            font-size: 15px;
            font-weight: 400;
        }
        .meeting-summary-card strong {
            display: inline-block;
            min-width: 160px;
            color: #1e3a8a;
            font-weight: 800;
            margin-right: 6px;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )
    cards = []
    for r in rows:
        value = (r.get(field_name) or '').strip()
        if not value:
            value = '未提出' if r.get('提出状況') == '未提出' else '記載なし'
        cards.append(
            "<div class='meeting-summary-card'>"
            f"<strong>{html.escape(r['施設名'])}</strong>"
            f"{html.escape(value).replace(chr(10), '<br>')}"
            "</div>"
        )
    st.markdown("<div class='meeting-summary-list'>" + "".join(cards) + "</div>", unsafe_allow_html=True)


def _reports_csv_bytes(rows):
    buf = io.StringIO()
    pd.DataFrame(rows).to_csv(buf, index=False)
    return buf.getvalue().encode('utf-8-sig')


# =============================================================
# タブ
# =============================================================
if _is_admin:
    tab_summary, tab_ratio, tab_yoy, tab_meeting, tab_report, tab_sga_trend = st.tabs([
        "📊 サマリ", "🔍 構成比", "📅 比較", "🏛️ 業績会議", "📝 報告書提出", "📈 販管費推移",
    ])
else:
    tab_summary, tab_ratio, tab_yoy, tab_sga_trend = st.tabs([
        "📊 サマリ", "🔍 構成比", "📅 比較", "📈 販管費推移",
    ])
    tab_meeting = None
    tab_report = None


def _ym_minus(ym: str, n: int) -> str:
    """'2026-03' から n ヶ月前を返す。"""
    y, m = int(ym[:4]), int(ym[5:7])
    total = y * 12 + (m - 1) - n
    return f"{total // 12}-{(total % 12) + 1:02d}"


# =============================================================
# TAB 1: サマリ
# =============================================================
with tab_summary:
    c1, c2 = st.columns([1, 2])
    with c1:
        period_mode = st.radio("期間モード", ["単月", "期間指定"], key='pl_summary_period_mode')
    with c2:
        if period_mode == "単月":
            sel_ym = st.selectbox("対象月", existing_yms, index=0, key='pl_summary_ym')
            target_yms = [sel_ym]
            period_label = sel_ym
        else:
            cs1, cs2 = st.columns(2)
            with cs1:
                start_ym = st.selectbox(
                    "開始月", existing_yms,
                    index=len(existing_yms) - 1, key='pl_summary_start',
                )
            with cs2:
                end_ym = st.selectbox(
                    "終了月", existing_yms, index=0, key='pl_summary_end',
                )
            target_yms = sorted([ym for ym in existing_yms if start_ym <= ym <= end_ym])
            period_label = f"{start_ym} 〜 {end_ym}"

    if not target_yms:
        st.warning("選択期間にデータがありません")
    else:
        entries = fetch_pl(target_yms, sel_group_id, sel_subunit_id)
        if not entries:
            st.warning("対象期間にデータがありません")
        else:
            # ヘッダ
            st.markdown(
                f"<div style='background:linear-gradient(90deg,#dbeafe 0%,#fef3c7 100%); "
                f"padding:14px 20px; border-radius:10px; border-left:6px solid #2563eb; "
                f"margin:16px 0;'>"
                f"<div style='font-size:13px; color:#475569; font-weight:600;'>表示中</div>"
                f"<div style='font-size:20px; color:#0f172a; font-weight:700; margin-top:4px;'>"
                f"{period_label}　／　{sel_display}</div></div>",
                unsafe_allow_html=True,
            )

            # 主要指標 (5項目: 売上高/人件費/経費/営業利益/経常利益)
            rev = sum_by_cat(entries, 'revenue_total')
            sga = sum_by_cat(entries, 'sga_total')
            pers = sum(e['amount'] for e in entries
                       if e['category'] == 'sga' and e['account_name'] in PERSONNEL_ACCOUNTS)
            exp = sga - pers
            op = sum_by_cat(entries, 'op_profit')
            ord_ = sum_by_cat(entries, 'ordinary_profit')

            c1, c2, c3, c4, c5 = st.columns(5)
            c1.metric("売上高", f"{rev:,} 円")
            c2.metric(
                "人件費", f"{pers:,} 円",
                delta=f"人件費率 {fmt_pct(pers, rev)}" if rev else None,
                delta_color="off",
            )
            c3.metric(
                "経費", f"{exp:,} 円",
                delta=f"経費率 {fmt_pct(exp, rev)}" if rev else None,
                delta_color="off",
            )
            c4.metric(
                "営業利益", f"{op:,} 円",
                delta=f"営業利益率 {fmt_pct(op, rev)}" if rev else None,
                delta_color="normal" if op >= 0 else "inverse",
            )
            c5.metric(
                "経常利益", f"{ord_:,} 円",
                delta=f"経常利益率 {fmt_pct(ord_, rev)}" if rev else None,
                delta_color="normal" if ord_ >= 0 else "inverse",
            )

            # ===== グループ別 / サブ部門別 比較 =====
            st.markdown("---")
            if sel_type == 'subunit':
                # 単一サブ部門のみ → 比較ではなく科目別内訳のみ表示
                st.markdown("### 販管費 科目別内訳")
                sga_entries = [e for e in entries if e['category'] == 'sga']
                if sga_entries:
                    acc_agg = defaultdict(int)
                    acc_order = {}
                    for e in sga_entries:
                        acc_agg[e['account_name']] += e['amount']
                        acc_order[e['account_name']] = e['account_order']
                    sga_rows = []
                    for acc_name, v in sorted(acc_agg.items(), key=lambda x: acc_order[x[0]]):
                        if v != 0:
                            sga_rows.append({'科目': acc_name, '金額': v, '対売上比': fmt_pct(v, rev)})
                    if sga_rows:
                        df_sga = pd.DataFrame(sga_rows)
                        st.dataframe(
                            df_sga.style.format({'金額': '{:,}'}),
                            hide_index=True, width='stretch',
                            height=min(35 * (len(df_sga) + 1) + 3, 600),
                        )
                    else:
                        st.info("販管費はすべて0円です。")
                else:
                    st.info("販管費データがありません。")
            elif sel_group_id is None:
                st.markdown("### グループ別 損益比較")
                grp_agg = defaultdict(lambda: {
                    'order': 0, 'name': '', 'rev': 0, 'pers': 0, 'exp': 0,
                    'op': 0, 'ord_': 0,
                })
                for e in entries:
                    g = grp_agg[e['group_id']]
                    g['order'] = e['group_order']
                    g['name'] = e['group_name']
                    cat = e['category']
                    if cat == 'revenue_total': g['rev'] += e['amount']
                    elif cat == 'sga':
                        if e['account_name'] in PERSONNEL_ACCOUNTS:
                            g['pers'] += e['amount']
                        else:
                            g['exp'] += e['amount']
                    elif cat == 'op_profit':   g['op'] += e['amount']
                    elif cat == 'ordinary_profit': g['ord_'] += e['amount']

                rows = []
                for _, v in sorted(grp_agg.items(), key=lambda x: x[1]['order']):
                    rows.append({
                        'グループ': v['name'],
                        '売上高': v['rev'], '人件費': v['pers'], '経費': v['exp'],
                        '営業利益': v['op'], '営業利益率': fmt_pct(v['op'], v['rev']),
                        '経常利益': v['ord_'], '経常利益率': fmt_pct(v['ord_'], v['rev']),
                        '人件費率': fmt_pct(v['pers'], v['rev']),
                    })
                df = pd.DataFrame(rows)
                styled = (df.style
                    .map(color_profit, subset=['営業利益', '経常利益'])
                    .format({c: '{:,}' for c in ['売上高', '人件費', '経費', '営業利益', '経常利益']})
                )
                st.dataframe(
                    styled, hide_index=True, width='stretch',
                    height=35 * (len(df) + 1) + 3,
                )
            else:
                st.markdown("### サブ部門別 損益比較")
                sub_agg = defaultdict(lambda: {
                    'order': 0, 'name': '', 'rev': 0, 'pers': 0, 'exp': 0,
                    'op': 0, 'ord_': 0,
                })
                for e in entries:
                    s = sub_agg[e['subunit_id']]
                    s['order'] = e['subunit_order']
                    s['name'] = e['subunit_name']
                    cat = e['category']
                    if cat == 'revenue_total': s['rev'] += e['amount']
                    elif cat == 'sga':
                        if e['account_name'] in PERSONNEL_ACCOUNTS:
                            s['pers'] += e['amount']
                        else:
                            s['exp'] += e['amount']
                    elif cat == 'op_profit':   s['op'] += e['amount']
                    elif cat == 'ordinary_profit': s['ord_'] += e['amount']

                rows = []
                for _, v in sorted(sub_agg.items(), key=lambda x: x[1]['order']):
                    rows.append({
                        'サブ部門': v['name'],
                        '売上高': v['rev'], '人件費': v['pers'], '経費': v['exp'],
                        '営業利益': v['op'], '営業利益率': fmt_pct(v['op'], v['rev']),
                        '経常利益': v['ord_'], '経常利益率': fmt_pct(v['ord_'], v['rev']),
                        '人件費率': fmt_pct(v['pers'], v['rev']),
                    })
                # 合計行
                if rows:
                    sums = {k: sum(r[k] for r in rows) for k in ['売上高', '人件費', '経費', '営業利益', '経常利益']}
                    rows.append({
                        'サブ部門': '【グループ合計】',
                        **sums,
                        '営業利益率': fmt_pct(sums['営業利益'], sums['売上高']),
                        '経常利益率': fmt_pct(sums['経常利益'], sums['売上高']),
                        '人件費率': fmt_pct(sums['人件費'], sums['売上高']),
                    })
                df = pd.DataFrame(rows)

                def bold_total(row):
                    if row['サブ部門'] == '【グループ合計】':
                        return ['background-color:#f1f5f9; font-weight:700'] * len(row)
                    return [''] * len(row)

                styled = (df.style
                    .apply(bold_total, axis=1)
                    .map(color_profit, subset=['営業利益', '経常利益'])
                    .format({c: '{:,}' for c in ['売上高', '人件費', '経費', '営業利益', '経常利益']})
                )
                st.dataframe(
                    styled, hide_index=True, width='stretch',
                    height=35 * (len(df) + 1) + 3,
                )

                # 販管費科目別
                st.markdown("---")
                st.markdown("### 販管費 科目別内訳")
                sga_entries = [e for e in entries if e['category'] == 'sga']
                if sga_entries:
                    sub_names_ordered = [
                        v['name'] for _, v in sorted(sub_agg.items(), key=lambda x: x[1]['order'])
                    ]
                    acc_agg = defaultdict(lambda: defaultdict(int))
                    acc_order = {}
                    for e in sga_entries:
                        acc_agg[e['account_name']][e['subunit_name']] += e['amount']
                        acc_order[e['account_name']] = e['account_order']

                    sga_rows = []
                    for acc_name, by_sub in sorted(acc_agg.items(), key=lambda x: acc_order[x[0]]):
                        row = {'科目': acc_name}
                        s = 0
                        for sn in sub_names_ordered:
                            v = by_sub.get(sn, 0)
                            row[sn] = v
                            s += v
                        row['合計'] = s
                        if s != 0:
                            sga_rows.append(row)
                    if sga_rows:
                        df_sga = pd.DataFrame(sga_rows)
                        fmt = {c: '{:,}' for c in df_sga.columns if c != '科目'}
                        st.dataframe(
                            df_sga.style.format(fmt),
                            hide_index=True, width='stretch',
                            height=min(35 * (len(df_sga) + 1) + 3, 600),
                        )

# =============================================================
# TAB 2: 構成比
# =============================================================
with tab_ratio:
    c1, c2 = st.columns([1, 2])
    with c1:
        ratio_period_mode = st.radio(
            "期間モード", ["単月", "期間指定", "期累計"], key='pl_ratio_period_mode',
        )
    with c2:
        if ratio_period_mode == "単月":
            ratio_ym = st.selectbox("対象月", existing_yms, index=0, key='pl_ratio_ym')
            ratio_yms = [ratio_ym]
            ratio_label = ratio_ym
        elif ratio_period_mode == "期間指定":
            cs1, cs2 = st.columns(2)
            with cs1:
                rs = st.selectbox("開始月", existing_yms, index=len(existing_yms) - 1, key='pl_ratio_start')
            with cs2:
                re_ = st.selectbox("終了月", existing_yms, index=0, key='pl_ratio_end')
            ratio_yms = sorted([ym for ym in existing_yms if rs <= ym <= re_])
            ratio_label = f"{rs} 〜 {re_}"
        else:
            fy_list = db.list_pl_fiscal_years(start_month=fy_start_month)
            ratio_fy = st.selectbox(
                "対象期", fy_list,
                format_func=fy_label,
                key='pl_ratio_fy',
            )
            ratio_yms = [
                m for m in db.pl_fiscal_year_months(ratio_fy, start_month=fy_start_month)
                if m in existing_yms
            ]
            ratio_label = fy_label(ratio_fy) + "(累計)"

    if not ratio_yms:
        st.warning("選択期間にデータがありません")
    else:
        entries = fetch_pl(ratio_yms, sel_group_id, sel_subunit_id)
        st.markdown(
            f"<div style='font-size:14px; color:#475569; margin-top:8px;'>対象: <b>{ratio_label}</b> ／ {sel_display}</div>",
            unsafe_allow_html=True,
        )
        rev_total = sum_by_cat(entries, 'revenue_total')
        sga_total = sum_by_cat(entries, 'sga_total')

        col_r, col_s = st.columns(2)

        with col_r:
            st.markdown("##### 売上の構成")
            rev_agg = defaultdict(int)
            rev_order = {}
            for e in entries:
                if e['category'] == 'revenue':
                    rev_agg[e['account_name']] += e['amount']
                    rev_order[e['account_name']] = e['account_order']
            rev_rows = []
            for name, v in sorted(rev_agg.items(), key=lambda x: rev_order[x[0]]):
                if v != 0:
                    rev_rows.append({
                        '科目': name, '金額': v,
                        '構成比': fmt_pct(v, rev_total),
                    })
            if rev_rows:
                rev_rows.append({'科目': '【売上高 計】', '金額': rev_total, '構成比': '100.0%'})
                df_r = pd.DataFrame(rev_rows)

                def bold_total_r(row):
                    if row['科目'] == '【売上高 計】':
                        return ['background-color:#f1f5f9; font-weight:700'] * len(row)
                    return [''] * len(row)

                st.dataframe(
                    df_r.style.apply(bold_total_r, axis=1).format({'金額': '{:,}'}),
                    hide_index=True, width='stretch',
                )
            else:
                st.info("売上データがありません。")

            # ===== 主要指標（人件費・経費・営業利益・経常利益） =====
            st.markdown("##### 主要指標")
            PERSONNEL_ACCOUNTS = {
                '管理者給', '指導員給', '法定福利費', '退職給付費用', '賞与', '事務員給',
            }
            personnel = sum(
                e['amount'] for e in entries
                if e['category'] == 'sga' and e['account_name'] in PERSONNEL_ACCOUNTS
            )
            other_expense = sga_total - personnel
            op_total = sum_by_cat(entries, 'op_profit')
            ord_total = sum_by_cat(entries, 'ordinary_profit')

            metrics_rows = [
                {'項目': '人件費', '金額': personnel,
                 '対売上比': fmt_pct(personnel, rev_total)},
                {'項目': '経費',   '金額': other_expense,
                 '対売上比': fmt_pct(other_expense, rev_total)},
                {'項目': '営業利益', '金額': op_total,
                 '対売上比': fmt_pct(op_total, rev_total)},
                {'項目': '経常利益', '金額': ord_total,
                 '対売上比': fmt_pct(ord_total, rev_total)},
            ]
            df_m = pd.DataFrame(metrics_rows)

            def style_metrics(row):
                styles = [''] * len(row)
                idx = list(row.index)
                amt_idx = idx.index('金額')
                pct_idx = idx.index('対売上比')
                # 人件費・経費はコスト寄り：金額の色付けはせず通常表示
                # 営業利益・経常利益は損益なので 赤/緑 で色分け
                if row['項目'] in ('営業利益', '経常利益'):
                    amt = row['金額']
                    pct = row['対売上比']
                    if isinstance(amt, (int, float)):
                        if amt < 0:
                            styles[amt_idx] = 'color:#dc2626; font-weight:700'
                        elif amt > 0:
                            styles[amt_idx] = 'color:#15803d; font-weight:700'
                    if isinstance(pct, str) and pct != '-':
                        if pct.startswith('-'):
                            styles[pct_idx] = 'color:#dc2626; font-weight:700'
                        else:
                            styles[pct_idx] = 'color:#15803d; font-weight:700'
                return styles

            st.dataframe(
                df_m.style.apply(style_metrics, axis=1).format({'金額': '{:,}'}),
                hide_index=True, width='stretch',
            )
            st.caption(
                "※ 人件費 = 管理者給 + 指導員給 + 法定福利費 + 退職給付費用 + 賞与 + 事務員給。"
                " 経費 = 販管費合計 − 人件費。"
            )

        with col_s:
            st.markdown("##### 販管費の構成")
            sga_agg = defaultdict(int)
            sga_order = {}
            for e in entries:
                if e['category'] == 'sga':
                    sga_agg[e['account_name']] += e['amount']
                    sga_order[e['account_name']] = e['account_order']
            sga_rows = []
            for name, v in sorted(sga_agg.items(), key=lambda x: -abs(x[1])):
                if v != 0:
                    sga_rows.append({
                        '科目': name, '金額': v,
                        '販管費比': fmt_pct(v, sga_total),
                        '対売上比': fmt_pct(v, rev_total),
                    })
            if sga_rows:
                sga_rows.append({
                    '科目': '【販管費 計】', '金額': sga_total,
                    '販管費比': '100.0%',
                    '対売上比': fmt_pct(sga_total, rev_total),
                })
                df_s = pd.DataFrame(sga_rows)

                def bold_total_s(row):
                    if row['科目'] == '【販管費 計】':
                        return ['background-color:#f1f5f9; font-weight:700'] * len(row)
                    return [''] * len(row)

                st.dataframe(
                    df_s.style.apply(bold_total_s, axis=1).format({'金額': '{:,}'}),
                    hide_index=True, width='stretch',
                    height=min(35 * (len(df_s) + 1) + 3, 600),
                )
            else:
                st.info("販管費データがありません。")

        # ===== A4 PDFレポート (構成比) =====
        st.markdown("---")
        st.markdown("### 📄 A4 PDFレポート")
        st.caption(
            "1施設=1ページの構成比レポートを生成します。"
            "個別施設は単独PDF、全親グループ／全子サブ部門は施設ごとに分かれた個別PDFをZIPでダウンロードできます（施設管理者に配布しやすい形式）。"
        )

        from lib import pdf_report as _pdf

        def _build_composition_for(scope_label: str, scope_yms: list,
                                     group_id: int | None = None,
                                     subunit_id: int | None = None):
            """指定スコープの 1施設(=A4 1ページ) 構成比レポート Flowables を返す。"""
            ent = db.fetch_pl_entries(
                year_months=scope_yms,
                group_ids=[group_id] if group_id else None,
                subunit_ids=[subunit_id] if subunit_id else None,
            ) if scope_yms else []

            rev_total_v = sum(e['amount'] for e in ent if e['category'] == 'revenue_total')
            sga_total_v = sum(e['amount'] for e in ent if e['category'] == 'sga_total')

            # 売上構成
            rev_agg2 = defaultdict(int)
            rev_order2 = {}
            for e in ent:
                if e['category'] == 'revenue':
                    rev_agg2[e['account_name']] += e['amount']
                    rev_order2[e['account_name']] = e['account_order']
            rev_data = []
            for nm, v in sorted(rev_agg2.items(), key=lambda x: rev_order2[x[0]]):
                if v != 0:
                    rev_data.append((nm, v, fmt_pct(v, rev_total_v)))
            if rev_data:
                rev_data.append(('【売上高 計】', rev_total_v, '100.0%'))

            # 主要指標
            personnel_v = sum(e['amount'] for e in ent
                                if e['category'] == 'sga' and e['account_name'] in PERSONNEL_ACCOUNTS)
            other_exp_v = sga_total_v - personnel_v
            op_v = sum(e['amount'] for e in ent if e['category'] == 'op_profit')
            ord_v = sum(e['amount'] for e in ent if e['category'] == 'ordinary_profit')
            metrics_data = [
                ('人件費', personnel_v, fmt_pct(personnel_v, rev_total_v)),
                ('経費', other_exp_v, fmt_pct(other_exp_v, rev_total_v)),
                ('営業利益', op_v, fmt_pct(op_v, rev_total_v)),
                ('経常利益', ord_v, fmt_pct(ord_v, rev_total_v)),
            ]

            # 販管費構成
            sga_agg2 = defaultdict(int)
            for e in ent:
                if e['category'] == 'sga':
                    sga_agg2[e['account_name']] += e['amount']
            sga_data = []
            for nm, v in sorted(sga_agg2.items(), key=lambda x: -abs(x[1])):
                if v != 0:
                    sga_data.append((
                        nm, v,
                        fmt_pct(v, sga_total_v),
                        fmt_pct(v, rev_total_v),
                    ))
            if sga_data:
                sga_data.append((
                    '【販管費 計】', sga_total_v,
                    '100.0%', fmt_pct(sga_total_v, rev_total_v),
                ))

            return _pdf.build_composition_page(
                facility_label=scope_label,
                period_label=f"対象: {ratio_label}",
                rev_data=rev_data,
                metrics_data=metrics_data,
                sga_data=sga_data,
            )

        cc1, cc2, cc3 = st.columns(3)
        with cc1:
            if st.button("📄 この施設のPDFレポート生成", key='ratio_pdf_single_btn'):
                page = _build_composition_for(
                    scope_label=sel_display, scope_yms=ratio_yms,
                    group_id=sel_group_id, subunit_id=sel_subunit_id,
                )
                pdf_bytes = _pdf.build_pdf([page])
                st.session_state['ratio_pdf_single'] = (pdf_bytes, sel_display)
            if 'ratio_pdf_single' in st.session_state:
                pdf_bytes, lbl = st.session_state['ratio_pdf_single']
                safe = lbl.replace('／', '_').replace('（', '(').replace('）', ')').replace('/', '_')
                st.download_button(
                    "⬇ 個別レポート (PDF)",
                    data=pdf_bytes,
                    file_name=f"構成比_{safe}_{ratio_label}.pdf",
                    mime='application/pdf',
                    key='ratio_pdf_single_dl',
                )

        with cc2:
            if st.button("📁 全親グループ 個別PDF (ZIP)", key='ratio_pdf_groups_btn'):
                items = []
                for g in db.list_pl_groups():
                    label = f"{g['code']}.{g['name']}"
                    if g.get('note'):
                        label += f"（{g['note']}）"
                    items.append((label, _build_composition_for(
                        scope_label=label, scope_yms=ratio_yms,
                        group_id=g['id'],
                    )))
                zip_bytes = _build_pdfs_zip(items, period_tag=ratio_label, prefix='構成比')
                st.session_state['ratio_pdf_groups'] = (zip_bytes, len(items))
            if 'ratio_pdf_groups' in st.session_state:
                zip_b, n = st.session_state['ratio_pdf_groups']
                st.download_button(
                    f"⬇ 親グループ個別PDF×{n} (ZIP)",
                    data=zip_b,
                    file_name=f"構成比_全親グループ_{_safe_filename(ratio_label)}.zip",
                    mime='application/zip',
                    key='ratio_pdf_groups_dl',
                )

        with cc3:
            if st.button("📁 全子サブ部門 個別PDF (ZIP)", key='ratio_pdf_subs_btn'):
                items = []
                for s in db.list_pl_subunits():
                    items.append((s['display_name'], _build_composition_for(
                        scope_label=s['display_name'], scope_yms=ratio_yms,
                        subunit_id=s['id'],
                    )))
                zip_bytes = _build_pdfs_zip(items, period_tag=ratio_label, prefix='構成比')
                st.session_state['ratio_pdf_subs'] = (zip_bytes, len(items))
            if 'ratio_pdf_subs' in st.session_state:
                zip_b, n = st.session_state['ratio_pdf_subs']
                st.download_button(
                    f"⬇ 子サブ部門個別PDF×{n} (ZIP)",
                    data=zip_b,
                    file_name=f"構成比_全子サブ部門_{_safe_filename(ratio_label)}.zip",
                    mime='application/zip',
                    key='ratio_pdf_subs_dl',
                )

        st.markdown("### 📊 CSV出力")
        st.caption(
            "損益計算書CSVは、選択中の期間・施設範囲だけを切り出します。"
            "仕訳帳を取り込んでいる場合は、同じ範囲の仕訳帳CSVも出力できます。"
        )
        has_journal_csv = bool(db.list_journal_imports(limit=1))
        c_csv1, c_csv2, c_csv3 = st.columns(3)
        with c_csv1:
            st.download_button(
                "⬇ この施設の損益CSV",
                data=_pl_csv_bytes(
                    sel_display, ratio_yms,
                    group_id=sel_group_id, subunit_id=sel_subunit_id,
                ),
                file_name=f"構成比_{_safe_filename(sel_display)}_{_safe_filename(ratio_label)}_損益.csv",
                mime='text/csv',
                key='ratio_csv_single_pl_dl',
            )
            if has_journal_csv:
                st.download_button(
                    "⬇ この施設の仕訳帳CSV",
                    data=_journal_csv_bytes(
                        sel_display, ratio_yms,
                        group_id=sel_group_id, subunit_id=sel_subunit_id,
                    ),
                    file_name=f"構成比_{_safe_filename(sel_display)}_{_safe_filename(ratio_label)}_仕訳帳.csv",
                    mime='text/csv',
                    key='ratio_csv_single_journal_dl',
                )
        with c_csv2:
            if st.button("📁 全親グループ 個別CSV (ZIP)", key='ratio_csv_groups_btn'):
                items = []
                for g in db.list_pl_groups():
                    label = f"{g['code']}.{g['name']}"
                    if g.get('note'):
                        label += f"（{g['note']}）"
                    items.append({
                        'label': label,
                        'year_months': ratio_yms,
                        'group_id': g['id'],
                    })
                st.session_state['ratio_csv_groups'] = (
                    _build_scope_csv_zip(
                        items, period_tag=ratio_label, prefix='構成比',
                        include_journal=has_journal_csv,
                    ),
                    len(items),
                )
            if 'ratio_csv_groups' in st.session_state:
                zip_b, n = st.session_state['ratio_csv_groups']
                st.download_button(
                    f"⬇ 親グループ個別CSV×{n} (ZIP)",
                    data=zip_b,
                    file_name=f"構成比_全親グループ_CSV_{_safe_filename(ratio_label)}.zip",
                    mime='application/zip',
                    key='ratio_csv_groups_dl',
                )
        with c_csv3:
            if st.button("📂 全子サブ部門 個別CSV (ZIP)", key='ratio_csv_subs_btn'):
                items = [
                    {
                        'label': s['display_name'],
                        'year_months': ratio_yms,
                        'subunit_id': s['id'],
                    }
                    for s in db.list_pl_subunits()
                ]
                st.session_state['ratio_csv_subs'] = (
                    _build_scope_csv_zip(
                        items, period_tag=ratio_label, prefix='構成比',
                        include_journal=has_journal_csv,
                    ),
                    len(items),
                )
            if 'ratio_csv_subs' in st.session_state:
                zip_b, n = st.session_state['ratio_csv_subs']
                st.download_button(
                    f"⬇ 子サブ部門個別CSV×{n} (ZIP)",
                    data=zip_b,
                    file_name=f"構成比_全子サブ部門_CSV_{_safe_filename(ratio_label)}.zip",
                    mime='application/zip',
                    key='ratio_csv_subs_dl',
                )

# =============================================================
# TAB 4: 比較 (当月 / 前月 / 前年 単月比較)
# =============================================================
with tab_yoy:
    fy_list = db.list_pl_fiscal_years(start_month=fy_start_month)
    if len(fy_list) < 1:
        st.info("データがありません。")
    else:
        c1, _ = st.columns([1, 2])
        with c1:
            curr_fy = st.selectbox(
                "対象期", fy_list,
                format_func=fy_label,
                key='pl_yoy_curr_fy',
            )

        # 当期 にあるデータの月リスト
        curr_months_all = [
            m for m in db.pl_fiscal_year_months(curr_fy, start_month=fy_start_month)
            if m in existing_yms
        ]
        if not curr_months_all:
            st.warning("当期データがありません")
        else:
            # 月選択 (単月のみ)
            month_idx_options = list(range(len(curr_months_all)))
            sel_idx = st.selectbox(
                "対象月",
                month_idx_options,
                index=len(curr_months_all) - 1,
                format_func=lambda i: curr_months_all[i],
                key='pl_yoy_month_idx',
            )
            curr_end = curr_months_all[sel_idx]

            # 前月・前年同月
            prev_month = _ym_minus(curr_end, 1)
            yoy_end = _ym_minus(curr_end, 12)

            curr_yms = [curr_end]
            prev_month_yms = [prev_month] if prev_month in existing_yms else []
            prev_yms = [yoy_end] if yoy_end in existing_yms else []

            curr_entries = fetch_pl(curr_yms, sel_group_id, sel_subunit_id) if curr_yms else []
            prev_month_entries = fetch_pl(prev_month_yms, sel_group_id, sel_subunit_id) if prev_month_yms else []
            prev_entries = fetch_pl(prev_yms, sel_group_id, sel_subunit_id) if prev_yms else []

            curr_label = curr_end
            prev_month_label = prev_month if prev_month_yms else f"{prev_month}(無)"
            prev_label = yoy_end if prev_yms else f"{yoy_end}(無)"

            st.markdown(
                f"<div style='font-size:14px; color:#475569; margin-top:8px;'>"
                f"当月 <b>{curr_label}</b> ／ 前月 <b>{prev_month_label}</b> "
                f"／ 前年 <b>{prev_label}</b> ／ {sel_display}"
                f"</div>", unsafe_allow_html=True,
            )

            if not prev_yms:
                st.warning(
                    f"⚠️ 前年同月({yoy_end})のデータがありません。\n\n"
                    "📥「Excel取込」から該当月を取り込むと前年比較が見られます。"
                )

            # 主要指標 (5項目: 売上高/人件費/経費/営業利益/経常利益)
            def aggs(entries):
                sga_total = sum_by_cat(entries, 'sga_total')
                pers = sum(e['amount'] for e in entries
                           if e['category'] == 'sga' and e['account_name'] in PERSONNEL_ACCOUNTS)
                return {
                    '売上高': sum_by_cat(entries, 'revenue_total'),
                    '人件費': pers,
                    '経費': sga_total - pers,
                    '営業利益': sum_by_cat(entries, 'op_profit'),
                    '経常利益': sum_by_cat(entries, 'ordinary_profit'),
                }
            curr_agg = aggs(curr_entries)
            prev_month_agg = aggs(prev_month_entries)
            prev_agg = aggs(prev_entries)

            yoy_rows = []
            for k in ['売上高', '人件費', '経費', '営業利益', '経常利益']:
                cv = curr_agg[k]
                mv = prev_month_agg[k]
                pv = prev_agg[k]
                yoy_rows.append({
                    '項目': k,
                    '当月': cv, '前月': mv, '前年': pv,
                    '前月比': cv - mv,
                    '前月比%': fmt_diff_pct(cv, mv),
                    '前年比': cv - pv,
                    '前年比%': fmt_diff_pct(cv, pv),
                })
            df_yoy = pd.DataFrame(yoy_rows)

            def color_diff_amount(val):
                """項目によって色を変える: 売上/粗利/利益はプラス＝緑、販管費はマイナス＝緑"""
                # 文字列セルの subset で行毎に判断するため、ここでは利益系の color_profit を使う
                # apply で行毎処理に切り替え
                return ''

            def style_row(row):
                styles = [''] * len(row)
                idx = list(row.index)
                k = row['項目']
                is_cost = k in ('人件費', '経費')

                def colorize(amt_col, pct_col):
                    a = row[amt_col]
                    p = row[pct_col]
                    if is_cost:
                        # コストは増えたら赤・減ったら緑
                        if isinstance(a, (int, float)):
                            if a > 0: styles[idx.index(amt_col)] = 'color:#dc2626; font-weight:700'
                            elif a < 0: styles[idx.index(amt_col)] = 'color:#15803d; font-weight:700'
                        if isinstance(p, str):
                            if p.startswith('-') and p != '-': styles[idx.index(pct_col)] = 'color:#15803d; font-weight:700'
                            elif p.startswith('+'): styles[idx.index(pct_col)] = 'color:#dc2626; font-weight:700'
                    else:
                        # 売上・利益は増えたら緑・減ったら赤
                        if isinstance(a, (int, float)):
                            if a > 0: styles[idx.index(amt_col)] = 'color:#15803d; font-weight:700'
                            elif a < 0: styles[idx.index(amt_col)] = 'color:#dc2626; font-weight:700'
                        if isinstance(p, str):
                            if p.startswith('+'): styles[idx.index(pct_col)] = 'color:#15803d; font-weight:700'
                            elif p.startswith('-') and p != '-': styles[idx.index(pct_col)] = 'color:#dc2626; font-weight:700'

                colorize('前月比', '前月比%')
                colorize('前年比', '前年比%')
                return styles

            styled = (df_yoy.style
                .apply(style_row, axis=1)
                .format({
                    '当月': '{:,}', '前月': '{:,}', '前年': '{:,}',
                    '前月比': '{:,}', '前年比': '{:,}',
                })
            )
            st.dataframe(styled, hide_index=True, width='stretch')

            # 主要販管費科目の比較
            if curr_entries and prev_entries:
                st.markdown("##### 主要 販管費科目 比較")
                curr_sga = defaultdict(int)
                prev_month_sga = defaultdict(int)
                prev_sga = defaultdict(int)
                acc_order_yoy = {}
                for e in curr_entries:
                    if e['category'] == 'sga':
                        curr_sga[e['account_name']] += e['amount']
                        acc_order_yoy[e['account_name']] = e['account_order']
                for e in prev_month_entries:
                    if e['category'] == 'sga':
                        prev_month_sga[e['account_name']] += e['amount']
                        acc_order_yoy[e['account_name']] = e['account_order']
                for e in prev_entries:
                    if e['category'] == 'sga':
                        prev_sga[e['account_name']] += e['amount']
                        acc_order_yoy[e['account_name']] = e['account_order']

                all_names = set(curr_sga.keys()) | set(prev_month_sga.keys()) | set(prev_sga.keys())
                yoy_sga_rows = []
                for name in all_names:
                    c = curr_sga.get(name, 0)
                    m = prev_month_sga.get(name, 0)
                    p = prev_sga.get(name, 0)
                    if c == 0 and m == 0 and p == 0:
                        continue
                    yoy_sga_rows.append({
                        '科目': name,
                        '当月': c, '前月': m, '前年': p,
                        '前月比': c - m,
                        '前月比%': fmt_diff_pct(c, m),
                        '前年比': c - p,
                        '前年比%': fmt_diff_pct(c, p),
                        '_order': acc_order_yoy.get(name, 999),
                    })
                yoy_sga_rows.sort(key=lambda r: -abs(r['当月']))
                df_yoy_sga = pd.DataFrame(yoy_sga_rows).drop(columns=['_order'])

                def style_sga_row(row):
                    styles = [''] * len(row)
                    idx = list(row.index)
                    # 販管費科目は コスト視点 (増=赤, 減=緑)
                    for amt_col, pct_col in [('前月比', '前月比%'), ('前年比', '前年比%')]:
                        a = row[amt_col]
                        p = row[pct_col]
                        if isinstance(a, (int, float)):
                            if a > 0: styles[idx.index(amt_col)] = 'color:#dc2626; font-weight:700'
                            elif a < 0: styles[idx.index(amt_col)] = 'color:#15803d; font-weight:700'
                        if isinstance(p, str):
                            if p.startswith('-') and p != '-': styles[idx.index(pct_col)] = 'color:#15803d; font-weight:700'
                            elif p.startswith('+'): styles[idx.index(pct_col)] = 'color:#dc2626; font-weight:700'
                    return styles

                st.dataframe(
                    df_yoy_sga.style
                        .apply(style_sga_row, axis=1)
                        .format({
                            '当月': '{:,}', '前月': '{:,}', '前年': '{:,}',
                            '前月比': '{:,}', '前年比': '{:,}',
                        }),
                    hide_index=True, width='stretch',
                    height=min(35 * (len(df_yoy_sga) + 1) + 3, 500),
                )

            # ===== 分析レポート(画面では非表示・印刷物には記載) =====
            yoy_top10 = []
            if curr_entries and prev_entries:
                # SGA 増減の絶対値で大きい順 TOP10
                sga_diffs = []
                for name in (set(curr_sga.keys()) | set(prev_month_sga.keys()) | set(prev_sga.keys())):
                    c = curr_sga.get(name, 0)
                    m = prev_month_sga.get(name, 0)
                    p = prev_sga.get(name, 0)
                    d = c - p
                    if d == 0:
                        continue
                    pct = (d / abs(p) * 100) if p else None
                    sga_diffs.append((name, c, m, p, d, pct))
                sga_diffs.sort(key=lambda x: -abs(x[4]))
                yoy_top10 = sga_diffs[:10]

            # 仕訳明細抽出のための subunit_ids を確定
            if sel_subunit_id:
                _journal_sub_ids = [sel_subunit_id]
            elif sel_group_id:
                _journal_sub_ids = [s['id'] for s in db.list_pl_subunits(group_id=sel_group_id)]
            else:
                _journal_sub_ids = None  # 全グループ → 制約なし

            with st.expander("📊 分析レポート（販管費 動きの大きい10科目 + 取引先明細）", expanded=False):
                if yoy_top10:
                    st.caption("印刷レポートにはこの分析セクションが含まれます。仕訳帳取込済の場合、各科目の取引先内訳も表示されます。")
                    rep_rows = []
                    for name, c, m, p, d, pct in yoy_top10:
                        direction = '増加' if d > 0 else '減少'
                        pct_str = f"{pct:+.1f}%" if pct is not None else "-"
                        m_pct_str = fmt_diff_pct(c, m)
                        rep_rows.append({
                            '科目': name,
                            '当月': c, '前月': m, '前年': p,
                            '前月比': c - m, '前月比%': m_pct_str,
                            '前年比': d, '前年比%': pct_str,
                            '方向': direction,
                        })
                    df_rep = pd.DataFrame(rep_rows)
                    def _style_rep(row):
                        s = [''] * len(row)
                        idx = list(row.index)
                        for col in ('前月比', '前年比'):
                            v = row[col]
                            if isinstance(v, (int, float)):
                                if v > 0:
                                    s[idx.index(col)] = 'color:#dc2626; font-weight:700'
                                elif v < 0:
                                    s[idx.index(col)] = 'color:#15803d; font-weight:700'
                        return s
                    st.dataframe(
                        df_rep.style.apply(_style_rep, axis=1).format(
                            {'当月': '{:,}', '前月': '{:,}', '前年': '{:,}',
                             '前月比': '{:,}', '前年比': '{:,}'}),
                        hide_index=True, width='stretch',
                    )

                    # ===== 各科目の何に差があるか (仕訳帳の品目/取引先で分析) =====
                    st.markdown("---")
                    st.markdown("##### 何に差があったか (仕訳帳より) — 取引先・品目別 当月/前年 比較")
                    if not db.list_journal_imports(limit=1):
                        st.info(
                            "仕訳帳が未取込です。"
                            "上の「📒 仕訳帳CSV取込」から取込むと、ここに具体的な内訳が表示されます。"
                        )
                    else:
                        st.caption(
                            "取引先・品目・備考のいずれかが入力されている行のみ抽出。"
                            "完全に不明な行(社内振替など)は除外。差額の絶対値が大きい順に上位5件。"
                        )
                        for name, c, m, p, d, pct in yoy_top10:
                            items = db.journal_item_diffs_for_account(
                                name, curr_yms=curr_yms, prev_yms=prev_yms,
                                subunit_ids=_journal_sub_ids, top_n=5,
                            )
                            if not items:
                                continue
                            arrow = '🔺' if d > 0 else '🔻'
                            color = '#dc2626' if d > 0 else '#15803d'
                            pct_str = f" ({pct:+.1f}%)" if pct is not None else ""
                            st.markdown(
                                f"**{arrow} {name}**　"
                                f"<span style='color:{color}; font-weight:700'>"
                                f"{int(d/1000):+,}千円</span>{pct_str}",
                                unsafe_allow_html=True,
                            )
                            df_i = pd.DataFrame([
                                {
                                    '取引先／品目': x['label'],
                                    '当月': x['curr_amt'],
                                    '前年': x['prev_amt'],
                                    '差額': x['diff'],
                                    '当月件': x['curr_cnt'],
                                    '前年件': x['prev_cnt'],
                                } for x in items
                            ])
                            def _style_diff(row):
                                s = [''] * len(row)
                                idx = list(row.index)
                                amt_i = idx.index('差額')
                                if isinstance(row['差額'], (int, float)):
                                    if row['差額'] > 0:
                                        s[amt_i] = 'color:#dc2626; font-weight:700'
                                    elif row['差額'] < 0:
                                        s[amt_i] = 'color:#15803d; font-weight:700'
                                return s
                            st.dataframe(
                                df_i.style.apply(_style_diff, axis=1).format({
                                    '当月': '{:,}', '前年': '{:,}', '差額': '{:,}',
                                }),
                                hide_index=True, width='stretch',
                            )
                else:
                    st.caption("前年データがないため分析できません。")

            # ===== A4 PDFレポート ダウンロード =====
            st.markdown("---")
            st.markdown("### 📄 A4 PDFレポート")
            st.caption(
                "1施設=1ページのPDFレポートを生成します。"
                "現在の選択スコープ(個別施設) または 全施設まとめ(各ページ1施設) を選択可能。"
            )

            from lib import pdf_report as _pdf

            def _build_facility_page_for(scope_label: str, scope_curr_yms: list,
                                          scope_prev_month_yms: list, scope_prev_yms: list,
                                          group_id: int | None = None, subunit_id: int | None = None,
                                          subunit_ids: list | None = None):
                """指定スコープの 1施設(=A4 1ページ) Flowables を返す。"""
                fetch_kwargs = {}
                if subunit_ids:
                    fetch_kwargs['subunit_ids'] = list(subunit_ids)
                elif subunit_id:
                    fetch_kwargs['subunit_ids'] = [subunit_id]
                elif group_id:
                    fetch_kwargs['group_ids'] = [group_id]

                ent_c = db.fetch_pl_entries(year_months=scope_curr_yms, **fetch_kwargs) if scope_curr_yms else []
                ent_m = db.fetch_pl_entries(year_months=scope_prev_month_yms, **fetch_kwargs) if scope_prev_month_yms else []
                ent_p = db.fetch_pl_entries(year_months=scope_prev_yms, **fetch_kwargs) if scope_prev_yms else []

                # account_id 単位で集計
                by_acc_c = defaultdict(int)
                by_acc_m = defaultdict(int)
                by_acc_p = defaultdict(int)
                for e in ent_c:
                    by_acc_c[e['account_id']] += e['amount']
                for e in ent_m:
                    by_acc_m[e['account_id']] += e['amount']
                for e in ent_p:
                    by_acc_p[e['account_id']] += e['amount']

                all_accs = db.list_pl_accounts()

                # 売上総額 (利益率の分母)
                rev_total_id = next(
                    (a['id'] for a in all_accs if a['category'] == 'revenue_total'), None
                )
                rev_curr = by_acc_c.get(rev_total_id, 0) if rev_total_id else 0
                rev_prev = by_acc_p.get(rev_total_id, 0) if rev_total_id else 0

                # 損益計算書 行データ (当月/前月/前年)
                pl_rows = _pdf.build_pl_rows(
                    by_acc_c, by_acc_m, by_acc_p, all_accs,
                    PERSONNEL_ACCOUNTS,
                )

                # 利益率
                def _agg_cat(data, cat):
                    return sum(data.get(a['id'], 0) for a in all_accs if a['category'] == cat)

                def _agg_personnel(data):
                    return sum(data.get(a['id'], 0) for a in all_accs
                               if a['category'] == 'sga' and a['name'] in PERSONNEL_ACCOUNTS)

                gross_c = _agg_cat(by_acc_c, 'gross_profit')
                gross_p = _agg_cat(by_acc_p, 'gross_profit')
                sga_total_c = _agg_cat(by_acc_c, 'sga_total')
                sga_total_p = _agg_cat(by_acc_p, 'sga_total')
                pers_c = _agg_personnel(by_acc_c)
                pers_p = _agg_personnel(by_acc_p)
                other_c = sga_total_c - pers_c
                other_p = sga_total_p - pers_p
                op_c = _agg_cat(by_acc_c, 'op_profit')
                op_p = _agg_cat(by_acc_p, 'op_profit')
                ord_c = _agg_cat(by_acc_c, 'ordinary_profit')
                ord_p = _agg_cat(by_acc_p, 'ordinary_profit')
                net_c = _agg_cat(by_acc_c, 'net_income')
                net_p = _agg_cat(by_acc_p, 'net_income')

                def _ratio(num_c, num_p):
                    cv = (num_c / rev_curr * 100) if rev_curr else None
                    pv = (num_p / rev_prev * 100) if rev_prev else None
                    return (cv, pv)

                ratios = {
                    '売上総利益率': _ratio(gross_c, gross_p),
                    '販管費率': _ratio(sga_total_c, sga_total_p),
                    '人件費率': _ratio(pers_c, pers_p),
                    '経費率': _ratio(other_c, other_p),
                    '営業利益率': _ratio(op_c, op_p),
                    '経常利益率': _ratio(ord_c, ord_p),
                    '当期純利益率': _ratio(net_c, net_p),
                }

                # SGA TOP10 (vs 前年 増減)
                sga_c = defaultdict(int)
                sga_m = defaultdict(int)
                sga_p = defaultdict(int)
                for e in ent_c:
                    if e['category'] == 'sga':
                        sga_c[e['account_name']] += e['amount']
                for e in ent_m:
                    if e['category'] == 'sga':
                        sga_m[e['account_name']] += e['amount']
                for e in ent_p:
                    if e['category'] == 'sga':
                        sga_p[e['account_name']] += e['amount']
                top = []
                for n in (set(sga_c.keys()) | set(sga_m.keys()) | set(sga_p.keys())):
                    c = sga_c.get(n, 0)
                    m = sga_m.get(n, 0)
                    p = sga_p.get(n, 0)
                    d = c - p
                    if d == 0:
                        continue
                    pct = (d / abs(p) * 100) if p else None
                    top.append((n, c, m, p, d, pct))
                top.sort(key=lambda x: -abs(x[4]))
                top10 = top[:10]

                # 期間ラベル
                period_lbl = (
                    f"当月: {scope_curr_yms[0] if scope_curr_yms else '-'} "
                    f"／ 前月: {scope_prev_month_yms[0] if scope_prev_month_yms else '-'} "
                    f"／ 前年: {scope_prev_yms[0] if scope_prev_yms else '-'}"
                )

                # 仕訳帳明細 (上位科目から差額の大きい取引先・品目を集めて、全体から TOP10)
                journal_top10 = []
                if db.list_journal_imports(limit=1):
                    j_sub_ids = (
                        list(subunit_ids) if subunit_ids else
                        ([subunit_id] if subunit_id else
                         ([s['id'] for s in db.list_pl_subunits(group_id=group_id)] if group_id else None))
                    )
                    merged = []
                    for name, *_ in top10:
                        items = db.journal_item_diffs_for_account(
                            name, curr_yms=scope_curr_yms, prev_yms=scope_prev_yms,
                            subunit_ids=j_sub_ids, top_n=8,
                        )
                        for it in items:
                            it2 = dict(it)
                            it2['account'] = name
                            merged.append(it2)
                    merged.sort(key=lambda x: -abs(x['diff']))
                    journal_top10 = merged[:15]

                return _pdf.build_facility_page(
                    facility_label=scope_label,
                    period_label=period_lbl,
                    pl_rows=pl_rows,
                    ratios=ratios,
                    sga_top=top10,
                    journal_top=journal_top10,
                )

            # 個別施設 PDF (現在の選択)
            cc_pdf1, cc_pdf2, cc_pdf3 = st.columns(3)
            with cc_pdf1:
                if st.button("📄 この施設のPDFレポート生成", key='pl_pdf_single_btn'):
                    page_elems = _build_facility_page_for(
                        scope_label=sel_display,
                        scope_curr_yms=curr_yms,
                        scope_prev_month_yms=prev_month_yms, scope_prev_yms=prev_yms,
                        group_id=sel_group_id, subunit_id=sel_subunit_id,
                    )
                    pdf_bytes = _pdf.build_pdf([page_elems])
                    st.session_state['pl_pdf_single'] = (pdf_bytes, sel_display)
                if 'pl_pdf_single' in st.session_state:
                    pdf_bytes, lbl = st.session_state['pl_pdf_single']
                    safe = lbl.replace('／', '_').replace('（', '(').replace('）', ')').replace('/', '_')
                    st.download_button(
                        "⬇ 個別レポート (PDF)",
                        data=pdf_bytes,
                        file_name=f"比較_{safe}_{curr_end}.pdf",
                        mime='application/pdf',
                        key='pl_pdf_single_dl',
                    )

            # 全親グループ 個別PDF (ZIP)
            with cc_pdf2:
                if st.button("📁 全親グループ 個別PDF (ZIP)", key='pl_pdf_all_btn'):
                    items = []
                    for g in db.list_pl_groups():
                        label = f"{g['code']}.{g['name']}"
                        if g.get('note'):
                            label += f"（{g['note']}）"
                        items.append((label, _build_facility_page_for(
                            scope_label=label,
                            scope_curr_yms=curr_yms,
                            scope_prev_month_yms=prev_month_yms, scope_prev_yms=prev_yms,
                            group_id=g['id'],
                        )))
                    zip_bytes = _build_pdfs_zip(items, period_tag=curr_end, prefix='比較')
                    st.session_state['pl_pdf_all'] = (zip_bytes, len(items))
                if 'pl_pdf_all' in st.session_state:
                    zip_b, n = st.session_state['pl_pdf_all']
                    st.download_button(
                        f"⬇ 親グループ個別PDF×{n} (ZIP)",
                        data=zip_b,
                        file_name=f"比較_全親グループ_{curr_end}.zip",
                        mime='application/zip',
                        key='pl_pdf_all_dl',
                    )

            # 全子サブ部門 個別PDF (ZIP, 業績会議の行構成準拠)
            with cc_pdf3:
                if st.button("📂 全子サブ部門 個別PDF (ZIP)", key='pl_pdf_sub_btn'):
                    # 業績会議用の行構成 (REPORT_STRUCTURE / EXCLUDED_ROWS / NPO_REPORT_STRUCTURE)
                    # を子レベル単位で展開。旧名+新名統合(例:シェア天理1.2+シェアホーム天理1.2)はそのまま合算。
                    items = []
                    for region, subrows in REPORT_STRUCTURE:
                        for label, excel_names in subrows:
                            sub_ids = [subs_by_excel[n] for n in excel_names if n in subs_by_excel]
                            if not sub_ids: continue
                            full_label = f"[EMIFULL_{region}] {label}"
                            items.append((full_label, _build_facility_page_for(
                                scope_label=full_label,
                                scope_curr_yms=curr_yms,
                                scope_prev_month_yms=prev_month_yms, scope_prev_yms=prev_yms,
                                subunit_ids=sub_ids,
                            )))
                    for label, excel_names in EXCLUDED_ROWS:
                        sub_ids = [subs_by_excel[n] for n in excel_names if n in subs_by_excel]
                        if not sub_ids: continue
                        full_label = f"[EMIFULL_別枠] {label}"
                        items.append((full_label, _build_facility_page_for(
                            scope_label=full_label,
                            scope_curr_yms=curr_yms,
                            scope_prev_month_yms=prev_month_yms, scope_prev_yms=prev_yms,
                            subunit_ids=sub_ids,
                        )))
                    for region, subrows in NPO_REPORT_STRUCTURE:
                        for label, excel_names in subrows:
                            sub_ids = [subs_by_excel[n] for n in excel_names if n in subs_by_excel]
                            if not sub_ids: continue
                            full_label = f"[NPO_{region}] {label}"
                            items.append((full_label, _build_facility_page_for(
                                scope_label=full_label,
                                scope_curr_yms=curr_yms,
                                scope_prev_month_yms=prev_month_yms, scope_prev_yms=prev_yms,
                                subunit_ids=sub_ids,
                            )))
                    if items:
                        zip_bytes = _build_pdfs_zip(items, period_tag=curr_end, prefix='比較')
                        st.session_state['pl_pdf_sub'] = (zip_bytes, len(items))
                if 'pl_pdf_sub' in st.session_state:
                    zip_b, n = st.session_state['pl_pdf_sub']
                    st.download_button(
                        f"⬇ 子サブ部門個別PDF×{n} (ZIP)",
                        data=zip_b,
                        file_name=f"比較_全子サブ部門_{curr_end}.zip",
                        mime='application/zip',
                        key='pl_pdf_sub_dl',
                    )

            st.caption(
                "※ ボタンを押すと施設ごとに分かれた個別PDFを ZIP にまとめて生成します。"
                " 全親グループ=12施設 / 全子サブ部門=23施設前後。"
                " ZIPを展開すると施設ごとの 1ページPDFが得られるので、施設管理者にそのまま配布できます。"
            )

            st.markdown("### 📊 CSV出力")
            st.caption(
                "比較CSVは、当月・前月・前年の損益データをまとめて出力します。"
                "仕訳帳を取り込んでいる場合は、同じ月・同じ施設範囲の仕訳帳CSVも出力できます。"
            )
            compare_csv_yms = _unique_yms(curr_yms, prev_month_yms, prev_yms)
            has_journal_csv = bool(db.list_journal_imports(limit=1))
            c_csv1, c_csv2, c_csv3 = st.columns(3)
            with c_csv1:
                st.download_button(
                    "⬇ この施設の損益CSV",
                    data=_pl_csv_bytes(
                        sel_display, compare_csv_yms,
                        group_id=sel_group_id, subunit_id=sel_subunit_id,
                    ),
                    file_name=f"比較_{_safe_filename(sel_display)}_{curr_end}_損益.csv",
                    mime='text/csv',
                    key='pl_csv_single_pl_dl',
                )
                if has_journal_csv:
                    st.download_button(
                        "⬇ この施設の仕訳帳CSV",
                        data=_journal_csv_bytes(
                            sel_display, compare_csv_yms,
                            group_id=sel_group_id, subunit_id=sel_subunit_id,
                        ),
                        file_name=f"比較_{_safe_filename(sel_display)}_{curr_end}_仕訳帳.csv",
                        mime='text/csv',
                        key='pl_csv_single_journal_dl',
                    )
            with c_csv2:
                if st.button("📁 全親グループ 個別CSV (ZIP)", key='pl_csv_all_btn'):
                    items = []
                    for g in db.list_pl_groups():
                        label = f"{g['code']}.{g['name']}"
                        if g.get('note'):
                            label += f"（{g['note']}）"
                        items.append({
                            'label': label,
                            'year_months': compare_csv_yms,
                            'group_id': g['id'],
                        })
                    st.session_state['pl_csv_all'] = (
                        _build_scope_csv_zip(
                            items, period_tag=curr_end, prefix='比較',
                            include_journal=has_journal_csv,
                        ),
                        len(items),
                    )
                if 'pl_csv_all' in st.session_state:
                    zip_b, n = st.session_state['pl_csv_all']
                    st.download_button(
                        f"⬇ 親グループ個別CSV×{n} (ZIP)",
                        data=zip_b,
                        file_name=f"比較_全親グループ_CSV_{curr_end}.zip",
                        mime='application/zip',
                        key='pl_csv_all_dl',
                    )
            with c_csv3:
                if st.button("📂 全子サブ部門 個別CSV (ZIP)", key='pl_csv_sub_btn'):
                    items = []
                    for region, subrows in REPORT_STRUCTURE:
                        for label, excel_names in subrows:
                            sub_ids = [subs_by_excel[n] for n in excel_names if n in subs_by_excel]
                            if not sub_ids:
                                continue
                            items.append({
                                'label': f"[EMIFULL_{region}] {label}",
                                'year_months': compare_csv_yms,
                                'subunit_ids': sub_ids,
                            })
                    for label, excel_names in EXCLUDED_ROWS:
                        sub_ids = [subs_by_excel[n] for n in excel_names if n in subs_by_excel]
                        if not sub_ids:
                            continue
                        items.append({
                            'label': f"[EMIFULL_別枠] {label}",
                            'year_months': compare_csv_yms,
                            'subunit_ids': sub_ids,
                        })
                    for region, subrows in NPO_REPORT_STRUCTURE:
                        for label, excel_names in subrows:
                            sub_ids = [subs_by_excel[n] for n in excel_names if n in subs_by_excel]
                            if not sub_ids:
                                continue
                            items.append({
                                'label': f"[NPO_{region}] {label}",
                                'year_months': compare_csv_yms,
                                'subunit_ids': sub_ids,
                            })
                    if items:
                        st.session_state['pl_csv_sub'] = (
                            _build_scope_csv_zip(
                                items, period_tag=curr_end, prefix='比較',
                                include_journal=has_journal_csv,
                            ),
                            len(items),
                        )
                if 'pl_csv_sub' in st.session_state:
                    zip_b, n = st.session_state['pl_csv_sub']
                    st.download_button(
                        f"⬇ 子サブ部門個別CSV×{n} (ZIP)",
                        data=zip_b,
                        file_name=f"比較_全子サブ部門_CSV_{curr_end}.zip",
                        mime='application/zip',
                        key='pl_csv_sub_dl',
                    )

            # ===== HTML版 (任意) =====
            def _fmt_int(v):
                try:
                    return f"{int(v):,}"
                except (TypeError, ValueError):
                    return ''

            def _fmt_signed(v):
                try:
                    n = int(v)
                except (TypeError, ValueError):
                    return ''
                return f"{n:+,}"

            def _build_yoy_html_report():
                """A4 1ページの HTMLレポートを生成 (当月/前月/前年/増減/増減率)。"""
                title = "施設別 損益レポート"
                subtitle = (
                    f"当月: {curr_label} ／ 前月: {prev_month_label} ／ 前年: {prev_label}"
                    f" ／ 対象: {sel_display}"
                )

                # 主要指標表 HTML (5項目: 売上高/人件費/経費/営業利益/経常利益)
                main_rows_html = []
                for k in ['売上高', '人件費', '経費', '営業利益', '経常利益']:
                    cv = curr_agg[k]
                    mv = prev_month_agg[k]
                    pv = prev_agg[k]
                    diff_m = cv - mv
                    diff_p = cv - pv
                    pct_m = fmt_diff_pct(cv, mv)
                    pct_p = fmt_diff_pct(cv, pv)
                    is_cost = k in ('人件費', '経費')

                    def _color(v):
                        if v == 0:
                            return '#0f172a'
                        if is_cost:
                            return '#dc2626' if v > 0 else '#15803d'
                        return '#15803d' if v > 0 else '#dc2626'
                    cm = _color(diff_m); cp = _color(diff_p)
                    main_rows_html.append(f"""
                    <tr>
                      <td>{k}</td>
                      <td class="num">{_fmt_int(cv)}</td>
                      <td class="num">{_fmt_int(mv)}</td>
                      <td class="num">{_fmt_int(pv)}</td>
                      <td class="num" style="color:{cm}; font-weight:700;">{_fmt_signed(diff_m)}</td>
                      <td class="num" style="color:{cm}; font-weight:700;">{pct_m}</td>
                      <td class="num" style="color:{cp}; font-weight:700;">{_fmt_signed(diff_p)}</td>
                      <td class="num" style="color:{cp}; font-weight:700;">{pct_p}</td>
                    </tr>
                    """)

                # 販管費科目 上位10 HTML (当月基準で前月比/前年比)
                top_rows_html = []
                for i, (name, c, m, p, d, pct) in enumerate(yoy_top10, 1):
                    diff_m = c - m
                    pct_m_str = fmt_diff_pct(c, m)
                    pct_p_str = f"{pct:+.1f}%" if pct is not None else "-"
                    cm = '#dc2626' if diff_m > 0 else ('#15803d' if diff_m < 0 else '#0f172a')
                    cp = '#dc2626' if d > 0 else ('#15803d' if d < 0 else '#0f172a')
                    top_rows_html.append(f"""
                    <tr>
                      <td class="rank">{i}</td>
                      <td>{name}</td>
                      <td class="num">{_fmt_int(c)}</td>
                      <td class="num">{_fmt_int(m)}</td>
                      <td class="num">{_fmt_int(p)}</td>
                      <td class="num" style="color:{cm}; font-weight:700;">{_fmt_signed(diff_m)}</td>
                      <td class="num" style="color:{cm}; font-weight:700;">{pct_m_str}</td>
                      <td class="num" style="color:{cp}; font-weight:700;">{_fmt_signed(d)}</td>
                      <td class="num" style="color:{cp}; font-weight:700;">{pct_p_str}</td>
                    </tr>
                    """)

                # 分析コメント (上位5件のみ — A4 1枚に収めるため)
                analysis_lines = []
                for name, c, m, p, d, pct in yoy_top10[:5]:
                    direction = '増' if d > 0 else '減'
                    diff_m = c - m
                    pct_p_str = f"{pct:+.1f}%" if pct is not None else ""
                    pct_m_str = fmt_diff_pct(c, m)
                    analysis_lines.append(
                        f"<li><b>{name}</b>: "
                        f"前月比 <span style='color:{'#dc2626' if diff_m > 0 else '#15803d'};font-weight:700;'>"
                        f"{_fmt_signed(diff_m)}円 ({pct_m_str})</span> ／ "
                        f"前年比 <span style='color:{'#dc2626' if d > 0 else '#15803d'};font-weight:700;'>"
                        f"{_fmt_signed(d)}円 ({pct_p_str})</span> → {direction}加要因の確認推奨</li>"
                    )

                from datetime import datetime
                generated = datetime.now().strftime('%Y-%m-%d %H:%M')

                html = f"""<!DOCTYPE html>
<html lang="ja">
<head>
<meta charset="UTF-8">
<title>{title} - {sel_display} - {curr_label}</title>
<style>
  @page {{ size: A4; margin: 10mm 8mm; }}
  body {{ font-family: 'Yu Gothic', 'Meiryo', sans-serif; font-size: 9pt; color: #0f172a; margin: 0; }}
  .container {{ max-width: 194mm; margin: 0 auto; }}
  h1 {{ font-size: 14pt; color: #1e3a8a; margin: 0 0 2px 0; padding-bottom: 2px; border-bottom: 1.5px solid #1e3a8a; }}
  .subtitle {{ font-size: 9pt; color: #475569; margin-bottom: 6px; }}
  h2 {{ font-size: 10.5pt; color: #1e40af; margin: 6px 0 3px; padding-left: 6px; border-left: 3px solid #1e40af; }}
  table {{ width: 100%; border-collapse: collapse; font-size: 8.5pt; margin-bottom: 4px; }}
  th, td {{ border: 1px solid #cbd5e1; padding: 2px 5px; }}
  th {{ background-color: #1e40af; color: #fff; font-weight: 600; }}
  td.num {{ text-align: right; font-variant-numeric: tabular-nums; }}
  td.rank {{ text-align: center; width: 24px; background:#f1f5f9; font-weight: 700; }}
  .analysis {{ font-size: 8.5pt; }}
  .analysis ul {{ margin: 2px 0 0 14px; padding: 0; }}
  .analysis li {{ margin-bottom: 2px; }}
  .footer {{ font-size: 7pt; color: #94a3b8; text-align: right; margin-top: 4px; }}
</style>
</head>
<body>
  <div class="container">
    <h1>{title}</h1>
    <div class="subtitle">{subtitle}</div>

    <h2>主要指標 比較</h2>
    <table>
      <thead>
        <tr>
          <th rowspan="2">項目</th>
          <th rowspan="2">当月</th>
          <th rowspan="2">前月</th>
          <th rowspan="2">前年</th>
          <th colspan="2" style="background:#1e3a8a;">前月比</th>
          <th colspan="2" style="background:#1e3a8a;">前年比</th>
        </tr>
        <tr>
          <th>増減</th><th>%</th>
          <th>増減</th><th>%</th>
        </tr>
      </thead>
      <tbody>
        {''.join(main_rows_html)}
      </tbody>
    </table>

    <h2>販管費 動きの大きい科目 (上位10)</h2>
    <table>
      <thead>
        <tr>
          <th rowspan="2" style="width:24px;">#</th>
          <th rowspan="2">科目</th>
          <th rowspan="2">当月</th>
          <th rowspan="2">前月</th>
          <th rowspan="2">前年</th>
          <th colspan="2" style="background:#1e3a8a;">前月比</th>
          <th colspan="2" style="background:#1e3a8a;">前年比</th>
        </tr>
        <tr>
          <th>増減</th><th>%</th>
          <th>増減</th><th>%</th>
        </tr>
      </thead>
      <tbody>
        {''.join(top_rows_html) if top_rows_html else '<tr><td colspan="9" style="text-align:center;">前年データなし</td></tr>'}
      </tbody>
    </table>

    <h2>分析メモ</h2>
    <div class="analysis">
      <ul>
        {''.join(analysis_lines) if analysis_lines else '<li>分析対象の差額がありません。</li>'}
      </ul>
    </div>

    <div class="footer">障がい事業部ダッシュボード — 出力 {generated}</div>
  </div>
</body>
</html>
"""
                return html

            with st.expander("HTML版もダウンロード(オプション)", expanded=False):
                html_report = _build_yoy_html_report()
                safe_name = sel_display.replace('／', '_').replace('（', '_').replace('）', '').replace('/', '_').replace('.', '_')
                st.download_button(
                    "📄 HTML版をダウンロード",
                    data=html_report.encode('utf-8'),
                    file_name=f"比較レポート_{safe_name}_{curr_end}.html",
                    mime='text/html',
                    key='pl_yoy_report_html_dl',
                )

# =============================================================
# TAB: 報告書提出
# =============================================================
if _is_admin and tab_report is not None:
  with tab_report:
    st.markdown("##### 報告書提出 — 収支の振り返り・対策")
    current = auth.current_user() or {}
    current_email = current.get('email') or ''
    current_user_for_report = db.get_user_by_email(current_email) if current_email else None
    default_reporter_name = (
        (current_user_for_report or {}).get('name')
        or current.get('name')
        or ''
    )
    default_reporter_role = (
        (current_user_for_report or {}).get('position')
        or current.get('position')
        or '一般職'
    )
    if default_reporter_role not in REPORTER_ROLES:
        default_reporter_role = '一般職'
    if st.session_state.get('profit_report_prefill_user') != current_email:
        st.session_state.profit_report_name = default_reporter_name
        st.session_state.profit_report_role = default_reporter_role
        st.session_state.profit_report_prefill_user = current_email

    default_month = _default_report_month()
    month_idx = existing_yms.index(default_month) if default_month in existing_yms else 0
    st.markdown(
        """
        <style>
        div[data-testid="stForm"] textarea {
            background: #ffffff !important;
            border: 1.5px solid #94a3b8 !important;
            border-radius: 8px !important;
            box-shadow: inset 0 0 0 1px rgba(148, 163, 184, 0.18) !important;
        }
        div[data-testid="stForm"] textarea:focus {
            border-color: #60a5fa !important;
            box-shadow: 0 0 0 3px rgba(96, 165, 250, 0.22) !important;
        }
        div[data-testid="stForm"] textarea::placeholder {
            color: #94a3b8 !important;
            opacity: 1 !important;
        }
        .emifull-sad {
            margin: 12px 0 18px;
            padding: 22px 24px;
            border: 2px solid #fecaca;
            border-radius: 10px;
            background: #fef2f2;
            color: #b91c1c;
            font-size: 34px;
            font-weight: 800;
            text-align: center;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )

    if st.button("文章欄だけクリア", key='profit_report_clear_texts'):
        for key in (
            'profit_report_previous', 'profit_report_issue',
            'profit_report_actions', 'profit_report_other',
        ):
            st.session_state[key] = ''
        st.rerun()

    with st.form("profit_report_form"):
        c1, c2 = st.columns(2)
        with c1:
            report_month = st.selectbox(
                "対象月", existing_yms, index=month_idx, key='profit_report_month'
            )
        with c2:
            facility_labels = [f['name'] for f in REPORT_FACILITIES]
            facility_name = st.selectbox(
                "施設選択", facility_labels, key='profit_report_facility'
            )
        selected_fac = REPORT_FACILITY_BY_NAME[facility_name]

        c3, c4 = st.columns([1, 2])
        with c3:
            reporter_role = st.selectbox(
                "役職",
                REPORTER_ROLES,
                index=REPORTER_ROLES.index(default_reporter_role),
                key='profit_report_role',
            )
        with c4:
            reporter_name = st.text_input(
                "氏名",
                key='profit_report_name',
            )

        previous_review = st.text_area(
            "① 前月の振り返り",
            height=140,
            key='profit_report_previous',
            placeholder="例：前月に決めた予約確認・欠席フォローを実施し、利用率は改善しました。一方で、送迎調整や記録業務に時間がかかり、現場負担が残りました。",
        )
        issue_review = st.text_area(
            "② 現在の課題",
            height=150,
            key='profit_report_issue',
            placeholder="例：欠席の偏り、稼働率のばらつき、人件費率の上昇、送迎・記録業務の負担など、今月時点で利益や運営に影響している課題を記入してください。",
        )
        next_actions = st.text_area(
            "③ 次月以降の対策",
            height=150,
            key='profit_report_actions',
            placeholder="例：予約枠の早期確認、欠席フォロー、シフト調整、経費の見直しを進めます。担当者・期限・確認方法が分かる形で書くと会議で共有しやすくなります。",
        )
        other_notes = st.text_area(
            "④ その他（任意）",
            height=110,
            key='profit_report_other',
            placeholder="例：採用状況、設備修繕、関係機関連携、利用者動向など、会議で共有したいことがあれば記入してください。",
        )
        love = st.selectbox(
            "送信前の確認：EMIFULLは大好きですか？",
            ["", "はい", "もちろんです", "いいえ"],
            key='profit_report_love',
        )

        same_reports = db.get_profit_reports(
            target_month=report_month,
            facility_name=facility_name,
            created_by_user=current_email if not _is_admin else None,
        )
        same_reporter = [
            r for r in same_reports
            if (r.get('reporter_name') or '').strip() == reporter_name.strip()
        ]
        duplicate_mode = "追加入力"
        if same_reporter:
            duplicate_mode = st.radio(
                "同じ月・施設・報告者の提出があります",
                ["追加入力", "上書き"],
                horizontal=True,
                key='profit_report_duplicate_mode',
            )

        submitted = st.form_submit_button("報告書を提出", type='primary')

    if submitted:
        if not love:
            st.warning("EMIFULL愛の確認がまだです。ここだけは外せません。")
        elif love == "いいえ":
            st.markdown("<div class='emifull-sad'>悲しいです。</div>", unsafe_allow_html=True)
        elif not reporter_name.strip():
            st.warning("氏名を入力してください。")
        elif not previous_review.strip() or not issue_review.strip() or not next_actions.strip():
            st.warning("前月の振り返り、現在の課題、次月以降の対策を入力してください。")
        else:
            report_id = same_reporter[0]['id'] if same_reporter and duplicate_mode == "上書き" else None
            saved_id = db.save_profit_report(
                fiscal_year_type=fy_mode_label,
                target_month=report_month,
                facility_id=selected_fac['id'],
                facility_name=facility_name,
                reporter_role=reporter_role,
                reporter_name=reporter_name.strip(),
                previous_review=previous_review.strip(),
                issue_review=issue_review.strip(),
                next_actions=next_actions.strip(),
                other_notes=other_notes.strip(),
                created_by_user=current_email,
                report_id=report_id,
            )
            st.success("一緒に人生咲かそう")

    st.markdown("---")
    st.markdown("##### 自分の提出内容")
    my_reports = db.get_profit_reports(created_by_user=current_email)
    if my_reports:
        my_df = pd.DataFrame(my_reports)[[
            'target_month', 'facility_name', 'reporter_role', 'reporter_name',
            'previous_review', 'issue_review', 'next_actions', 'other_notes', 'ai_summary', 'created_at'
        ]].rename(columns={
            'target_month': '対象月',
            'facility_name': '施設名',
            'reporter_role': '役職',
            'reporter_name': '報告者',
            'previous_review': '前月の振り返り',
            'issue_review': '現在の課題',
            'next_actions': '次月以降の対策',
            'other_notes': 'その他',
            'ai_summary': 'AI要約',
            'created_at': '提出日時',
        })
        st.dataframe(my_df, hide_index=True, width='stretch')
        st.markdown("###### 原文を開いて確認")
        for r in my_reports:
            label = f"{r['target_month']} / {r['facility_name']} / {r['reporter_role']} {r['reporter_name']}"
            with st.expander(label, expanded=False):
                st.markdown("**前月の振り返り**")
                st.text_area(
                    "前月の振り返り 原文",
                    value=r.get('previous_review') or '',
                    height=140,
                    label_visibility='collapsed',
                    disabled=True,
                    key=f"my_previous_raw_{r['id']}",
                )
                st.markdown("**現在の課題**")
                st.text_area(
                    "現在の課題 原文",
                    value=r.get('issue_review') or '',
                    height=140,
                    label_visibility='collapsed',
                    disabled=True,
                    key=f"my_issue_raw_{r['id']}",
                )
                st.markdown("**次月以降の対策**")
                st.text_area(
                    "次月以降の対策 原文",
                    value=r.get('next_actions') or '',
                    height=180,
                    label_visibility='collapsed',
                    disabled=True,
                    key=f"my_actions_raw_{r['id']}",
                )
                if r.get('other_notes'):
                    st.markdown("**その他**")
                    st.text_area(
                        "その他 原文",
                        value=r.get('other_notes') or '',
                        height=130,
                        label_visibility='collapsed',
                        disabled=True,
                        key=f"my_other_raw_{r['id']}",
                    )
    else:
        st.info("まだ提出はありません。")

    editable_reports = [
        r for r in db.get_profit_reports()
        if _is_admin or r.get('created_by_user') == current_email
    ]
    if editable_reports:
        with st.expander("提出済み報告書を修正", expanded=False):
            edit_options = {
                f"{r['target_month']} / {r['facility_name']} / {r['reporter_name']} / #{r['id']}": r
                for r in editable_reports
            }
            edit_key = st.selectbox("修正する報告書", list(edit_options.keys()), key='profit_report_edit_select')
            edit_r = edit_options[edit_key]
            with st.form("profit_report_edit_form"):
                ec1, ec2 = st.columns(2)
                with ec1:
                    edit_month_idx = existing_yms.index(edit_r['target_month']) if edit_r['target_month'] in existing_yms else month_idx
                    emonth = st.selectbox(
                        "対象月",
                        existing_yms,
                        index=edit_month_idx,
                        key='profit_report_edit_month',
                    )
                with ec2:
                    facility_labels = [f['name'] for f in REPORT_FACILITIES]
                    edit_fac_idx = facility_labels.index(edit_r['facility_name']) if edit_r['facility_name'] in facility_labels else 0
                    efacility_name = st.selectbox(
                        "施設選択",
                        facility_labels,
                        index=edit_fac_idx,
                        key='profit_report_edit_facility',
                    )
                edit_fac = REPORT_FACILITY_BY_NAME[efacility_name]
                erole = st.selectbox(
                    "役職", REPORTER_ROLES,
                    index=REPORTER_ROLES.index(edit_r['reporter_role']) if edit_r['reporter_role'] in REPORTER_ROLES else 0,
                    key='profit_report_edit_role',
                )
                ename = st.text_input("氏名", value=edit_r['reporter_name'], key='profit_report_edit_name')
                eprevious = st.text_area("① 前月の振り返り", value=edit_r.get('previous_review') or '', height=120, key='profit_report_edit_previous')
                eissue = st.text_area("② 現在の課題", value=edit_r['issue_review'], height=140, key='profit_report_edit_issue')
                eactions = st.text_area("③ 次月以降の対策", value=edit_r['next_actions'], height=140, key='profit_report_edit_actions')
                eother = st.text_area("④ その他", value=edit_r.get('other_notes') or '', height=100, key='profit_report_edit_other')
                delete_confirm = st.checkbox("この報告書を削除する内容を確認しました", key='profit_report_delete_confirm')
                bc1, bc2 = st.columns([1, 1])
                with bc1:
                    edit_submit = st.form_submit_button("修正を保存", type='primary')
                with bc2:
                    delete_submit = st.form_submit_button("この報告書を削除")
            if edit_submit:
                old_month = edit_r['target_month']
                old_facility = edit_r['facility_name']
                db.save_profit_report(
                    fiscal_year_type=edit_r.get('fiscal_year_type') or fy_mode_label,
                    target_month=emonth,
                    facility_id=edit_fac['id'],
                    facility_name=efacility_name,
                    reporter_role=erole,
                    reporter_name=ename.strip(),
                    previous_review=eprevious.strip(),
                    issue_review=eissue.strip(),
                    next_actions=eactions.strip(),
                    other_notes=eother.strip(),
                    ai_summary=None,
                    created_by_user=edit_r.get('created_by_user'),
                    report_id=edit_r['id'],
                )
                st.success("修正を保存しました。")
                st.rerun()
            if delete_submit:
                if not delete_confirm:
                    st.warning("削除する場合は確認チェックを入れてください。")
                else:
                    old_month = edit_r['target_month']
                    old_facility = edit_r['facility_name']
                    deleted = db.delete_profit_report(edit_r['id'])
                    if deleted:
                        st.success("報告書を削除しました。")
                        st.rerun()
                    else:
                        st.warning("削除対象が見つかりませんでした。")

    if _is_admin:
        st.markdown("---")
        st.markdown("##### 管理者確認")
        ac1, ac2 = st.columns(2)
        with ac1:
            admin_month = st.selectbox(
                "対象月", existing_yms, index=month_idx, key='profit_report_admin_month'
            )
        with ac2:
            admin_facility = st.selectbox(
                "施設フィルタ", ["全施設"] + [f['name'] for f in REPORT_FACILITIES],
                key='profit_report_admin_facility',
            )

        admin_facilities = REPORT_FACILITIES if admin_facility == "全施設" else [REPORT_FACILITY_BY_NAME[admin_facility]]
        admin_rows = _report_preview_rows(admin_month, admin_facilities)
        status_df = pd.DataFrame(admin_rows)
        st.dataframe(status_df, hide_index=True, width='stretch')

        missing = [r['施設名'] for r in admin_rows if r['提出状況'] == '未提出']
        with st.container(border=True):
            st.markdown("###### 未提出施設一覧")
            if missing:
                st.write("、".join(missing))
            else:
                st.success("すべて提出済みです。")

        csv_bytes = _reports_csv_bytes(admin_rows)
        st.download_button(
            "📥 対象月の報告一覧をCSVダウンロード",
            data=csv_bytes,
            file_name=f"損益報告一覧_{admin_month}.csv",
            mime="text/csv",
            key="profit_reports_csv_dl",
        )

        if st.button("AI要約を作成・更新", type='secondary', key='profit_report_ai_build'):
            grouped = _reports_for_month_by_facility(admin_month)
            updated = 0
            for fac in admin_facilities:
                reports = grouped.get(fac['name'], [])
                if not reports:
                    continue
                summary = build_facility_report_summary(
                    admin_month, fac['name'], fac['excel_names'], reports, force_ai=True
                )
                for r in reports:
                    db.update_profit_report_ai_summary(r['id'], summary)
                    updated += 1
            st.success(f"AI要約を更新しました（{updated}件）。")


# =============================================================
# TAB 5: 業績会議 (単月 / 千円切捨 / 実績・前月比・前年比)
# =============================================================
# 上の共通定数セクションで定義済み: PERSONNEL_ACCOUNTS / REPORT_STRUCTURE /
#                                  EXCLUDED_ROWS / NPO_REPORT_STRUCTURE / subs_by_excel


if _is_admin and tab_meeting is not None:
  with tab_meeting:
    st.markdown("##### 業績会議 — 単月 / 千円(切捨)")

    sel_meeting_ym = st.selectbox(
        "対象月", existing_yms, index=0, key='pl_meeting_ym',
    )

    prev_ym = _ym_minus(sel_meeting_ym, 1)
    yoy_ym = _ym_minus(sel_meeting_ym, 12)
    has_prev = prev_ym in existing_yms
    has_yoy = yoy_ym in existing_yms

    st.caption(
        f"対象: **{sel_meeting_ym}** ／ 前月: {prev_ym} "
        f"({'有' if has_prev else '無'}) ／ 前年同月: {yoy_ym} "
        f"({'有' if has_yoy else '無'}) ／ 単位: 千円(切捨)"
    )

    # 一括取得して month-単位に分配
    yms_to_fetch = [sel_meeting_ym]
    if has_prev: yms_to_fetch.append(prev_ym)
    if has_yoy: yms_to_fetch.append(yoy_ym)
    all_e_for_meeting = db.fetch_pl_entries(year_months=yms_to_fetch)
    by_ym = {ym: [e for e in all_e_for_meeting if e['year_month'] == ym] for ym in yms_to_fetch}

    # excel_name → subunit_id 解決マップ
    subs_by_excel = {s['excel_name']: s['id'] for s in db.list_pl_subunits()}

    def _metrics(entries_list, excel_names):
        """指定 entries 配列から、与えられた excel_names に該当する subunit のみで4指標を集計。"""
        sub_ids = {subs_by_excel[n] for n in excel_names if n in subs_by_excel}
        ent = [e for e in entries_list if e['subunit_id'] in sub_ids]
        rev = sum(e['amount'] for e in ent if e['category'] == 'revenue_total')
        sga = sum(e['amount'] for e in ent if e['category'] == 'sga_total')
        pers = sum(e['amount'] for e in ent
                   if e['category'] == 'sga' and e['account_name'] in PERSONNEL_ACCOUNTS)
        exp = sga - pers
        op = sum(e['amount'] for e in ent if e['category'] == 'op_profit')
        return {'rev': rev, 'exp': exp, 'pers': pers, 'op': op}

    def _add(a, b):
        return {k: a[k] + b[k] for k in a}

    def _sub_or_none(c, p):
        if p is None:
            return None
        return {k: c[k] - p[k] for k in c}

    def _to_k(v):
        """円 → 千円(切捨)。Noneは保持。"""
        if v is None:
            return None
        if v >= 0:
            return v // 1000
        # 負数は数学的には -(abs(v) // 1000) で「切捨」(0方向ではなく負の無限へ)。
        # ただし会議資料の慣習として 0 方向への切り捨て(=truncate)が一般的なので
        # int(v / 1000) を使う。
        return int(v / 1000)

    # 4指標 × 3列 = 12列 + 部門名列
    metric_cols = [
        ('rev',  '売上'),
        ('exp',  '販管費(経費)'),
        ('pers', '販管費(人件費)'),
        ('op',   '営業利益'),
    ]

    # 数値表示: 0/Noneは空白、それ以外はカンマ区切り
    def _fmt(v):
        if v is None or (isinstance(v, float) and pd.isna(v)):
            return '－'
        if str(v) == 'None':
            return '－'
        if isinstance(v, float):
            return f"{v:.1f}%"
        try:
            n = int(v)
        except (TypeError, ValueError):
            return str(v)
        if n == 0:
            return ''
        return f"{n:,}"

    BG_REGION = '#fef9c3'
    BG_GRAND = '#fde68a'
    BG_EXCL = '#f1f5f9'
    RED = '#dc2626'
    GREEN = '#15803d'

    def render_meeting_section(section_title, structure, excluded_rows, dl_suffix, highlight_top_n=5):
        """1セクション(EMIFULL or NPO)の業績会議表を描画"""
        rows = []
        grand_curr = {'rev': 0, 'exp': 0, 'pers': 0, 'op': 0}
        grand_prev = {'rev': 0, 'exp': 0, 'pers': 0, 'op': 0}
        grand_yoy = {'rev': 0, 'exp': 0, 'pers': 0, 'op': 0}

        for region_label, subrows in structure:
            region_curr = {'rev': 0, 'exp': 0, 'pers': 0, 'op': 0}
            region_prev = {'rev': 0, 'exp': 0, 'pers': 0, 'op': 0}
            region_yoy = {'rev': 0, 'exp': 0, 'pers': 0, 'op': 0}
            for row_label, excel_names in subrows:
                m_c = _metrics(by_ym[sel_meeting_ym], excel_names)
                m_p = _metrics(by_ym[prev_ym], excel_names) if has_prev else None
                m_y = _metrics(by_ym[yoy_ym], excel_names) if has_yoy else None
                rows.append(('detail', row_label, m_c, m_p, m_y))
                region_curr = _add(region_curr, m_c)
                grand_curr = _add(grand_curr, m_c)
                if m_p:
                    region_prev = _add(region_prev, m_p)
                    grand_prev = _add(grand_prev, m_p)
                if m_y:
                    region_yoy = _add(region_yoy, m_y)
                    grand_yoy = _add(grand_yoy, m_y)
            rows.append((
                'region_total', f'合計({region_label})',
                region_curr,
                region_prev if has_prev else None,
                region_yoy if has_yoy else None,
            ))

        if excluded_rows:
            rows.append((
                'grand_above', '総計(上記)',
                grand_curr,
                grand_prev if has_prev else None,
                grand_yoy if has_yoy else None,
            ))
            excluded_curr = {'rev': 0, 'exp': 0, 'pers': 0, 'op': 0}
            excluded_prev = {'rev': 0, 'exp': 0, 'pers': 0, 'op': 0}
            excluded_yoy = {'rev': 0, 'exp': 0, 'pers': 0, 'op': 0}
            for row_label, excel_names in excluded_rows:
                m_c = _metrics(by_ym[sel_meeting_ym], excel_names)
                m_p = _metrics(by_ym[prev_ym], excel_names) if has_prev else None
                m_y = _metrics(by_ym[yoy_ym], excel_names) if has_yoy else None
                rows.append(('detail_excl', row_label, m_c, m_p, m_y))
                excluded_curr = _add(excluded_curr, m_c)
                if m_p: excluded_prev = _add(excluded_prev, m_p)
                if m_y: excluded_yoy = _add(excluded_yoy, m_y)
            total_curr = _add(grand_curr, excluded_curr)
            total_prev = _add(grand_prev, excluded_prev) if has_prev else None
            total_yoy = _add(grand_yoy, excluded_yoy) if has_yoy else None
            rows.append(('grand_full', '総計(全体)', total_curr, total_prev, total_yoy))
        else:
            rows.append((
                'grand_full', '総計',
                grand_curr,
                grand_prev if has_prev else None,
                grand_yoy if has_yoy else None,
            ))

        columns = [('', '部門')]
        for _, cat_label in metric_cols:
            columns.extend([(cat_label, '実績'), (cat_label, '前月比'), (cat_label, '前年比')])
            if cat_label == '営業利益':
                columns.append((cat_label, '営業利益率'))

        data_rows = []
        local_row_types = []
        for row_type, label, m_c, m_p, m_y in rows:
            diff_p = _sub_or_none(m_c, m_p)
            diff_y = _sub_or_none(m_c, m_y)
            values = [label]
            for k, _ in metric_cols:
                values.append(_to_k(m_c[k]))
                values.append(_to_k(diff_p[k]) if diff_p is not None else None)
                if diff_y is None or (k != 'rev' and m_y and m_y['rev'] == 0):
                    values.append('－')
                else:
                    values.append(_to_k(diff_y[k]))
                if k == 'op':
                    values.append((m_c[k] / m_c['rev'] * 100) if m_c['rev'] else '－')
            data_rows.append(values)
            local_row_types.append(row_type)

        df_section = pd.DataFrame(data_rows, columns=pd.MultiIndex.from_tuples(columns))
        df_section = df_section.replace({None: '－', 'None': '－'})

        def _style_row(row):
            rt = local_row_types[row.name]
            styles = [''] * len(row)
            if rt == 'region_total':
                for i in range(len(styles)):
                    styles[i] = f'background-color:{BG_REGION}; font-weight:700'
            elif rt in ('grand_above', 'grand_full'):
                for i in range(len(styles)):
                    styles[i] = f'background-color:{BG_GRAND}; font-weight:700'
            elif rt == 'detail_excl':
                for i in range(len(styles)):
                    styles[i] = f'background-color:{BG_EXCL}'
            for i, col in enumerate(list(row.index)):
                if col == ('', '部門'):
                    continue
                v = row[col]
                if v is None or pd.isna(v):
                    continue
                cat_label, kind = col
                try:
                    num = float(v) if kind == '営業利益率' else int(v)
                except (TypeError, ValueError):
                    continue
                color = ''
                if num < 0:
                    color = RED
                elif cat_label == '営業利益' and kind in ('実績', '営業利益率') and num > 0:
                    color = GREEN
                if color:
                    styles[i] += f'; color:{color}; font-weight:700'
            return styles

        fmt_dict = {col: _fmt for col in df_section.columns if col != ('', '部門')}
        styled_section = (
            df_section.style.apply(_style_row, axis=1).format(fmt_dict, na_rep='－')
        )
        row_count = len(df_section)
        height = min(38 * (row_count + 2) + 5, 900)

        st.markdown(f"##### {section_title}")
        st.dataframe(styled_section, hide_index=True, width='stretch', height=height)
        report_facilities = _report_facilities_for_structure(structure, excluded_rows)
        report_preview_rows = _report_preview_rows(sel_meeting_ym, report_facilities)
        prev_report_preview_rows = (
            _report_preview_rows(prev_ym, report_facilities) if has_prev else []
        )

        def _build_meeting_excel():
            import openpyxl as _opx
            from openpyxl.styles import Font as _Font, PatternFill as _Fill, Alignment as _Align, Border as _Border, Side as _Side
            from openpyxl.utils import get_column_letter

            wb = _opx.Workbook()
            ws = wb.active
            ws.title = f"業績会議_{dl_suffix}"

            excel_columns = list(df_section.columns)
            op_rate_idx = excel_columns.index(('営業利益', '営業利益率'))
            usage_columns = [
                ('1日平均利用者数', '当月'),
                ('1日平均利用者数', '前月'),
                ('1日平均利用者数', '前年'),
            ]
            excel_columns = excel_columns[:op_rate_idx + 1] + usage_columns + excel_columns[op_rate_idx + 1:]

            thin = _Side(border_style='thin', color='D9DEE8')
            thick = _Side(border_style='medium', color='94A3B8')
            border = _Border(left=thin, right=thin, top=thin, bottom=thin)
            group_end_border = _Border(left=thin, right=thick, top=thin, bottom=thin)
            header_fill = _Fill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')
            region_fill = _Fill(start_color='FEF9C3', end_color='FEF9C3', fill_type='solid')
            grand_fill = _Fill(start_color='FDE68A', end_color='FDE68A', fill_type='solid')
            excl_fill = _Fill(start_color='F1F5F9', end_color='F1F5F9', fill_type='solid')
            excel_font_name = 'BIZ UDPゴシック'
            base_font = _Font(name=excel_font_name, size=10)
            red_font = _Font(name=excel_font_name, size=10, color='DC2626', bold=True)
            green_font = _Font(name=excel_font_name, size=10, color='15803D', bold=True)
            bold_font = _Font(name=excel_font_name, size=10, bold=True)
            header_font = _Font(name=excel_font_name, size=10, bold=True, color='64748B')

            ws.cell(row=1, column=1, value=section_title).font = _Font(name=excel_font_name, bold=True, size=14)
            ws.cell(row=2, column=1, value=f"対象: {sel_meeting_ym} ／ 単位: 千円(切捨)").font = _Font(name=excel_font_name, color='64748B')

            header_row_1 = 4
            header_row_2 = 5
            for col_idx, (top, sub) in enumerate(excel_columns, start=1):
                c1 = ws.cell(row=header_row_1, column=col_idx, value=top)
                c2 = ws.cell(row=header_row_2, column=col_idx, value=sub)
                for c in (c1, c2):
                    c.fill = header_fill
                    c.font = header_font
                    c.alignment = _Align(horizontal='center', vertical='center')
                    c.border = border

            merge_start = 1
            current_top = excel_columns[0][0]
            for idx, (top, _) in enumerate(excel_columns + [(None, None)], start=1):
                if top != current_top:
                    if current_top != '' and idx - merge_start > 1:
                        ws.merge_cells(start_row=header_row_1, start_column=merge_start,
                                       end_row=header_row_1, end_column=idx - 1)
                    merge_start = idx
                    current_top = top
            ws.merge_cells(start_row=header_row_1, start_column=1, end_row=header_row_2, end_column=1)
            ws.cell(row=header_row_1, column=1, value='部門')

            for row_idx, (_, row) in enumerate(df_section.iterrows(), start=6):
                rt = local_row_types[row_idx - 6]
                if rt == 'region_total':
                    row_fill = region_fill
                elif rt in ('grand_above', 'grand_full'):
                    row_fill = grand_fill
                elif rt == 'detail_excl':
                    row_fill = excl_fill
                else:
                    row_fill = None

                for col_idx, col in enumerate(excel_columns, start=1):
                    value = '' if col in usage_columns else row[col]
                    if value is None or (isinstance(value, float) and pd.isna(value)) or str(value) == 'None':
                        value = '－'
                    if col == ('営業利益', '営業利益率') and isinstance(value, (int, float)):
                        value = value / 100
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.font = base_font
                    cell.border = border
                    cell.alignment = _Align(horizontal='right' if col_idx > 1 else 'left', vertical='center')
                    if row_fill:
                        cell.fill = row_fill
                        cell.font = bold_font
                    if isinstance(value, (int, float)):
                        cell.number_format = '0.0%' if col == ('営業利益', '営業利益率') else '#,##0'
                        if value < 0:
                            cell.font = red_font
                        elif col[0] == '営業利益' and col[1] in ('実績', '営業利益率') and value > 0:
                            cell.font = green_font
                    if col in usage_columns:
                        cell.fill = row_fill if row_fill else _Fill(fill_type=None)

            group_end_cols = []
            for idx, col in enumerate(excel_columns, start=1):
                next_top = excel_columns[idx][0] if idx < len(excel_columns) else None
                if col[0] != next_top:
                    group_end_cols.append(idx)
            for row in ws.iter_rows(min_row=header_row_1, max_row=ws.max_row):
                for cell in row:
                    if cell.column in group_end_cols:
                        cell.border = group_end_border

            widths = {1: 22}
            for idx, col in enumerate(excel_columns[1:], start=2):
                widths[idx] = 9
            for col_idx, width in widths.items():
                ws.column_dimensions[get_column_letter(col_idx)].width = width
            ws.freeze_panes = 'B6'

            section_fill = _Fill(start_color='E2F0D9', end_color='E2F0D9', fill_type='solid')
            section_font = _Font(name=excel_font_name, size=10, bold=True)
            wrap = _Align(vertical='top', wrap_text=True)

            def _write_report_block(title, values):
                start = ws.max_row + 2
                ws.cell(row=start, column=1, value=title)
                ws.merge_cells(start_row=start, start_column=1, end_row=start, end_column=8)
                for col_idx in range(1, 9):
                    cell = ws.cell(row=start, column=col_idx)
                    cell.fill = section_fill
                    cell.font = section_font
                    cell.border = border
                    cell.alignment = _Align(vertical='center')
                row_idx = start + 1
                for facility, value in values:
                    ws.cell(row=row_idx, column=1, value=facility)
                    ws.cell(row=row_idx, column=2, value=value or '')
                    ws.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=8)
                    for col_idx in range(1, 9):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        cell.font = base_font
                        cell.border = border
                        cell.alignment = wrap
                    text_len = len(str(value or ''))
                    ws.row_dimensions[row_idx].height = min(180, max(36, 24 + (text_len // 45) * 16))
                    row_idx += 1

            preview_by_fac = {r['施設名']: r for r in report_preview_rows}
            prev_by_fac = {r['施設名']: r for r in prev_report_preview_rows}
            facility_names = [f['name'] for f in report_facilities]

            _write_report_block("2: 前月の振り返り", [
                (name, preview_by_fac.get(name, {}).get('前月の振り返り要約') or '未提出')
                for name in facility_names
            ])
            _write_report_block("3: 現在の課題", [
                (name, preview_by_fac.get(name, {}).get('現在の課題要約') or '未提出')
                for name in facility_names
            ])
            _write_report_block("4: 次月以降の対策", [
                (name, preview_by_fac.get(name, {}).get('対策要約') or '未提出')
                for name in facility_names
            ])
            _write_report_block("5: その他", [
                (name, preview_by_fac.get(name, {}).get('その他要約') or ('未提出' if preview_by_fac.get(name, {}).get('提出状況') == '未提出' else ''))
                for name in facility_names
            ])
            _write_report_block("6: 前回TODO", [
                (name, " / ".join(filter(None, [
                    prev_by_fac.get(name, {}).get('対策'),
                    prev_by_fac.get(name, {}).get('その他'),
                ])))
                for name in facility_names
            ])

            ws.column_dimensions['A'].width = 22
            for col_letter in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
                ws.column_dimensions[col_letter].width = 18

            output = io.BytesIO()
            wb.save(output)
            return output.getvalue()

        excel_data = _build_meeting_excel()
        st.download_button(
            f"📥 この表をExcelダウンロード ({dl_suffix})",
            data=excel_data,
            file_name=f"業績会議_{dl_suffix}_{sel_meeting_ym}.xlsx",
            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            key=f'pl_meeting_dl_{dl_suffix}',
        )

        st.markdown("##### 報告書プレビュー")
        with st.container(border=True):
            _render_summary_cards("① 前月の振り返り", report_preview_rows, '前月の振り返り要約')
        with st.container(border=True):
            _render_summary_cards("② 現在の課題", report_preview_rows, '現在の課題要約')
        with st.container(border=True):
            _render_summary_cards("③ 次月以降の対策", report_preview_rows, '対策要約')
        with st.container(border=True):
            _render_summary_cards("④ その他", report_preview_rows, 'その他要約')
        preview_df = pd.DataFrame(report_preview_rows)[[
            '対象月', '施設名', '提出状況', 'AI要約',
            '前月の振り返り要約', '現在の課題要約', '対策要約', 'その他要約',
            '報告者', '提出日時'
        ]]
        st.dataframe(
            preview_df,
            hide_index=True,
            width='stretch',
            height=min(38 * (len(report_preview_rows) + 1), 520),
        )

        # ===== 分析ハイライト =====
        section_excel_names = []
        for _, subrows in structure:
            for _, names in subrows:
                section_excel_names.extend(names)
        for _, names in excluded_rows:
            section_excel_names.extend(names)
        section_sub_ids = {subs_by_excel[n] for n in section_excel_names if n in subs_by_excel}

        def _section_breakdown(entries_list, category):
            """このセクションのサブ部門に絞って科目別合計を返す。"""
            agg = defaultdict(int)
            for e in entries_list:
                if e['subunit_id'] in section_sub_ids and e['category'] == category:
                    agg[e['account_name']] += e['amount']
            return agg

        def _section_breakdown_per_subunit(entries_list, category):
            """{(subunit_id, account_name): amount} - サブ部門×科目別の値"""
            agg = defaultdict(int)
            for e in entries_list:
                if e['subunit_id'] in section_sub_ids and e['category'] == category:
                    agg[(e['subunit_id'], e['account_name'])] += e['amount']
            return agg

        def _calc_diffs(curr_dict, base_dict):
            """{科目: (当月, 比較月, 差額, 差率%)} を 差額 abs 降順で返す。"""
            diffs = []
            for acc in set(curr_dict) | set(base_dict):
                c = curr_dict.get(acc, 0)
                b = base_dict.get(acc, 0)
                d = c - b
                if d == 0:
                    continue
                pct = (d / abs(b) * 100) if b else None
                diffs.append((acc, c, b, d, pct))
            diffs.sort(key=lambda x: -abs(x[3]))
            return diffs

        # サブ部門 ID → 表示名(報告書ラベル)。
        # 報告書では excel_name が複数の row_label にまとまる場合(例 シェア天理1.2 = 旧名+新名)があるので、
        # report_label を優先表示する。
        excel_to_report_label = {}
        for _, subrows in structure:
            for row_label, names in subrows:
                for n in names:
                    excel_to_report_label[n] = row_label
        for row_label, names in excluded_rows:
            for n in names:
                excel_to_report_label[n] = row_label
        sub_id_to_excel = {s['id']: s['excel_name'] for s in db.list_pl_subunits()}
        sub_id_to_label = {
            sid: excel_to_report_label.get(en, en)
            for sid, en in sub_id_to_excel.items()
        }

        def _top_contributors(curr_per, base_per, account_name, top_n=2):
            """指定科目のサブ部門別差額を集計し、(報告書ラベル, 差額) を 絶対値降順で返す。
            同じ報告書ラベル（例 シェア天理1.2 = 旧名+新名）はマージ。"""
            by_label = defaultdict(int)
            for (sid, acc), amt in curr_per.items():
                if acc != account_name:
                    continue
                label = sub_id_to_label.get(sid)
                if label:
                    by_label[label] += amt
            for (sid, acc), amt in base_per.items():
                if acc != account_name:
                    continue
                label = sub_id_to_label.get(sid)
                if label:
                    by_label[label] -= amt
            diffs = [(lab, d) for lab, d in by_label.items() if d != 0]
            diffs.sort(key=lambda x: -abs(x[1]))
            return diffs[:top_n]

        def _render_movers(diffs, base_per_dict, top_n=5):
            if not diffs:
                st.caption("差額0以外の変動なし")
                return
            for acc, c, b, d, pct in diffs[:top_n]:
                arrow = '🔺' if d > 0 else '🔻'
                color = '#dc2626' if d > 0 else '#15803d'   # 増=赤(注意), 減=緑
                d_k = int(d / 1000)
                pct_str = f" <span style='color:#64748b;'>({pct:+.1f}%)</span>" if pct is not None else ""
                line = (
                    f"{arrow} <b>{acc}</b>　"
                    f"<span style='color:{color}; font-weight:700'>{d_k:+,}千円</span>"
                    f"{pct_str}"
                )
                st.markdown(f"- {line}", unsafe_allow_html=True)
                # 主に動いた施設(サブ部門)を併記
                contribs = _top_contributors(sga_curr_per, base_per_dict, acc, top_n=3)
                if contribs:
                    parts = []
                    for label, dd in contribs:
                        cc = '#dc2626' if dd > 0 else '#15803d'
                        parts.append(
                            f"<b>{label}</b> "
                            f"<span style='color:{cc}; font-weight:700'>{int(dd/1000):+,}千円</span>"
                        )
                    st.markdown(
                        "　　<span style='color:#64748b; font-size:12px;'>↳ 主な動き: "
                        + "　/　".join(parts)
                        + "</span>",
                        unsafe_allow_html=True,
                    )

        sga_curr = _section_breakdown(by_ym[sel_meeting_ym], 'sga')
        sga_prev = _section_breakdown(by_ym[prev_ym], 'sga') if has_prev else {}
        sga_yoy = _section_breakdown(by_ym[yoy_ym], 'sga') if has_yoy else {}
        sga_curr_per = _section_breakdown_per_subunit(by_ym[sel_meeting_ym], 'sga')
        sga_prev_per = _section_breakdown_per_subunit(by_ym[prev_ym], 'sga') if has_prev else {}
        sga_yoy_per = _section_breakdown_per_subunit(by_ym[yoy_ym], 'sga') if has_yoy else {}

        with st.expander("🔍 分析ハイライト — 注意・確認したい変動 (クリックで表示)", expanded=False):
            # 全体サマリ
            def _agg(entries_list, cat):
                return sum(e['amount'] for e in entries_list
                           if e['subunit_id'] in section_sub_ids and e['category'] == cat)

            rev_c = _agg(by_ym[sel_meeting_ym], 'revenue_total')
            op_c = _agg(by_ym[sel_meeting_ym], 'op_profit')
            rev_p = _agg(by_ym[prev_ym], 'revenue_total') if has_prev else None
            op_p = _agg(by_ym[prev_ym], 'op_profit') if has_prev else None
            rev_y = _agg(by_ym[yoy_ym], 'revenue_total') if has_yoy else None
            op_y = _agg(by_ym[yoy_ym], 'op_profit') if has_yoy else None

            def _summary_line(label, c, p, y):
                parts = [f"**{label}** 当月 {int(c/1000):,}千円"]
                if p is not None:
                    d = c - p
                    color = '#15803d' if d >= 0 else '#dc2626'
                    parts.append(
                        f"前月比 <span style='color:{color}; font-weight:700'>{int(d/1000):+,}千円</span>"
                    )
                if y is not None:
                    d = c - y
                    color = '#15803d' if d >= 0 else '#dc2626'
                    parts.append(
                        f"前年比 <span style='color:{color}; font-weight:700'>{int(d/1000):+,}千円</span>"
                    )
                return "　／　".join(parts)

            st.markdown(_summary_line("売上", rev_c, rev_p, rev_y), unsafe_allow_html=True)
            # 営業利益は逆色(赤=損, 緑=益)で意味が同じなので普通に
            def _summary_op(c, p, y):
                parts = [f"**営業利益** 当月 {int(c/1000):+,}千円"]
                if p is not None:
                    d = c - p
                    color = '#15803d' if d >= 0 else '#dc2626'
                    parts.append(f"前月比 <span style='color:{color}; font-weight:700'>{int(d/1000):+,}千円</span>")
                if y is not None:
                    d = c - y
                    color = '#15803d' if d >= 0 else '#dc2626'
                    parts.append(f"前年比 <span style='color:{color}; font-weight:700'>{int(d/1000):+,}千円</span>")
                return "　／　".join(parts)
            st.markdown(_summary_op(op_c, op_p, op_y), unsafe_allow_html=True)

            st.markdown("---")

            cc1, cc2 = st.columns(2)
            with cc1:
                if has_prev:
                    st.markdown(f"**📈 vs 前月 ({prev_ym})** — 販管費 大きい順 TOP{highlight_top_n}")
                    _render_movers(
                        _calc_diffs(sga_curr, sga_prev),
                        sga_prev_per,
                        top_n=highlight_top_n,
                    )
                else:
                    st.info("前月データなし")
            with cc2:
                if has_yoy:
                    st.markdown(f"**📅 vs 前年同月 ({yoy_ym})** — 販管費 大きい順 TOP{highlight_top_n}")
                    _render_movers(
                        _calc_diffs(sga_curr, sga_yoy),
                        sga_yoy_per,
                        top_n=highlight_top_n,
                    )
                else:
                    st.info("前年同月データなし")

            st.caption(
                "※ 🔺増加=赤(コスト増は要因確認)、🔻減少=緑。"
                "売上連動・人員増などで自然増になる場合もあるため、"
                "数字だけで判断せず背景を確認してください。"
            )

    # ===== 描画: 医療法人社団EMIFULL =====
    render_meeting_section(
        "医療法人社団EMIFULL",
        REPORT_STRUCTURE,
        EXCLUDED_ROWS,
        "EMIFULL",
        highlight_top_n=10,
    )

    st.markdown("---")

    # ===== 描画: NPO法人EMIFULL (旧: のじぎく高砂) =====
    render_meeting_section(
        "NPO法人EMIFULL（旧: のじぎく高砂、2026/1改称）",
        NPO_REPORT_STRUCTURE,
        [],   # 別枠なし
        "NPO",
        highlight_top_n=5,
    )

    st.caption(
        "※ 人件費 = 管理者給 + 指導員給 + 法定福利費 + 退職給付費用 + 賞与 + 事務員給。"
        " 経費 = 販管費合計 − 人件費。前月比/前年比 = 当月実績 − 比較対象月実績。"
        " 単位: 千円(切捨)。0は空白表示。"
    )

    # =========================================================
    # 全施設 損益計算書 (当月/前月比/前年比 横並び)
    # =========================================================
    st.markdown("---")
    st.markdown("### 📊 全施設 損益計算書 — 当月／前月比／前年比 横並び")

    # 列グループ定義: (列ラベル, [excel_names])
    emifull_cols = []
    for _region, subrows in REPORT_STRUCTURE:
        for label, names in subrows:
            emifull_cols.append((label, list(names)))
    for label, names in EXCLUDED_ROWS:
        emifull_cols.append((label, list(names)))
    emifull_all_excel = [n for _, names in emifull_cols for n in names]

    npo_cols = []
    for _region, subrows in NPO_REPORT_STRUCTURE:
        for label, names in subrows:
            npo_cols.append((label, list(names)))
    npo_all_excel = [n for _, names in npo_cols for n in names]

    columns_def = (
        emifull_cols
        + [('医療法人 合計', emifull_all_excel)]
        + npo_cols
        + [('NPO 合計', npo_all_excel)]
        + [('総合計', emifull_all_excel + npo_all_excel)]
    )

    # 全データを取得
    all_accounts = db.list_pl_accounts()
    yms_for_pl = [sel_meeting_ym]
    if has_prev: yms_for_pl.append(prev_ym)
    if has_yoy: yms_for_pl.append(yoy_ym)
    pl_entries_all = db.fetch_pl_entries(year_months=yms_for_pl)

    by_key = defaultdict(int)  # (ym, sub_id, acc_id) -> amount
    for e in pl_entries_all:
        by_key[(e['year_month'], e['subunit_id'], e['account_id'])] += e['amount']

    # MultiIndex columns
    multi_columns = [('科目', '')]
    for label, _ in columns_def:
        multi_columns.append((label, '当月'))
        multi_columns.append((label, '前月比'))
        multi_columns.append((label, '前年比'))

    # 行定義: (label, is_total, [acc_id, ...] = 集計対象、is_subtotal_synthetic)
    # 通常行は account_id 1個。小計行(人件費小計/経費小計)は複数 acc_id を合算。
    sga_personnel_accs = [a for a in all_accounts
                          if a['category'] == 'sga' and a['name'] in PERSONNEL_ACCOUNTS]
    sga_other_accs = [a for a in all_accounts
                      if a['category'] == 'sga' and a['name'] not in PERSONNEL_ACCOUNTS]

    # 行リストを再構築:
    #  非SGA → 元の display_order
    #  SGA を 人件費グループ→人件費小計→経費グループ→経費小計 の順に並べ替え
    row_defs = []   # list of dict: {label, is_total, account_ids, source_acc(optional)}
    sga_inserted = False
    for acc in all_accounts:
        if acc['category'] == 'sga':
            # SGA は別途処理済みなのでスキップ（最初の sga 出現タイミングで一括挿入）
            if not sga_inserted:
                # ヘッダ風の見出し
                row_defs.append({
                    'label': '──── 人件費 ────', 'is_total': False, 'is_section_header': True,
                    'account_ids': [],
                })
                for a in sga_personnel_accs:
                    row_defs.append({
                        'label': a['name'], 'is_total': False, 'is_section_header': False,
                        'account_ids': [a['id']], 'source_acc': a,
                    })
                row_defs.append({
                    'label': '小計: 人件費', 'is_total': True, 'is_section_header': False,
                    'account_ids': [a['id'] for a in sga_personnel_accs], 'is_synthetic': True,
                })
                row_defs.append({
                    'label': '──── 経費 ────', 'is_total': False, 'is_section_header': True,
                    'account_ids': [],
                })
                for a in sga_other_accs:
                    row_defs.append({
                        'label': a['name'], 'is_total': False, 'is_section_header': False,
                        'account_ids': [a['id']], 'source_acc': a,
                    })
                row_defs.append({
                    'label': '小計: その他経費', 'is_total': True, 'is_section_header': False,
                    'account_ids': [a['id'] for a in sga_other_accs], 'is_synthetic': True,
                })
                sga_inserted = True
            continue
        # 非SGAはそのまま
        row_defs.append({
            'label': acc['name'], 'is_total': bool(acc['is_total']),
            'is_section_header': False,
            'account_ids': [acc['id']], 'source_acc': acc,
        })

    # データ行構築
    pl_rows = []
    pl_row_meta = []   # parallel meta for styling
    for rd in row_defs:
        row_values = [rd['label']]
        any_nonzero = False
        if rd.get('is_section_header'):
            # セクション見出し行: 値は出さない（空文字でフォーマット時の '—' を回避）
            for _ in columns_def:
                row_values.extend(['', '', ''])
            pl_rows.append(row_values)
            pl_row_meta.append(rd)
            continue
        for label, excel_names in columns_def:
            sub_ids_set = {subs_by_excel[n] for n in excel_names if n in subs_by_excel}
            curr_v = 0
            prev_v = 0
            yoy_v = 0
            for aid in rd['account_ids']:
                for sid in sub_ids_set:
                    curr_v += by_key.get((sel_meeting_ym, sid, aid), 0)
                    if has_prev:
                        prev_v += by_key.get((prev_ym, sid, aid), 0)
                    if has_yoy:
                        yoy_v += by_key.get((yoy_ym, sid, aid), 0)
            row_values.append(_to_k(curr_v))
            row_values.append(_to_k(curr_v - prev_v) if has_prev else '－')
            row_values.append(_to_k(curr_v - yoy_v) if has_yoy else '－')
            if curr_v != 0 or (has_prev and prev_v != 0) or (has_yoy and yoy_v != 0):
                any_nonzero = True
        if any_nonzero or rd['is_total']:
            pl_rows.append(row_values)
            pl_row_meta.append(rd)

    df_full_pl = pd.DataFrame(pl_rows, columns=pd.MultiIndex.from_tuples(multi_columns))
    df_full_pl = df_full_pl.replace({None: '－', 'None': '－'})

    # スタイル
    BG_TOTAL = '#fde68a'    # 集計行(売上高計, 販管費計, 営業損益, 経常損益 等)
    BG_KEYTOTAL = '#fef3c7' # 売上高計, 販管費計
    RED2 = '#dc2626'

    def _style_full_pl(row):
        idx = row.name
        rd = pl_row_meta[idx]
        styles = [''] * len(row)
        if rd.get('is_section_header'):
            for i in range(len(styles)):
                styles[i] = 'background-color:#dbeafe; font-weight:700; color:#1e3a8a'
        elif rd['is_total']:
            for i in range(len(styles)):
                styles[i] = f'background-color:{BG_TOTAL}; font-weight:700'
        # 負の値は赤
        for i, col in enumerate(row.index):
            if col == ('科目', ''):
                continue
            v = row[col]
            if v is None or pd.isna(v):
                continue
            try:
                num = int(v)
            except (TypeError, ValueError):
                continue
            if num < 0:
                styles[i] += f'; color:{RED2}; font-weight:700'
        return styles

    fmt_full = {col: _fmt for col in df_full_pl.columns if col != ('科目', '')}
    styled_full_pl = (
        df_full_pl.style.apply(_style_full_pl, axis=1).format(fmt_full, na_rep='－')
    )

    # 高さは行数 × 35 (上限1000)
    h = min(35 * (len(df_full_pl) + 3) + 10, 1000)
    st.dataframe(styled_full_pl, hide_index=True, width='stretch', height=h)

    # ===== Excelダウンロード(色付き) =====
    import io
    import openpyxl as _opx
    from openpyxl.styles import Font as _Font, PatternFill as _Fill, Alignment as _Align, Border as _Border, Side as _Side

    def _build_excel():
        wb = _opx.Workbook()
        ws = wb.active
        ws.title = f"損益計算書_{sel_meeting_ym}"

        thin = _Side(border_style='thin', color='BBBBBB')
        thick = _Side(border_style='medium', color='000000')   # 施設区切り用の太線
        border = _Border(left=thin, right=thin, top=thin, bottom=thin)
        border_facility_end = _Border(left=thin, right=thick, top=thin, bottom=thin)
        border_kamoku = _Border(left=thin, right=thick, top=thin, bottom=thin)  # 科目列右(太)

        def _border_for(col_idx):
            """col_idx: Excel列番号(1-based)。
            - 1: 科目列 → 右が太
            - 2以降: facility ブロックの最終列(当月/前月比/前年比 の3列目)で右が太
            """
            if col_idx == 1:
                return border_kamoku
            # 各 facility は 3 列。col=2,3,4 が1施設目。col=4 が右端。
            if (col_idx - 2) % 3 == 2:
                return border_facility_end
            return border
        header_fill = _Fill(start_color='1E40AF', end_color='1E40AF', fill_type='solid')
        header_font = _Font(color='FFFFFF', bold=True, size=10)
        sub_header_fill = _Fill(start_color='3B82F6', end_color='3B82F6', fill_type='solid')
        sub_header_font = _Font(color='FFFFFF', bold=True, size=9)
        total_fill = _Fill(start_color='FDE68A', end_color='FDE68A', fill_type='solid')
        bold_font = _Font(bold=True)
        red_font = _Font(color='DC2626', bold=True)
        red_total_font = _Font(color='DC2626', bold=True)

        # Title row
        ws.cell(row=1, column=1, value=f"全施設 損益計算書  対象月: {sel_meeting_ym}  単位: 千円(切捨)").font = _Font(bold=True, size=12)
        ws.row_dimensions[1].height = 18

        # 行 2: 親列ヘッダ (列ラベル: '科目', 各 facility, 合計)
        # 行 3: 子列ヘッダ ('当月', '前月比', '前年比')
        ws.cell(row=2, column=1, value='科目').fill = header_fill
        ws.cell(row=2, column=1).font = header_font
        ws.cell(row=2, column=1).alignment = _Align(horizontal='center', vertical='center')
        ws.cell(row=2, column=1).border = _border_for(1)
        ws.merge_cells(start_row=2, start_column=1, end_row=3, end_column=1)
        # マージ後の row=3 col=1 は MergedCell (read-only) なので value/fill 設定はスキップ。

        col_idx = 2
        for label, _ in columns_def:
            # 親
            top_cell = ws.cell(row=2, column=col_idx, value=label)
            top_cell.fill = header_fill
            top_cell.font = header_font
            top_cell.alignment = _Align(horizontal='center', vertical='center')
            ws.merge_cells(start_row=2, start_column=col_idx, end_row=2, end_column=col_idx + 2)
            # 子 + 罫線
            for j, kind in enumerate(['当月', '前月比', '前年比']):
                c_col = col_idx + j
                c = ws.cell(row=3, column=c_col, value=kind)
                c.fill = sub_header_fill
                c.font = sub_header_font
                c.alignment = _Align(horizontal='center')
                c.border = _border_for(c_col)
            # 親セルの右側を太線にしたいので、親の最終列(col_idx+2) に太線右
            ws.cell(row=2, column=col_idx + 2).border = _border_for(col_idx + 2)
            col_idx += 3

        # データ行
        section_header_fill = _Fill(start_color='DBEAFE', end_color='DBEAFE', fill_type='solid')
        section_header_font = _Font(color='1E3A8A', bold=True)
        for r_idx, (rd, values) in enumerate(zip(pl_row_meta, pl_rows)):
            xl_row = r_idx + 4
            is_section = rd.get('is_section_header', False)
            is_total = rd['is_total']

            # 科目
            name_cell = ws.cell(row=xl_row, column=1, value=values[0])
            if is_section:
                name_cell.fill = section_header_fill
                name_cell.font = section_header_font
            elif is_total:
                name_cell.fill = total_fill
                name_cell.font = bold_font
            name_cell.border = _border_for(1)
            name_cell.alignment = _Align(horizontal='left', vertical='center')

            # 数値
            for c_off, v in enumerate(values[1:], start=2):
                cell = ws.cell(row=xl_row, column=c_off)
                # 値設定: None / 空文字 はセル空、それ以外は int 化
                if v is None or v == '':
                    cell.value = None
                else:
                    try:
                        cell.value = int(v)
                        cell.number_format = '#,##0;-#,##0;""'
                    except (TypeError, ValueError):
                        cell.value = v   # 念のためフォールバック
                cell.border = _border_for(c_off)
                cell.alignment = _Align(horizontal='right')
                if is_section:
                    cell.fill = section_header_fill
                    cell.font = section_header_font
                elif is_total:
                    cell.fill = total_fill
                if not is_section and isinstance(v, (int, float)) and v < 0:
                    cell.font = red_total_font if is_total else red_font
                elif is_total:
                    cell.font = bold_font

        # 列幅
        ws.column_dimensions['A'].width = 22
        for c in range(2, 2 + len(columns_def) * 3):
            from openpyxl.utils import get_column_letter
            ws.column_dimensions[get_column_letter(c)].width = 11
        ws.freeze_panes = 'B4'  # 科目列とヘッダ2行を固定

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.getvalue()

    excel_bytes = _build_excel()
    st.download_button(
        "📥 Excelダウンロード(色付き・列ヘッダ固定)",
        data=excel_bytes,
        file_name=f"全施設損益計算書_{sel_meeting_ym}.xlsx",
        mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        key='pl_full_xlsx_dl',
    )

    # CSV ダウンロード(プレーン、色なし)
    csv_full = df_full_pl.to_csv(index=False, encoding='utf-8-sig')
    st.download_button(
        "📥 CSVダウンロード(プレーン)",
        data=csv_full.encode('utf-8-sig'),
        file_name=f"全施設損益計算書_{sel_meeting_ym}.csv",
        mime='text/csv',
        key='pl_full_csv_dl',
    )

    st.caption(
        "※ 単位: 千円(0方向への切捨)。集計行(売上高計・販管費計・営業損益・経常損益 等)は黄色背景。"
        " 全列で値0の科目行は自動的に非表示。"
        " 色付きで見たいときはExcel版をご利用ください。"
    )

# =============================================================
# TAB 5: 販管費推移 (勘定科目別 / 1年推移 + 仕訳帳内訳)
# =============================================================
with tab_sga_trend:
    st.markdown("##### 販管費推移 — 勘定科目別 / 1年推移")
    st.caption(
        "会計年度の基準は上の「法人決算・現場評価」の選択に連動します。"
        "金額推移は損益データ、内訳は仕訳帳データを参考に表示します。"
    )

    sga_accounts = [
        a for a in db.list_pl_accounts(category='sga', include_total=False)
        if not a.get('is_total')
    ]
    if not sga_accounts:
        st.info("販管費の勘定科目がありません。")
    else:
        fy_list_trend = db.list_pl_fiscal_years(start_month=fy_start_month)
        if not fy_list_trend:
            st.info("損益データがありません。")
        else:
            c1, c2 = st.columns([1, 2])
            with c1:
                trend_fy = st.selectbox(
                    "対象年度",
                    fy_list_trend,
                    format_func=fy_label,
                    key='pl_sga_trend_fy',
                )
            with c2:
                account_names = [a['name'] for a in sga_accounts]
                default_account = '水道光熱費'
                default_idx = account_names.index(default_account) if default_account in account_names else 0
                trend_account = st.selectbox(
                    "勘定科目",
                    account_names,
                    index=default_idx,
                    key='pl_sga_trend_account',
                )

            trend_months = db.pl_fiscal_year_months(trend_fy, start_month=fy_start_month)
            fetch_months = [m for m in trend_months if m in existing_yms]
            trend_entries = fetch_pl(fetch_months, sel_group_id, sel_subunit_id) if fetch_months else []
            amount_by_month = {m: 0 for m in trend_months}
            for e in trend_entries:
                if e['category'] == 'sga' and e['account_name'] == trend_account:
                    amount_by_month[e['year_month']] += e['amount']

            total_amount = sum(amount_by_month.values())
            avg_amount = int(total_amount / 12) if trend_months else 0
            max_ym = max(amount_by_month, key=lambda m: amount_by_month[m]) if trend_months else None
            min_ym = min(amount_by_month, key=lambda m: amount_by_month[m]) if trend_months else None

            kpi1, kpi2, kpi3 = st.columns(3)
            kpi1.metric("年間合計", f"{total_amount:,.0f} 円")
            kpi2.metric("月平均", f"{avg_amount:,.0f} 円")
            if max_ym and min_ym:
                kpi3.metric("最大月 / 最小月", f"{max_ym} / {min_ym}")

            journal_subunit_ids = None
            if sel_subunit_id:
                journal_subunit_ids = [sel_subunit_id]
            elif sel_group_id:
                journal_subunit_ids = [s['id'] for s in all_subs if s['group_id'] == sel_group_id]
            has_journal = bool(db.list_journal_imports(limit=1))

            left, right = st.columns([1.45, 1])
            with left:
                chart_df = pd.DataFrame({
                    '月': trend_months,
                    '金額': [amount_by_month[m] for m in trend_months],
                }).set_index('月')
                st.line_chart(chart_df, height=340)

                monthly_rows = []
                for ym in trend_months:
                    details = db.journal_summary_for_account(
                        trend_account, ym, subunit_ids=journal_subunit_ids, top_n=3,
                    ) if has_journal else []
                    top_text = " / ".join(
                        f"{d['vendor']} {int(d['total_amount']):,}円"
                        for d in details
                    )
                    monthly_rows.append({
                        '月': ym,
                        '金額': amount_by_month[ym],
                        '主な中身（仕訳帳）': top_text or '－',
                    })
                df_monthly = pd.DataFrame(monthly_rows)
                st.dataframe(
                    df_monthly.style.format({'金額': '{:,.0f}'}),
                    hide_index=True,
                    width='stretch',
                    height=min(36 * (len(df_monthly) + 1), 520),
                )

            with right:
                st.markdown("##### 仕訳帳から見た中身")
                if not has_journal:
                    st.info("仕訳帳CSVを取り込むと、取引先・品目・備考ごとの中身が表示されます。")
                else:
                    period_items = db.journal_summary_for_period(
                        trend_account,
                        trend_months,
                        subunit_ids=journal_subunit_ids,
                        top_n=10,
                    )
                    if not period_items:
                        st.info("この科目・期間・対象範囲に該当する仕訳帳明細がありません。")
                    else:
                        detail_rows = []
                        for item in period_items:
                            detail_rows.append({
                                '取引先・品目': item.get('label') or item.get('vendor') or '－',
                                '金額': item.get('total_amount') or 0,
                                '件数': item.get('cnt') or 0,
                                '備考・品目': item.get('memos') or '－',
                            })
                        df_detail = pd.DataFrame(detail_rows)
                        st.dataframe(
                            df_detail.style.format({'金額': '{:,.0f}'}),
                            hide_index=True,
                            width='stretch',
                            height=420,
                        )

                st.markdown("##### 見方")
                st.markdown(
                    "- 折れ線は選択した勘定科目の月別金額です。\n"
                    "- 右側は仕訳帳から集計した取引先・備考の上位です。\n"
                    "- 支払手数料、水道光熱費など、動きが大きい科目の中身確認に使えます。"
                )

# =============================================================
# 削除（管理者用 / フッタ）
# =============================================================
if _is_admin:
    st.markdown("---")
    with st.expander("🗑️ データ削除（再取込用）"):
        st.warning("選択月のデータを削除します。再取込前のクリーンアップ用。")
        c1, c2 = st.columns([1, 1])
        with c1:
            del_ym = st.selectbox(
                "削除する月",
                ["（選択してください）"] + existing_yms,
                key='pl_del_ym',
            )
        with c2:
            del_grp_options = {"全グループ": None}
            for g in groups_all:
                del_grp_options[f"{g['code']}: {g['name']}"] = g['id']
            del_grp_label = st.selectbox(
                "対象グループ", list(del_grp_options.keys()), key='pl_del_grp',
            )
            del_grp_id = del_grp_options[del_grp_label]

        if del_ym != "（選択してください）":
            cnt = db.count_pl_entries(year_month=del_ym, group_id=del_grp_id)
            st.write(f"対象エントリ数: **{cnt}** 件")
            if cnt > 0 and st.button("⚠️ 削除を実行", type='secondary', key='pl_del_run'):
                n = db.delete_pl_entries(year_month=del_ym, group_id=del_grp_id)
                st.success(f"{n} 件削除しました。")
                st.rerun()

auth.render_sidebar_user_box()
